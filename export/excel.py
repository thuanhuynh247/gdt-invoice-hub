"""Excel generation for invoice export."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

import openpyxl
try:
    openpyxl.LXML = False
    import openpyxl.xml.functions
    from et_xmlfile import xmlfile
    openpyxl.xml.functions.xmlfile = xmlfile
    import openpyxl.worksheet._writer
    openpyxl.worksheet._writer.xmlfile = xmlfile
except Exception:
    pass
from openpyxl import Workbook




from export.formatter import (
    HEADER_FILL,
    WARNING_FILL,
    HIGH_RISK_FILL,
    MED_RISK_FILL,
    OK_FILL,
    auto_adjust_column_widths,
    format_header_row,
)


def generate_excel_workbook(invoices: list[dict]) -> bytes:
    """Build an Excel workbook containing summary invoice data."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"

    headers = [
        "ID",
        "Date",
        "Amount",
        "Status",
        "Issuer",
        "Description",
        "Cancellation Date",
        "Cancellation Reason",
    ]
    worksheet.append(headers)
    format_header_row(worksheet)

    for invoice in invoices:
        row = [
            invoice["id"],
            _format_display_date(invoice["date"]),
            invoice["amount"],
            invoice["status"],
            invoice["issuer"],
            invoice["description"],
            _format_display_date(invoice.get("cancellation_date")),
            invoice.get("cancellation_reason"),
        ]
        worksheet.append(row)
        if invoice.get("is_cancelled"):
            for cell in worksheet[worksheet.max_row]:
                cell.fill = WARNING_FILL

    for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0 "VND"'

    auto_adjust_column_widths(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def generate_local_excel_workbook(invoices: list[dict]) -> bytes:
    """Build a professionally formatted Excel workbook for locally audited invoices featuring an AI compliance dashboard."""

    workbook = Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: AI COMPLIANCE SUMMARY DASHBOARD
    # ----------------------------------------------------
    dash_sheet = workbook.active
    dash_sheet.title = "Tổng Quan AI Compliance"
    dash_sheet.views.sheetView[0].showGridLines = True

    # Title Banner
    dash_sheet.append(["BÁO CÁO TỔNG QUAN TUÂN THỦ THUẾ & AI AUDIT"])
    dash_sheet.merge_cells("A1:D1")
    title_cell = dash_sheet["A1"]
    title_cell.font = openpyxl.styles.Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    dash_sheet.row_dimensions[1].height = 40

    # Sub-title
    dash_sheet.append(["Xuất bản tự động bởi Gemma-4 Tax Intelligence Engine", "", "", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    dash_sheet.merge_cells("A2:C2")
    dash_sheet["A2"].font = openpyxl.styles.Font(italic=True, color="555555")
    dash_sheet["D2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    dash_sheet.append([])  # Blank row

    # Calculate statistics
    total_count = len(invoices)
    total_healthy = sum(1 for inv in invoices if inv.get("is_valid", True) and len(inv.get("ai_warnings", [])) == 0)
    total_flagged = total_count - total_healthy
    avg_t_score = sum(inv.get("t_score", 100) for inv in invoices) / total_count if total_count > 0 else 100.0

    # Count warning types
    warning_counts = {
        "personal_purchase": 0,
        "price_anomaly": 0,
        "invoice_timing": 0,
        "cash_payment_risk": 0,
        "tax_rate_mismatch": 0,
        "suspicious_transaction": 0
    }
    for inv in invoices:
        for w in inv.get("ai_warnings", []):
            w_type = w.get("warning_type")
            if w_type in warning_counts:
                warning_counts[w_type] += 1

    # KPI Tables
    dash_sheet.append(["CHỈ SỐ ĐO LƯỜNG CHÍNH", "", "", ""])
    dash_sheet.merge_cells("A4:D4")
    dash_sheet["A4"].font = openpyxl.styles.Font(bold=True, size=12, color="1F4E78")
    dash_sheet.append([])

    kpi_headers = ["Chỉ Số Tuân Thủ", "Giá Trị", "Đánh Giá Trạng Thái", "Mô Tả"]
    dash_sheet.append(kpi_headers)
    format_header_row(dash_sheet)

    dash_sheet.append(["Tổng số hóa đơn đã kiểm toán", total_count, "Hoàn tất", "Số lượng hóa đơn đã chạy qua AI Auditor"])
    dash_sheet.append(["Hóa đơn an toàn (Compliant)", total_healthy, f"{total_healthy/total_count*100:.1f}%" if total_count > 0 else "100%", "Không phát hiện lỗi hoặc rủi ro thuế"])
    dash_sheet.append(["Hóa đơn phát hiện rủi ro", total_flagged, f"{total_flagged/total_count*100:.1f}%" if total_count > 0 else "0%", "Cần kế toán rà soát/loại trừ chi phí"])
    dash_sheet.append(["Điểm tín nhiệm trung bình (T-Score)", f"{avg_t_score:.1f}", "A++ (Xuất sắc)" if avg_t_score >= 95 else ("A (Tốt)" if avg_t_score >= 85 else "B (Trung bình)"), "Thang điểm uy tín thuế doanh nghiệp (0 - 100)"])

    # Highlight KPI Status Fills
    for row_idx in range(7, 11):
        dash_sheet[f"A{row_idx}"].font = openpyxl.styles.Font(bold=True)
        dash_sheet[f"B{row_idx}"].font = openpyxl.styles.Font(bold=True)
        dash_sheet[f"C{row_idx}"].font = openpyxl.styles.Font(bold=True)
        dash_sheet[f"B{row_idx}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        dash_sheet[f"C{row_idx}"].alignment = openpyxl.styles.Alignment(horizontal="center")

    dash_sheet["B8"].fill = OK_FILL
    dash_sheet["C8"].fill = OK_FILL
    if total_flagged > 0:
        dash_sheet["B9"].fill = HIGH_RISK_FILL
        dash_sheet["C9"].fill = HIGH_RISK_FILL
    else:
        dash_sheet["B9"].fill = OK_FILL
        dash_sheet["C9"].fill = OK_FILL

    if avg_t_score >= 90:
        dash_sheet["B10"].fill = OK_FILL
        dash_sheet["C10"].fill = OK_FILL
    elif avg_t_score >= 70:
        dash_sheet["B10"].fill = MED_RISK_FILL
        dash_sheet["C10"].fill = MED_RISK_FILL
    else:
        dash_sheet["B10"].fill = HIGH_RISK_FILL
        dash_sheet["C10"].fill = HIGH_RISK_FILL

    dash_sheet.append([])
    dash_sheet.append([])

    # Warning type breakdown
    dash_sheet.append(["PHÂN TÍCH RỦI RO THUẾ CHI TIẾT (GEMMA-4 TAX TAXONOMY)", "", "", ""])
    dash_sheet.merge_cells("A13:D13")
    dash_sheet["A13"].font = openpyxl.styles.Font(bold=True, size=12, color="1F4E78")
    dash_sheet.append([])

    risk_headers = ["Mã Rủi Ro AI", "Loại Cảnh Báo Thuế", "Số Lượng Phát Hiện", "Mức Độ Rủi Ro Pháp Lý"]
    dash_sheet.append(risk_headers)
    for cell in dash_sheet[dash_sheet.max_row]:
        cell.fill = HEADER_FILL
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    dash_sheet.append(["personal_purchase", "Mua sắm tiêu dùng cá nhân / Ô tô vượt 1.6 tỷ", warning_counts["personal_purchase"], "Cao (Không được khấu trừ thuế GTGT & TNDN)"])
    dash_sheet.append(["price_anomaly", "Đơn giá bất thường (>20% trung bình lịch sử)", warning_counts["price_anomaly"], "Trung bình (Rủi ro chuyển giá / Nghi vấn nâng khống)"])
    dash_sheet.append(["invoice_timing", "Hóa đơn sai thời điểm lập & ký số (NĐ 123)", warning_counts["invoice_timing"], "Trung bình (Xử phạt hành chính chậm xuất hóa đơn)"])
    dash_sheet.append(["cash_payment_risk", "Thanh toán tiền mặt cho hóa đơn >= 20 triệu", warning_counts["cash_payment_risk"], "Cao (Mất quyền khấu trừ thuế GTGT đầu vào)"])
    dash_sheet.append(["tax_rate_mismatch", "Áp sai mức thuế suất ưu đãi giảm 8% (NĐ 72)", warning_counts["tax_rate_mismatch"], "Cao (Truy thu thuế & Phạt hành vi kê khai sai)"])
    dash_sheet.append(["suspicious_transaction", "Nhà cung cấp ngừng hoạt động / Rủi ro cao", warning_counts["suspicious_transaction"], "Nghiêm trọng (Mua bán hóa đơn khống / Trốn thuế)"])

    # Color Risk Breakdown Rows
    for row_idx in range(16, 22):
        dash_sheet[f"C{row_idx}"].font = openpyxl.styles.Font(bold=True)
        dash_sheet[f"C{row_idx}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        dash_sheet[f"D{row_idx}"].font = openpyxl.styles.Font(bold=True)
        dash_sheet[f"A{row_idx}"].font = openpyxl.styles.Font(name="Courier New", size=9)
        count_val = dash_sheet[f"C{row_idx}"].value
        if count_val > 0:
            dash_sheet[f"C{row_idx}"].fill = HIGH_RISK_FILL
            severity = dash_sheet[f"D{row_idx}"].value
            if "Nghiêm trọng" in severity or "Cao" in severity:
                dash_sheet[f"D{row_idx}"].fill = HIGH_RISK_FILL
            else:
                dash_sheet[f"D{row_idx}"].fill = MED_RISK_FILL
        else:
            dash_sheet[f"C{row_idx}"].fill = OK_FILL
            dash_sheet[f"D{row_idx}"].fill = OK_FILL

    auto_adjust_column_widths(dash_sheet)

    # ----------------------------------------------------
    # SHEET 2: DETAIL AUDITED INVOICES
    # ----------------------------------------------------
    worksheet = workbook.create_sheet(title="Hóa Đơn Chi Tiết")
    worksheet.views.sheetView[0].showGridLines = True

    headers = [
        "Mã Hóa Đơn",
        "Ký Hiệu",
        "Số Hóa Đơn",
        "Ngày Lập",
        "MST Người Bán",
        "Tên Người Bán",
        "MST Người Mua",
        "Tên Người Mua",
        "Trước Thuế",
        "Tiền Thuế",
        "Tổng Cộng",
        "Hình Thức TT",
        "T-Score",
        "T-Rating",
        "Đánh Giá",
        "Chi Tiết Cảnh Báo AI (Gemma-4)",
        "Lỗi Kỹ Thuật / XML Mismatch",
    ]
    worksheet.append(headers)
    format_header_row(worksheet)

    for invoice in invoices:
        is_valid = invoice.get("is_valid", True)
        t_score = invoice.get("t_score", 100)
        t_rating = invoice.get("t_rating", "A++")
        
        # Build detailed AI warnings string
        ai_warnings = invoice.get("ai_warnings", [])
        ai_warnings_str = "; ".join([f"[{w.get('warning_type')}] {w.get('explanation')}" for w in ai_warnings]) if ai_warnings else ""
        
        # Build traditional parsing errors string
        trad_warnings_str = "; ".join(invoice.get("warnings", [])) if invoice.get("warnings") else ""

        # Overall status label
        if not is_valid or len(ai_warnings) > 0 or t_score < 80:
            audit_status = "Phát hiện rủi ro"
        else:
            audit_status = "Đạt yêu cầu"

        row = [
            invoice.get("id", ""),
            invoice.get("symbol", ""),
            invoice.get("number", ""),
            _format_display_date(invoice.get("date")),
            invoice.get("seller_mst", ""),
            invoice.get("seller_name", ""),
            invoice.get("buyer_mst", ""),
            invoice.get("buyer_name", ""),
            invoice.get("amount_before_tax", 0.0),
            invoice.get("tax_amount", 0.0),
            invoice.get("total_amount", 0.0),
            invoice.get("payment_method", ""),
            t_score,
            t_rating,
            audit_status,
            ai_warnings_str,
            trad_warnings_str,
        ]
        worksheet.append(row)

        # Multi-risk color coding rows based on Gemma-4 Audit Outcomes
        row_cells = worksheet[worksheet.max_row]
        
        # Check critical AI warnings
        critical_ai_alert = any(w.get("warning_type") in ["suspicious_transaction", "personal_purchase", "tax_rate_mismatch", "cash_payment_risk"] for w in ai_warnings)
        
        if t_score < 60 or critical_ai_alert:
            # Critical High-Risk: Soft Red
            for cell in row_cells:
                cell.fill = HIGH_RISK_FILL
        elif t_score < 90 or len(ai_warnings) > 0 or trad_warnings_str:
            # Warning / Medium-Risk: Soft Yellow
            for cell in row_cells:
                cell.fill = MED_RISK_FILL
        else:
            # Perfect compliance: Soft Green
            for cell in row_cells:
                cell.fill = OK_FILL

    # Format numeric column values (Trước Thuế, Tiền Thuế, Tổng Cộng)
    for col_idx in [9, 10, 11]:
        for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0 "VND"'

    # Format T-Score center alignment
    for row in worksheet.iter_rows(min_row=2, min_col=13, max_col=14):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            cell.font = openpyxl.styles.Font(bold=True)

    auto_adjust_column_widths(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _format_display_date(value: str | None) -> str:
    """Convert ISO date strings into DD/MM/YYYY for Excel readability."""

    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(value)

