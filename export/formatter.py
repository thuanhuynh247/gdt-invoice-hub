"""Formatting helpers for the Excel workbook."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
HIGH_RISK_FILL = PatternFill(fill_type="solid", start_color="FADBD8", end_color="FADBD8")  # Soft red/pink
MED_RISK_FILL = PatternFill(fill_type="solid", start_color="FCF3CF", end_color="FCF3CF")   # Soft yellow
OK_FILL = PatternFill(fill_type="solid", start_color="D5F5E3", end_color="D5F5E3")         # Soft green


def format_header_row(worksheet) -> None:
    """Apply consistent styling to the first row of the workbook."""

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


from openpyxl.utils import get_column_letter

def auto_adjust_column_widths(worksheet) -> None:
    """Resize columns to fit header and content length."""

    for column_cells in worksheet.columns:
        first_cell = column_cells[0]
        column_letter = get_column_letter(first_cell.column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)
