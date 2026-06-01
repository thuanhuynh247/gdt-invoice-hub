"""Invoice-facing routes for search, download and export."""

from __future__ import annotations

from io import BytesIO

from flask import Blueprint, Response, current_app, jsonify, redirect, render_template, request, send_file, session, url_for

from export.excel import generate_excel_workbook, generate_local_excel_workbook
from invoices.parser import DateValidationError, validate_date_range
from invoices.service import (
    GDTIntegrationNotReadyError,
    InvoiceQuery,
    build_invoice_lookup,
    download_invoice_xml,
    fetch_invoices,
    resolve_live_download_name,
    fetch_invoice_line_items,
    extract_partners_from_invoices,
    generate_tax_usage_report,
)

from extensions import db
from auth.decorators import roles_required

invoices_blueprint = Blueprint("invoices", __name__)

import os
import uuid
import threading
from datetime import datetime

DOWNLOAD_TASKS = {}
DOWNLOAD_TASKS_LOCK = threading.Lock()



def _ensure_logged_in():
    """Return a 401 JSON response when the session is missing."""

    if not session.get("logged_in"):
        return jsonify({"error": "Phien dang nhap da het han. Vui long dang nhap lai."}), 401
    return None


@invoices_blueprint.get("/invoices")
def invoices_page():
    """Render the invoice search screen for authenticated users."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("invoices.html")

@invoices_blueprint.get("/cashflow")
def cashflow_page():
    """Render the cashflow oracle dashboard."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("cashflow.html")


@invoices_blueprint.get("/harness")
def harness_page():
    """Render the Harness Agent Control Center."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("harness.html")


@invoices_blueprint.get("/tax-bctc")
def tax_bctc_page():
    """Render the V17 Tax and BCTC services screen for authenticated users."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("tax_bctc.html")


@invoices_blueprint.get("/api/config")
def api_config():
    """Return small frontend configuration flags."""

    return jsonify({"mock_mode": current_app.config["GDT_USE_MOCK"], "locale": "vi-VN"})


@invoices_blueprint.get("/api/invoices")
def api_invoices():
    """Return invoices in JSON for the requested date range."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        cancelled_only = request.args.get("cancelled_only", "false").lower() == "true"
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, cancelled_only, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        session["invoice_lookup"] = build_invoice_lookup(invoices) if "invoices" in locals() else {}
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"total_count": len(invoices), "invoices": invoices})


@invoices_blueprint.get("/api/sync/events")
def sse_sync_stream():
    """SSE endpoint to stream real-time sync progress to the frontend."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.sync_daemon import get_sse_stream
    from flask import Response
    
    return Response(get_sse_stream(), mimetype="text/event-stream")

@invoices_blueprint.post("/api/reconciliation/upload")
@roles_required("admin", "auditor")
def api_reconciliation_upload():
    """Upload bank statement CSV and perform automated reconciliation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file or not file.filename.endswith(".csv"):
        return jsonify({"error": "Only CSV files are supported"}), 400

    try:
        content = file.read().decode("utf-8")
        from invoices.reconciliation_service import ReconciliationEngine
        engine = ReconciliationEngine()
        
        # 1. Parse and save
        transactions = engine.process_csv(content)
        
        # 2. Run matching engine
        results = engine.run_matching()
        
        return jsonify({
            "status": "success",
            "message": "Bank reconciliation completed successfully.",
            "results": results
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reconciliation/results")
@roles_required("admin", "auditor", "viewer")
def api_reconciliation_results():
    """Get all parsed bank transactions and their matched status."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    from invoices.models import BankTransaction
    txns = BankTransaction.query.order_by(BankTransaction.transaction_date.desc()).all()
    
    return jsonify({
        "status": "success",
        "transactions": [t.to_dict() for t in txns]
    })

@invoices_blueprint.post("/api/invoices/vision-upload")
@roles_required("admin", "auditor", "viewer")
def api_vision_upload():
    """Process an uploaded image/pdf using Vision OCR and return structured invoice data."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Empty file"}), 400

    allowed_exts = [".jpg", ".jpeg", ".png", ".pdf"]
    if not any(file.filename.lower().endswith(ext) for ext in allowed_exts):
        return jsonify({"error": "Only JPG, PNG, and PDF files are supported"}), 400

    try:
        file_bytes = file.read()
        mime_type = file.mimetype or "image/jpeg"
        
        from invoices.vision_service import VisionOCRService
        vision_service = VisionOCRService()
        
        # Call OCR service
        extracted_data = vision_service.extract_invoice_data(file_bytes, file.filename, mime_type)
        
        # In a real scenario we would save this to the database here as a Draft/Pending Invoice.
        # For now, we return it to the frontend to review and save.
        return jsonify({
            "status": "success",
            "message": "OCR Extraction successful (Needs human verification)",
            "data": extracted_data
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/cancelled-invoices")

def api_cancelled_invoices():
    """Return cancelled invoices using the same date filters."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, True, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        session["invoice_lookup"] = build_invoice_lookup(invoices) if "invoices" in locals() else {}
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"total_count": len(invoices), "cancelled_invoices": invoices})


@invoices_blueprint.get("/api/invoices/<invoice_id>/download")
def api_download_invoice(invoice_id: str):
    """Download one invoice XML file."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
        xml_bytes = download_invoice_xml(invoice_id)
        invoice = (current_app.config.get("CURRENT_INVOICE_LOOKUP") or {}).get(invoice_id)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except NotImplementedError as error:
        return jsonify({"error": str(error)}), 501
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    filename = resolve_live_download_name(invoice) if invoice else f"invoice_{invoice_id}.xml"
    return Response(
        xml_bytes,
        mimetype="application/zip" if filename.endswith(".zip") else "application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@invoices_blueprint.get("/api/invoices/<invoice_id>/details")
def api_invoice_details(invoice_id: str):
    """Return the detailed line items for a specific invoice."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    session_inv = None
    try:
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
        line_items = fetch_invoice_line_items(invoice_id)
        session_inv = (current_app.config.get("CURRENT_INVOICE_LOOKUP") or {}).get(invoice_id)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except NotImplementedError as error:
        return jsonify({"error": str(error)}), 501
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    from invoices.service import get_local_invoices
    local_invoices = get_local_invoices()
    local_inv = None
    for inv in local_invoices:
        if inv["id"] == invoice_id:
            local_inv = inv
            break

    warnings = local_inv.get("warnings", []) if local_inv else []
    is_valid = local_inv.get("is_valid", True) if local_inv else True

    payment_method = ""
    if local_inv:
        payment_method = local_inv.get("payment_method", "")
    elif session_inv:
        payment_method = session_inv.get("payment_method") or session_inv.get("raw", {}).get("htttoan") or ""

    from invoices.models import Invoice
    invoice = db.session.get(Invoice, invoice_id)
    ai_warnings = []
    if invoice:
        ai_warnings = [w.to_dict() for w in invoice.ai_audit_results]

    return jsonify({
        "invoice_id": invoice_id,
        "line_items": line_items,
        "warnings": warnings,
        "is_valid": is_valid,
        "payment_method": payment_method,
        "ai_warnings": ai_warnings,
        "ai_audited": invoice.ai_audited if invoice else False,
        "signature_details": invoice.signature_details if invoice else None
    })



@invoices_blueprint.get("/api/invoices/stats")
def api_invoices_stats():
    """Return financial statistics and aggregations for the requested date range."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        
        # Check hybrid stats cache first (US-124)
        mst = session.get("tax_code")
        from_str = parsed_from.isoformat()
        to_str = parsed_to.isoformat()
        
        from invoices.stats_cache import get_cached_stats, set_cached_stats
        cached_result = get_cached_stats(mst, from_str, to_str, direction)
        if cached_result is not None:
            return jsonify(cached_result)

        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    total_spend = 0.0
    total_tax = 0.0
    active_count = 0
    cancelled_count = 0

    vendor_stats = {}
    tax_breakdown = {"0%": 0.0, "5%": 0.0, "8%": 0.0, "10%": 0.0, "khac": 0.0}

    for inv in invoices:
        amount = inv.get("amount", 0.0)
        is_cancelled = inv.get("is_cancelled", False)

        if is_cancelled:
            cancelled_count += 1
        else:
            active_count += 1
            total_spend += amount

            vendor = inv.get("issuer", "Khong ro")
            if vendor not in vendor_stats:
                vendor_stats[vendor] = {"spend": 0.0, "count": 0}
            vendor_stats[vendor]["spend"] += amount
            vendor_stats[vendor]["count"] += 1

            line_items = inv.get("line_items", [])
            for item in line_items:
                rate = str(item.get("tax_rate", "10%")).strip()
                tax_amt = item.get("tax_amount", 0.0)
                total_tax += tax_amt

                if "10" in rate:
                    tax_breakdown["10%"] += tax_amt
                elif "8" in rate:
                    tax_breakdown["8%"] += tax_amt
                elif "5" in rate:
                    tax_breakdown["5%"] += tax_amt
                elif "0" in rate:
                    tax_breakdown["0%"] += tax_amt
                else:
                    tax_breakdown["khac"] += tax_amt

    top_vendors = []
    for vendor, data in vendor_stats.items():
        top_vendors.append({"name": vendor, "spend": data["spend"], "count": data["count"]})
    top_vendors.sort(key=lambda x: x["spend"], reverse=True)
    top_vendors = top_vendors[:5]

    response_payload = {
        "total_spend": total_spend,
        "total_tax": total_tax,
        "active_count": active_count,
        "cancelled_count": cancelled_count,
        "top_vendors": top_vendors,
        "tax_breakdown": tax_breakdown,
    }
    
    # Store calculated stats in hybrid cache
    set_cached_stats(mst, from_str, to_str, direction, response_payload)

    return jsonify(response_payload)





@invoices_blueprint.get("/api/export-excel")
def api_export_excel():
    """Export invoice search results to an Excel workbook download."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        cancelled_only = request.args.get("cancelled_only", "false").lower() == "true"
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, cancelled_only, direction))
        workbook_bytes = generate_excel_workbook(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    filename = f"invoices_{parsed_from.isoformat()}_{parsed_to.isoformat()}.xlsx"
    return send_file(
        BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@invoices_blueprint.get("/api/erp/export/misa")
def api_erp_export_misa():
    """Export selected or all invoices to a MISA-compatible Excel template."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    ids_str = request.args.get("ids", "")
    from invoices.models import Invoice
    if ids_str:
        invoices = Invoice.query.filter(Invoice.id.in_(ids_str.split(","))).all()
    else:
        invoices = Invoice.query.all()

    from invoices.erp_service import generate_misa_export
    try:
        excel_bytes = generate_misa_export(invoices)
        filename = "misa_export.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất MISA: {str(e)}"}), 500


@invoices_blueprint.get("/api/erp/export/odoo")
def api_erp_export_odoo():
    """Export selected or all invoices to an Odoo-compatible CSV template."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    ids_str = request.args.get("ids", "")
    from invoices.models import Invoice
    if ids_str:
        invoices = Invoice.query.filter(Invoice.id.in_(ids_str.split(","))).all()
    else:
        invoices = Invoice.query.all()

    from invoices.erp_service import generate_odoo_export
    try:
        csv_str = generate_odoo_export(invoices)
        filename = "odoo_export.csv"
        return send_file(
            BytesIO(csv_str.encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Odoo: {str(e)}"}), 500


@invoices_blueprint.get("/api/partners")
def api_partners():
    """Extract and return corporate business partners and their statistics."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        partners = extract_partners_from_invoices(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"partners": partners})


@invoices_blueprint.get("/api/partners/<mst>/status")
def api_partner_status(mst):
    """Force an on-demand MST tax status verification."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.mst_service import check_mst_status
    try:
        result = check_mst_status(mst, force_refresh=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/reports/usage")
def api_reports_usage():
    """Aggregate and return BC26 Vietnamese tax invoice usage compliance tables."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "sold")  # Default to sold for business output tracking
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        report = generate_tax_usage_report(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"report": report})


@invoices_blueprint.get("/api/invoices/<invoice_id>/pdf-view")
def api_invoice_pdf_view(invoice_id):
    """Render a beautiful, printable official-style HTML/CSS Vietnamese electronic invoice."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    # Locate invoice
    current_app.config["CURRENT_JWT"] = session.get("jwt")
    current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
    try:
        from datetime import date
        invoice = current_app.config["CURRENT_INVOICE_LOOKUP"].get(invoice_id)
        
        if not invoice:
            invoices_purchase = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "purchase"))
            invoice = build_invoice_lookup(invoices_purchase).get(invoice_id)
            
        if not invoice:
            invoices_sold = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "sold"))
            invoice = build_invoice_lookup(invoices_sold).get(invoice_id)

        if not invoice:
            from invoices.service import get_local_invoices
            local_db = get_local_invoices()
            for item in local_db:
                if item["id"] == invoice_id:
                    invoice = item
                    break
            
        if not invoice:
            return "Khong tim thay hoa don yeu cau.", 404

        line_items = fetch_invoice_line_items(invoice_id)
    except FileNotFoundError:
        return "Khong tim thay hoa don yeu cau.", 404
    except Exception as error:
        return f"Loi he thong: {str(error)}", 500
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    # Calculate sums
    sum_before_tax = sum(item.get("amount_before_tax", 0.0) for item in line_items)
    sum_tax = sum(item.get("tax_amount", 0.0) for item in line_items)
    total_payable = sum_before_tax + sum_tax

    # Auto buyer/seller properties
    user_company = {
        "name": "CONG TY CO PHAN CONG NGHE GDT INVOICE HUB",
        "mst": "0109999999",
        "address": "Toa nha Technopark, Gia Lam, TP. Ha Noi",
        "phone": "1900 8888",
    }
    
    partner_details = {
        "Cong ty A": {"mst": "0101234567", "address": "So 10 Pho Hue, Quan Hai Ba Trung, Ha Noi"},
        "Cong ty B": {"mst": "0209876543", "address": "250 Nguyen Thi Minh Khai, Quan 3, TP. Ho Chi Minh"},
        "Cong ty C": {"mst": "0301122334", "address": "15 Le Loi, Quan Hai Chau, Da Nang"},
    }
    
    if "seller_name" in invoice:
        seller = {
            "name": invoice.get("seller_name", ""),
            "mst": invoice.get("seller_mst", ""),
            "address": invoice.get("seller_address", ""),
            "phone": invoice.get("seller_phone", ""),
        }
        buyer = {
            "name": invoice.get("buyer_name", ""),
            "mst": invoice.get("buyer_mst", ""),
            "address": invoice.get("buyer_address", ""),
        }
    else:
        issuer = invoice.get("issuer", "Doi tac khac")
        partner = partner_details.get(
            issuer,
            {
                "mst": f"0{abs(hash(issuer)) % 1000000000:09d}",
                "address": f"Khu cong nghiep Binh Duong, Tinh Binh Duong",
            }
        )
        partner["name"] = issuer

        # If it is a purchase invoice, issuer is Seller, user_company is Buyer
        if invoice.get("direction", "purchase") == "purchase":
            seller = partner
            buyer = user_company
        else:
            seller = user_company
            buyer = partner

    return render_template(
        "invoice_pdf.html",
        invoice=invoice,
        line_items=line_items,
        seller=seller,
        buyer=buyer,
        sum_before_tax=sum_before_tax,
        sum_tax=sum_tax,
        total_payable=total_payable,
    )


from flask import send_file
import io

@invoices_blueprint.post("/api/invoices/batch-download")
def api_batch_download_invoices():
    """Fetch and package all GDT invoices for a month in a ZIP archive (Asynchronous)."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    month = payload.get("month", "").strip()  # Format: YYYY-MM
    direction = payload.get("direction", "purchase").strip()
    duplicate_strategy = payload.get("duplicate_strategy", "overwrite").strip()

    if not month:
        return jsonify({"error": "Vui long chon thang can tai."}), 400

    task_id = str(uuid.uuid4())

    with DOWNLOAD_TASKS_LOCK:
        DOWNLOAD_TASKS[task_id] = {
            "task_id": task_id,
            "status": "pending",
            "progress": 0,
            "completed_count": 0,
            "total": 0,
            "imported_count": 0,
            "skipped_count": 0,
            "overwritten_count": 0,
            "failed_count": 0,
            "error": None,
            "zip_bytes": None,
            "created_at": datetime.now().isoformat()
        }

    # Fetch configuration for child thread execution
    jwt_token = session.get("jwt")
    username = session.get("username")
    encrypted_password = session.get("encrypted_password")
    invoice_lookup = session.get("invoice_lookup", {})
    gdt_use_mock = current_app.config.get("GDT_USE_MOCK", True)
    gdt_base_url = current_app.config.get("GDT_BASE_URL")
    gdt_timeout = current_app.config.get("GDT_TIMEOUT_SECONDS", 10)
    app_instance = current_app._get_current_object()

    def run_task():
        with app_instance.app_context():
            # Inject active context configs
            app_instance.config["CURRENT_JWT"] = jwt_token
            app_instance.config["CURRENT_USERNAME"] = username
            app_instance.config["CURRENT_ENCRYPTED_PASSWORD"] = encrypted_password
            app_instance.config["CURRENT_INVOICE_LOOKUP"] = invoice_lookup
            app_instance.config["GDT_USE_MOCK"] = gdt_use_mock
            app_instance.config["GDT_BASE_URL"] = gdt_base_url
            app_instance.config["GDT_TIMEOUT_SECONDS"] = gdt_timeout

            def on_progress(completed, total, status, error=None, zip_bytes=None, imported=0, skipped=0, overwritten=0, failed=0):
                with DOWNLOAD_TASKS_LOCK:
                    if task_id in DOWNLOAD_TASKS:
                        task = DOWNLOAD_TASKS[task_id]
                        task["completed_count"] = completed
                        task["total"] = total
                        task["status"] = status
                        task["error"] = error
                        task["imported_count"] = imported
                        task["skipped_count"] = skipped
                        task["overwritten_count"] = overwritten
                        task["failed_count"] = failed
                        if zip_bytes:
                            task["zip_bytes"] = zip_bytes
                        if total > 0:
                            task["progress"] = int((completed / total) * 100)

            try:
                from invoices.service import batch_download_invoices
                batch_download_invoices(month, direction, on_progress=on_progress, duplicate_strategy=duplicate_strategy)
            except Exception as e:
                on_progress(0, 0, "failed", error=str(e))
            finally:
                # Clean thread configurations
                app_instance.config["CURRENT_JWT"] = None
                app_instance.config["CURRENT_USERNAME"] = None
                app_instance.config["CURRENT_ENCRYPTED_PASSWORD"] = None
                app_instance.config["CURRENT_INVOICE_LOOKUP"] = {}


    thread = threading.Thread(target=run_task, name=f"BatchDownloadThread-{task_id}")
    thread.daemon = True
    thread.start()

    return jsonify({"task_id": task_id, "status": "pending"}), 202


@invoices_blueprint.get("/api/invoices/batch-download/status/<task_id>")
def api_batch_download_status(task_id):
    """Check progress of a batch download task."""

    with DOWNLOAD_TASKS_LOCK:
        task = DOWNLOAD_TASKS.get(task_id)
        if not task:
            return jsonify({"error": "Khong tim thay thong tin tien trinh tai."}), 404

        return jsonify({
            "task_id": task["task_id"],
            "status": task["status"],
            "progress": task["progress"],
            "completed_count": task["completed_count"],
            "total": task["total"],
            "imported_count": task.get("imported_count", 0),
            "skipped_count": task.get("skipped_count", 0),
            "overwritten_count": task.get("overwritten_count", 0),
            "failed_count": task.get("failed_count", 0),
            "error": task["error"]
        })


@invoices_blueprint.get("/api/invoices/batch-download/download/<task_id>")
def api_batch_download_retrieve(task_id):
    """Retrieve the generated ZIP file for a completed batch download task."""

    with DOWNLOAD_TASKS_LOCK:
        task = DOWNLOAD_TASKS.get(task_id)
        if not task:
            return jsonify({"error": "Khong tim thay file tai ve cho phien nay."}), 404

        if task["status"] != "completed" or not task.get("zip_bytes"):
            return jsonify({"error": f"Tien trinh chua hoan thanh. Trang thai: {task['status']}"}), 400

        zip_bytes = task["zip_bytes"]
        # Clear ZIP memory to prevent leaks
        del DOWNLOAD_TASKS[task_id]

    import io
    return send_file(
        io.BytesIO(zip_bytes),
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"GDT_Invoices_Batch.zip"
    )


@invoices_blueprint.post("/api/invoices/upload")
@roles_required("admin", "auditor")
def api_upload_invoices():
    """Import drag-and-drop XML/ZIP invoices and run smart MISA audits."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "files" not in request.files:
        return jsonify({"error": "Khong tim thay tep tin duoc tai len."}), 400

    files = request.files.getlist("files")
    if not files or files[0].filename == "":
        return jsonify({"error": "Khong co tep tin nao duoc chon."}), 400

    duplicate_strategy = request.form.get("duplicate_strategy", "overwrite").strip()

    from invoices.service import import_xml_invoice
    import zipfile

    imported_count = 0
    skipped_count = 0
    overwritten_count = 0
    errors = []

    for file in files:
        filename = file.filename
        try:
            file_bytes = file.read()
            if filename.lower().endswith(".xml"):
                res = import_xml_invoice(file_bytes, filename, duplicate_strategy=duplicate_strategy)
                status = res.get("import_status", "imported")
                if status == "skipped":
                    skipped_count += 1
                elif status == "overwritten":
                    overwritten_count += 1
                else:
                    imported_count += 1
            elif filename.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                    for zinfo in z.infolist():
                        if zinfo.filename.lower().endswith(".xml") and not zinfo.is_dir():
                            xml_content = z.read(zinfo.filename)
                            base_xml_name = os.path.basename(zinfo.filename)
                            res = import_xml_invoice(xml_content, base_xml_name, duplicate_strategy=duplicate_strategy)
                            status = res.get("import_status", "imported")
                            if status == "skipped":
                                skipped_count += 1
                            elif status == "overwritten":
                                overwritten_count += 1
                            else:
                                imported_count += 1
            else:
                errors.append(f"Tep {filename} khong dung dinh dang XML hoac ZIP.")
        except Exception as e:
            errors.append(f"Loi khi nhap tep {filename}: {str(e)}")

    return jsonify({
        "status": "success",
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "overwritten_count": overwritten_count,
        "errors": errors
    })


@invoices_blueprint.get("/api/invoices/local")
def api_get_local_invoices():
    """Retrieve all smart-audited locally stored invoices."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.service import get_local_invoices
    mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    if mst == "all" or not mst:
        mst = None
    return jsonify({"invoices": get_local_invoices(mst)})


@invoices_blueprint.get("/api/invoices/local/export-excel")
@roles_required("admin", "auditor")
def api_export_local_excel():
    """Export the local audited database to an Excel workbook download, filtered by active corporate taxpayer profile."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from invoices.service import get_local_invoices
        mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
        if mst == "all" or not mst:
            mst = None
        invoices = get_local_invoices(mst)
        workbook_bytes = generate_local_excel_workbook(invoices)
    except Exception as error:
        return jsonify({"error": str(error)}), 500

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"audited_invoices_{timestamp}.xlsx"
    return send_file(
        BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@invoices_blueprint.get("/api/invoices/local/items")
def api_search_local_items():
    """Global search across line items of locally imported invoices, filtered by active corporate taxpayer profile."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    q = request.args.get("q", "").strip()
    mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    if mst == "all" or not mst:
        mst = None
    from invoices.service import search_local_items
    return jsonify({"items": search_local_items(q, mst)})


@invoices_blueprint.delete("/api/invoices/local/clear")
@roles_required("admin", "auditor")
def api_clear_local_invoices():
    """Clear all records from local SQLite database and remove XML storage files."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice, LineItem
    from invoices.service import XML_DIR
    import shutil
    import os

    try:
        # Delete all records from Invoice and LineItem tables
        LineItem.query.delete()
        Invoice.query.delete()
        db.session.commit()

        from invoices.security_audit_service import log_security_event
        log_security_event("DELETE", "Cleared all records from local SQLite database and removed XML storage files.")

        if os.path.exists(XML_DIR):
            shutil.rmtree(XML_DIR)
            os.makedirs(XML_DIR, exist_ok=True)

        return jsonify({"status": "success", "message": "Da lam sach co so du lieu cuc bo."})

    except Exception as e:
        return jsonify({"error": f"Loi khi lam sach du lieu: {str(e)}"}), 500


@invoices_blueprint.delete("/api/invoices/local/<invoice_id>")
@roles_required("admin", "auditor")
def api_delete_local_invoice(invoice_id):
    """Delete a single local invoice and its XML file."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.service import delete_local_invoice
    success = delete_local_invoice(invoice_id)
    if not success:
        return jsonify({"error": "Không tìm thấy hóa đơn cần xóa."}), 404

    from invoices.security_audit_service import log_security_event
    log_security_event("DELETE", f"Deleted local invoice: {invoice_id}")

    return jsonify({"status": "success", "message": "Đã xóa hóa đơn thành công."})


@invoices_blueprint.patch("/api/invoices/local/<invoice_id>")
def api_adjust_local_invoice(invoice_id):
    """Adjust fields of a local invoice and update its smart auditing warnings."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}

    from invoices.service import adjust_local_invoice
    try:
        updated_invoice = adjust_local_invoice(invoice_id, payload)
        return jsonify({
            "status": "success",
            "message": "Đã điều chỉnh hóa đơn thành công.",
            "invoice": updated_invoice
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Lỗi không xác định: {str(e)}"}), 500


@invoices_blueprint.get("/api/settings")
@roles_required("admin", "auditor")
def api_get_settings():
    """Retrieve current scheduler and SMTP settings with masked passwords."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.scheduler import load_scheduler_settings
    settings = load_scheduler_settings()

    # Mask sensitive credentials
    if settings.get("smtp_pass"):
        settings["smtp_pass"] = "••••••••"
    if settings.get("gdt_password"):
        settings["gdt_password"] = "••••••••"
    if settings.get("ai_api_key"):
        settings["ai_api_key"] = "••••••••"
    if settings.get("telegram_bot_token"):
        settings["telegram_bot_token"] = "••••••••"
    if settings.get("gdrive_client_secret"):
        settings["gdrive_client_secret"] = "••••••••"
    if settings.get("gdrive_refresh_token"):
        settings["gdrive_refresh_token"] = "••••••••"
    if settings.get("onedrive_client_secret"):
        settings["onedrive_client_secret"] = "••••••••"
    # Mask sensitive credentials
    if settings.get("smtp_pass"):
        settings["smtp_pass"] = "••••••••"
    if settings.get("gdt_password"):
        settings["gdt_password"] = "••••••••"
    if settings.get("ai_api_key"):
        settings["ai_api_key"] = "••••••••"
    if settings.get("telegram_bot_token"):
        settings["telegram_bot_token"] = "••••••••"
    if settings.get("gdrive_client_secret"):
        settings["gdrive_client_secret"] = "••••••••"
    if settings.get("gdrive_refresh_token"):
        settings["gdrive_refresh_token"] = "••••••••"
    if settings.get("onedrive_client_secret"):
        settings["onedrive_client_secret"] = "••••••••"
    if settings.get("onedrive_refresh_token"):
        settings["onedrive_refresh_token"] = "••••••••"
    if settings.get("erp_auth_token"):
        settings["erp_auth_token"] = "••••••••"
    if settings.get("webhook_secret"):
        settings["webhook_secret"] = "••••••••"

    return jsonify(settings)


@invoices_blueprint.post("/api/settings")
@roles_required("admin")
def api_post_settings():
    """Save scheduler and SMTP settings."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}

    try:
        smtp_port = int(payload.get("smtp_port", 587))
        schedule_weekday = int(payload.get("schedule_weekday", 0))
        realtime_interval = int(payload.get("realtime_sync_interval", 15))
    except ValueError:
        return jsonify({"error": "Dữ liệu cấu hình không hợp lệ."}), 400

    from invoices.scheduler import save_scheduler_settings
    save_scheduler_settings({
        "smtp_host": payload.get("smtp_host", "").strip(),
        "smtp_port": smtp_port,
        "smtp_user": payload.get("smtp_user", "").strip(),
        "smtp_pass": payload.get("smtp_pass", ""),
        "smtp_use_tls": bool(payload.get("smtp_use_tls", True)),
        "recipient_email": payload.get("recipient_email", "").strip(),
        "schedule_enabled": bool(payload.get("schedule_enabled", False)),
        "schedule_interval": payload.get("schedule_interval", "daily"),
        "schedule_time": payload.get("schedule_time", "08:00").strip(),
        "schedule_weekday": schedule_weekday,
        "gdt_username": payload.get("gdt_username", "").strip(),
        "gdt_password": payload.get("gdt_password", ""),
        "ai_enabled": bool(payload.get("ai_enabled", False)),
        "ai_provider": payload.get("ai_provider", "ollama").strip(),
        "ai_ollama_endpoint": payload.get("ai_ollama_endpoint", "http://localhost:11434").strip(),
        "ai_api_key": payload.get("ai_api_key", ""),
        "ai_model_name": payload.get("ai_model_name", "gemma-4").strip(),
        "ai_system_prompt": payload.get("ai_system_prompt", "").strip(),
        "telegram_enabled": bool(payload.get("telegram_enabled", False)),
        "telegram_bot_token": payload.get("telegram_bot_token", ""),
        "telegram_chat_id": payload.get("telegram_chat_id", "").strip(),
        "audit_agent_enabled": bool(payload.get("audit_agent_enabled", False)),
        "audit_agent_schedule_time": payload.get("audit_agent_schedule_time", "23:00").strip(),
        "gdrive_enabled": bool(payload.get("gdrive_enabled", False)),
        "gdrive_client_id": payload.get("gdrive_client_id", "").strip(),
        "gdrive_client_secret": payload.get("gdrive_client_secret", ""),
        "gdrive_refresh_token": payload.get("gdrive_refresh_token", ""),
        "gdrive_folder_id": payload.get("gdrive_folder_id", "").strip(),
        "onedrive_enabled": bool(payload.get("onedrive_enabled", False)),
        "onedrive_client_id": payload.get("onedrive_client_id", "").strip(),
        "onedrive_client_secret": payload.get("onedrive_client_secret", ""),
        "onedrive_refresh_token": payload.get("onedrive_refresh_token", ""),
        "onedrive_folder_path": payload.get("onedrive_folder_path", "HoaDon_DienTu").strip(),
        "erp_enabled": bool(payload.get("erp_enabled", False)),
        "erp_type": payload.get("erp_type", "none").strip(),
        "erp_api_url": payload.get("erp_api_url", "").strip(),
        "erp_auth_token": payload.get("erp_auth_token", ""),
        "realtime_sync_enabled": bool(payload.get("realtime_sync_enabled", False)),
        "realtime_sync_interval": realtime_interval,
        "webhook_enabled": bool(payload.get("webhook_enabled", False)),
        "webhook_url": payload.get("webhook_url", "").strip(),
        "webhook_secret": payload.get("webhook_secret", ""),
        "signature_filter_enabled": bool(payload.get("signature_filter_enabled", True)),
        "blacklist_filter_enabled": bool(payload.get("blacklist_filter_enabled", True))
    })

    from invoices.security_audit_service import log_security_event
    log_security_event("UPDATE", "Updated application settings (SMTP, scheduler, AI, cloud sync, ERP, and webhooks).")

    return jsonify({"status": "success", "message": "Đã lưu thiết lập thành công."})


@invoices_blueprint.post("/api/invoices/realtime/trigger")
@roles_required("admin", "auditor")
def api_trigger_realtime_sync():
    """Manually trigger background real-time sync immediately."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.scheduler import load_scheduler_settings
    settings = load_scheduler_settings()

    # Trigger async thread so we return immediately and keep the UI responsive!
    from flask import current_app
    from invoices.scheduler import _scheduler_thread

    if _scheduler_thread and _scheduler_thread.is_alive():
        # Spin up a daemon thread to run the sync
        import threading
        def worker(app):
            with app.app_context():
                try:
                    _scheduler_thread.execute_realtime_sync(settings)
                except Exception as ex:
                    app.logger.error(f"Manual real-time sync failed: {ex}")

        t = threading.Thread(target=worker, args=(current_app._get_current_object(),), daemon=True)
        t.start()
        return jsonify({"status": "success", "message": "Đã kích hoạt đồng bộ hóa thời gian thực chạy ngầm."})
    else:
        return jsonify({"error": "Không thể kết nối với dịch vụ background scheduler."}), 500


@invoices_blueprint.get("/api/invoices/realtime/stream")
def api_realtime_stream():
    """SSE streaming endpoint to push downloaded invoice events in real-time."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import queue
    from invoices.scheduler import REALTIME_CLIENT_QUEUES, REALTIME_QUEUES_LOCK

    q = queue.Queue(maxsize=100)
    with REALTIME_QUEUES_LOCK:
        REALTIME_CLIENT_QUEUES.append(q)

    def event_generator():
        try:
            # Send initial keepalive
            yield f"data: {json.dumps({'event': 'connected'})}\n\n"
            while True:
                try:
                    # Wait for an event with a 20-second timeout for heartbeat/keepalive
                    event_data = q.get(timeout=20)
                    yield f"data: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                except queue.Empty:
                    # Send keepalive heartbeat to prevent connection timeout
                    yield "data: {\"event\": \"keepalive\"}\n\n"
        finally:
            with REALTIME_QUEUES_LOCK:
                if q in REALTIME_CLIENT_QUEUES:
                    REALTIME_CLIENT_QUEUES.remove(q)

    from flask import Response
    return Response(event_generator(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "Transfer-Encoding": "chunked",
        "Connection": "keep-alive"
    })


@invoices_blueprint.get("/api/blacklist")
@roles_required("admin", "auditor")
def api_list_blacklist():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import BlacklistedMST
    blacklist = BlacklistedMST.query.all()
    return jsonify([item.to_dict() for item in blacklist])


@invoices_blueprint.post("/api/blacklist")
@roles_required("admin", "auditor")
def api_add_blacklist():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    payload = request.json or {}
    mst = payload.get("mst", "").strip()
    reason = payload.get("reason", "").strip()
    if not mst:
        return jsonify({"error": "Mã số thuế không được để trống"}), 400
    
    from invoices.models import BlacklistedMST
    from extensions import db
    import datetime
    
    existing = db.session.get(BlacklistedMST, mst)
    if existing:
        existing.reason = reason
        existing.blacklisted_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    else:
        item = BlacklistedMST(
            mst=mst,
            reason=reason,
            blacklisted_at=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(item)
    db.session.commit()
    return jsonify({"status": "success", "message": "Đã thêm mã số thuế vào danh sách đen."})


@invoices_blueprint.delete("/api/blacklist/<mst>")
@roles_required("admin", "auditor")
def api_delete_blacklist(mst):
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import BlacklistedMST
    from extensions import db
    
    item = db.session.get(BlacklistedMST, mst)
    if not item:
        return jsonify({"error": "Không tìm thấy mã số thuế trong danh sách đen."}), 404
        
    db.session.delete(item)
    db.session.commit()
    return jsonify({"status": "success", "message": "Đã xóa mã số thuế khỏi danh sách đen."})


@invoices_blueprint.post("/api/settings/test-email")
@roles_required("admin")
def api_test_email():
    """Trigger a manual test email with the provided SMTP parameters."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}

    from invoices.scheduler import load_scheduler_settings, SchedulerThread
    current_settings = load_scheduler_settings()

    smtp_host = payload.get("smtp_host", "").strip()
    try:
        smtp_port = int(payload.get("smtp_port", 587))
    except ValueError:
        return jsonify({"error": "Cổng SMTP không hợp lệ."}), 400

    smtp_user = payload.get("smtp_user", "").strip()
    smtp_pass = payload.get("smtp_pass", "").strip()
    smtp_use_tls = payload.get("smtp_use_tls", True)
    recipient = payload.get("recipient_email", "").strip()

    if not smtp_host or not smtp_user or not recipient:
        return jsonify({"error": "Vui lòng nhập đầy đủ SMTP Host, SMTP User và Email nhận."}), 400

    # Retrieve existing encrypted password if they passed the mask
    if smtp_pass == "••••••••" or not smtp_pass:
        from auth.crypto import decrypt_password
        enc_pass = current_settings.get("smtp_pass", "")
        if enc_pass:
            try:
                smtp_pass = decrypt_password(enc_pass)
            except Exception:
                return jsonify({"error": "Không thể giải mã mật khẩu SMTP đã lưu."}), 500
        else:
            return jsonify({"error": "Mật khẩu SMTP trống."}), 400

    # Build a simple text test message
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = recipient
    msg["Subject"] = "[GDT Invoice Hub] Kiểm tra kết nối SMTP thành công"
    msg.attach(MIMEText("Kết nối SMTP từ GDT Invoice Hub của bạn hoạt động bình thường!", "plain"))

    try:
        SchedulerThread.send_smtp_message(smtp_host, smtp_port, smtp_user, smtp_pass, smtp_use_tls, recipient, msg)
        return jsonify({"status": "success", "message": "Đã gửi email thử nghiệm thành công!"})
    except Exception as e:
        return jsonify({"error": f"Lỗi gửi email thử nghiệm: {str(e)}"}), 500


@invoices_blueprint.post("/api/settings/test-audit")
@roles_required("admin")
def api_test_audit():
    """Trigger a manual run of the autonomous AI audit agent immediately."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.scheduler import load_scheduler_settings, SchedulerThread
    settings = load_scheduler_settings()
    # Force audit agent to run synchronously by ignoring its scheduled time check
    thread = SchedulerThread(current_app)
    try:
        thread.execute_autonomous_audit(settings)
        return jsonify({"status": "success", "message": "Đã chạy kiểm toán tự động thành công!"})
    except Exception as e:
        return jsonify({"error": f"Lỗi chạy kiểm toán tự động: {str(e)}"}), 500


@invoices_blueprint.get("/api/settings/logs")
@roles_required("admin", "auditor")
def api_get_settings_logs():
    """Retrieve history of background scheduler executions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.scheduler import get_scheduler_logs
    return jsonify(get_scheduler_logs())


@invoices_blueprint.post("/api/invoices/local/<invoice_id>/ai-audit")
@roles_required("admin", "auditor")
def api_trigger_ai_audit(invoice_id):
    """Trigger manual AI auditing for a specific local invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import Invoice
    from invoices.ai_service import AIComplianceAuditor
    
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return jsonify({"error": "Không tìm thấy hóa đơn cần phân tích."}), 404

    # Run AI audit (which returns AIAuditResult objects and commits them to the DB)
    auditor = AIComplianceAuditor()
    results = auditor.audit_invoice(invoice)

    return jsonify({
        "status": "success",
        "ai_warnings": [w.to_dict() for w in results]
    })


@invoices_blueprint.get("/api/invoices/local/<invoice_id>/correction-proposals")
def api_get_invoice_correction_proposals(invoice_id):
    """Get all correction proposals for a specific invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import InvoiceCorrectionProposal
    proposals = InvoiceCorrectionProposal.query.filter_by(invoice_id=invoice_id).order_by(InvoiceCorrectionProposal.id.desc()).all()
    return jsonify([p.to_dict() for p in proposals])


@invoices_blueprint.get("/api/correction-proposals")
def api_get_all_correction_proposals():
    """Get all correction proposals filtered by current taxpayer/tenant MST."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import InvoiceCorrectionProposal
    mst = session.get("tax_code")
    
    query = InvoiceCorrectionProposal.query
    if mst:
        query = query.filter_by(taxpayer_mst=mst)
        
    proposals = query.order_by(InvoiceCorrectionProposal.id.desc()).all()
    return jsonify([p.to_dict() for p in proposals])


@invoices_blueprint.post("/api/invoices/local/<invoice_id>/correction-proposals/generate")
@roles_required("admin", "auditor")
def api_generate_invoice_correction_proposals(invoice_id):
    """Manually trigger AI Auditor correction proposals generation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import Invoice
    from invoices.ai_service import AIComplianceAuditor

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return jsonify({"error": "Không tìm thấy hóa đơn cần đề xuất hiệu chỉnh."}), 404

    auditor = AIComplianceAuditor()
    if not invoice.ai_audited:
        auditor.audit_invoice(invoice)
        
    proposals = auditor.generate_correction_proposals(invoice)
    return jsonify({
        "status": "success",
        "proposals": [p.to_dict() for p in proposals]
    })


@invoices_blueprint.post("/api/correction-proposals/<int:proposal_id>/approve")
@roles_required("admin", "auditor")
def api_approve_correction_proposal(proposal_id):
    """Approve a correction proposal and apply changes directly to the invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import InvoiceCorrectionProposal
    from invoices.ai_service import apply_correction_proposal

    proposal = db.session.get(InvoiceCorrectionProposal, proposal_id)
    if not proposal:
        return jsonify({"error": "Không tìm thấy đề xuất hiệu chỉnh."}), 404

    if proposal.status != "pending":
        return jsonify({"error": f"Đề xuất đã ở trạng thái: {proposal.status}. Không thể phê duyệt lại."}), 400

    success = apply_correction_proposal(proposal)
    if success:
        return jsonify({
            "status": "success",
            "message": "Đã phê duyệt và áp dụng đề xuất hiệu chỉnh thành công.",
            "proposal": proposal.to_dict()
        })
    else:
        return jsonify({"error": "Thất bại khi áp dụng đề xuất hiệu chỉnh lên hóa đơn."}), 500


@invoices_blueprint.post("/api/correction-proposals/<int:proposal_id>/reject")
@roles_required("admin", "auditor")
def api_reject_correction_proposal(proposal_id):
    """Reject a correction proposal."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import InvoiceCorrectionProposal
    from datetime import datetime

    proposal = db.session.get(InvoiceCorrectionProposal, proposal_id)
    if not proposal:
        return jsonify({"error": "Không tìm thấy đề xuất hiệu chỉnh."}), 404

    if proposal.status != "pending":
        return jsonify({"error": f"Đề xuất đã ở trạng thái: {proposal.status}. Không thể từ chối."}), 400

    proposal.status = "rejected"
    proposal.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.session.commit()

    return jsonify({
        "status": "success",
        "message": "Đã từ chối đề xuất hiệu chỉnh.",
        "proposal": proposal.to_dict()
    })



@invoices_blueprint.post("/api/invoices/local/<invoice_id>/mitigation-letter")
@roles_required("admin", "auditor")
def api_generate_mitigation_letter(invoice_id):
    """Generate professional Vietnamese explanation letter (Công văn giải trình) using AI."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import Invoice
    from invoices.ai_service import AIComplianceAuditor

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return jsonify({"error": "Không tìm thấy hóa đơn cần giải trình."}), 404

    try:
        auditor = AIComplianceAuditor()
        letter_content = auditor.generate_mitigation_letter(invoice)
        return jsonify({
            "status": "success",
            "letter": letter_content
        })
    except Exception as e:
        return jsonify({"error": f"Lỗi tạo giải trình AI: {str(e)}"}), 500


@invoices_blueprint.post("/api/invoices/local/<invoice_id>/mitigation-letter/export")
@roles_required("admin", "auditor")
def api_export_mitigation_letter(invoice_id):
    """Export explanation letter to Word (.doc) or PDF."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    letter_text = payload.get("letter", "").strip()
    export_format = payload.get("format", "doc").lower()  # 'doc' or 'pdf'

    if not letter_text:
        return jsonify({"error": "Nội dung giải trình trống."}), 400

    filename = f"Giai_trinh_hoa_don_{invoice_id}"

    formatted_letter = letter_text.replace('\n', '<br>')

    if export_format == "pdf":
        html_content = f"""
        <html>
        <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: a4;
                margin: 2.5cm 2cm 2.5cm 2.5cm;
            }}
            body {{
                font-family: 'Arial', sans-serif;
                font-size: 12px;
                line-height: 1.6;
            }}
            .bold {{ font-weight: bold; }}
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
            .title {{ font-size: 14px; font-weight: bold; text-align: center; margin-top: 15px; margin-bottom: 15px; }}
            p {{ margin-bottom: 8px; text-align: justify; text-indent: 1cm; }}
            .no-indent {{ text-indent: 0; }}
        </style>
        </head>
        <body>
        {formatted_letter}
        </body>
        </html>
        """
        try:
            pdf_buf = render_html_to_pdf(html_content)
            return send_file(
                pdf_buf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"{filename}.pdf"
            )
        except Exception as e:
            return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

    else:
        # Default to Word compatible HTML (.doc)
        html_content = f"""
        <html xmlns:o="urn:schemas-microsoft-com:office:office"
              xmlns:w="urn:schemas-microsoft-com:office:word"
              xmlns="http://www.w3.org/TR/REC-html40">
        <head>
        <meta charset="utf-8">
        <!--[if gte mso 9]>
        <xml>
        <w:WordDocument>
        <w:View>Print</w:View>
        <w:Zoom>100</w:Zoom>
        <w:DoNotOptimizeForBrowser/>
        </w:WordDocument>
        </xml>
        <![endif]-->
        <style>
            @page {{
                size: 8.27in 11.69in; /* A4 */
                margin: 1.0in 0.79in 1.0in 1.18in; /* Top Right Bottom Left in Word format */
                mso-header-margin: .5in;
                mso-footer-margin: .5in;
                mso-paper-source: 0;
            }}
            body {{
                font-family: 'Times New Roman', serif;
                font-size: 12pt;
                line-height: 1.5;
            }}
            p {{
                margin-top: 0;
                margin-bottom: 6pt;
                text-align: justify;
            }}
            .bold {{ font-weight: bold; }}
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
        </style>
        </head>
        <body>
        {formatted_letter}
        </body>
        </html>
        """
        doc_bytes = html_content.encode("utf-8")
        return send_file(
            BytesIO(doc_bytes),
            mimetype="application/msword",
            as_attachment=True,
            download_name=f"{filename}.doc"
        )


def render_html_to_pdf(html_content: str) -> io.BytesIO:
    """Helper to compile HTML content to PDF using xhtml2pdf with pre-registered Vietnamese fonts."""
    # Find a suitable Vietnamese TrueType Font
    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    font_path = None
    for path in candidates:
        if os.path.exists(path):
            font_path = path
            break
            
    if font_path:
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            pdfmetrics.registerFont(TTFont('Arial', font_path))
            
            # Also attempt to register Bold font
            bold_path = font_path.replace("arial.ttf", "arialbd.ttf").replace("tahoma.ttf", "tahomabd.ttf").replace("Arial.ttf", "Arial Bold.ttf")
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont('Arial-Bold', bold_path))
            else:
                pdfmetrics.registerFont(TTFont('Arial-Bold', font_path))
        except Exception:
            pass

    pdf_buffer = io.BytesIO()
    from xhtml2pdf import pisa
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
    if pisa_status.err:
        raise RuntimeError("xhtml2pdf rendering failed.")
    pdf_buffer.seek(0)
    return pdf_buffer


@invoices_blueprint.get("/api/invoices/<invoice_id>/pdf")
@roles_required("admin", "auditor")
def api_invoice_pdf_download(invoice_id):
    """Download printable official-style PDF electronic invoice using xhtml2pdf."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    # Locate invoice
    current_app.config["CURRENT_JWT"] = session.get("jwt")
    current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
    try:
        from datetime import date
        invoice = current_app.config["CURRENT_INVOICE_LOOKUP"].get(invoice_id)
        
        if not invoice:
            invoices_purchase = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "purchase"))
            invoice = build_invoice_lookup(invoices_purchase).get(invoice_id)
            
        if not invoice:
            invoices_sold = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "sold"))
            invoice = build_invoice_lookup(invoices_sold).get(invoice_id)

        if not invoice:
            from invoices.service import get_local_invoices
            local_db = get_local_invoices()
            for item in local_db:
                if item["id"] == invoice_id:
                    invoice = item
                    break
            
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn yêu cầu."}), 404

        line_items = fetch_invoice_line_items(invoice_id)
    except FileNotFoundError:
        return jsonify({"error": "Không tìm thấy hóa đơn yêu cầu."}), 404
    except Exception as error:
        return jsonify({"error": f"Lỗi hệ thống: {str(error)}"}), 500
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    # Calculate sums
    sum_before_tax = sum(item.get("amount_before_tax", 0.0) for item in line_items)
    sum_tax = sum(item.get("tax_amount", 0.0) for item in line_items)
    total_payable = sum_before_tax + sum_tax

    # Convert total payable to words
    from invoices.service import doc_so_tien_vietnam
    total_payable_words = doc_so_tien_vietnam(total_payable)

    # Auto buyer/seller properties
    user_company = {
        "name": "CONG TY CO PHAN CONG NGHE GDT INVOICE HUB",
        "mst": "0109999999",
        "address": "Toa nha Technopark, Gia Lam, TP. Ha Noi",
        "phone": "1900 8888",
    }
    
    partner_details = {
        "Cong ty A": {"mst": "0101234567", "address": "So 10 Pho Hue, Quan Hai Ba Trung, Ha Noi"},
        "Cong ty B": {"mst": "0209876543", "address": "250 Nguyen Thi Minh Khai, Quan 3, TP. Ho Chi Minh"},
        "Cong ty C": {"mst": "0301122334", "address": "15 Le Loi, Quan Hai Chau, Da Nang"},
    }
    
    if "seller_name" in invoice:
        seller = {
            "name": invoice.get("seller_name", ""),
            "mst": invoice.get("seller_mst", ""),
            "address": invoice.get("seller_address", ""),
            "phone": invoice.get("seller_phone", ""),
        }
        buyer = {
            "name": invoice.get("buyer_name", ""),
            "mst": invoice.get("buyer_mst", ""),
            "address": invoice.get("buyer_address", ""),
        }
    else:
        issuer = invoice.get("issuer", "Doi tac khac")
        partner = partner_details.get(
            issuer,
            {
                "mst": f"0{abs(hash(issuer)) % 1000000000:09d}",
                "address": f"Khu cong nghiep Binh Duong, Tinh Binh Duong",
            }
        )
        partner["name"] = issuer

        # If it is a purchase invoice, issuer is Seller, user_company is Buyer
        if invoice.get("direction", "purchase") == "purchase":
            seller = partner
            buyer = user_company
        else:
            seller = user_company
            buyer = partner

    html_content = render_template(
        "invoice_pdf_export.html",
        invoice=invoice,
        line_items=line_items,
        seller=seller,
        buyer=buyer,
        sum_before_tax=sum_before_tax,
        sum_tax=sum_tax,
        total_payable=total_payable,
        total_payable_words=total_payable_words
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"invoice_{invoice_id}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500


@invoices_blueprint.get("/api/reports/partners/pdf")
@roles_required("admin", "auditor")
def api_reports_partners_pdf():
    """Export the Business Partner Directory matching the dashboard as a PDF report."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        partners = extract_partners_from_invoices(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    date_now = datetime.now().strftime("%d/%m/%Y %H:%M")
    html_content = render_template(
        "report_partners_pdf.html",
        partners=partners,
        from_date=parsed_from.strftime("%d/%m/%Y"),
        to_date=parsed_to.strftime("%d/%m/%Y"),
        direction=direction,
        date_now=date_now
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"partner_directory_{parsed_from.isoformat()}_{parsed_to.isoformat()}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500


@invoices_blueprint.get("/api/reports/usage/pdf")
@roles_required("admin", "auditor")
def api_reports_usage_pdf():
    """Export the BC26 Tax Compliance and invoice usage report as a PDF report."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "sold")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        report = generate_tax_usage_report(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    date_now = datetime.now().strftime("%d/%m/%Y %H:%M")
    html_content = render_template(
        "report_usage_pdf.html",
        report=report,
        from_date=parsed_from.strftime("%d/%m/%Y"),
        to_date=parsed_to.strftime("%d/%m/%Y"),
        direction=direction,
        date_now=date_now
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"bc26_usage_report_{parsed_from.isoformat()}_{parsed_to.isoformat()}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500


@invoices_blueprint.get("/api/ai/chat/sessions")
def api_chat_sessions():
    """Retrieve all conversational sessions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession
    try:
        sessions = AIChatSession.query.order_by(AIChatSession.created_at.desc()).all()
        # Return list directly to satisfy legacy unit test (we will make main.js support both formats)
        return jsonify([s.to_dict() for s in sessions]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/chat/sessions")
def api_create_chat_session():
    """Create a new chat session."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession
    import uuid
    from datetime import datetime
    try:
        data = request.get_json() or {}
        title = data.get("title", "Cuộc hội thoại mới")
        session_id = str(uuid.uuid4())
        new_session = AIChatSession(
            id=session_id,
            title=title,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(new_session)
        db.session.commit()
        # Dual compatibility: return details both directly and nested under 'session'
        res_dict = new_session.to_dict()
        res_dict["session"] = new_session.to_dict()
        return jsonify(res_dict), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/chat/sessions/<session_id>/message")
def api_send_chat_message(session_id):
    """Send a user message to the session and get the AI assistant response."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession, AIChatMessage
    from invoices.ai_service import AIChatAgent
    from datetime import datetime
    try:
        session = db.session.get(AIChatSession, session_id)
        if not session:
            return jsonify({"error": "Không tìm thấy phiên hội thoại."}), 404

        data = request.get_json() or {}
        content = data.get("message", "").strip()
        if not content:
            return jsonify({"error": "Nội dung tin nhắn trống."}), 400

        # Save user message
        user_msg = AIChatMessage(
            session_id=session_id,
            role="user",
            content=content,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(user_msg)
        db.session.commit()

        # Update session title if it was default
        if session.title == "Cuộc hội thoại mới" or session.title == "Cuộc trò chuyện mới":
            session.title = content[:30] + ("..." if len(content) > 30 else "")
            db.session.commit()

        # Call AI assistant agent
        agent = AIChatAgent()
        ai_response = agent.ask(session_id, content)

        # Save assistant response
        assistant_msg = AIChatMessage(
            session_id=session_id,
            role="assistant",
            content=ai_response,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(assistant_msg)
        db.session.commit()

        # Dual compatibility: return both legacy fields and 'reply' field
        return jsonify({
            "user_message": user_msg.to_dict(),
            "assistant_message": assistant_msg.to_dict(),
            "session_title": session.title,
            "reply": ai_response
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.delete("/api/ai/chat/sessions/<session_id>")
def api_delete_chat_session(session_id):
    """Delete a chat session."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession
    try:
        session = db.session.get(AIChatSession, session_id)
        if not session:
            return jsonify({"error": "Không tìm thấy phiên hội thoại."}), 404
        db.session.delete(session)
        db.session.commit()
        return jsonify({"success": True, "message": "Đã xóa phiên hội thoại thành công."}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/classify-items")
@roles_required("admin", "auditor")
def api_classify_invoice_items():
    """Classify items inside an invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice, LineItem
    from invoices.ai_service import AIExpenseClassifier

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        if not invoice_id:
            return jsonify({"error": "Thiếu mã hóa đơn invoice_id."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        classifier = AIExpenseClassifier()
        items = LineItem.query.filter_by(invoice_id=invoice.id).all()
        if not items:
            return jsonify({"success": True, "classified_items": []})

        classifications = classifier.classify_line_items(items)

        # Save results to DB
        for item in items:
            if item.id in classifications:
                item.expense_category = classifications[item.id]
        db.session.commit()

        return jsonify({
            "success": True,
            "classified_items": [
                {
                    "item_id": item.id,
                    "item_name": item.item_name,
                    "category": item.expense_category
                } for item in items
            ]
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/update-item-category")
@roles_required("admin", "auditor")
def api_update_item_category():
    """Manually update the expense category of a specific line item."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import LineItem
    try:
        data = request.get_json() or {}
        item_id = data.get("item_id")
        new_category = data.get("category")

        if not item_id or not new_category:
            return jsonify({"error": "Thiếu thông tin item_id hoặc category."}), 400

        item = db.session.get(LineItem, item_id)
        if not item:
            return jsonify({"error": "Không tìm thấy mặt hàng."}), 404

        # Validate category matches standard list
        from invoices.ai_service import AIExpenseClassifier
        if new_category not in AIExpenseClassifier.CATEGORIES:
            return jsonify({"error": "Danh mục chi phí không hợp lệ."}), 400

        item.expense_category = new_category
        db.session.commit()

        return jsonify({"success": True, "item_id": item.id, "category": item.expense_category})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/repair-metadata")
@roles_required("admin", "auditor")
def api_repair_metadata():
    """Analyze and generate AI suggestions for repairing invoice metadata."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice
    from invoices.ai_service import AIDataRepairer

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        if not invoice_id:
            return jsonify({"error": "Thiếu mã hóa đơn invoice_id."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        repairer = AIDataRepairer()
        suggestions = repairer.repair_metadata(invoice)

        before = {
            "seller_name": invoice.seller_name or "",
            "buyer_name": invoice.buyer_name or "",
            "buyer_address": invoice.buyer_address or "",
            "amount_in_words": invoice.amount_in_words or ""
        }
        
        differences = []
        for key in ["seller_name", "buyer_name", "buyer_address", "amount_in_words"]:
            val_before = before[key].strip()
            val_after = suggestions.get(key, "").strip()
            if val_before != val_after and val_after:
                differences.append(key)

        return jsonify({
            "success": True,
            "invoice_id": invoice.id,
            "before": before,
            "after": suggestions,
            "differences": differences
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ai/apply-repair")
@roles_required("admin", "auditor")
def api_apply_repair():
    """Apply selected AI repair suggestions to persistent SQLite database."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        fields_to_apply = data.get("fields", [])

        if not invoice_id or not fields_to_apply:
            return jsonify({"error": "Thiếu thông tin invoice_id hoặc fields để áp dụng."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        allowed_fields = ["seller_name", "buyer_name", "buyer_address", "amount_in_words"]
        applied = []
        for field in fields_to_apply:
            if field in allowed_fields:
                val = data.get(field)
                if val:
                    setattr(invoice, field, val)
                    applied.append(field)

        if applied:
            db.session.commit()
            from invoices.security_audit_service import log_security_event
            log_security_event("REPAIR", f"Applied AI repair to invoice {invoice_id} for fields: {', '.join(applied)}")

        return jsonify({
            "success": True,
            "invoice_id": invoice.id,
            "applied_fields": applied
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/invoices/summary-by-seller")
def api_summary_by_seller():
    """Aggregate input invoices by month or quarter, grouped by seller."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice
    from sqlalchemy import func

    try:
        period_type = request.args.get("period_type", "monthly")
        year_filter = request.args.get("year", "")

        # Base query for all invoices
        query = db.session.query(
            Invoice.seller_mst,
            Invoice.seller_name,
            Invoice.date,
            func.count(Invoice.id).label("invoice_count"),
            func.sum(Invoice.amount_before_tax).label("total_before_tax"),
            func.sum(Invoice.tax_amount).label("total_tax"),
            func.sum(Invoice.total_amount).label("total_amount")
        )

        # Apply year filter if provided
        if year_filter:
            query = query.filter(Invoice.date.like(f"{year_filter}-%"))
        
        # Pull raw grouped results and aggregate/group them cleanly in Python
        results = query.group_by(
            Invoice.seller_mst,
            Invoice.seller_name,
            func.substr(Invoice.date, 1, 7)
        ).all()

        period_map = {}

        for row in results:
            mst, name, date_val, count, before_tax, tax, total = row
            if not date_val or len(date_val) < 7:
                continue
            
            row_year = date_val[0:4]
            if year_filter and row_year != year_filter:
                continue

            row_month = date_val[5:7]

            if period_type == "quarterly":
                try:
                    m_int = int(row_month)
                except ValueError:
                    m_int = 1
                if m_int in [1, 2, 3]:
                    period = f"Quý 1 / {row_year}"
                elif m_int in [4, 5, 6]:
                    period = f"Quý 2 / {row_year}"
                elif m_int in [7, 8, 9]:
                    period = f"Quý 3 / {row_year}"
                else:
                    period = f"Quý 4 / {row_year}"
            else:
                period = f"Tháng {row_month} / {row_year}"

            if period not in period_map:
                period_map[period] = {}

            # Aggregate sellers within the same period
            seller_key = mst or "UNKNOWN"
            if seller_key not in period_map[period]:
                period_map[period][seller_key] = {
                    "seller_mst": mst or "Không rõ",
                    "seller_name": name or "Không rõ",
                    "invoice_count": 0,
                    "total_before_tax": 0.0,
                    "total_tax": 0.0,
                    "total_amount": 0.0
                }

            entry = period_map[period][seller_key]
            entry["invoice_count"] += count
            entry["total_before_tax"] += before_tax or 0.0
            entry["total_tax"] += tax or 0.0
            entry["total_amount"] += total or 0.0

        # Format and sort periods
        data = []
        for period, sellers_dict in period_map.items():
            sellers_list = list(sellers_dict.values())
            # Sort sellers by total amount descending
            sellers_list.sort(key=lambda x: x["total_amount"], reverse=True)
            data.append({
                "period": period,
                "sellers": sellers_list,
                "total_before_tax": sum(s["total_before_tax"] for s in sellers_list),
                "total_tax": sum(s["total_tax"] for s in sellers_list),
                "total_amount": sum(s["total_amount"] for s in sellers_list)
            })

        # Sort periods. Monthly: "Tháng 12 / 2026" -> "Tháng 01 / 2026" descending.
        # Format key is period title itself.
        data.sort(key=lambda x: x["period"], reverse=True)

        return jsonify({
            "success": True,
            "period_type": period_type,
            "year": year_filter or "Tất cả",
            "data": data
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/reports/vat-declaration")
def api_reports_vat_declaration():
    """Generate a draft of the Vietnamese VAT Return Mẫu 01/GTGT and list of disputed/high-risk input invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            # Default to current or last completed month/quarter
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        # Standardizing month string to "02" instead of "2" for monthly
        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice, LineItem

        query = Invoice.query.filter(Invoice.is_cancelled == False)

        # Apply date filters
        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        # Initialize output VAT rate aggregates (Sold)
        output_exempt_val = 0.0
        output_0_val = 0.0
        output_5_val = 0.0
        output_5_vat = 0.0
        output_10_val = 0.0
        output_10_vat = 0.0
        
        # Input Aggregates (Purchase)
        input_total_value = 0.0
        input_total_vat = 0.0
        input_deductible_vat = 0.0
        
        disputed_invoices = []

        for inv in invoices:
            if inv.invoice_type == "sold":
                has_items = len(inv.items) > 0
                if has_items:
                    for item in inv.items:
                        rate = (item.tax_rate or "").strip().lower()
                        val = item.amount_before_tax or 0.0
                        tax = item.tax_amount or 0.0
                        
                        if "không chịu" in rate or "khong chiu" in rate:
                            output_exempt_val += val
                        elif "0%" in rate or rate == "0":
                            output_0_val += val
                        elif "5%" in rate or rate == "5":
                            output_5_val += val
                            output_5_vat += tax
                        else:
                            # 8% and 10% grouped into standard rate
                            output_10_val += val
                            output_10_vat += tax
                else:
                    # Fallback to invoice totals if no line items exist
                    val = inv.amount_before_tax or 0.0
                    tax = inv.tax_amount or 0.0
                    output_10_val += val
                    output_10_vat += tax

            elif inv.invoice_type == "purchase":
                val = inv.amount_before_tax or 0.0
                tax = inv.tax_amount or 0.0
                
                input_total_value += val
                input_total_vat += tax
                
                # Combine traditional parsing warnings and Gemma-4 AI auditor warnings
                warnings = list(inv.warnings) if inv.warnings else []
                ai_warnings = [f"[AI: {w.warning_type}] {w.explanation}" for w in inv.ai_audit_results]
                warnings.extend(ai_warnings)
                
                is_disputed = len(warnings) > 0
                
                if is_disputed:
                    warning_msg = "; ".join(warnings)
                    disputed_invoices.append({
                        "id": inv.id,
                        "number": inv.number or "Không số",
                        "date": inv.date or "Không ngày",
                        "seller_name": inv.seller_name or "Không rõ",
                        "seller_mst": inv.seller_mst or "Không rõ",
                        "amount_before_tax": val,
                        "tax_amount": tax,
                        "total_amount": inv.total_amount or (val + tax),
                        "warning": warning_msg
                    })
                else:
                    input_deductible_vat += tax

        output_taxable_val = output_0_val + output_5_val + output_10_val
        output_total_value = output_exempt_val + output_taxable_val
        output_total_vat = output_5_vat + output_10_vat

        vat_payable = max(0.0, output_total_vat - input_deductible_vat)
        vat_carried_forward = max(0.0, input_deductible_vat - output_total_vat)

        return jsonify({
            "success": True,
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "outputs": {
                "exempt_val": output_exempt_val,
                "tax_0_val": output_0_val,
                "tax_5_val": output_5_val,
                "tax_5_vat": output_5_vat,
                "tax_10_val": output_10_val,
                "tax_10_vat": output_10_vat,
                "taxable_val": output_taxable_val,
                "total_val": output_total_value,
                "total_vat": output_total_vat
            },
            "inputs": {
                "total_value": input_total_value,
                "total_vat": input_total_vat,
                "deductible_vat": input_deductible_vat
            },
            "calculations": {
                "vat_payable": vat_payable,
                "vat_carried_forward": vat_carried_forward
            },
            "disputed_invoices": disputed_invoices
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Analytics Pro: Supplier Price Trends
# ---------------------------------------------------------------------------

@invoices_blueprint.get("/api/analytics/top-items")
def analytics_top_items():
    """Return top 20 most-purchased line item names for autocomplete."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from extensions import db
        from invoices.models import LineItem
        rows = (
            db.session.query(LineItem.item_name, db.func.count(LineItem.id).label("cnt"))
            .group_by(db.func.lower(LineItem.item_name))
            .order_by(db.desc("cnt"))
            .limit(20)
            .all()
        )
        items = [{"name": r[0], "count": r[1]} for r in rows]
        return jsonify({"success": True, "items": items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/analytics/supplier-price-trends")
def analytics_supplier_price_trends():
    """Return monthly unit price data for a given item, grouped by seller."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    item_name = request.args.get("item_name", "").strip()
    year_filter = request.args.get("year", "").strip()

    if not item_name:
        return jsonify({"error": "Vui lòng cung cấp tên mặt hàng (item_name)."}), 400

    try:
        from extensions import db
        from invoices.models import LineItem, Invoice as Inv

        query = (
            db.session.query(
                LineItem.item_name,
                LineItem.unit_price,
                LineItem.amount_before_tax,
                Inv.date,
                Inv.seller_name,
                Inv.seller_mst,
            )
            .join(Inv, LineItem.invoice_id == Inv.id)
            .filter(db.func.lower(LineItem.item_name).like(f"%{item_name.lower()}%"))
            .filter(Inv.is_cancelled == False)
        )

        if year_filter:
            query = query.filter(Inv.date.like(f"{year_filter}%"))

        rows = query.all()

        if not rows:
            return jsonify({"success": True, "item_name": item_name, "sellers": [], "months": [], "series": [], "anomalies": []})

        # Compute global average unit price for anomaly detection
        all_prices = [r[1] for r in rows if r[1] and r[1] > 0]
        avg_global = sum(all_prices) / len(all_prices) if all_prices else 0

        # Build month × seller matrix
        seller_map = {}  # seller_mst → {name, months: {YYYY-MM: [prices]}}
        months_set = set()

        for item_name_val, unit_price, amount, inv_date, seller_name, seller_mst in rows:
            if not inv_date:
                continue
            month_key = inv_date[:7]  # YYYY-MM
            months_set.add(month_key)
            mst = seller_mst or "unknown"
            if mst not in seller_map:
                seller_map[mst] = {"seller_mst": mst, "seller_name": seller_name or mst, "months": {}}
            seller_map[mst]["months"].setdefault(month_key, []).append(unit_price or 0)

        months_sorted = sorted(months_set)

        # Build series for chart
        series = []
        anomalies = []
        for mst, info in seller_map.items():
            prices_by_month = []
            for m in months_sorted:
                month_prices = info["months"].get(m, [])
                if month_prices:
                    avg_m = sum(month_prices) / len(month_prices)
                    prices_by_month.append(round(avg_m, 0))
                    if avg_global > 0 and avg_m > avg_global * 1.20:
                        anomalies.append({
                            "month": m,
                            "seller_name": info["seller_name"],
                            "seller_mst": mst,
                            "price": round(avg_m, 0),
                            "avg_global": round(avg_global, 0),
                            "pct_above": round((avg_m / avg_global - 1) * 100, 1),
                        })
                else:
                    prices_by_month.append(None)

            # Summary stats
            flat = [p for p in prices_by_month if p is not None]
            series.append({
                "seller_mst": mst,
                "seller_name": info["seller_name"],
                "prices": prices_by_month,
                "avg_price": round(sum(flat) / len(flat), 0) if flat else 0,
                "min_price": round(min(flat), 0) if flat else 0,
                "max_price": round(max(flat), 0) if flat else 0,
                "purchase_count": sum(len(info["months"].get(m, [])) for m in months_sorted),
            })

        # Sort series by avg_price ascending (cheapest first)
        series.sort(key=lambda s: s["avg_price"])

        return jsonify({
            "success": True,
            "item_name": item_name,
            "avg_global": round(avg_global, 0),
            "months": months_sorted,
            "sellers": [{"seller_mst": s["seller_mst"], "seller_name": s["seller_name"]} for s in series],
            "series": series,
            "anomalies": anomalies,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Analytics Pro: VAT Forecast
# ---------------------------------------------------------------------------

@invoices_blueprint.get("/api/analytics/vat-forecast")
def analytics_vat_forecast():
    """Return monthly actual VAT net (output-input) and 2-month linear forecast."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import datetime as _dt
    year_filter = request.args.get("year", str(_dt.now().year)).strip()

    try:
        from extensions import db
        from invoices.models import Invoice
        invoices_all = Invoice.query.filter(Invoice.is_cancelled == False).all()

        # Build monthly buckets for the selected year
        monthly = {}
        for inv in invoices_all:
            if not inv.date or not inv.date.startswith(year_filter):
                continue
            month_key = inv.date[:7]  # YYYY-MM
            if month_key not in monthly:
                monthly[month_key] = {"output_vat": 0.0, "input_vat": 0.0}
            if inv.invoice_type == "sold":
                monthly[month_key]["output_vat"] += inv.tax_amount or 0.0
            elif inv.invoice_type == "purchase":
                monthly[month_key]["input_vat"] += inv.tax_amount or 0.0

        # All 12 months of the selected year
        all_months = [f"{year_filter}-{m:02d}" for m in range(1, 13)]
        actual = []
        for m in all_months:
            b = monthly.get(m, {"output_vat": 0.0, "input_vat": 0.0})
            net = b["output_vat"] - b["input_vat"]
            actual.append({
                "month": m,
                "output_vat": round(b["output_vat"], 0),
                "input_vat": round(b["input_vat"], 0),
                "net_vat": round(net, 0),
                "has_data": m in monthly,
            })

        # Identify last N months with real data for trend computation
        real_months = [a for a in actual if a["has_data"]]
        forecast = []

        if len(real_months) >= 2:
            # Linear trend: avg delta over last min(3, n) real months
            window = real_months[-3:] if len(real_months) >= 3 else real_months
            deltas = [window[i]["net_vat"] - window[i - 1]["net_vat"] for i in range(1, len(window))]
            avg_delta = sum(deltas) / len(deltas) if deltas else 0
            last_net = real_months[-1]["net_vat"]
            last_month_str = real_months[-1]["month"]

            # Build next 2 months after the last real month
            last_dt = _dt.strptime(last_month_str, "%Y-%m")
            for i in range(1, 3):
                if last_dt.month + i <= 12:
                    fdt = last_dt.replace(month=last_dt.month + i)
                else:
                    fdt = last_dt.replace(year=last_dt.year + 1, month=(last_dt.month + i) - 12)
                projected_net = last_net + avg_delta * i
                prev_net = last_net + avg_delta * (i - 1)
                warning = prev_net > 0 and projected_net > prev_net * 1.30
                forecast.append({
                    "month": fdt.strftime("%Y-%m"),
                    "net_vat_forecast": round(projected_net, 0),
                    "warning": warning,
                })

        # Compute year summary
        total_output = sum(a["output_vat"] for a in actual if a["has_data"])
        total_input = sum(a["input_vat"] for a in actual if a["has_data"])
        total_net = total_output - total_input

        return jsonify({
            "success": True,
            "year": year_filter,
            "actual": actual,
            "forecast": forecast,
            "summary": {
                "total_output_vat": round(total_output, 0),
                "total_input_vat": round(total_input, 0),
                "total_net_vat": round(total_net, 0),
                "months_with_data": len(real_months),
            },
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# US-035: Budget Monitor & Spending Alerts
# ---------------------------------------------------------------------------

@invoices_blueprint.get("/api/budget/config")
def budget_config_get():
    """Return saved budget configuration for a given month."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    month = request.args.get("month", "").strip()
    if not month:
        month = datetime.now().strftime("%Y-%m")

    try:
        from invoices.models import SystemConfig
        key = f"budget_config_{month}"
        cfg = db.session.get(SystemConfig, key)
        if cfg:
            import json as _json
            configs = _json.loads(cfg.value)
        else:
            configs = []
        return jsonify({"success": True, "month": month, "configs": configs})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/budget/config")
@roles_required("admin", "auditor")
def budget_config_save():
    """Save budget configuration for a given month."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        import json as _json
        body = request.get_json(force=True) or {}
        month = body.get("month", "").strip()
        configs = body.get("configs", [])
        if not month:
            month = datetime.now().strftime("%Y-%m")

        from extensions import db
        from invoices.models import SystemConfig
        key = f"budget_config_{month}"
        cfg = db.session.get(SystemConfig, key)
        if cfg:
            cfg.value = _json.dumps(configs, ensure_ascii=False)
        else:
            cfg = SystemConfig(key=key, value=_json.dumps(configs, ensure_ascii=False))
            db.session.add(cfg)
        db.session.commit()
        return jsonify({"success": True, "month": month, "saved": len(configs)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/budget/actuals")
def budget_actuals():
    """Return actual spending per expense_category for a given month with budget vs. actual comparison."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    month = request.args.get("month", "").strip()
    if not month:
        month = datetime.now().strftime("%Y-%m")

    try:
        import json as _json
        from extensions import db
        from invoices.models import LineItem, Invoice as _Inv, SystemConfig

        # Aggregate actual spending by expense_category for the month
        rows = (
            db.session.query(
                LineItem.expense_category,
                db.func.sum(LineItem.amount_before_tax).label("actual_vnd"),
            )
            .join(_Inv, LineItem.invoice_id == _Inv.id)
            .filter(
                _Inv.invoice_type == "purchase",
                _Inv.is_cancelled == False,
                _Inv.date.like(f"{month}%"),
            )
            .group_by(LineItem.expense_category)
            .all()
        )

        actuals_map = {}
        for category, actual_vnd in rows:
            cat = category or "Chưa phân loại"
            actuals_map[cat] = round(actual_vnd or 0, 0)

        # Load budget config
        key = f"budget_config_{month}"
        cfg_rec = db.session.get(SystemConfig, key)
        budget_configs = _json.loads(cfg_rec.value) if cfg_rec else []
        budget_map = {c["category"]: c["limit_vnd"] for c in budget_configs}

        # Build response combining actuals + budgets
        all_categories = set(actuals_map.keys()) | set(budget_map.keys())
        actuals = []
        for cat in sorted(all_categories):
            actual_vnd = actuals_map.get(cat, 0)
            limit_vnd = budget_map.get(cat)
            if limit_vnd and limit_vnd > 0:
                pct = round(actual_vnd / limit_vnd * 100, 1)
                if pct >= 100:
                    status = "over_budget"
                elif pct >= 70:
                    status = "warning"
                else:
                    status = "ok"
            else:
                pct = None
                status = "no_budget"
            actuals.append({
                "category": cat,
                "actual_vnd": actual_vnd,
                "limit_vnd": limit_vnd,
                "pct_used": pct,
                "status": status,
            })

        any_over = any(a["status"] == "over_budget" for a in actuals)
        any_warning = any(a["status"] == "warning" for a in actuals)
        alert_level = "over_budget" if any_over else ("warning" if any_warning else "ok")

        return jsonify({
            "success": True,
            "month": month,
            "actuals": actuals,
            "alert_level": alert_level,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# US-036: Invoice Aging & Receivable Tracker
# ---------------------------------------------------------------------------

_AGING_BUCKETS = [
    ("1–30 ngày",   1,  30),
    ("31–60 ngày",  31, 60),
    ("61–90 ngày",  61, 90),
    (">90 ngày",    91, None),
]


@invoices_blueprint.get("/api/aging/summary")
def aging_summary():
    """
    Return outstanding sold (receivables) and bought (payables) invoices classified into aging buckets.

    Buckets: Current / 1-30 / 31-60 / 61-90 / >90 days overdue.
    - Excludes invoices with paid_date set (already paid).
    - Excludes cancelled invoices.
    - Falls back to invoice.date when due_date is absent.
    - Accepts optional ?as_of=YYYY-MM-DD (default: today).
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import date as _date
    as_of_str = request.args.get("as_of", "").strip()
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    
    try:
        as_of = _date.fromisoformat(as_of_str) if as_of_str else _date.today()
    except ValueError:
        return jsonify({"error": "Định dạng as_of không hợp lệ. Dùng YYYY-MM-DD."}), 400

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv

        # Fetch all outstanding sold and bought invoices for active taxpayer
        query = _Inv.query.filter(
            _Inv.is_cancelled == False,
            _Inv.paid_date == None,
        )
        if mst:
            query = query.filter(_Inv.taxpayer_mst == mst)
        invoices = query.all()

        # Build empty bucket structures for receivables and payables
        def empty_buckets():
            return [
                {"label": "Chưa quá hạn (Current)", "min_days": None, "max_days": 0, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "1–30 ngày", "min_days": 1, "max_days": 30, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "31–60 ngày", "min_days": 31, "max_days": 60, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "61–90 ngày", "min_days": 61, "max_days": 90, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": ">90 ngày", "min_days": 91, "max_days": None, "count": 0, "total_amount": 0.0, "invoices": []}
            ]

        ar_buckets = empty_buckets()
        ap_buckets = empty_buckets()

        for inv in invoices:
            # Determine reference date for aging (due_date preferred, fall back to invoice date)
            ref_date_str = inv.due_date or inv.date
            if not ref_date_str:
                continue
            try:
                ref_date = _date.fromisoformat(ref_date_str)
            except ValueError:
                continue

            age_days = (as_of - ref_date).days

            # Phase 3: Autonomous Categorization rules
            cat = "OPEX"
            if inv.seller_name and any(kw in inv.seller_name.lower() for kw in ["điện", "nước", "utility"]):
                cat = "UTILITIES"
            elif inv.invoice_type == "sold":
                cat = "REVENUE"

            inv_data = {
                "id": inv.id,
                "date": inv.date,
                "due_date": inv.due_date,
                "invoice_type": inv.invoice_type,
                "seller_name": inv.seller_name or "",
                "buyer_name": inv.buyer_name or "",
                "buyer_mst": inv.buyer_mst or "",
                "amount_before_tax": inv.amount_before_tax or 0.0,
                "total_amount": inv.total_amount or 0.0,
                "age_days": age_days,
                "ai_category": cat
            }

            target_buckets = ar_buckets if inv.invoice_type == "sold" else ap_buckets

            # Assign to correct bucket based on age_days
            if age_days <= 0:
                target_buckets[0]["count"] += 1
                target_buckets[0]["total_amount"] += inv.total_amount or 0.0
                target_buckets[0]["invoices"].append(inv_data)
            else:
                for bucket in target_buckets[1:]:
                    mn, mx = bucket["min_days"], bucket["max_days"]
                    if age_days >= mn and (mx is None or age_days <= mx):
                        bucket["count"] += 1
                        bucket["total_amount"] += inv.total_amount or 0.0
                        bucket["invoices"].append(inv_data)
                        break

        # Calculate totals
        total_ar = sum(b["total_amount"] for b in ar_buckets)
        total_ap = sum(b["total_amount"] for b in ap_buckets)

        # For backwards compatibility with standard dashboard charts, return overdue-only AR buckets
        legacy_buckets = []
        for label, mn, mx in _AGING_BUCKETS:
            # Find the corresponding ar_bucket
            corr = next((b for b in ar_buckets if b["label"] == label), None)
            if corr:
                legacy_buckets.append(corr)
            else:
                legacy_buckets.append({
                    "label": label,
                    "min_days": mn,
                    "max_days": mx,
                    "count": 0,
                    "total_amount": 0.0,
                    "invoices": []
                })

        return jsonify({
            "success": True,
            "as_of": as_of.isoformat(),
            "total_outstanding": total_ar,  # legacy name for AR total
            "total_count": sum(b["count"] for b in ar_buckets[1:]),
            "buckets": legacy_buckets,      # legacy overdue sold buckets
            "receivables": {
                "total_amount": total_ar,
                "total_count": sum(b["count"] for b in ar_buckets),
                "buckets": ar_buckets
            },
            "payables": {
                "total_amount": total_ap,
                "total_count": sum(b["count"] for b in ap_buckets),
                "buckets": ap_buckets
            }
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/cashflow/projection")
def cashflow_projection_endpoint():
    """
    US-083: Predictive Cash Projection & What-If late payment simulation.
    Projects optimistic vs pessimistic cash balance daily over the next 90 days.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import date as _date, timedelta
    as_of_str = request.args.get("as_of", "").strip()
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    
    # Custom simulation parameters
    late_days = int(request.args.get("late_days", "0").strip() or "0")
    client_mst = request.args.get("client_mst", "").strip()
    
    # Base balance override
    base_balance_str = request.args.get("base_balance", "").strip()
    
    try:
        as_of = _date.fromisoformat(as_of_str) if as_of_str else _date.today()
    except ValueError:
        return jsonify({"error": "Định dạng as_of không hợp lệ. Dùng YYYY-MM-DD."}), 400

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv, BankTransaction as _Tx

        # 1. Base Balance Dynamic Calculation
        if base_balance_str:
            try:
                base_balance = float(base_balance_str)
            except ValueError:
                base_balance = 500000000.0
        else:
            # Calculate base balance from actual bank transaction history
            tx_query = _Tx.query
            if mst:
                tx_query = tx_query.filter(_Tx.taxpayer_mst == mst)
            txs = tx_query.all()
            if txs:
                base_balance = sum(tx.amount for tx in txs)
            else:
                base_balance = 500000000.0  # Default fallback if no bank tx

        # 2. Fetch all unpaid outstanding invoices
        query = _Inv.query.filter(
            _Inv.is_cancelled == False,
            _Inv.paid_date == None
        )
        if mst:
            query = query.filter(_Inv.taxpayer_mst == mst)
        invoices = query.all()

        # 3. Simulate day-by-day cashflows for the next 90 days
        projections = []
        current_opt = base_balance
        current_sim = base_balance

        # We pre-calculate expected flows per day
        opt_inflows = {}
        sim_inflows = {}
        outflows = {}

        for i in range(91):
            day = as_of + timedelta(days=i)
            day_str = day.isoformat()
            opt_inflows[day_str] = 0.0
            sim_inflows[day_str] = 0.0
            outflows[day_str] = 0.0

        for inv in invoices:
            ref_date_str = inv.due_date or inv.date
            if not ref_date_str:
                continue
            try:
                ref_date = _date.fromisoformat(ref_date_str)
            except ValueError:
                continue

            amount = inv.total_amount or 0.0

            if inv.invoice_type == "sold":
                # Sales -> Inflows
                # Optimistic Expected Date (always due_date/date)
                opt_day = ref_date
                if opt_day < as_of:
                    opt_day = as_of # Already overdue: assume collected today
                opt_str = opt_day.isoformat()
                if opt_str in opt_inflows:
                    opt_inflows[opt_str] += amount

                # Simulated/Pessimistic Expected Date
                sim_day = ref_date
                if late_days > 0:
                    if not client_mst or inv.buyer_mst == client_mst:
                        sim_day = sim_day + timedelta(days=late_days)
                
                if sim_day < as_of:
                    sim_day = as_of # Overdue shift
                
                sim_str = sim_day.isoformat()
                if sim_str in sim_inflows:
                    sim_inflows[sim_str] += amount
            else:
                # Purchases -> Outflows
                out_day = ref_date
                if out_day < as_of:
                    out_day = as_of # Overdue payables must be paid today
                out_str = out_day.isoformat()
                if out_str in outflows:
                    outflows[out_str] += amount

        for i in range(91):
            day = as_of + timedelta(days=i)
            day_str = day.isoformat()

            in_opt = opt_inflows.get(day_str, 0.0)
            in_sim = sim_inflows.get(day_str, 0.0)
            out = outflows.get(day_str, 0.0)

            current_opt += (in_opt - out)
            current_sim += (in_sim - out)

            projections.append({
                "date": day_str,
                "day_label": day.strftime("%d/%m"),
                "inflow_opt": in_opt,
                "inflow_sim": in_sim,
                "outflow": out,
                "balance_opt": current_opt,
                "balance_sim": current_sim
            })

        return jsonify({
            "success": True,
            "base_balance": base_balance,
            "projections": projections
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/analytics/cashflow_forecast")
def cashflow_forecast():
    """US-062: Cashflow Forecasting API
    Predicts inflow/outflow based on unpaid invoices over the next 30 days.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from datetime import datetime, timedelta
        mst = request.args.get("mst") or session.get("active_taxpayer_mst")
        
        # Query unpaid invoices
        query = Invoice.query.filter(
            Invoice.is_cancelled == False,
            Invoice.paid_date == None
        )
        if mst:
            query = query.filter(Invoice.taxpayer_mst == mst)
            
        invoices = query.all()
        
        # Bucket by day (0 to 30 days ahead)
        today = datetime.now().date()
        forecast = []
        
        base_liquidity = 100000000  # Default base cash reserve
        current_liquidity = base_liquidity
        
        for i in range(30):
            target_date = today + timedelta(days=i)
            target_date_str = target_date.isoformat()
            
            inflow = 0
            outflow = 0
            
            for inv in invoices:
                due = inv.due_date or inv.date
                if due == target_date_str:
                    if inv.invoice_type == "sold":
                        inflow += inv.total_amount
                    else:
                        outflow += inv.total_amount
                        
            current_liquidity += (inflow - outflow)
            
            forecast.append({
                "date": target_date_str,
                "day_label": target_date.strftime("%d/%m"),
                "inflow": inflow,
                "outflow": outflow,
                "net_liquidity": current_liquidity
            })
            
        return jsonify({
            "success": True,
            "base_liquidity": base_liquidity,
            "forecast": forecast
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.patch("/api/invoices/<string:invoice_id>/payment")
def update_invoice_payment(invoice_id: str):
    """
    Update due_date and/or paid_date on an invoice.

    Body: {due_date?: "YYYY-MM-DD", paid_date?: "YYYY-MM-DD"}
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv

        inv = db.session.get(_Inv, invoice_id)
        if not inv:
            return jsonify({"error": f"Không tìm thấy hóa đơn: {invoice_id}"}), 404

        body = request.get_json(force=True) or {}
        updated = False

        if "due_date" in body:
            inv.due_date = body["due_date"] or None
            updated = True
        if "paid_date" in body:
            inv.paid_date = body["paid_date"] or None
            updated = True

        if updated:
            inv.updated_at = datetime.now().isoformat()
            db.session.commit()

        return jsonify({
            "success": True,
            "id": invoice_id,
            "due_date": inv.due_date,
            "paid_date": inv.paid_date,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/profiles")
@roles_required("admin", "auditor", "viewer")
def get_taxpayer_profiles():
    """Retrieve all stored taxpayer profiles."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    try:
        profiles = TaxpayerProfile.query.order_by(TaxpayerProfile.mst).all()
        return jsonify([p.to_dict() for p in profiles])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/profiles")
@roles_required("admin", "auditor")
def create_taxpayer_profile():
    """Create or update a taxpayer profile with encrypted credentials."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    from auth.crypto import encrypt_password
    try:
        body = request.get_json() or {}
        mst = body.get("mst")
        company_name = body.get("company_name")
        gdt_username = body.get("gdt_username")
        gdt_password = body.get("gdt_password")

        if not all([mst, company_name, gdt_username, gdt_password]):
            return jsonify({"error": "Thiếu các thông tin bắt buộc."}), 400

        mst = str(mst).strip()
        if len(mst) not in [10, 14]:
            return jsonify({"error": "Mã số thuế không đúng định dạng (phải có 10 hoặc 14 ký tự)."}), 400

        encrypted_password = encrypt_password(gdt_password)

        profile = db.session.get(TaxpayerProfile, mst)
        is_update = profile is not None
        if profile:
            profile.company_name = company_name
            profile.gdt_username = gdt_username
            profile.gdt_password_encrypted = encrypted_password
        else:
            profile = TaxpayerProfile(
                mst=mst,
                company_name=company_name,
                gdt_username=gdt_username,
                gdt_password_encrypted=encrypted_password,
                is_active=True,
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.session.add(profile)

        db.session.commit()

        from invoices.security_audit_service import log_security_event
        if is_update:
            log_security_event("PROFILE", f"Updated taxpayer profile for MST: {mst}")
        else:
            log_security_event("PROFILE", f"Created new taxpayer profile for MST: {mst}")

        return jsonify({"success": True, "profile": profile.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.delete("/api/profiles/<mst>")
@roles_required("admin", "auditor")
def delete_taxpayer_profile(mst):
    """Delete a taxpayer profile and cascade delete all its invoices."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    try:
        profile = db.session.get(TaxpayerProfile, mst)
        if not profile:
            return jsonify({"error": f"Không tìm thấy hồ sơ mã số thuế {mst}."}), 404

        db.session.delete(profile)
        db.session.commit()

        from invoices.security_audit_service import log_security_event
        log_security_event("PROFILE", f"Deleted taxpayer profile for MST: {mst}")

        return jsonify({"success": True, "message": f"Đã xóa hồ sơ và tất cả hóa đơn của MST {mst}."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/profiles/switch")
@roles_required("admin", "auditor", "viewer")
def switch_taxpayer_profile():
    """Switch the current active taxpayer profile in session."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        mst = body.get("mst")

        from invoices.security_audit_service import log_security_event
        if mst == "all" or not mst:
            session["active_taxpayer_mst"] = None
            log_security_event("PROFILE", "Switched active taxpayer profile to all (no filter)")
            return jsonify({"success": True, "active_taxpayer_mst": None})

        from invoices.models import TaxpayerProfile
        profile = db.session.get(TaxpayerProfile, mst)
        if not profile:
            return jsonify({"error": f"Không tìm thấy hồ sơ mã số thuế {mst}."}), 404

        session["active_taxpayer_mst"] = mst
        log_security_event("PROFILE", f"Switched active taxpayer profile to MST: {mst}")
        return jsonify({"success": True, "active_taxpayer_mst": mst})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/invoices/<invoice_id>/post-erp")
@roles_required("admin", "auditor")
def api_post_invoice_to_erp(invoice_id):
    """Manually post an invoice to the configured ERP system."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import Invoice
    from invoices.erp_service import post_invoice_to_erp

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return jsonify({"error": "Không tìm thấy hóa đơn cần đồng bộ."}), 404

    success = post_invoice_to_erp(invoice)
    if success:
        return jsonify({
            "success": True,
            "message": "Đồng bộ hóa đơn lên ERP thành công.",
            "erp_synced": invoice.erp_synced,
            "erp_sync_date": invoice.erp_sync_date
        })
    else:
        return jsonify({
            "success": False,
            "message": "Đồng bộ hóa đơn lên ERP thất bại.",
            "error": invoice.erp_sync_error
        }), 400


def classify_fct_item(item_name: str, seller_name: str) -> tuple[str, float, float]:
    """
    Classify transaction based on Circular 103/2014/TT-BTC for e-commerce and digital services.
    Returns: (Category, VAT_rate, CIT_rate)
    """
    name = (item_name or "").lower() + " " + (seller_name or "").lower()
    
    # 1. Digital Advertising (e.g. Google Ads, Meta Ads)
    if any(k in name for k in ["ads", "quảng cáo", "quang cao", "marketing", "facebook ads", "google ads", "adwords"]):
        return "Dịch vụ Quảng cáo trực tuyến (Online Advertising)", 0.05, 0.05
        
    # 2. Cloud & Hosting (e.g. AWS, Azure)
    if any(k in name for k in ["cloud", "hosting", "aws", "amazon web services", "azure", "server", "vps", "lưu trữ", "digitalocean"]):
        return "Dịch vụ Điện toán đám mây & Lưu trữ (Cloud & Hosting)", 0.05, 0.05
        
    # 3. Software License / SaaS (VAT Exempt under VN Tax Law, CIT 5%)
    if any(k in name for k in ["phần mềm", "phan mem", "software", "license", "bản quyền", "ban quyen", "zoom", "slack", "subscription"]):
        return "Bản quyền & Phần mềm SaaS (Software & License)", 0.00, 0.05
        
    # Default fallback (General Digital Services)
    return "Thương mại điện tử & Dịch vụ số khác", 0.05, 0.05


def generate_fct_excel(fct_data: dict) -> bytes:
    """Generate a highly polished Excel sheet conforming to Mẫu số 01/NTNN layout."""
    import openpyxl
    from openpyxl import Workbook
    from export.formatter import auto_adjust_column_widths, HEADER_FILL, MED_RISK_FILL, OK_FILL

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Tờ khai 01-NTNN"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.append(["TỜ KHAI THUẾ NHÀ THẦU NƯỚC NGOÀI (Mẫu số 01/NTNN)"])
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.font = openpyxl.styles.Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Period Info
    period_str = f"Tháng {fct_data['period_value']}" if fct_data['period_type'] == "monthly" else f"Quý {fct_data['period_value']}"
    ws.append([f"Kỳ kê khai: {period_str} năm {fct_data['year']}", "", "", "", "", "", "", "", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    ws.merge_cells("A2:H2")
    ws["A2"].font = openpyxl.styles.Font(italic=True, color="555555")
    ws["I2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws.append([])  # Blank row

    # Headers
    headers = [
        "STT",
        "Tên Nhà thầu nước ngoài",
        "Mã số thuế",
        "Nội dung dịch vụ",
        "Doanh thu tính thuế (₫)",
        "Tỷ lệ GTGT (%)",
        "Thuế GTGT phải nộp (₫)",
        "Tỷ lệ TNDN (%)",
        "Thuế TNDN phải nộp (₫)",
    ]
    ws.append(headers)
    
    # Format Headers
    for col_idx in range(1, 10):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Data Rows
    invoices = fct_data.get("fct_invoices", [])
    for idx, inv in enumerate(invoices, 1):
        row = [
            idx,
            inv["seller_name"],
            inv["seller_mst"],
            inv["category"],
            inv["amount"],
            f"{inv['vat_rate'] * 100}%",
            inv["vat_withheld"],
            f"{inv['cit_rate'] * 100}%",
            inv["cit_withheld"]
        ]
        ws.append(row)
        
        # Zebra styling
        row_cells = ws[ws.max_row]
        for cell in row_cells:
            cell.fill = OK_FILL if idx % 2 == 0 else openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.font = openpyxl.styles.Font(size=10)

    # Total Row
    ws.append([
        "TỔNG CỘNG", "", "", "",
        fct_data["total_revenue"],
        "",
        fct_data["total_vat_withheld"],
        "",
        fct_data["total_cit_withheld"]
    ])
    ws.merge_cells(f"A{ws.max_row}:D{ws.max_row}")
    
    for col_idx in range(1, 10):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True, size=11, color="1F4E78")
        cell.fill = MED_RISK_FILL
    ws.row_dimensions[ws.max_row].height = 24

    # Alignments & Formats
    for row in range(5, ws.max_row + 1):
        ws[f"A{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"C{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"F{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"H{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        
        for col_idx in [5, 7, 9]:
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '#,##0 "VND"'
            cell.alignment = openpyxl.styles.Alignment(horizontal="right")

    auto_adjust_column_widths(ws)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


@invoices_blueprint.get("/api/reports/fct-declaration")
def api_reports_fct_declaration():
    """
    Generate a draft of the Vietnamese Foreign Contractor Tax (FCT) Return Mẫu 01/NTNN.
    Identifies foreign e-commerce/digital giant suppliers and calculates withholding VAT/CIT.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice

        query = Invoice.query.filter(Invoice.is_cancelled == False, Invoice.invoice_type == "purchase")

        # Apply date filters
        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        fct_invoices = []
        total_revenue = 0.0
        total_vat_withheld = 0.0
        total_cit_withheld = 0.0

        for inv in invoices:
            seller_mst = (inv.seller_mst or "").strip()
            seller_name = (inv.seller_name or "").strip()
            
            # Match 900xxxxxxx or digital giants names
            is_fct = (
                seller_mst.startswith("900") or
                any(k in seller_name.lower() for k in ["google", "facebook", "meta", "amazon", "aws", "netflix", "zoom", "slack", "microsoft", "github", "digitalocean"])
            )
            
            if not is_fct:
                continue

            amount = inv.amount_before_tax or 0.0
            category = "Thương mại điện tử & Dịch vụ số khác"
            vat_rate = 0.05
            cit_rate = 0.05
            
            if inv.items and len(inv.items) > 0:
                first_item = inv.items[0].item_name
                category, vat_rate, cit_rate = classify_fct_item(first_item, seller_name)
            else:
                category, vat_rate, cit_rate = classify_fct_item("", seller_name)

            vat_withheld = amount * vat_rate
            cit_withheld = amount * cit_rate
            fct_total = vat_withheld + cit_withheld

            fct_invoices.append({
                "id": inv.id,
                "number": inv.number or "Không số",
                "date": inv.date or "Không ngày",
                "seller_name": seller_name,
                "seller_mst": seller_mst,
                "category": category,
                "amount": amount,
                "vat_rate": vat_rate,
                "vat_withheld": vat_withheld,
                "cit_rate": cit_rate,
                "cit_withheld": cit_withheld,
                "fct_total": fct_total
            })

            total_revenue += amount
            total_vat_withheld += vat_withheld
            total_cit_withheld += cit_withheld

        return jsonify({
            "success": True,
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "total_revenue": total_revenue,
            "total_vat_withheld": total_vat_withheld,
            "total_cit_withheld": total_cit_withheld,
            "total_fct_payable": total_vat_withheld + total_cit_withheld,
            "fct_invoices": fct_invoices
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/reports/fct-declaration/export-excel")
def api_reports_fct_declaration_export_excel():
    """Export the Mẫu 01/NTNN draft to a beautifully formatted Excel spreadsheet."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice
        query = Invoice.query.filter(Invoice.is_cancelled == False, Invoice.invoice_type == "purchase")

        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        fct_invoices = []
        total_revenue = 0.0
        total_vat_withheld = 0.0
        total_cit_withheld = 0.0

        for inv in invoices:
            seller_mst = (inv.seller_mst or "").strip()
            seller_name = (inv.seller_name or "").strip()
            
            is_fct = (
                seller_mst.startswith("900") or
                any(k in seller_name.lower() for k in ["google", "facebook", "meta", "amazon", "aws", "netflix", "zoom", "slack", "microsoft", "github", "digitalocean"])
            )
            
            if not is_fct:
                continue

            amount = inv.amount_before_tax or 0.0
            category = "Thương mại điện tử & Dịch vụ số khác"
            vat_rate = 0.05
            cit_rate = 0.05
            
            if inv.items and len(inv.items) > 0:
                first_item = inv.items[0].item_name
                category, vat_rate, cit_rate = classify_fct_item(first_item, seller_name)
            else:
                category, vat_rate, cit_rate = classify_fct_item("", seller_name)

            vat_withheld = amount * vat_rate
            cit_withheld = amount * cit_rate
            fct_total = vat_withheld + cit_withheld

            fct_invoices.append({
                "seller_name": seller_name,
                "seller_mst": seller_mst,
                "category": category,
                "amount": amount,
                "vat_rate": vat_rate,
                "vat_withheld": vat_withheld,
                "cit_rate": cit_rate,
                "cit_withheld": cit_withheld,
                "fct_total": fct_total
            })

            total_revenue += amount
            total_vat_withheld += vat_withheld
            total_cit_withheld += cit_withheld

        fct_data = {
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "total_revenue": total_revenue,
            "total_vat_withheld": total_vat_withheld,
            "total_cit_withheld": total_cit_withheld,
            "total_fct_payable": total_vat_withheld + total_cit_withheld,
            "fct_invoices": fct_invoices
        }

        excel_bytes = generate_fct_excel(fct_data)
        
        filename = f"ToKhai_01NTNN_{year}_{period_value}.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/reports/vat-refund-eligibility")
@roles_required("admin", "auditor")
def api_vat_refund_eligibility():
    """Calculates taxpayer eligibility for VAT refund and returns a structured breakdown."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    if not mst:
        # Fallback to the first taxpayer profile
        from invoices.models import TaxpayerProfile
        first_profile = TaxpayerProfile.query.first()
        if first_profile:
            mst = first_profile.mst
        else:
            return jsonify({"error": "Không có mã số thuế hoạt động hoặc hồ sơ doanh nghiệp nào được đăng ký."}), 400

    try:
        from invoices.refund_service import VATRefundEligibilityEngine
        engine = VATRefundEligibilityEngine()
        result = engine.get_eligibility(mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi tính toán hoàn thuế: {str(e)}"}), 500


@invoices_blueprint.post("/api/reports/vat-refund-eligibility/dossier")
@roles_required("admin", "auditor")
def api_vat_refund_dossier():
    """Generates Circular 80 Mẫu 01/HT refund dossier and AI justification letter."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    mst = payload.get("mst") or session.get("active_taxpayer_mst")
    
    if not mst:
        from invoices.models import TaxpayerProfile
        first_profile = TaxpayerProfile.query.first()
        if first_profile:
            mst = first_profile.mst
        else:
            return jsonify({"error": "Không có mã số thuế hoạt động hoặc hồ sơ doanh nghiệp nào được đăng ký."}), 400

    try:
        from invoices.refund_service import VATRefundEligibilityEngine
        engine = VATRefundEligibilityEngine()
        result = engine.generate_dossier(mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi soạn hồ sơ AI: {str(e)}"}), 500


@invoices_blueprint.post("/api/reports/vat-refund-eligibility/dossier/export")
@roles_required("admin", "auditor")
def api_export_vat_refund_dossier():
    """Exports the generated dossier or justification letter to Word (.doc) or PDF."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    content = payload.get("content", "").strip()
    export_format = payload.get("format", "doc").lower()  # 'doc' or 'pdf'
    document_type = payload.get("type", "dossier")  # 'dossier' or 'justification'

    if not content:
        return jsonify({"error": "Nội dung tài liệu trống."}), 400

    filename = "Mau_01_HT_De_Nghi_Hoan_Thue" if document_type == "dossier" else "Bao_Cao_Bien_Phap_Bao_Ve_Ho_So"

    if export_format == "pdf":
        html_content = f"""
        <html>
        <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: a4;
                margin: 2cm;
            }}
            body {{
                font-family: 'Arial', sans-serif;
                font-size: 11px;
                line-height: 1.5;
            }}
            .bold {{ font-weight: bold; }}
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
            .title {{ font-size: 13px; font-weight: bold; text-align: center; margin-top: 15px; margin-bottom: 15px; }}
            p {{ margin-bottom: 6px; text-align: justify; }}
            pre {{
                font-family: 'Arial', sans-serif;
                white-space: pre-wrap;
                word-wrap: break-word;
            }}
        </style>
        </head>
        <body>
        <pre>{content}</pre>
        </body>
        </html>
        """
        try:
            pdf_buf = render_html_to_pdf(html_content)
            return send_file(
                pdf_buf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"{filename}.pdf"
            )
        except Exception as e:
            return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

    else:
        # Default to Word compatible HTML (.doc)
        html_content = f"""
        <html xmlns:o="urn:schemas-microsoft-com:office:office"
              xmlns:w="urn:schemas-microsoft-com:office:word"
              xmlns="http://www.w3.org/TR/REC-html40">
        <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: 8.27in 11.69in; /* A4 */
                margin: 1.0in 0.79in 1.0in 1.18in;
            }}
            body {{
                font-family: 'Times New Roman', serif;
                font-size: 11pt;
                line-height: 1.5;
            }}
            pre {{
                font-family: 'Times New Roman', serif;
                white-space: pre-wrap;
                word-wrap: break-word;
            }}
        </style>
        </head>
        <body>
        <pre>{content}</pre>
        </body>
        </html>
        """
        buf = BytesIO(html_content.encode("utf-8"))
        return send_file(
            buf,
            mimetype="application/msword",
            as_attachment=True,
            download_name=f"{filename}.doc"
        )


@invoices_blueprint.post("/api/bank/reconcile/upload")
def api_bank_reconcile_upload():
    """Ingests a Techcombank/Vietcombank Excel statement and stores transactions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = request.form.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    bank_name = request.form.get("bank_name", "Generic")
    account_number = request.form.get("account_number", "")

    if not taxpayer_mst:
        return jsonify({"error": "Mã số thuế hoạt động trống."}), 400

    if "file" not in request.files:
        return jsonify({"error": "Không có tệp tải lên."}), 400

    uploaded_file = request.files["file"]
    if not uploaded_file.filename:
        return jsonify({"error": "Tên tệp không hợp lệ."}), 400

    # Save to a local temporary location in our designated temp dir
    temp_dir = os.path.join(current_app.root_path, "data", "temp")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, f"statement_{int(datetime.now().timestamp())}.xlsx")
    uploaded_file.save(temp_path)

    try:
        from invoices.bank_reconcile_service import parse_bank_statement
        parsed_txs = parse_bank_statement(temp_path, bank_name)
        
        from invoices.models import BankTransaction
        imported_count = 0
        skipped_count = 0
        
        for p in parsed_txs:
            # Check for reference duplicate
            exists = BankTransaction.query.filter_by(id=p["id"]).first()
            if exists:
                skipped_count += 1
                continue
                
            tx = BankTransaction(
                id=p["id"],
                taxpayer_mst=taxpayer_mst,
                bank_name=p["bank_name"],
                account_number=account_number,
                transaction_date=p["transaction_date"],
                reference_number=p["reference_number"],
                description=p["description"],
                amount=p["amount"],
                status="unreconciled",
                imported_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.session.add(tx)
            imported_count += 1
            
        if imported_count > 0:
            db.session.commit()
            
        return jsonify({
            "status": "success",
            "imported_count": imported_count,
            "skipped_count": skipped_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi xử lý tệp sổ phụ: {str(e)}"}), 500
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


@invoices_blueprint.get("/api/bank/reconcile/transactions")
def api_bank_reconcile_transactions():
    """Retrieve bank transactions list for active MST."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    status_filter = request.args.get("status", "all")  # 'all', 'unreconciled', 'matched'

    if not taxpayer_mst:
        return jsonify([])

    from invoices.models import BankTransaction, Invoice
    query = BankTransaction.query.filter_by(taxpayer_mst=taxpayer_mst)
    if status_filter != "all":
        query = query.filter_by(status=status_filter)

    transactions = query.order_by(BankTransaction.transaction_date.desc()).all()
    
    result = []
    for tx in transactions:
        tx_dict = tx.to_dict()
        if tx.matched_invoice_id:
            inv = Invoice.query.get(tx.matched_invoice_id)
            if inv:
                tx_dict["invoice_number"] = inv.number
                tx_dict["partner_name"] = inv.buyer_name if tx.amount > 0 else inv.seller_name
        else:
            tx_dict["invoice_number"] = ""
            tx_dict["partner_name"] = ""
        result.append(tx_dict)
        
    return jsonify(result)


@invoices_blueprint.post("/api/bank/reconcile/auto")
def api_bank_reconcile_auto():
    """Triggers autonomous Soundex/Phonetic matching reconciliation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    taxpayer_mst = payload.get("taxpayer_mst") or session.get("active_taxpayer_mst")

    if not taxpayer_mst:
        return jsonify({"error": "Mã số thuế hoạt động trống."}), 400

    try:
        from invoices.bank_reconcile_service import execute_auto_reconciliation
        res = execute_auto_reconciliation(taxpayer_mst)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": f"Lỗi đối chiếu tự động: {str(e)}"}), 500


@invoices_blueprint.post("/api/bank/reconcile/manual")
def api_bank_reconcile_manual():
    """Manual reconciliation override by ledger accountant."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    transaction_id = payload.get("transaction_id")
    invoice_id = payload.get("invoice_id")

    if not transaction_id or not invoice_id:
        return jsonify({"error": "Thiếu mã giao dịch hoặc mã hóa đơn khớp."}), 400

    from invoices.models import BankTransaction, Invoice
    tx = BankTransaction.query.get(transaction_id)
    inv = Invoice.query.get(invoice_id)

    if not tx or not inv:
        return jsonify({"error": "Không tìm thấy giao dịch ngân hàng hoặc hóa đơn tương ứng."}), 404

    try:
        tx.matched_invoice_id = invoice_id
        tx.confidence_score = 1.0  # Manual matching has 100% confidence
        tx.status = "matched"
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Đối chiếu thủ công thành công.",
            "details": {
                "transaction_id": tx.id,
                "matched_invoice_id": invoice_id,
                "invoice_number": inv.number,
                "partner_name": inv.buyer_name if tx.amount > 0 else inv.seller_name
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi đối chiếu thủ công: {str(e)}"}), 500


# ---------------------------------------------------------------------------
# US-084: Compliant E-Invoice Generator & Digital Signing Bridge (US-084, US-085)
# ---------------------------------------------------------------------------

@invoices_blueprint.get("/issue-invoice")
def issue_invoice_page():
    """Render the e-invoice draft builder page."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("issue_invoice.html")


@invoices_blueprint.post("/api/invoices/issue/draft")
def api_issue_draft():
    """Create a new e-invoice draft."""
    import os
    from datetime import datetime

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    buyer_mst = payload.get("buyer_mst", "").strip()
    buyer_name = payload.get("buyer_name", "").strip()
    buyer_address = payload.get("buyer_address", "").strip()
    items_data = payload.get("items") or []

    if not buyer_mst or not buyer_name:
        return jsonify({"error": "Thiếu mã số thuế hoặc tên đơn vị mua hàng."}), 400

    if not items_data:
        return jsonify({"error": "Danh sách hàng hóa dịch vụ không được trống."}), 400

    # Get active taxpayer
    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        return jsonify({"error": "Vui lòng chọn doanh nghiệp hoạt động trước."}), 400

    from invoices.models import TaxpayerProfile, Invoice, LineItem
    seller = TaxpayerProfile.query.filter_by(mst=active_mst).first()
    seller_name = seller.company_name if seller else "Công ty Phát hành Mẫu"
    seller_address = "123 Đường Phát Hành, Hà Nội"

    # Auto-increment number
    symbol = payload.get("symbol", "1C26TYY").strip()
    
    # Query count to auto-increment number
    count = Invoice.query.filter(Invoice.seller_mst == active_mst, Invoice.symbol == symbol).count()
    number = f"{count + 1:07d}"
    
    invoice_id = f"{active_mst}-{symbol}-{number}"

    # Calculate totals
    amount_before_tax = 0.0
    tax_amount = 0.0
    
    line_items = []
    
    for idx, item in enumerate(items_data):
        name = item.get("item_name", "").strip()
        unit = item.get("unit", "").strip()
        try:
            qty = float(item.get("quantity") or 0.0)
            price = float(item.get("unit_price") or 0.0)
        except ValueError:
            return jsonify({"error": f"Số lượng hoặc đơn giá của mục {idx+1} không hợp lệ."}), 400
            
        tax_rate_str = item.get("tax_rate", "10%")
        
        # Calculate item totals
        item_amt = qty * price
        
        # Calculate tax
        if "10" in tax_rate_str:
            item_tax = 0.10 * item_amt
        elif "8" in tax_rate_str:
            item_tax = 0.08 * item_amt
        elif "5" in tax_rate_str:
            item_tax = 0.05 * item_amt
        else:
            item_tax = 0.0
            
        amount_before_tax += item_amt
        tax_amount += item_tax
        
        line_items.append({
            "item_name": name,
            "unit": unit,
            "quantity": qty,
            "unit_price": price,
            "amount_before_tax": item_amt,
            "tax_rate": tax_rate_str,
            "tax_amount": item_tax
        })

    total_amount = amount_before_tax + tax_amount
    
    # Spell money in Vietnamese
    from invoices.ai_service import spell_money_vietnamese
    amount_in_words = spell_money_vietnamese(total_amount)

    try:
        inv = Invoice(
            id=invoice_id,
            filename=f"invoice_{invoice_id}.xml",
            invoice_type="sold",
            template_code="1",
            symbol=symbol,
            number=number,
            date=datetime.now().strftime("%Y-%m-%d"),
            currency="VND",
            seller_name=seller_name,
            seller_mst=active_mst,
            seller_address=seller_address,
            buyer_name=buyer_name,
            buyer_mst=buyer_mst,
            buyer_address=buyer_address,
            amount_before_tax=amount_before_tax,
            tax_amount=tax_amount,
            total_amount=total_amount,
            has_signature=False,
            amount_in_words=amount_in_words,
            imported_at=datetime.now().isoformat(),
            import_status="draft",
            invoice_status="draft",
            taxpayer_mst=active_mst
        )
        db.session.add(inv)
        
        for item_data in line_items:
            li = LineItem(
                invoice_id=invoice_id,
                item_name=item_data["item_name"],
                unit=item_data["unit"],
                quantity=item_data["quantity"],
                unit_price=item_data["unit_price"],
                amount_before_tax=item_data["amount_before_tax"],
                tax_rate=item_data["tax_rate"],
                tax_amount=item_data["tax_amount"],
                expense_category="REVENUE"
            )
            db.session.add(li)
            
        db.session.commit()
        return jsonify({
            "status": "success",
            "message": "Tạo hóa đơn nháp thành công.",
            "invoice": inv.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi lưu hóa đơn nháp: {str(e)}"}), 500


@invoices_blueprint.post("/api/invoices/issue/sign")
def api_issue_sign():
    """Digital sign draft e-invoice using mock USB Token."""
    import os
    import json
    from datetime import datetime

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_id = payload.get("invoice_id")

    if not invoice_id:
        return jsonify({"error": "Thiếu mã hóa đơn cần ký."}), 400

    from invoices.models import Invoice, LineItem
    inv = Invoice.query.get(invoice_id)
    if not inv:
        return jsonify({"error": "Không tìm thấy hóa đơn tương ứng."}), 404

    if inv.invoice_status != "draft":
        return jsonify({"error": "Hóa đơn này đã được phát hành và ký số."}), 400

    try:
        # Build GDT Circular 78 Compliant XML DSHDon list
        items_xml = ""
        for idx, item in enumerate(inv.items):
            items_xml += f"""        <HHDVu>
          <TChat>1</TChat>
          <STT>{idx + 1}</STT>
          <Ten>{item.item_name}</Ten>
          <DVT>{item.unit or 'Lần'}</DVT>
          <SLuong>{item.quantity}</SLuong>
          <DGia>{item.unit_price}</DGia>
          <ThTien>{item.amount_before_tax}</ThTien>
          <TSuat>{item.tax_rate}</TSuat>
          <TThue>{item.tax_amount}</TThue>
        </HHDVu>"""

        # Compile Canonical DLHDon XML
        dlhdon_xml = f"""<DLHDon Id="HD_{inv.id}">
      <TTChung>
        <PBan>1.0.0</PBan>
        <THDon>Hóa đơn giá trị gia tăng</THDon>
        <KHHDon>{inv.symbol}</KHHDon>
        <SHDon>{inv.number}</SHDon>
        <NLap>{inv.date}</NLap>
        <DVTTe>{inv.currency or 'VND'}</DVTTe>
        <TGia>1.0</TGia>
      </TTChung>
      <NDHDon>
        <NBan>
          <Ten>{inv.seller_name}</Ten>
          <MST>{inv.seller_mst}</MST>
          <DChi>{inv.seller_address or ''}</DChi>
        </NBan>
        <NMua>
          <Ten>{inv.buyer_name}</Ten>
          <MST>{inv.buyer_mst}</MST>
          <DChi>{inv.buyer_address or ''}</DChi>
        </NMua>
        <DSHDon>
{items_xml}
        </DSHDon>
        <TToan>
          <TgTCThue>{inv.amount_before_tax}</TgTCThue>
          <TgTThue>{inv.tax_amount}</TgTThue>
          <TgTTTBSo>{inv.total_amount}</TgTTTBSo>
          <TgTTTBChu>{inv.amount_in_words}</TgTTTBChu>
        </TToan>
      </NDHDon>
    </DLHDon>"""

        # Perform SHA-256 + RSA-2048 mock USB token cryptographic signing
        import hashlib
        import base64
        
        # Calculate digest
        digest = hashlib.sha256(dlhdon_xml.encode("utf-8")).digest()
        digest_b64 = base64.b64encode(digest).decode("utf-8")
        
        # Simulate USB Token RSA signature
        sig_b64 = base64.b64encode(hashlib.sha256(digest).digest() * 2).decode("utf-8")[:172] + "=="
        
        # Mock certificate x509
        mock_cert = "MIIDuTCCAqGgAwIBAgIUdT6ySjZ+N...MOCK_GDT_CIRCULAR_78_CERTIFICATE..."
        
        # Complete Circular 78 XML Package
        full_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<HDon>
    {dlhdon_xml}
    <Signature xmlns="http://www.w3.org/2000/09/xmldsig#">
        <SignedInfo>
            <CanonicalizationMethod Algorithm="http://www.w3.org/TR/2001/REC-xml-c14n-20010315"/>
            <SignatureMethod Algorithm="http://www.w3.org/2000/09/xmldsig#rsa-sha256"/>
            <Reference URI="#HD_{inv.id}">
                <Transforms>
                    <Transform Algorithm="http://www.w3.org/2000/09/xmldsig#enveloped-signature"/>
                </Transforms>
                <DigestMethod Algorithm="http://www.w3.org/2001/04/xmlenc#sha256"/>
                <DigestValue>{digest_b64}</DigestValue>
            </Reference>
        </SignedInfo>
        <SignatureValue>{sig_b64}</SignatureValue>
        <KeyInfo>
            <X509Data>
                <X509Certificate>{mock_cert}</X509Certificate>
            </X509Data>
        </KeyInfo>
    </Signature>
</HDon>"""

        # Store signed XML file locally
        from invoices.service import XML_DIR
        safe_filename = f"invoice_{inv.id}.xml"
        xml_path = os.path.join(XML_DIR, safe_filename)
        with open(xml_path, "w", encoding="utf-8") as f:
            f.write(full_xml)

        # Update database model state
        inv.has_signature = True
        inv.signing_date = datetime.now().strftime("%Y-%m-%d")
        inv.import_status = "imported"
        inv.invoice_status = "Gốc"
        
        # Also generate mock signature JSON for frontend display
        inv.signature_details_json = json.dumps({
            "subject": f"C=VN, ST=Hanoi, O={inv.seller_name}, CN={inv.seller_name}",
            "issuer": "VNPT CA / GDT Root CA",
            "valid_from": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "valid_to": "2029-12-31 23:59:59",
            "serial": "18392098487293847293847",
            "algo": "sha256RSA"
        }, ensure_ascii=False)
        
        db.session.commit()
        
        # Trigger an SSE update stream event for newly issued invoice
        try:
            from invoices.sync_daemon import push_sync_event
            push_sync_event("invoice_downloaded", {
                "id": inv.id,
                "seller": inv.seller_name,
                "buyer": inv.buyer_name,
                "amount": inv.total_amount
            })
        except Exception:
            pass

        return jsonify({
            "status": "success",
            "message": "Ký số hóa đơn thành công thông qua USB Token.",
            "invoice_id": inv.id,
            "xml_preview": full_xml[:1000] + "..."
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi ký số hóa đơn: {str(e)}"}), 500


# ── Version 6.0.0: Cryptographic Audit Ledger API ─────────────────

@invoices_blueprint.get("/api/audit/ledger")
def get_audit_ledger():
    """List audit ledger blocks with pagination (US-090)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.models import AuditBlock

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    per_page = min(per_page, 200)

    query = AuditBlock.query.order_by(AuditBlock.block_id.desc())
    total = query.count()
    blocks = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "blocks": [b.to_dict() for b in blocks],
    })


@invoices_blueprint.post("/api/audit/verify")
def verify_audit_ledger():
    """Run full-chain cryptographic integrity verification (US-091)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.audit_ledger_service import verify_ledger_integrity
    from invoices.models import AuditBlock

    is_valid, corrupted_id, error_msg = verify_ledger_integrity()
    total_blocks = AuditBlock.query.count()

    return jsonify({
        "is_valid": is_valid,
        "total_blocks": total_blocks,
        "corrupted_block_id": corrupted_id,
        "error_message": error_msg,
    })


@invoices_blueprint.get("/api/audit/stats")
def get_audit_stats():
    """Return summary statistics for the audit ledger dashboard (US-091)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.models import AuditBlock
    from sqlalchemy import func

    total = AuditBlock.query.count()
    action_counts = (
        db.session.query(AuditBlock.action_type, func.count(AuditBlock.block_id))
        .group_by(AuditBlock.action_type)
        .all()
    )

    latest = AuditBlock.query.order_by(AuditBlock.block_id.desc()).first()

    return jsonify({
        "total_blocks": total,
        "action_breakdown": {action: count for action, count in action_counts},
        "latest_block": latest.to_dict() if latest else None,
    })


@invoices_blueprint.post("/api/analytics/forecast")
@roles_required("admin", "auditor", "viewer")
def api_forecast_tax():
    """Forecast future tax liability using moving averages (US-110, US-111)."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        historical_data = body.get("historical_data")
        projected_period = body.get("projected_period", datetime.now().strftime("%Y-%m"))
        alpha = body.get("alpha", 0.7)
        window_size = body.get("window_size", 3)
        budget_limit = body.get("budget_limit", 500000000.0)

        active_mst = session.get("active_taxpayer_mst")
        if not active_mst:
            from invoices.models import TaxpayerProfile
            prof = TaxpayerProfile.query.filter_by(is_active=True).first()
            if prof:
                active_mst = prof.mst

        # Query from DB if not provided in request body
        if historical_data is None:
            if not active_mst:
                return jsonify({"error": "Không tìm thấy mã số thuế hoạt động để truy vấn dữ liệu."}), 400

            from invoices.models import Invoice
            sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
            purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

            from collections import defaultdict
            period_map = defaultdict(lambda: {"output_vat": 0.0, "input_vat": 0.0})
            
            for s in sales:
                if not s.date or len(s.date) < 7:
                    continue
                period_map[s.date[:7]]["output_vat"] += s.tax_amount
            for p in purchases:
                if not p.date or len(p.date) < 7:
                    continue
                period_map[p.date[:7]]["input_vat"] += p.tax_amount

            historical_data = []
            for p in sorted(period_map.keys()):
                historical_data.append({
                    "period": p,
                    "output_vat": period_map[p]["output_vat"],
                    "input_vat": period_map[p]["input_vat"]
                })

        from invoices.tax_forecaster import forecast_next_period_tax, TaxAlertManager
        forecast = forecast_next_period_tax(
            historical_data,
            projected_period=projected_period,
            alpha=alpha,
            window_size=window_size
        )

        # Run alerts evaluation
        alert_manager = TaxAlertManager(budget_limit=budget_limit)
        forecast_evaluated = alert_manager.evaluate_forecast(forecast)

        return jsonify({
            "taxpayer_mst": active_mst,
            "forecast": forecast_evaluated.to_dict()
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/invoices/batch-parse")
@roles_required("admin", "auditor")
def api_batch_parse_invoices():
    """Concurrently parse XML invoices, decompress if zipped, and import to DB (US-112, US-113)."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        invoice_items = body.get("invoices", [])
        duplicate_strategy = body.get("duplicate_strategy", "overwrite")
        active_mst = session.get("active_taxpayer_mst")

        # 1. Prepare batch (decompress if needed)
        parsed_batch_inputs = []
        for idx, item in enumerate(invoice_items):
            filename = item.get("filename", f"invoice_{idx}.xml")
            content = item.get("content", "")
            compressed = item.get("compressed", False)

            if not content:
                continue

            try:
                if compressed:
                    import base64
                    try:
                        byte_data = base64.b64decode(content)
                    except Exception:
                        byte_data = bytes.fromhex(content)
                    
                    from invoices.batch_parser import decompress_xml
                    xml_str = decompress_xml(byte_data)
                else:
                    xml_str = content

                parsed_batch_inputs.append((filename, xml_str))
            except Exception as e:
                pass

        # 2. Parallel Parse
        from invoices.batch_parser import parse_batch_xml
        parse_results = parse_batch_xml(parsed_batch_inputs)

        # 3. Serial Import to DB
        from invoices.service import import_xml_invoice
        db_results = []
        for res, (filename, xml_str) in zip(parse_results, parsed_batch_inputs):
            if not res.success:
                db_results.append({
                    "filename": filename,
                    "success": False,
                    "error_message": res.error_message
                })
                continue
            
            try:
                xml_bytes = xml_str.encode("utf-8")
                imported_dict = import_xml_invoice(
                    xml_bytes,
                    filename,
                    duplicate_strategy=duplicate_strategy,
                    taxpayer_mst=active_mst
                )
                db_results.append({
                    "filename": filename,
                    "success": True,
                    "invoice_id": imported_dict.get("id"),
                    "invoice_number": imported_dict.get("number"),
                    "total_amount": imported_dict.get("total_amount")
                })
            except Exception as e:
                db_results.append({
                    "filename": filename,
                    "success": False,
                    "error_message": str(e)
                })

        return jsonify({
            "total_processed": len(db_results),
            "results": db_results
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/analytics/kpis")
@roles_required("admin", "auditor", "viewer")
def api_get_financial_kpis():
    """Retrieve financial health metrics (Gross Margin, Tax Ratios, Clearance times) (US-114)."""
    err = _ensure_logged_in()
    if err:
        return err

    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        from invoices.models import TaxpayerProfile
        prof = TaxpayerProfile.query.filter_by(is_active=True).first()
        if prof:
            active_mst = prof.mst

    if not active_mst:
        return jsonify({"error": "Không tìm thấy mã số thuế hoạt động. Vui lòng chọn hồ sơ MST."}), 400

    from invoices.models import Invoice
    sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
    purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

    sales_dicts = [
        {"id": s.id, "amount_before_tax": s.amount_before_tax, "tax_amount": s.tax_amount, "date": s.date}
        for s in sales
    ]
    purchases_dicts = [
        {"id": p.id, "amount_before_tax": p.amount_before_tax, "tax_amount": p.tax_amount, "date": p.date}
        for p in purchases
    ]

    clearances = []
    for inv in sales + purchases:
        if inv.paid_date:
            clearances.append({
                "invoice_id": inv.id,
                "clearance_date": inv.paid_date
            })

    from invoices.financial_kpi import calculate_financial_kpis
    kpi = calculate_financial_kpis(sales_dicts, purchases_dicts, clearances)

    from collections import defaultdict
    sales_by_month = defaultdict(list)
    purchases_by_month = defaultdict(list)
    clearances_by_month = defaultdict(list)

    for s in sales_dicts:
        month = s["date"][:7] if s.get("date") else "unknown"
        sales_by_month[month].append(s)
    for p in purchases_dicts:
        month = p["date"][:7] if p.get("date") else "unknown"
        purchases_by_month[month].append(p)
    for c in clearances:
        inv_date = None
        for s in sales_dicts:
            if s["id"] == c["invoice_id"]:
                inv_date = s["date"]
                break
        if not inv_date:
            for p in purchases_dicts:
                if p["id"] == c["invoice_id"]:
                    inv_date = p["date"]
                    break
        month = inv_date[:7] if inv_date else "unknown"
        clearances_by_month[month].append(c)

    all_months = set(sales_by_month.keys()).union(purchases_by_month.keys())
    all_months.discard("unknown")
    
    monthly_trends = {}
    for month in sorted(all_months):
        m_kpi = calculate_financial_kpis(
            sales_by_month[month],
            purchases_by_month[month],
            clearances_by_month[month]
        )
        monthly_trends[month] = m_kpi.to_dict()

    return jsonify({
        "taxpayer_mst": active_mst,
        "overall": kpi.to_dict(),
        "monthly_trends": monthly_trends
    })


@invoices_blueprint.get("/api/analytics/kpis/export")
@roles_required("admin", "auditor", "viewer")
def api_export_financial_kpis():
    """Export monthly financial KPIs to a downloadable CSV file (US-115)."""
    err = _ensure_logged_in()
    if err:
        return err

    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        from invoices.models import TaxpayerProfile
        prof = TaxpayerProfile.query.filter_by(is_active=True).first()
        if prof:
            active_mst = prof.mst

    if not active_mst:
        return jsonify({"error": "Không tìm thấy mã số thuế hoạt động. Vui lòng chọn hồ sơ MST."}), 400

    from invoices.models import Invoice
    sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
    purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

    sales_dicts = [
        {"id": s.id, "amount_before_tax": s.amount_before_tax, "tax_amount": s.tax_amount, "date": s.date}
        for s in sales
    ]
    purchases_dicts = [
        {"id": p.id, "amount_before_tax": p.amount_before_tax, "tax_amount": p.tax_amount, "date": p.date}
        for p in purchases
    ]

    clearances = []
    for inv in sales + purchases:
        if inv.paid_date:
            clearances.append({
                "invoice_id": inv.id,
                "clearance_date": inv.paid_date
            })

    from collections import defaultdict
    sales_by_month = defaultdict(list)
    purchases_by_month = defaultdict(list)
    clearances_by_month = defaultdict(list)

    for s in sales_dicts:
        month = s["date"][:7] if s.get("date") else "unknown"
        sales_by_month[month].append(s)
    for p in purchases_dicts:
        month = p["date"][:7] if p.get("date") else "unknown"
        purchases_by_month[month].append(p)
    for c in clearances:
        inv_date = None
        for s in sales_dicts:
            if s["id"] == c["invoice_id"]:
                inv_date = s["date"]
                break
        if not inv_date:
            for p in purchases_dicts:
                if p["id"] == c["invoice_id"]:
                    inv_date = p["date"]
                    break
        month = inv_date[:7] if inv_date else "unknown"
        clearances_by_month[month].append(c)

    all_months = set(sales_by_month.keys()).union(purchases_by_month.keys())
    all_months.discard("unknown")
    
    from invoices.financial_kpi import calculate_financial_kpis, export_kpi_to_csv
    
    period_metrics = {}
    for month in sorted(all_months):
        period_metrics[month] = calculate_financial_kpis(
            sales_by_month[month],
            purchases_by_month[month],
            clearances_by_month[month]
        )

    csv_content = export_kpi_to_csv(period_metrics)

    from flask import Response
    response = Response(csv_content, mimetype="text/csv")
    filename = f"kpi_report_{active_mst}_{datetime.now().strftime('%Y%m%d')}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@invoices_blueprint.get("/api/compliance/rulebook")
@roles_required("admin", "auditor", "viewer")
def api_get_compliance_rulebook():
    """Retrieve the currently active compliance rulebook (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import json
    from invoices.models import ComplianceRulebook
    active_mst = session.get("taxpayer_mst")
    
    rulebook = None
    if active_mst:
        rulebook = ComplianceRulebook.query.filter_by(taxpayer_mst=active_mst, is_active=True).first()
    
    if not rulebook:
        rulebook = ComplianceRulebook.query.filter_by(id="rulebook_default").first()

    if not rulebook:
        default_rulebook_json = {
            "name": "Default Compliance Rulebook",
            "rules": [
                {
                    "id": "rule_cash_limit",
                    "name": "Verify cash transactions over 20M limit",
                    "severity": "critical",
                    "channels": ["in_app"],
                    "expression": {
                        "and": [
                            {"field": "payment_method", "op": "contains", "value": "Tiền mặt"},
                            {"field": "total_amount", "op": ">=", "value": 20000000}
                        ]
                    }
                }
            ]
        }
        return jsonify({
            "status": "success",
            "rulebook": default_rulebook_json
        })

    try:
        data = json.loads(rulebook.rulebook_json)
    except Exception:
        data = {}

    return jsonify({
        "status": "success",
        "rulebook": data,
        "updated_at": rulebook.updated_at
    })


@invoices_blueprint.post("/api/compliance/rulebook")
@roles_required("admin", "auditor")
def api_update_compliance_rulebook():
    """Update or upload the active compliance rulebook DSL (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    rulebook_data = payload.get("rulebook")
    if not rulebook_data:
        return jsonify({"error": "Dữ liệu rulebook trống."}), 400

    import json
    from invoices.compliance_hub import validate_rulebook_dsl
    ok, err = validate_rulebook_dsl(rulebook_data)
    if not ok:
        return jsonify({"error": f"Lỗi cú pháp DSL Rulebook: {err}"}), 400

    from invoices.models import ComplianceRulebook
    active_mst = session.get("taxpayer_mst")
    rulebook_id = f"rulebook_{active_mst}" if active_mst else "rulebook_default"
    
    rulebook = db.session.get(ComplianceRulebook, rulebook_id)
    now = datetime.now().isoformat()
    
    if not rulebook:
        rulebook = ComplianceRulebook(
            id=rulebook_id,
            taxpayer_mst=active_mst,
            name=rulebook_data.get("name", "Custom Rulebook"),
            rulebook_json=json.dumps(rulebook_data, ensure_ascii=False),
            is_active=True,
            updated_at=now
        )
        db.session.add(rulebook)
    else:
        rulebook.name = rulebook_data.get("name", rulebook.name)
        rulebook.rulebook_json = json.dumps(rulebook_data, ensure_ascii=False)
        rulebook.updated_at = now

    db.session.commit()

    from invoices.security_audit_service import log_security_event
    log_security_event("UPDATE", f"Updated compliance rulebook DSL: {rulebook.name}")

    return jsonify({
        "status": "success",
        "message": "Cập nhật DSL Rulebook thành công.",
        "updated_at": now
    })


@invoices_blueprint.post("/api/compliance/evaluate")
@roles_required("admin", "auditor", "viewer")
def api_evaluate_compliance():
    """Evaluate compliance of specified invoices against the active rulebook (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_ids = payload.get("invoice_ids", [])
    
    if not invoice_ids:
        return jsonify({"error": "Danh sách invoice_ids trống."}), 400

    from invoices.models import Invoice, ComplianceRulebook
    from invoices.compliance_hub import ComplianceEngine
    import json

    active_mst = session.get("taxpayer_mst")
    rulebook = None
    if active_mst:
        rulebook = ComplianceRulebook.query.filter_by(taxpayer_mst=active_mst, is_active=True).first()
    if not rulebook:
        rulebook = ComplianceRulebook.query.filter_by(id="rulebook_default").first()

    engine = ComplianceEngine()
    if rulebook:
        try:
            rulebook_data = json.loads(rulebook.rulebook_json)
            engine.set_rulebook(rulebook_data)
        except Exception:
            pass

    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
    all_alerts = []
    
    for inv in invoices:
        # Convert invoice model to dictionary format suited for ComplianceEngine
        inv_dict = inv.to_dict()
        alerts = engine.evaluate_invoice(inv_dict)
        all_alerts.extend([a.to_dict() for a in alerts])

    return jsonify({
        "status": "success",
        "alerts": all_alerts
    })


@invoices_blueprint.post("/api/compliance/map-ifrs")
@roles_required("admin", "auditor", "viewer")
def api_map_ifrs_compliance():
    """Map invoices to standard IFRS & calculate FCT liabilities dynamically (US-121)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_ids = payload.get("invoice_ids", [])
    reporting_currency = payload.get("reporting_currency", "USD").strip().upper()
    fct_category = payload.get("fct_category", "services").strip().lower()

    if not invoice_ids:
        return jsonify({"error": "Danh sách invoice_ids trống."}), 400

    from invoices.models import Invoice
    from invoices.tax_mapping import TaxMappingEngine
    from dataclasses import asdict

    engine = TaxMappingEngine()
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
    
    mapped_results = []
    for inv in invoices:
        inv_dict = inv.to_dict()
        mapping = engine.map_to_ifrs(inv_dict, reporting_currency=reporting_currency, fct_category=fct_category)
        mapped_results.append(asdict(mapping))

    return jsonify({
        "status": "success",
        "reporting_currency": reporting_currency,
        "mapped_invoices": mapped_results
    })


@invoices_blueprint.get("/api/reports/tax-risk-scoreboard")
@roles_required("admin", "auditor", "viewer")
def api_tax_risk_scoreboard():
    """Retrieve tax compliance audit warning distribution and supplier risk scoreboard."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import json
    from invoices.service import get_local_invoices
    from invoices.models import BlacklistedMST

    active_mst = session.get("taxpayer_mst")
    
    # 1. Fetch all local invoices for the current taxpayer
    invoices = get_local_invoices(active_mst)
    
    # 2. Get all blacklisted MSTs
    try:
        blacklisted_msts = {b.mst for b in BlacklistedMST.query.all()}
    except Exception:
        blacklisted_msts = set()

    total_analyzed = len(invoices)
    total_with_warnings = 0
    total_value_at_risk = 0.0

    blacklist_warnings_count = 0
    signature_violations_count = 0
    payment_type_flags_count = 0

    supplier_groups = {}

    for inv in invoices:
        warnings = inv.get("warnings") or []
        # Support both a string list or a json list
        if isinstance(warnings, str):
            try:
                warnings = json.loads(warnings)
            except Exception:
                warnings = []
        
        has_warning = False
        is_blacklisted_inv = False
        is_sig_violation = False
        is_payment_flag = False

        seller_mst = inv.get("seller_mst") or ""
        seller_name = inv.get("seller_name") or "Không rõ"
        total_amount = inv.get("total_amount") or 0.0
        payment_method = inv.get("payment_method") or ""
        has_signature = inv.get("has_signature", True)

        # A. Blacklist check
        if seller_mst in blacklisted_msts:
            is_blacklisted_inv = True
        else:
            for w in warnings:
                w_lower = str(w).lower()
                if "blacklist" in w_lower or "đen" in w_lower or "rủi ro" in w_lower:
                    is_blacklisted_inv = True
                    break

        # B. Signature violation check
        if not has_signature:
            is_sig_violation = True
        else:
            for w in warnings:
                w_lower = str(w).lower()
                if "signature" in w_lower or "chữ ký" in w_lower or "ký số" in w_lower:
                    is_sig_violation = True
                    break

        # C. Payment type flag check
        pay_method_upper = payment_method.upper()
        is_cash = any(x in pay_method_upper for x in ["TM", "TIỀN MẶT", "CASH"])
        if total_amount >= 20000000.0 and is_cash:
            is_payment_flag = True
        else:
            for w in warnings:
                w_lower = str(w).lower()
                if any(x in w_lower for x in ["tiền mặt", "phương thức thanh toán", "hạch toán", "không dùng tiền mặt"]):
                    is_payment_flag = True
                    break

        if is_blacklisted_inv or is_sig_violation or is_payment_flag or len(warnings) > 0:
            has_warning = True

        if is_blacklisted_inv:
            blacklist_warnings_count += 1
        if is_sig_violation:
            signature_violations_count += 1
        if is_payment_flag:
            payment_type_flags_count += 1

        if has_warning:
            total_with_warnings += 1
            total_value_at_risk += total_amount

        # Group by supplier
        if seller_mst:
            if seller_mst not in supplier_groups:
                supplier_groups[seller_mst] = {
                    "supplier_mst": seller_mst,
                    "supplier_name": seller_name,
                    "invoice_count": 0,
                    "warnings_count": 0,
                    "total_value": 0.0,
                    "total_t_score": 0,
                    "is_blacklisted": seller_mst in blacklisted_msts
                }
            
            supplier_groups[seller_mst]["invoice_count"] += 1
            supplier_groups[seller_mst]["total_value"] += total_amount
            supplier_groups[seller_mst]["total_t_score"] += inv.get("t_score", 100)
            if has_warning:
                supplier_groups[seller_mst]["warnings_count"] += 1

    suppliers_list = []
    for mst, s in supplier_groups.items():
        avg_score = s["total_t_score"] / s["invoice_count"] if s["invoice_count"] > 0 else 100
        suppliers_list.append({
            "supplier_mst": s["supplier_mst"],
            "supplier_name": s["supplier_name"],
            "invoice_count": s["invoice_count"],
            "warnings_count": s["warnings_count"],
            "total_value": s["total_value"],
            "average_t_score": round(avg_score, 1),
            "is_blacklisted": s["is_blacklisted"]
        })

    # Sort suppliers: highest warnings count first, then by total transaction value descending
    suppliers_list.sort(key=lambda x: (x["warnings_count"], x["total_value"]), reverse=True)

    # Filter to only list suppliers that have warnings or are blacklisted
    high_risk_suppliers = [s for s in suppliers_list if s["warnings_count"] > 0 or s["is_blacklisted"]]

    return jsonify({
        "status": "success",
        "summary": {
            "total_analyzed": total_analyzed,
            "total_with_warnings": total_with_warnings,
            "total_value_at_risk": total_value_at_risk,
            "blacklist_warnings_count": blacklist_warnings_count,
            "signature_violations_count": signature_violations_count,
            "payment_type_flags_count": payment_type_flags_count
        },
        "suppliers": high_risk_suppliers
    })



def get_harness_db():
    import sqlite3
    import os
    
    conn = sqlite3.connect("harness.db", timeout=10.0)
    conn.text_factory = lambda x: str(x, encoding="utf-8", errors="replace")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA journal_mode = WAL;")
    
    # Auto-initialize tables if story table doesn't exist
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='story';")
    if not cur.fetchone():
        conn.execute("""
        CREATE TABLE IF NOT EXISTS story (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            risk_lane TEXT NOT NULL,
            contract_doc TEXT,
            status TEXT NOT NULL,
            notes TEXT,
            unit_proof INTEGER DEFAULT 0,
            integration_proof INTEGER DEFAULT 0,
            e2e_proof INTEGER DEFAULT 0,
            platform_proof INTEGER DEFAULT 0,
            evidence TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS decision (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            status TEXT NOT NULL,
            doc_path TEXT,
            verify_command TEXT,
            last_verified_at TIMESTAMP,
            last_verified_result TEXT,
            predicted_impact TEXT,
            actual_outcome TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS backlog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            discovered_while TEXT,
            current_pain TEXT,
            suggested_improvement TEXT,
            risk TEXT,
            status TEXT NOT NULL,
            predicted_impact TEXT,
            actual_outcome TEXT,
            implemented_at TIMESTAMP,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS trace (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_summary TEXT NOT NULL,
            intake_id TEXT,
            story_id TEXT,
            agent TEXT,
            actions_taken TEXT,
            files_read TEXT,
            files_changed TEXT,
            decisions_made TEXT,
            errors TEXT,
            outcome TEXT,
            duration_seconds INTEGER,
            token_estimate INTEGER,
            harness_friction TEXT,
            notes TEXT,
            git_hash TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
    return conn


@invoices_blueprint.get("/api/harness/summary")
def api_harness_summary():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        conn = get_harness_db()
        cur = conn.cursor()

        # Get status counts for stories
        cur.execute("SELECT status, COUNT(*) as cnt FROM story GROUP BY status")
        story_status_rows = cur.fetchall()
        story_status = {r["status"]: r["cnt"] for r in story_status_rows}

        # Get risk lane counts for stories
        cur.execute("SELECT risk_lane, COUNT(*) as cnt FROM story GROUP BY risk_lane")
        story_lane_rows = cur.fetchall()
        story_lane = {r["risk_lane"]: r["cnt"] for r in story_lane_rows}

        # Get decision status counts
        cur.execute("SELECT status, COUNT(*) as cnt FROM decision GROUP BY status")
        decision_status_rows = cur.fetchall()
        decision_status = {r["status"]: r["cnt"] for r in decision_status_rows}

        # Get backlog status counts
        cur.execute("SELECT status, COUNT(*) as cnt FROM backlog GROUP BY status")
        backlog_status_rows = cur.fetchall()
        backlog_status = {r["status"]: r["cnt"] for r in backlog_status_rows}

        # Get trace counts
        cur.execute("SELECT COUNT(*) as cnt FROM trace")
        trace_count = cur.fetchone()["cnt"]

        # Fetch all stories
        cur.execute("SELECT id, title, created_at, risk_lane, contract_doc, status, unit_proof, integration_proof, e2e_proof, platform_proof, evidence, notes FROM story ORDER BY id DESC")
        stories = [dict(r) for r in cur.fetchall()]

        # Fetch all decisions
        cur.execute("SELECT id, title, created_at, status, doc_path, verify_command, last_verified_at, last_verified_result, predicted_impact, actual_outcome, notes FROM decision ORDER BY id DESC")
        decisions = [dict(r) for r in cur.fetchall()]

        # Fetch recent traces (last 30)
        cur.execute("SELECT id, created_at, task_summary, intake_id, story_id, agent, actions_taken, files_read, files_changed, decisions_made, errors, outcome, duration_seconds, token_estimate, harness_friction, notes, git_hash FROM trace ORDER BY id DESC LIMIT 30")
        traces = [dict(r) for r in cur.fetchall()]

        # Fetch all backlog items
        cur.execute("SELECT id, created_at, title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, actual_outcome, implemented_at, notes FROM backlog ORDER BY id DESC")
        backlog = [dict(r) for r in cur.fetchall()]

        # Build stats struct
        stats = {
            "stories": {
                "total": sum(story_status.values()),
                "status": story_status,
                "lanes": story_lane
            },
            "decisions": {
                "total": sum(decision_status.values()),
                "status": decision_status
            },
            "backlog": {
                "total": sum(backlog_status.values()),
                "status": backlog_status
            },
            "traces": {
                "total": trace_count
            }
        }

        return jsonify({
            "stats": stats,
            "stories": stories,
            "decisions": decisions,
            "traces": traces,
            "backlog": backlog
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@invoices_blueprint.post("/api/harness/story")
def api_harness_story_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        story_id = body.get("id", "").strip()
        title = body.get("title", "").strip()
        lane = body.get("lane", "normal").strip()
        contract = (body.get("contract") or body.get("contract_doc") or "").strip()
        status = body.get("status", "planned").strip()
        notes = body.get("notes", "").strip()

        if not story_id or not title:
            return jsonify({"error": "Mã (id) và Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO story (id, title, risk_lane, contract_doc, status, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (story_id, title, lane, contract, status, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Story {story_id} added successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@invoices_blueprint.post("/api/harness/story/update")
def api_harness_story_update():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        story_id = body.get("id", "").strip()
        status = body.get("status", "planned").strip()
        evidence = body.get("evidence", "").strip()
        
        proofs = body.get("proofs") or {}
        unit = body.get("unit") if body.get("unit") is not None else proofs.get("unit")
        integration = body.get("integration") if body.get("integration") is not None else proofs.get("integration")
        e2e = body.get("e2e") if body.get("e2e") is not None else proofs.get("e2e")
        platform = body.get("platform") if body.get("platform") is not None else proofs.get("platform")

        if not story_id:
            return jsonify({"error": "Mã (id) là bắt buộc."}), 400

        # convert potential empty strings or convert type
        try:
            unit = int(unit) if unit is not None and str(unit).strip() != "" else None
        except Exception:
            unit = None
        try:
            integration = int(integration) if integration is not None and str(integration).strip() != "" else None
        except Exception:
            integration = None
        try:
            e2e = int(e2e) if e2e is not None and str(e2e).strip() != "" else None
        except Exception:
            e2e = None
        try:
            platform = int(platform) if platform is not None and str(platform).strip() != "" else None
        except Exception:
            platform = None

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE story
            SET status = ?, evidence = COALESCE(?, evidence),
                unit_proof = COALESCE(?, unit_proof),
                integration_proof = COALESCE(?, integration_proof),
                e2e_proof = COALESCE(?, e2e_proof),
                platform_proof = COALESCE(?, platform_proof)
            WHERE id = ?
            """,
            (status, evidence, unit, integration, e2e, platform, story_id)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Story {story_id} updated successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@invoices_blueprint.post("/api/harness/decision")
def api_harness_decision_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        decision_id = body.get("id", "").strip()
        title = body.get("title", "").strip()
        status = body.get("status", "proposed").strip()
        doc = (body.get("doc") or body.get("doc_path") or "").strip()
        verify = (body.get("verify") or body.get("verify_command") or body.get("verify_cmd") or "").strip()
        predicted = (body.get("predicted") or body.get("predicted_impact") or "").strip()
        notes = body.get("notes", "").strip()

        if not decision_id or not title:
            return jsonify({"error": "Mã (id) và Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO decision (id, title, status, doc_path, verify_command, predicted_impact, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (decision_id, title, status, doc, verify, predicted, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Decision {decision_id} recorded successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@invoices_blueprint.post("/api/harness/backlog")
def api_harness_backlog_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        title = body.get("title", "").strip()
        discovered_while = body.get("discovered_while", "").strip()
        current_pain = body.get("current_pain", "").strip()
        suggested_improvement = body.get("suggested_improvement", "").strip()
        risk = body.get("risk", "normal").strip()
        status = body.get("status", "open").strip()
        predicted_impact = body.get("predicted_impact", "").strip()
        notes = body.get("notes", "").strip()

        if not title:
            return jsonify({"error": "Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO backlog (title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": "Backlog item added successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@invoices_blueprint.get("/api/harness/agent/stream")
def api_agent_stream():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    provider = request.args.get("provider", "gemini").strip()
    model = request.args.get("model", "gemini-2.5-flash").strip()
    goal = request.args.get("goal", "").strip()
    story_id = request.args.get("story_id", "").strip()

    if not goal:
        return jsonify({"error": "Goal is required"}), 400

    from flask import Response

    def generate():
        import subprocess
        import os
        import json

        env = os.environ.copy()
        env["AGENT_PROVIDER"] = provider
        env["AGENT_MODEL"] = model
        env["AGENT_GOAL"] = goal
        if story_id:
            env["AGENT_STORY_ID"] = story_id

        cmd = ["node", "scripts/agent-harness/run-agent.js"]

        proc = subprocess.Popen(
            cmd,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1
        )

        while True:
            line = proc.stdout.readline()
            if not line:
                break
            line_str = line.strip()
            if line_str:
                yield f"data: {line_str}\n\n"

        err = proc.stderr.read()
        if err:
            try:
                # Try parsing as JSON error from run-agent.js
                err_data = json.loads(err.strip())
                yield f"data: {json.dumps(err_data)}\n\n"
            except Exception:
                yield f"data: {json.dumps({'type': 'error', 'message': err.strip()})}\n\n"

        proc.wait()

        if proc.returncode != 0:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Process exited with code {proc.returncode}'})}\n\n"
        else:
            if story_id:
                try:
                    git_hash = "unknown"
                    try:
                        git_hash = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True).strip()
                        status_out = subprocess.check_output(["git", "status", "--porcelain"], text=True).strip()
                        if status_out:
                            git_hash += " (dirty)"
                    except Exception:
                        pass

                    conn = None
                    try:
                        conn = get_harness_db()
                        cur = conn.cursor()
                        cur.execute(
                            """
                            INSERT INTO trace (task_summary, story_id, agent, outcome, git_hash, created_at, actions_taken, notes)
                            VALUES (?, ?, ?, 'completed', ?, datetime('now'), ?, ?)
                            """,
                            (f"Autonomous Run: {goal[:50]}...", story_id, "SkawldAgent", git_hash, '["run-agent.js"]', f"Goal: {goal}")
                        )
                        conn.commit()
                    finally:
                        if conn:
                            conn.close()
                except Exception as db_err:
                    print(f"Error logging trace to DB: {db_err}")

    return Response(generate(), mimetype="text/event-stream")


@invoices_blueprint.post("/api/bctc/compile")
@roles_required("admin", "auditor")
def api_bctc_compile():
    """Compile BCTC B01-DN, B02-DN, B03-DN from Trial Balance ledger data (US-200)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    balances = {}
    metadata = request.json or {}
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
                # Populate metadata from request form fields if present
                metadata = {
                    "mst": request.form.get("mst", "0109998887"),
                    "company_name": request.form.get("company_name", "CONG TY TNHH MOCK"),
                    "year": int(request.form.get("year", datetime.now().year)),
                    "reporting_period_type": request.form.get("reporting_period_type", "N"),
                    "dividends_paid": float(request.form.get("dividends_paid", 0.0))
                }
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        balances = metadata.get("balances", {})
        
    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400
        
    try:
        from invoices.bctc_service import compile_bctc
        xml_str, warnings = compile_bctc(balances, metadata)
        return jsonify({
            "status": "success" if not warnings else "warning",
            "xml": xml_str,
            "warnings": warnings
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/bctc/audit-ledger")
@roles_required("admin", "auditor")
def api_bctc_audit_ledger():
    """Cross-reference General Ledger entries with e-invoices for compliance (US-201)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    balances = {}
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
                taxpayer_mst = request.form.get("taxpayer_mst") or taxpayer_mst
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        balances = payload.get("balances", {})
        taxpayer_mst = payload.get("taxpayer_mst") or taxpayer_mst
        
    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400
        
    try:
        from invoices.bctc_service import audit_ledger_against_invoices
        report = audit_ledger_against_invoices(balances, taxpayer_mst)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/payments/tax-slip")
@roles_required("admin", "auditor")
def api_payments_tax_slip():
    """Generate GDT Form 711/MB Tax Payment Slip XML and VietQR code (US-202)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    payload = request.json or {}
    mst = payload.get("mst") or session.get("taxpayer_mst") or "0109998887"
    company_name = payload.get("company_name", "CONG TY TNHH MOCK")
    tax_type = payload.get("tax_type")
    amount = payload.get("amount")
    
    if not tax_type or not amount:
        return jsonify({"error": "Thieu thong tin loai thue hoac so tien."}), 400
        
    try:
        amount_val = float(amount)
    except ValueError:
        return jsonify({"error": "So tien khong hop le."}), 400
        
    chapter_type = payload.get("chapter_type", "domestic_private")
    treasury_name = payload.get("treasury_name", "Kho bac Nha nuoc Quan Cau Giay")
    treasury_account = payload.get("treasury_account", "111222333444")
    bank_bin = payload.get("bank_bin", "970415")
    
    try:
        from invoices.tax_payment_service import generate_tax_payment_slip
        slip = generate_tax_payment_slip(
            mst=mst,
            company_name=company_name,
            tax_type=tax_type,
            amount=amount_val,
            chapter_type=chapter_type,
            treasury_name=treasury_name,
            treasury_account=treasury_account,
            bank_bin=bank_bin
        )
        return jsonify(slip)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/payments/bank-recon")
@roles_required("admin", "auditor")
def api_payments_bank_recon():
    """Standard bank statement parsing, fuzzy matching, and cash payment compliance auditing (US-203)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    # Process uploaded bank statement if present
    results = {}
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                content = file.read().decode("utf-8")
                from invoices.reconciliation_service import ReconciliationEngine
                engine = ReconciliationEngine()
                engine.process_csv(content)
                results = engine.run_matching()
            except Exception as e:
                return jsonify({"error": f"Loi xu ly file sao ke: {str(e)}"}), 400
                
    # Run cash compliance checks for invoices >= 20M VND
    try:
        from invoices.bank_reconcile_service import check_cash_payment_compliance
        compliance_flags = check_cash_payment_compliance(taxpayer_mst)
        return jsonify({
            "status": "success",
            "reconciliation_summary": results,
            "compliance_warnings": compliance_flags
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/ecommerce/sync")
@roles_required("admin", "auditor")
def api_ecommerce_sync():
    """Parse platform reports and record daily consolidated revenue & fees in the database (US-204)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    platform = request.args.get("platform", "shopee").strip()
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    orders = []
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.ecommerce_service import parse_ecommerce_sheet
                orders = parse_ecommerce_sheet(file.read(), platform)
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        orders = payload.get("orders", [])
        taxpayer_mst = payload.get("taxpayer_mst") or taxpayer_mst
        platform = payload.get("platform") or platform
        
    if not orders:
        return jsonify({"error": "Thieu du lieu don hang e-commerce."}), 400
        
    try:
        from invoices.ecommerce_service import sync_ecommerce_orders
        res = sync_ecommerce_orders(orders, taxpayer_mst, platform)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/ecommerce/reconcile")
@roles_required("admin", "auditor", "viewer")
def api_ecommerce_reconcile():
    """Reconcile Shopee/TikTok Shop order logs with output invoices (US-205)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    import json
    orders_json = request.args.get("orders")
    platform_orders = []
    if orders_json:
        try:
            platform_orders = json.loads(orders_json)
        except Exception:
            pass
            
    if not platform_orders:
        platform_orders = [
            {"order_id": "ORD-SHOPEE-1001", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 500000.0, "commission_fee": 15000.0, "service_fee": 5000.0},
            {"order_id": "ORD-SHOPEE-1002", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 1200000.0, "commission_fee": 36000.0, "service_fee": 12000.0},
            {"order_id": "ORD-SHOPEE-1003", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 850000.0, "commission_fee": 25500.0, "service_fee": 8500.0}
        ]
        
    try:
        from invoices.ecommerce_service import reconcile_ecommerce_tax
        report = reconcile_ecommerce_tax(taxpayer_mst, platform_orders)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# US-141: Audit Trail Viewer UI & Export
# ---------------------------------------------------------------------------

@invoices_blueprint.get("/audit-trail")
@roles_required("admin", "auditor")
def audit_trail_page():
    """Render the Audit Trail Viewer UI."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return redirect(url_for("index"))
    return render_template("audit_trail.html",
                           logged_in=session.get("logged_in"),
                           session_username=session.get("display_name") or session.get("username"))


@invoices_blueprint.get("/api/audit-logs")
@roles_required("admin", "auditor")
def api_get_audit_logs():
    """Retrieve security audit logs with optional filtering."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog

    try:
        query = SecurityAuditLog.query

        # Filters
        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)

        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))

        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))

        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)

        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")

        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        # Pagination
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)
        per_page = min(per_page, 200)

        total = query.count()
        logs = query.order_by(SecurityAuditLog.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            "total": total,
            "page": page,
            "per_page": per_page,
            "logs": [
                {
                    "id": log.id,
                    "timestamp": log.timestamp,
                    "username": log.username,
                    "tax_code": log.tax_code,
                    "event_category": log.event_category,
                    "ip_address": log.ip_address,
                    "event_details": log.event_details,
                }
                for log in logs
            ],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/audit-logs/export/csv")
@roles_required("admin", "auditor")
def api_export_audit_logs_csv():
    """Export filtered audit logs as CSV file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog
    import csv
    import io

    try:
        query = SecurityAuditLog.query

        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)
        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))
        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))
        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)
        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")
        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        logs = query.order_by(SecurityAuditLog.id.desc()).limit(10000).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Timestamp", "Username", "Tax Code", "Category", "IP Address", "Details"])
        for log in logs:
            writer.writerow([
                log.id, log.timestamp, log.username, log.tax_code or "",
                log.event_category, log.ip_address or "", log.event_details or "",
            ])

        from datetime import datetime
        filename = f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/audit-logs/export/pdf")
@roles_required("admin", "auditor")
def api_export_audit_logs_pdf():
    """Export filtered audit logs as PDF file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog
    from datetime import datetime

    try:
        query = SecurityAuditLog.query

        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)
        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))
        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))
        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)
        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")
        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        logs = query.order_by(SecurityAuditLog.id.desc()).limit(5000).all()

        # Generate HTML-based PDF using render_template
        html = render_template("audit_trail_pdf.html",
                               logs=logs,
                               generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                               total_records=len(logs),
                               filters={
                                   "category": category,
                                   "username": username,
                                   "tax_code": tax_code,
                                   "date_from": date_from,
                                   "date_to": date_to,
                                   "keyword": keyword,
                               })

        filename = f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        return Response(
            html,
            mimetype="text/html",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/reports/signed-compliance")
@roles_required("admin", "auditor")
def api_export_signed_compliance():
    """Export audited compliance report with embedded cryptographic signature."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        
        # Get invoices from service
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        
        # Get system secret key for hashing
        secret_key = current_app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
        
        from invoices.compliance_report_service import generate_signed_excel_report
        excel_bytes = generate_signed_excel_report(invoices, secret_key)
        
        # Log this administrative export event in the security audit ledger
        from invoices.security_audit_service import log_security_event
        log_security_event(
            username=session.get("username", "admin"),
            event_category="EXPORT",
            tax_code=session.get("tax_code", ""),
            ip_address=request.remote_addr,
            event_details=f"Exported cryptographically signed compliance report for period {parsed_from} to {parsed_to} ({len(invoices)} invoices)."
        )

        filename = f"signed_compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        current_app.config["CURRENT_JWT"] = None


@invoices_blueprint.post("/api/reports/verify-signed")
@roles_required("admin", "auditor")
def api_verify_signed_report():
    """Upload and verify a signed compliance report file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy tệp tin báo cáo được tải lên."}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Tên tệp tin không hợp lệ."}), 400

    try:
        file_bytes = file.read()
        secret_key = current_app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
        
        from invoices.compliance_report_service import verify_excel_report
        result = verify_excel_report(file_bytes, secret_key)
        
        # Log security audit verification event
        status_str = "SUCCESS" if result.get("verified") else "FAILED"
        from invoices.security_audit_service import log_security_event
        log_security_event(
            username=session.get("username", "admin"),
            event_category="VERIFY",
            tax_code=session.get("tax_code", ""),
            ip_address=request.remote_addr,
            event_details=f"Performed cryptographic verification of compliance report file '{file.filename}'. Result: {status_str} ({result.get('invoices_count', 0)} invoices parsed)."
        )

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi xử lý xác minh báo cáo: {str(e)}"}), 500


@invoices_blueprint.get("/api/sync/health")
@roles_required("admin", "auditor")
def api_sync_health():
    """Retrieve CAPTCHA solver statistics and overall crawler status."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from auth.captcha_solver import captcha_analytics

        # CAPTCHA metrics
        stats = captcha_analytics.get_stats()

        # Crawler status
        crawler_status = "idle"
        queue_instance = current_app.extensions.get("resilient_sync_queue")
        if queue_instance:
            with queue_instance._lock:
                if any(job.status == "running" for job in queue_instance.jobs.values()):
                    crawler_status = "running"

        return jsonify({
            "status": "healthy",
            "crawler_status": crawler_status,
            "solver": stats,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/consolidated-dashboard")
@roles_required("admin", "auditor")
def consolidated_dashboard_page():
    """Render the corporate multi-entity consolidated dashboard."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("consolidated.html")


@invoices_blueprint.route("/api/tenant/groups", methods=["GET", "POST"])
@roles_required("admin", "auditor")
def api_tenant_groups():
    """GET/POST API to fetch or create corporate tenant groups."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import TenantGroup
    import json

    username = session.get("username", "admin")

    if request.method == "POST":
        try:
            data = request.get_json() or {}
            group_name = data.get("group_name")
            taxpayer_msts = data.get("taxpayer_msts", [])

            if not group_name:
                return jsonify({"error": "Tên tập đoàn không được để trống."}), 400
            if not isinstance(taxpayer_msts, list):
                return jsonify({"error": "Danh sách MST phải là một mảng."}), 400

            # Validate MSTs
            taxpayer_msts = [str(mst).strip() for mst in taxpayer_msts if mst]

            # Upsert group
            group = TenantGroup.query.filter_by(group_name=group_name).first()
            if group:
                group.taxpayer_msts = json.dumps(taxpayer_msts)
                group.admin_username = username
            else:
                group = TenantGroup(
                    group_name=group_name,
                    admin_username=username,
                    taxpayer_msts=json.dumps(taxpayer_msts)
                )
                db.session.add(group)

            db.session.commit()
            return jsonify({"status": "success", "group": group.to_dict()})
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    # GET
    try:
        groups = TenantGroup.query.filter_by(admin_username=username).all()
        # Fallback to all groups if admin or no group found
        if not groups and username == "admin":
            groups = TenantGroup.query.all()
        return jsonify([g.to_dict() for g in groups])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.get("/api/tenant/consolidated")
@roles_required("admin", "auditor")
def api_tenant_consolidated():
    """Retrieve consolidated financial metrics and risk scores across a group's MSTs."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import TenantGroup
    from invoices.multitenant_service import get_tenant_consolidated_stats
    import json

    username = session.get("username", "admin")
    group_id = request.args.get("group_id")

    try:
        # 1. Resolve Group
        group = None
        if group_id:
            group = TenantGroup.query.get(group_id)
        else:
            group = TenantGroup.query.filter_by(admin_username=username).first()
            if not group and username == "admin":
                group = TenantGroup.query.first()

        if not group:
            return jsonify({
                "group_id": None,
                "group_name": "Không có nhóm",
                "summary": {
                    "total_invoices": 0,
                    "total_revenue": 0.0,
                    "vat_output": 0.0,
                    "vat_input": 0.0,
                    "average_t_score": 100.0
                },
                "entities": []
            })

        # 2. Query each member MST
        mst_list = group.get_mst_list()
        entities = []
        for mst in mst_list:
            stats = get_tenant_consolidated_stats(mst)
            entities.append(stats)

        # 3. Aggregate totals
        total_invoices = sum(e["total_invoices"] for e in entities)
        total_revenue = sum(e["total_revenue"] for e in entities)
        vat_output = sum(e["vat_output"] for e in entities)
        vat_input = sum(e["vat_input"] for e in entities)
        
        # Weighted average for T-Score
        t_score_sum = 0.0
        t_score_count = 0
        for e in entities:
            if e["total_invoices"] > 0:
                t_score_sum += e["average_t_score"] * e["total_invoices"]
                t_score_count += e["total_invoices"]
            else:
                t_score_sum += e["average_t_score"]
                t_score_count += 1
                
        average_t_score = round(t_score_sum / t_score_count, 1) if t_score_count > 0 else 100.0

        return jsonify({
            "group_id": group.id,
            "group_name": group.group_name,
            "summary": {
                "total_invoices": total_invoices,
                "total_revenue": total_revenue,
                "vat_output": vat_output,
                "vat_input": vat_input,
                "average_t_score": average_t_score
            },
            "entities": entities
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/cit/finalize")
@roles_required("admin", "auditor")
def api_cit_finalize():
    """US-180: Compile CIT Finalization and generate Form 03/TNDN XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import datetime
    balances = {}
    metadata = {
        "mst": request.form.get("mst") or session.get("taxpayer_mst") or "0109998887",
        "company_name": request.form.get("company_name", "CONG TY TNHH MOCK"),
        "year": int(request.form.get("year", datetime.now().year)),
        "non_deductible_manual": float(request.form.get("non_deductible_manual", 0.0)),
        "loss_carry_forward": float(request.form.get("loss_carry_forward", 0.0)),
        "rd_allowance": float(request.form.get("rd_allowance", 0.0))
    }

    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        balances = payload.get("balances", {})
        metadata.update(payload.get("metadata", {}))

    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400

    try:
        from invoices.cit_service import finalize_cit
        result = finalize_cit(balances, metadata)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@invoices_blueprint.post("/api/cit/simulate-scenario")
@roles_required("admin", "auditor")
def api_cit_simulate_scenario():
    """US-181: Simulate what-if tax scenarios based on slider adjustments."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        payload = request.get_json() or {}
        base_data = payload.get("base_data", {})
        adjustments = payload.get("adjustments", {})
        
        from invoices.cit_service import simulate_cit_scenario
        result = simulate_cit_scenario(base_data, adjustments)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# US-150: Smart Cash Flow Predictor — rolling 30/60/90-day projections
# US-151: Interactive Scenario Simulator — what-if stress testing
# ---------------------------------------------------------------------------


@invoices_blueprint.route("/api/finance/cashflow")
def api_finance_cashflow():
    """US-150: Return rolling 30/60/90-day cash-flow projections."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("mst")
        from invoices.cashflow_service import calculate_cashflow_projection
        result = calculate_cashflow_projection(taxpayer_mst=taxpayer_mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@invoices_blueprint.route("/api/finance/simulate", methods=["POST"])
def api_finance_simulate():
    """US-151: Stateless scenario simulation with adjustable parameters."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        payload = request.get_json() or {}
        delay_days = int(payload.get("delay_days", 0))
        rejection_rate = float(payload.get("rejection_rate", 0.0))
        vat_adjustment = float(payload.get("vat_adjustment", 0.0))
        taxpayer_mst = session.get("active_taxpayer_mst") or payload.get("mst")

        from invoices.cashflow_service import simulate_scenario
        result = simulate_scenario(
            taxpayer_mst=taxpayer_mst,
            delay_days=delay_days,
            rejection_rate=rejection_rate,
            vat_adjustment=vat_adjustment,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


