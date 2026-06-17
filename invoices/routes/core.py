from __future__ import annotations
from io import BytesIO
from flask import Blueprint, Response, current_app, jsonify, redirect, render_template, request, send_file, session, url_for
from export.excel import generate_excel_workbook, generate_local_excel_workbook
from invoices.parser import DateValidationError, validate_date_range
from invoices.service import (
    GDTIntegrationNotReadyError,
    InvoiceQuery,
    build_invoice_lookup,
    download_invoice_xml,
    fetch_invoices,
    resolve_live_download_name,
    fetch_invoice_line_items,
    extract_partners_from_invoices,
    generate_tax_usage_report,
)
from extensions import db
from auth.decorators import roles_required
import os
import uuid
import threading
from datetime import datetime
from flask import send_file
import io
from invoices.routes.shared import invoices_blueprint, DOWNLOAD_TASKS, DOWNLOAD_TASKS_LOCK
from invoices.routes.helpers import (
    _ensure_logged_in,
    get_supplier_pivot_data,
    _AGING_BUCKETS,
    classify_fct_item,
    generate_fct_excel,
    require_api_signature,
    get_harness_db,
    render_html_to_pdf
)

@invoices_blueprint.get("/invoices")
def invoices_page():
    """Render the invoice search screen for authenticated users."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("invoices.html")

@invoices_blueprint.get("/cashflow")
def cashflow_page():
    """Render the cashflow oracle dashboard."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("cashflow.html")

@invoices_blueprint.get("/harness")
def harness_page():
    """Render the Harness Agent Control Center."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("harness.html")

@invoices_blueprint.get("/tax-bctc")
def tax_bctc_page():
    """Render the V17 Tax and BCTC services screen for authenticated users."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("tax_bctc.html")

@invoices_blueprint.get("/compliance-concept-map")
def compliance_concept_map_page():
    """Render the Compliance Concept Map Explorer page."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("compliance_concept_map.html")

@invoices_blueprint.get("/api/compliance/concept-map")
def api_compliance_concept_map():
    """Return JSON configuration for compliance concept map nodes and edges."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    mst = session.get("taxpayer_mst") or "0102030405"
    
    # Base configuration for nodes and links
    nodes = [
        {"id": "v26", "label": "CIT Compliance (v26)", "group": "income", "risk": "high", "url": "/v26-compliance", "status": "active"},
        {"id": "v27", "label": "E-Invoice Format (v27)", "group": "vat", "risk": "high", "url": "/v27-compliance", "status": "active"},
        {"id": "v28", "label": "Import-Export VAT (v28)", "group": "vat", "risk": "medium", "url": "/v28-compliance", "status": "active"},
        {"id": "v29", "label": "Contractor Tax (v29)", "group": "income", "risk": "medium", "url": "/v29-compliance", "status": "active"},
        {"id": "v30", "label": "Household Tax (v30)", "group": "income", "risk": "low", "url": "/v30-compliance", "status": "active"},
        {"id": "v31", "label": "Environmental Tax (v31)", "group": "environmental", "risk": "medium", "url": "/v31-compliance", "status": "active"},
        {"id": "v44", "label": "Compliance Hub (v44)", "group": "core", "risk": "low", "url": "/v44-compliance-hub", "status": "active"},
        {"id": "v45", "label": "CIT & TP Hub (v45)", "group": "income", "risk": "high", "url": "/v45-compliance-hub", "status": "active"},
        {"id": "v46", "label": "Error & Conversion (v46)", "group": "core", "risk": "medium", "url": "/v46-compliance-hub", "status": "active"},
        {"id": "v47", "label": "VAT Rate Hub (v47)", "group": "vat", "risk": "high", "url": "/v47-compliance-hub", "status": "active"},
        {"id": "v48", "label": "Agricultural Hub (v48)", "group": "vat", "risk": "medium", "url": "/v48-compliance-hub", "status": "active"},
        {"id": "v49", "label": "CIT Law 67 Hub (v49)", "group": "income", "risk": "medium", "url": "/v49-compliance-hub", "status": "active"},
        {"id": "v50", "label": "PIT Law 109 Hub (v50)", "group": "income", "risk": "medium", "url": "/v50-compliance-hub", "status": "active"},
        {"id": "v51", "label": "Admin Law 108 Hub (v51)", "group": "core", "risk": "low", "url": "/v51-compliance-hub", "status": "active"},
        {"id": "v52", "label": "SCT Law 66 Hub (v52)", "group": "special", "risk": "high", "url": "/v52-compliance-hub", "status": "active"},
        {"id": "v53", "label": "EP Tax Hub (v53)", "group": "environmental", "risk": "high", "url": "/v53-compliance-hub", "status": "active"},
        {"id": "v54", "label": "NRT Hub (v54)", "group": "special", "risk": "medium", "url": "/v54-compliance-hub", "status": "active"},
        {"id": "v55", "label": "IET Hub (v55)", "group": "special", "risk": "medium", "url": "/v55-compliance-hub", "status": "active"},
        {"id": "v56", "label": "License Fee Hub (v56)", "group": "fee", "risk": "low", "url": "/v56-compliance-hub", "status": "active"},
        {"id": "v57", "label": "Registration Fee Hub (v57)", "group": "fee", "risk": "low", "url": "/v57-compliance-hub", "status": "active"},
        {"id": "v58", "label": "NR Tax Hub (v58)", "group": "special", "risk": "medium", "url": "/v58-compliance-hub", "status": "active"},
        {"id": "v59", "label": "NALUT Hub (v59)", "group": "fee", "risk": "medium", "url": "/v59-compliance-hub", "status": "active"},
        {"id": "v60", "label": "ALUT Hub (v60)", "group": "fee", "risk": "medium", "url": "/v60-compliance-hub", "status": "active"},
        {"id": "v61", "label": "EP Wastewater Hub (v61)", "group": "environmental", "risk": "medium", "url": "/v61-compliance-hub", "status": "active"},
        {"id": "v62", "label": "EP Emissions Hub (v62)", "group": "environmental", "risk": "medium", "url": "/v62-compliance-hub", "status": "active"},
        {"id": "v63", "label": "EP Mineral Hub (v63)", "group": "environmental", "risk": "medium", "url": "/v63-compliance-hub", "status": "active"},
        {"id": "v64", "label": "EP Solid Waste Hub (v64)", "group": "environmental", "risk": "medium", "url": "/v64-compliance-hub", "status": "active"},
        {"id": "v65", "label": "EPR Recycling Hub (v65)", "group": "environmental", "risk": "medium", "url": "/v65-compliance-hub", "status": "active"},
        {"id": "v66", "label": "GHG Emissions Hub (v66)", "group": "environmental", "risk": "high", "url": "/v66-compliance-hub", "status": "active"},
        {"id": "v67", "label": "Scrap Import Deposit (v67)", "group": "environmental", "risk": "medium", "url": "/v67-compliance-hub", "status": "active"},
        {"id": "v68", "label": "Biodiversity Hub (v68)", "group": "environmental", "risk": "low", "url": "/v68-compliance-hub", "status": "active"},
        {"id": "v69", "label": "Oil Spill Hub (v69)", "group": "environmental", "risk": "high", "url": "/v69-compliance-hub", "status": "active"},
        {"id": "v70", "label": "ODS Quota Hub (v70)", "group": "environmental", "risk": "medium", "url": "/v70-compliance-hub", "status": "active"},
    ]
    
    links = [
        {"source": "v47", "target": "v28", "type": "dependency", "label": "VAT Baseline"},
        {"source": "v47", "target": "v41", "type": "dependency", "label": "Export Eligibility"},
        {"source": "v26", "target": "v45", "type": "dependency", "label": "CIT Base"},
        {"source": "v45", "target": "v49", "type": "dependency", "label": "Transfer Pricing Limit"},
        {"source": "v53", "target": "v31", "type": "dependency", "label": "EP Tax Baseline"},
        {"source": "v53", "target": "v61", "type": "subset", "label": "Wastewater Regulation"},
        {"source": "v53", "target": "v62", "type": "subset", "label": "Emissions Regulation"},
        {"source": "v53", "target": "v63", "type": "subset", "label": "Mineral Exploitation"},
        {"source": "v53", "target": "v64", "type": "subset", "label": "Solid Waste Management"},
        {"source": "v53", "target": "v65", "type": "subset", "label": "Recycling Obligation"},
        {"source": "v53", "target": "v66", "type": "subset", "label": "GHG Inventory"},
        {"source": "v53", "target": "v67", "type": "subset", "label": "Scrap Import Security"},
        {"source": "v53", "target": "v68", "type": "subset", "label": "Biodiversity Conservation"},
        {"source": "v53", "target": "v69", "type": "subset", "label": "Oil Spill Response"},
        {"source": "v53", "target": "v70", "type": "subset", "label": "ODS Quotas Control"},
    ]

    violations = {n["id"]: 0 for n in nodes}
    has_fuel = False
    has_coal = False
    has_plastic = False
    has_ods = False

    try:
        from invoices.models import Invoice
        invoices = Invoice.query.filter(Invoice.taxpayer_mst == mst).all()
        for inv in invoices:
            # Check invoice-level warnings
            for w in (inv.warnings or []):
                w_lower = w.lower()
                if any(k in w_lower for k in ["format", "xml", "chữ ký", "signature"]):
                    violations["v27"] += 1
                if any(k in w_lower for k in ["thuế suất", "tax rate", "mismatch", "8%", "10%", "vat", "giá trị gia tăng"]):
                    violations["v47"] += 1
                if any(k in w_lower for k in ["cit", "tp", "giao dịch liên kết", "chi phí", "deductibility", "personal_purchase"]):
                    violations["v26"] += 1
                    violations["v45"] += 1
                if any(k in w_lower for k in ["môi trường", "ep_tax"]):
                    violations["v53"] += 1

            # Check line items keywords
            for item in (inv.items or []):
                name_lower = item.item_name.lower()
                if any(k in name_lower for k in ["xăng", "dầu", "diesel", "fuel", "gasoil", "petrol"]):
                    has_fuel = True
                if any(k in name_lower for k in ["than", "coal", "anthracite", "lignite"]):
                    has_coal = True
                if any(k in name_lower for k in ["túi ni-lông", "túi nhựa", "plastic bag"]):
                    has_plastic = True
                if any(k in name_lower for k in ["ods", "hcfc", "cfc", "methyl bromide", "tầng ô-dôn"]):
                    has_ods = True
    except Exception:
        pass

    # Update node violations and dynamic risks
    for node in nodes:
        node["violations_count"] = violations.get(node["id"], 0)
        if node["violations_count"] > 3:
            node["risk"] = "high"
        elif node["violations_count"] > 0:
            node["risk"] = "medium"

    # AI recommendation links
    suggested_links = []
    if has_fuel or has_coal or has_plastic:
        suggested_links.append({
            "source": "v47",
            "target": "v53",
            "type": "recommended",
            "label": "AI: Fuel EP Tax Link",
            "reason": "Phát hiện giao dịch nhiên liệu hoặc túi nhựa chịu thuế Bảo vệ Môi trường. Đề xuất kiểm toán liên kết VAT -> EP Tax."
        })
    if has_ods:
        suggested_links.append({
            "source": "v53",
            "target": "v70",
            "type": "recommended",
            "label": "AI: ODS Quotas Link",
            "reason": "Phát hiện hóa chất suy giảm tầng ô-dôn (ODS). Đề xuất liên kết hạn ngạch Nghị định 06/2022/NĐ-CP."
        })

    return jsonify({
        "nodes": nodes,
        "links": links,
        "suggested_links": suggested_links
    })

@invoices_blueprint.get("/api/compliance/concept-map/expand/<version_id>")
def api_compliance_concept_map_expand(version_id):
    """US-STORY-CONCEPT-MAP-EXPANDER-INTEGRATION: Dynamic concept map expander returning 7-page field guide."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = session.get("taxpayer_mst") or "0102030405"
    base_data_dir = current_app.config.get("BASE_DATA_DIR")
    
    # Query database stats if available
    db_stats = {
        "fuel_logs_count": 0,
        "coal_logs_count": 0,
        "plastic_bag_logs_count": 0,
        "chemical_logs_count": 0,
        "total_violations": 0
    }
    
    try:
        from invoices.multitenant_service import get_tenant_db_path
        import sqlite3
        db_path = get_tenant_db_path(mst, base_data_dir)
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ep_tax_fuel_logs'")
            if cursor.fetchone():
                cursor.execute("SELECT COUNT(*) FROM ep_tax_fuel_logs")
                db_stats["fuel_logs_count"] = cursor.fetchone()[0]
                
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ep_tax_coal_logs'")
            if cursor.fetchone():
                cursor.execute("SELECT COUNT(*) FROM ep_tax_coal_logs")
                db_stats["coal_logs_count"] = cursor.fetchone()[0]
                
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ep_tax_plastic_bag_logs'")
            if cursor.fetchone():
                cursor.execute("SELECT COUNT(*) FROM ep_tax_plastic_bag_logs")
                db_stats["plastic_bag_logs_count"] = cursor.fetchone()[0]
                
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ep_tax_chemical_logs'")
            if cursor.fetchone():
                cursor.execute("SELECT COUNT(*) FROM ep_tax_chemical_logs")
                db_stats["chemical_logs_count"] = cursor.fetchone()[0]
                
            conn.close()
    except Exception:
        pass

    try:
        from invoices.models import Invoice
        invoices_with_warnings = Invoice.query.filter(Invoice.taxpayer_mst == mst).all()
        db_stats["total_violations"] = sum(len(inv.warnings) for inv in invoices_with_warnings if inv.warnings)
    except Exception:
        pass

    # Customize content based on version
    v_clean = version_id.lower().strip()
    
    valid_nodes = {
        "v26", "v27", "v28", "v29", "v30", "v31", "v44", "v45", "v46", "v47", "v48", "v49",
        "v50", "v51", "v52", "v53", "v54", "v55", "v56", "v57", "v58", "v59", "v60", "v61",
        "v62", "v63", "v64", "v65", "v66", "v67", "v68", "v69", "v70"
    }
    
    if v_clean not in valid_nodes:
        return jsonify({
            "status": "error",
            "message": f"Compliance node {version_id} not found."
        }), 404
        
    pages = []
    if v_clean == "v53":
        pages = [
            {
                "title": "1. Định Hướng (Orientation)",
                "content": f"### Mục tiêu chính (True Purpose)\nHiểu rõ và vận hành tự động tính toán, kiểm toán thuế Bảo vệ Môi trường (Environmental Protection Tax - EP Tax) theo Luật số 57/2010/QH12 đối với các mặt hàng xăng dầu, than đá, túi ni-lông và hóa chất HCFC.\n\n### Câu hỏi trọng tâm (Focus Question)\n*Làm thế nào để hệ thống GDT Invoice Hub đối soát tự động hóa đơn xăng dầu, than đá, túi ni-lông và hóa chất nhằm phát hiện nhanh các sai sót và áp dụng đúng quy định miễn thuế?*\n\n### Lời hứa của bản đồ (Map Promise)\nBản đồ này cung cấp đầy đủ danh mục biểu thuế suất tuyệt đối, cơ chế miễn trừ (transit, phát điện, phân hủy sinh học), và liên kết dữ liệu thời gian thực giúp kế toán doanh nghiệp tự động hóa 100% khâu kiểm tra biểu thuế bảo vệ môi trường trên hóa đơn đầu vào."
            },
            {
                "title": "2. Mô Hình Lõi (Core Model)",
                "content": f"### Các Thực Thể & Thuộc Tính Chính\n\n| Thực thể (Entities) | Thuộc tính chính (Attributes) | Loại dữ liệu (Data Type) |\n| :--- | :--- | :--- |\n| **EP Tax Fuel Log** | `fuel_type`, `quantity_litres`, `ep_tax_rate`, `ep_tax_amount`, `is_exempt` | Cấu trúc dữ liệu xăng dầu |\n| **EP Tax Coal Log** | `coal_type`, `quantity_tonnes`, `ep_tax_rate`, `ep_tax_amount`, `is_exempt`, `usage` | Cấu trúc dữ liệu than |\n| **Plastic Bag Log** | `bag_name`, `weight_kg`, `ep_tax_rate`, `is_certified_biodegradable` | Cấu trúc dữ liệu túi nhựa |\n| **Chemical Log** | `chemical_name`, `weight_kg`, `ep_tax_rate`, `ep_tax_amount` | Cấu trúc dữ liệu hóa chất |\n\n### Luồng Xử Lý Chính\n1. Kiểm tra mã hàng hóa hoặc tên sản phẩm trên hóa đơn đầu vào.\n2. Trích xuất số lượng (lít, kg, tấn).\n3. Định tuyến loại hàng hóa đến dịch vụ tính thuế bảo vệ môi trường tương ứng.\n4. Thực hiện đối so sánh chéo với các điều kiện miễn thuế."
            },
            {
                "title": "3. Phân Vùng Phạm Vi (Scope Rings)",
                "content": "### Các Phân Lớp Phạm Vi\n\n*   **Vùng Lõi (Core)**: Tính toán thuế suất tuyệt đối theo Luật EP Tax (Xăng: 2,000đ/l, Dầu diesel: 1,000đ/l, Kerosene: 600đ/l, Túi ni-lông: 50,000đ/kg, Hóa chất HCFC: 5,000đ/kg). Áp dụng các quy tắc miễn thuế đối với hàng tạm nhập tái xuất hoặc than dùng cho phát điện.\n*   **Vùng Cận Biên (Adjacent)**: Cơ chế đa chi nhánh (multitenant DB isolation) và liên kết hóa đơn với các đối tác cung ứng.\n*   **Vùng Biên Giới (Frontier)**: Sử dụng các mô hình AI/NLP để tự động phân loại mặt hàng dựa vào chuỗi văn bản không cấu trúc trên hóa đơn.\n*   **Ngoài Phạm Vi (Out-of-Scope)**: Theo dõi thực tế lượng khí thải môi trường tại nhà máy hoặc kiểm tra chứng chỉ phân hủy sinh học thực địa."
            },
            {
                "title": "4. Ngữ Pháp Liên Kết (Relation Grammar)",
                "content": "### Các Mối Quan Hệ Nguyên Tắc (Relational Propositions)\n\n*   **Định nghĩa (Definition)**: Thuế Bảo vệ Môi trường v53 là thuế gián thu, thu vào sản phẩm, hàng hóa khi sử dụng gây tác động xấu đến môi trường.\n*   **Cơ chế (Mechanism)**: Số thuế phải nộp = Số lượng đơn vị hàng hóa tính thuế × Mức thuế tuyệt đối trên một đơn vị hàng hóa.\n*   **Ràng buộc (Constraint)**: Túi ni-lông chỉ được miễn thuế **nếu và chỉ nếu** có chứng chỉ tự phân hủy sinh học hợp chuẩn được cấp bởi Bộ Tài nguyên và Môi trường.\n*   **Đánh đổi (Trade-off)**: Than đá sử dụng cho mục đích phát điện hoặc xuất khẩu được miễn thuế, tuy nhiên doanh nghiệp phải lưu trữ đầy đủ hồ sơ chứng minh mục đích sử dụng để giải trình khi quyết toán thuế."
            },
            {
                "title": "5. Cơ Chế Vận Hành (Mechanism & Dynamics)",
                "content": "### Quy Trình Vận Hành Quy Tắc Thuế Bảo Vệ Môi Trường\n\n```mermaid\ngraph TD\n    A[Nhận hóa đơn đầu vào] --> B{Phân loại mặt hàng}\n    B -->|Xăng dầu| C[Áp dụng mức tuyệt đối VND/lít]\n    B -->|Than đá| D{Kiểm tra mục đích sử dụng}\n    D -->|Phát điện/Xuất khẩu| E[Miễn thuế 100%]\n    D -->|Khác| F[Áp dụng VND/tấn]\n    B -->|Túi nhựa| G{Có chứng chỉ phân hủy?}\n    G -->|Có| H[Miễn thuế]\n    G -->|Không| I[Áp dụng 50.000 VND/kg]\n```\n\n### Biểu Phí Thuế Tuyệt Đối Tham Chiếu\n*   **Petrol (Xăng)**: 2,000 VND / lít\n*   **Diesel (Dầu Diesel)**: 1,000 VND / lít\n*   **Kerosene (Dầu hỏa)**: 600 VND / lít\n*   **Anthracite Coal (Than Antracit)**: 30,000 VND / tấn\n*   **Lignite Coal (Than nâu)**: 20,000 VND / tấn\n*   **Plastic Bag (Túi ni-lông)**: 50,000 VND / kg\n*   **HCFC Chemical (Hóa chất HCFC)**: 5,000 VND / kg"
            },
            {
                "title": "6. Giới Hạn & Lỗi Thường Gặp (Boundaries & Failure Cases)",
                "content": f"### Dữ Liệu Thực Tế Hệ Thống (Live Telemetry Statistics)\n\n*   MST Đang Xem: **{mst}**\n*   Số bản ghi Xăng dầu đã xử lý: **{db_stats['fuel_logs_count']}**\n*   Số bản ghi Than đá đã xử lý: **{db_stats['coal_logs_count']}**\n*   Số bản ghi Túi nhựa đã kiểm tra: **{db_stats['plastic_bag_logs_count']}**\n*   Số bản ghi Hóa chất đã xử lý: **{db_stats['chemical_logs_count']}**\n*   Tổng số cảnh báo lỗi/vi phạm phát hiện: **{db_stats['total_violations']}**\n\n### Các Tình Huống Sai Sót Thường Gặp (Failure Modes)\n1. **Sai lệch đơn vị tính**: Hóa đơn túi ni-lông ghi đơn vị tính là 'Cái' thay vì 'kg', dẫn đến lỗi không tính được khối lượng tính thuế.\n2. **Khai báo miễn thuế không hợp lệ**: Tích chọn miễn thuế xăng dầu nhưng không có hồ sơ chứng minh xuất khẩu/tạm nhập tái xuất.\n3. **Sai mã hóa chất**: Tên hóa chất chứa HCFC nhưng viết sai định dạng viết tắt, làm trôi lọt kiểm tra kiểm toán."
            },
            {
                "title": "7. Ứng Dụng & Lộ Trình Học Tập (Application & Learning Path)",
                "content": "### Lộ Trình Áp Dụng Trong Thực Tế\n*   **Bước 1**: Tích hợp API đối soát v53 vào luồng hóa đơn đầu vào của bộ phận Mua hàng.\n*   **Bước 2**: Thiết lập cảnh báo sớm trên Dashboard khi tỷ lệ thuế EP Tax trên đơn giá hàng hóa vượt ngưỡng an toàn.\n*   **Bước 3**: Chạy hậu kiểm định kỳ cuối tháng đối với toàn bộ tờ khai thuế Bảo vệ Môi trường mẫu 01/TBVMT.\n\n### Tài Liệu Nghiên Cứu Đề Xuất\n1. *Luật Thuế bảo vệ môi trường số 57/2010/QH12*\n2. *Thông tư số 152/2011/TT-BTC hướng dẫn thi hành Luật Thuế bảo vệ môi trường*\n3. *Nghị quyết số 579/2018/UBTVQH14 về biểu thuế bảo vệ môi trường*"
            }
        ]
    elif v_clean == "v70":
        pages = [
            {
                "title": "1. Định Hướng (Orientation)",
                "content": "### Mục tiêu chính (True Purpose)\nQuản lý và kiểm soát hạn ngạch nhập khẩu, sản xuất và sử dụng các chất làm suy giảm tầng ô-dôn (Ozone-Depleting Substances - ODS) theo quy định tại Nghị định số 06/2022/NĐ-CP của Chính phủ.\n\n### Câu hỏi trọng tâm (Focus Question)\n*Làm thế nào để hệ thống giám sát tự động hạn ngạch tiêu thụ ODS của doanh nghiệp dựa trên hóa đơn nhập khẩu và giấy phép đăng ký hạn ngạch hàng năm?*\n\n### Lời hứa của bản đồ (Map Promise)\nCung cấp góc nhìn toàn diện về quy trình cấp phép ODS, ngưỡng hạn ngạch tối đa và cảnh báo tự động khi doanh nghiệp tiệm cận giới hạn cho phép nhằm tránh các chế tài pháp lý nghiêm khắc."
            },
            {
                "title": "2. Mô Hình Lõi (Core Model)",
                "content": "### Các Thực Thể Quản Lý ODS\n*   **Taxpayer Quota (Hạn ngạch MST)**: Lưu trữ hạn ngạch ODS tối đa được Bộ TNMT cấp phép trong năm.\n*   **ODS Consumption (Lượng tiêu thụ)**: Lượng ODS thực tế nhập khẩu/mua vào thông qua hóa đơn đầu vào.\n*   **License Verification (Giấy phép)**: Trạng thái và thời hạn giấy phép nhập khẩu ODS tương ứng.\n\n### Biểu thức tính hạn ngạch còn lại\n`Lượng hạn ngạch còn lại = Hạn ngạch được cấp - Tổng lượng nhập khẩu đã đối soát`"
            },
            {
                "title": "3. Phân Vùng Phạm Vi (Scope Rings)",
                "content": "### Phân Lớp Quản Lý ODS v70\n*   **Vùng Lõi (Core)**: Quản lý lượng hạn ngạch cấp phép, trừ lùi hạn ngạch tự động qua hóa đơn đầu vào, cảnh báo khi lượng mua vượt quá hạn ngạch cho phép.\n*   **Vùng Cận Biên (Adjacent)**: Đồng bộ dữ liệu tờ khai hải quan nhập khẩu (customs declaration) để so sánh chéo khối lượng thực nhập.\n*   **Vùng Biên Giới (Frontier)**: Tự động dự báo xu hướng tiêu thụ ODS dựa trên kế hoạch sản xuất để đề xuất xin thêm hạn ngạch sớm.\n*   **Ngoài Phạm Vi (Out-of-Scope)**: Đo đạc nồng độ hóa chất bay hơi trực tiếp trong nhà xưởng hoặc kiểm tra hiện trường rò rỉ khí gas."
            },
            {
                "title": "4. Ngữ Pháp Liên Kết (Relation Grammar)",
                "content": "### Các Mối Quan Hệ Hạn Ngạch ODS\n*   **Định nghĩa (Definition)**: ODS bao gồm các chất chứa clo, brom gây suy giảm tầng ô-dôn như CFC, Halon, HCFC, và methyl bromide.\n*   **Ràng buộc (Constraint)**: Việc nhập khẩu ODS **phải** được Bộ Tài nguyên và Môi trường phân bổ hạn ngạch nhập khẩu.\n*   **Cơ chế (Mechanism)**: Lượng ODS thực tế tính theo tấn khí tương đương CO2 hoặc trọng lượng thuần tùy loại hóa chất quy định trong danh mục phụ lục Nghị định 06/2022/NĐ-CP."
            },
            {
                "title": "5. Cơ Chế Vận Hành (Mechanism & Dynamics)",
                "content": "### Luồng Duyệt Giấy Phép & Hạn Ngạch ODS\n1. Doanh nghiệp tải tờ khai hải quan XML hoặc hóa đơn nhập khẩu ODS lên GDT Invoice Hub.\n2. Hệ thống đọc mã HS của hóa chất và đối chiếu với danh mục chất kiểm soát ODS v70.\n3. Hệ thống tính toán lượng tiêu thụ thực tế quy đổi.\n4. Thực hiện kiểm tra hạn ngạch còn lại:\n   - Nếu lượng tiêu thụ vượt hạn ngạch: Kích hoạt cảnh báo **Nguy cấp (Critical Block)**.\n   - Nếu hạn ngạch còn dưới 10%: Kích hoạt cảnh báo **Cận giới hạn (Warning)**."
            },
            {
                "title": "6. Giới Hạn & Lỗi Thường Gặp (Boundaries & Failure Cases)",
                "content": "### Thống Kê & Cảnh Báo Hạn Ngạch ODS\n*   Trạng thái cấp hạn ngạch năm hiện tại: **Đang hoạt động**\n*   Số lượng hóa đơn ODS đã ghi nhận: **0 hóa đơn**\n*   Các vi phạm hạn ngạch phát hiện: **0 cảnh báo**\n\n### Các Tình Huống Vi Phạm Lỗi\n1. **Nhập khẩu không giấy phép**: Doanh nghiệp khai báo mua hóa chất HCFC nhưng giấy phép nhập khẩu đã hết hạn hoặc chưa được duyệt.\n2. **Sai hệ số quy đổi**: Khai báo trọng lượng khí hóa lỏng không đúng thể tích nén thực tế dẫn đến tính sai lượng hạn ngạch tiêu hao."
            },
            {
                "title": "7. Ứng Dụng & Lộ Trình Học Tập (Application & Learning Path)",
                "content": "### Kế Hoạch Triển Khai\n*   **Tháng 1**: Khai báo và cấu hình định mức hạn ngạch ODS được cấp vào Profile của Doanh nghiệp trên GDT Hub.\n*   **Tháng 2**: Kích hoạt bộ lọc cảnh báo tự động trên phân hệ Hải quan / Mua vào.\n*   **Tháng 3**: Kết xuất báo cáo sử dụng chất ODS định kỳ gửi Cục Biến đổi khí hậu.\n\n### Tài liệu tham khảo chính\n*   *Nghị định số 06/2022/NĐ-CP quy định chi tiết giảm nhẹ phát thải khí nhà kính và bảo vệ tầng ô-dôn*\n*   *Thông tư số 01/2022/TT-BTNMT quy định chi tiết thi hành Luật Bảo vệ môi trường về ứng phó với biến đổi khí hậu*"
            }
        ]
    elif v_clean == "v26":
        pages = [
            {
                "title": "1. Định Hướng (Orientation)",
                "content": "### Mục tiêu chính (True Purpose)\nTối ưu hóa quyết toán Thuế thu nhập doanh nghiệp (Corporate Income Tax - CIT) v26 và quản lý chặt chẽ các chi phí không được trừ khi tính thuế CIT.\n\n### Câu hỏi trọng tâm (Focus Question)\n*Làm thế nào để hệ thống tự động nhận diện các hóa đơn có rủi ro chi phí không được trừ (ví dụ: hóa đơn khống, mua sắm cá nhân, chi phí vượt định mức) phục vụ quyết toán CIT v26?*\n\n### Lời hứa của bản đồ (Map Promise)\nGiúp kế toán trưởng và giám đốc tài chính nắm bắt toàn bộ sơ đồ logic xác định thu nhập tính thuế CIT, cấu trúc các khoản chi phí hợp lý hợp lệ, và đối chiếu tờ khai quyết toán thuế mẫu 03/TNDN nhanh chóng."
            },
            {
                "title": "2. Mô Hình Lõi (Core Model)",
                "content": "### Cấu Trúc Mô Hình Tính CIT\n*   **Thu nhập chịu thuế**: = Doanh thu - Chi phí được trừ + Các khoản thu nhập khác.\n*   **Thu nhập tính thuế**: = Thu nhập chịu thuế - Thu nhập được miễn thuế - Các khoản lỗ được kết chuyển.\n*   **Thuế CIT phải nộp**: = Thu nhập tính thuế × Thuế suất (Mặc định 20% hoặc mức ưu đãi).\n\n### Phân loại chi phí trên hệ thống\n*   **Deductible (Chi phí được trừ)**: Đầy đủ hóa đơn, chứng từ thanh toán không dùng tiền mặt nếu từ 20 triệu đồng trở lên.\n*   **Non-deductible (Chi phí không được trừ)**: Hóa đơn mua sắm cá nhân, chi phí lãi vay vượt trần EBITDA 30% (Nghị định 132), chi phí không phục vụ sản xuất kinh doanh."
            },
            {
                "title": "3. Phân Vùng Phạm Vi (Scope Rings)",
                "content": "### Phạm Vi Ứng Dụng Thuế CIT v26\n*   **Vùng Lõi (Core)**: Tính toán thuế suất CIT cơ bản (20%), phân loại hóa đơn đầu vào hợp lệ/không hợp lệ, ghi nhận chi phí được trừ.\n*   **Vùng Cận Biên (Adjacent)**: Liên kết với phân hệ Ngân hàng để tự động kiểm soát chứng từ thanh toán không dùng tiền mặt đối với các hóa đơn giá trị từ 20 triệu đồng.\n*   **Vùng Biên Giới (Frontier)**: Phân tích dự báo số thuế CIT tạm nộp hàng quý nhằm tối ưu dòng tiền doanh nghiệp.\n*   **Ngoài Phạm Vi (Out-of-Scope)**: Kế toán quản trị nội bộ hoặc lập báo cáo tài chính quốc tế IFRS."
            },
            {
                "title": "4. Ngữ Pháp Liên Kết (Relation Grammar)",
                "content": "### Các Quy Tắc Liên Kết Thuế CIT\n*   **Ràng buộc chi phí (Constraint)**: Khoản chi có hóa đơn từ 20 triệu đồng trở lên **bắt buộc** phải có chứng từ thanh toán không dùng tiền mặt để được tính là chi phí được trừ.\n*   **Đánh đổi ưu đãi (Trade-off)**: Hưởng thuế suất ưu đãi CIT tại khu công nghiệp yêu cầu doanh nghiệp phải hạch toán độc lập doanh thu và chi phí của dự án đầu tư ưu đãi đó."
            },
            {
                "title": "5. Cơ Chế Vận Hành (Mechanism & Dynamics)",
                "content": "### Luồng Đối Soát Chi Phí Quyết Toán CIT\n1. Phân tích hóa đơn mua vào tự động.\n2. Kiểm tra giá trị hóa đơn:\n   - Nếu giá trị < 20.000.000 VND: Chấp nhận hạch toán chi phí hợp lệ thông thường.\n   - Nếu giá trị >= 20.000.000 VND: Yêu cầu đối chiếu với dữ liệu ngân hàng để tìm chứng từ chuyển khoản tương ứng.\n3. Phát hiện rủi ro doanh nghiệp ma: Đối chiếu MST người bán với danh sách doanh nghiệp tạm ngừng hoạt động hoặc bỏ địa chỉ kinh doanh.\n4. Kết xuất báo cáo chi phí không được trừ ước tính cuối kỳ."
            },
            {
                "title": "6. Giới Hạn & Lỗi Thường Gặp (Boundaries & Failure Cases)",
                "content": f"### Chỉ Số Kiểm Toán CIT v26\n*   Thuế suất CIT mặc định áp dụng: **20%**\n*   Tổng chi phí nghi ngờ (Không được trừ): **{db_stats['total_violations'] * 1250000:,.0f} VND**\n*   Số hóa đơn đầu vào cần đối soát thanh toán: **{db_stats['fuel_logs_count'] + db_stats['coal_logs_count']} hóa đơn**\n\n### Tình huống lỗi điển hình\n*   **Thanh toán sai phương thức**: Trả tiền mặt cho hóa đơn mua xăng dầu có tổng giá trị thanh toán 25 triệu đồng (bao gồm VAT).\n*   **Lãi vay vượt trần**: Chi phí lãi vay vượt mức 30% EBITDA do không theo dõi quan hệ liên kết của các công ty mẹ con."
            },
            {
                "title": "7. Ứng Dụng & Lộ Trình Học Tập (Application & Learning Path)",
                "content": "### Lộ Trình Học Tập & Triển Khai\n*   **Bước 1**: Rà soát lại toàn bộ hóa đơn mua vào cuối mỗi quý bằng chức năng đối soát tự động của GDT Invoice Hub.\n*   **Bước 2**: Lập bảng kê các khoản chi phí không được trừ để điều chỉnh trên chỉ tiêu B4 của tờ khai quyết toán thuế CIT.\n\n### Luật tham chiếu\n*   **Luật Thuế thu nhập doanh nghiệp số 14/2008/QH12** và các luật sửa đổi bổ sung.\n*   **Thông tư số 78/2014/TT-BTC** hướng dẫn thi hành Luật Thuế thu nhập doanh nghiệp."
            }
        ]
    else:
        # Default fallback structure for any other nodes
        label = version_id
        pages = [
            {
                "title": "1. Định Hướng (Orientation)",
                "content": f"### Mục tiêu chính (True Purpose)\nTài liệu hướng dẫn và vận hành chi tiết phân hệ **{label}** trong hệ thống GDT Invoice Hub.\n\n### Câu hỏi trọng tâm (Focus Question)\n*Làm thế nào để ứng dụng và tích hợp quy tắc {label} nhằm tăng cường tính tuân thủ thuế và quản lý rủi ro hóa đơn?*\n\n### Lời hứa của bản đồ (Map Promise)\nCung cấp các thông tin nền tảng về khái niệm, cách thiết lập cấu hình và biểu đồ vận hành của phân hệ này."
            },
            {
                "title": "2. Mô Hình Lõi (Core Model)",
                "content": f"### Cấu Trúc Khái Niệm Phân Hệ {label}\nPhân hệ này đảm nhận việc xử lý các ràng buộc nghiệp vụ liên quan đến **{label}**, thực hiện thu thập dữ liệu hóa đơn đầu vào, kiểm tra định dạng XML, đối sánh tham số và kích hoạt các cảnh báo rủi ro tương ứng."
            },
            {
                "title": "3. Phân Vùng Phạm Vi (Scope Rings)",
                "content": f"### Phân Lớp Phạm Vi của {label}\n*   **Vùng Lõi (Core)**: Các quy tắc cơ bản trực tiếp ảnh hưởng đến trạng thái hợp lệ của hóa đơn.\n*   **Vùng Cận Biên (Adjacent)**: Liên kết dữ liệu giữa các phân hệ quản lý thuế phụ thuộc.\n*   **Vùng Biên Giới (Frontier)**: Các tính năng mở rộng ứng dụng AI tự động hóa kiểm tra.\n*   **Ngoài Phạm Vi (Out-of-Scope)**: Quy trình kế toán độc lập bên ngoài hệ thống."
            },
            {
                "title": "4. Ngữ Pháp Liên Kết (Relation Grammar)",
                "content": f"### Các Nguyên Tắc Mối Quan Hệ\n*   **Định nghĩa (Definition)**: {label} là một phân hệ thành phần của hệ thống quản lý tuân thủ thuế GDT Invoice Hub.\n*   **Ràng buộc (Constraint)**: Mọi dữ liệu hóa đơn nhập vào phân hệ phải tuân thủ đúng định dạng chuẩn XML quy định bởi Tổng cục Thuế."
            },
            {
                "title": "5. Cơ Chế Vận Hành (Mechanism & Dynamics)",
                "content": f"### Luồng Vận Hành Chung\n1. Nhận thông tin hóa đơn từ luồng đồng bộ daemon.\n2. Phân tích các thẻ dữ liệu cấu trúc tương ứng.\n3. Áp dụng tập luật kiểm toán của {label}.\n4. Ghi nhận nhật ký cảnh báo và cập nhật điểm xếp hạng tín nhiệm thuế của doanh nghiệp."
            },
            {
                "title": "6. Giới Hạn & Lỗi Thường Gặp (Boundaries & Failure Cases)",
                "content": "### Chỉ số hoạt động\n*   Trạng thái phân hệ: **Đang hoạt động**\n*   Mức độ rủi ro cấu hình: **Trung bình**\n*   Số lỗi phát hiện trong kỳ: **0 cảnh báo**"
            },
            {
                "title": "7. Ứng Dụng & Lộ Trình Học Tập (Application & Learning Path)",
                "content": f"### Lộ trình áp dụng\n*   **Bước 1**: Đọc tài liệu đặc tả nghiệp vụ phân hệ {label}.\n*   **Bước 2**: Thực hiện cấu hình tham số kiểm soát tương thích với mô hình kinh doanh của doanh nghiệp.\n*   **Bước 3**: Theo dõi định kỳ báo cáo kiểm toán rủi ro trên trang Dashboard chính."
            }
        ]
        
    return jsonify({
        "status": "success",
        "version_id": version_id,
        "mst": mst,
        "pages": pages
    })

@invoices_blueprint.get("/api/config")
def api_config():
    """Return small frontend configuration flags."""

    return jsonify({"mock_mode": current_app.config["GDT_USE_MOCK"], "locale": "vi-VN"})

@invoices_blueprint.get("/api/invoices")
def api_invoices():
    """Return invoices in JSON for the requested date range."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        cancelled_only = request.args.get("cancelled_only", "false").lower() == "true"
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, cancelled_only, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        session["invoice_lookup"] = build_invoice_lookup(invoices) if "invoices" in locals() else {}
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"total_count": len(invoices), "invoices": invoices})

@invoices_blueprint.get("/api/sync/events")
def sse_sync_stream():
    """SSE endpoint to stream real-time sync progress to the frontend."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.sync_daemon import get_sse_stream
    from flask import Response
    
    return Response(get_sse_stream(), mimetype="text/event-stream")

@invoices_blueprint.get("/api/cancelled-invoices")

def api_cancelled_invoices():
    """Return cancelled invoices using the same date filters."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, True, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        session["invoice_lookup"] = build_invoice_lookup(invoices) if "invoices" in locals() else {}
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"total_count": len(invoices), "cancelled_invoices": invoices})

@invoices_blueprint.get("/api/invoices/<invoice_id>/download")
def api_download_invoice(invoice_id: str):
    """Download one invoice XML file."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
        xml_bytes = download_invoice_xml(invoice_id)
        invoice = (current_app.config.get("CURRENT_INVOICE_LOOKUP") or {}).get(invoice_id)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except NotImplementedError as error:
        return jsonify({"error": str(error)}), 501
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    filename = resolve_live_download_name(invoice) if invoice else f"invoice_{invoice_id}.xml"
    return Response(
        xml_bytes,
        mimetype="application/zip" if filename.endswith(".zip") else "application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@invoices_blueprint.get("/api/invoices/<invoice_id>/details")
def api_invoice_details(invoice_id: str):
    """Return the detailed line items for a specific invoice."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    session_inv = None
    try:
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
        line_items = fetch_invoice_line_items(invoice_id)
        session_inv = (current_app.config.get("CURRENT_INVOICE_LOOKUP") or {}).get(invoice_id)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except NotImplementedError as error:
        return jsonify({"error": str(error)}), 501
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    from invoices.service import get_local_invoices
    local_invoices = get_local_invoices()
    local_inv = None
    for inv in local_invoices:
        if inv["id"] == invoice_id:
            local_inv = inv
            break

    warnings = local_inv.get("warnings", []) if local_inv else []
    is_valid = local_inv.get("is_valid", True) if local_inv else True

    payment_method = ""
    if local_inv:
        payment_method = local_inv.get("payment_method", "")
    elif session_inv:
        payment_method = session_inv.get("payment_method") or session_inv.get("raw", {}).get("htttoan") or ""

    from invoices.models import Invoice
    invoice = db.session.get(Invoice, invoice_id)
    ai_warnings = []
    if invoice:
        ai_warnings = [w.to_dict() for w in invoice.ai_audit_results]

    return jsonify({
        "invoice_id": invoice_id,
        "line_items": line_items,
        "warnings": warnings,
        "is_valid": is_valid,
        "payment_method": payment_method,
        "ai_warnings": ai_warnings,
        "ai_audited": invoice.ai_audited if invoice else False,
        "signature_details": invoice.signature_details if invoice else None
    })

@invoices_blueprint.get("/api/invoices/stats")
def api_invoices_stats():
    """Return financial statistics and aggregations for the requested date range."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        
        # Check hybrid stats cache first (US-124)
        mst = session.get("tax_code")
        from_str = parsed_from.isoformat()
        to_str = parsed_to.isoformat()
        
        from invoices.stats_cache import get_cached_stats, set_cached_stats
        cached_result = get_cached_stats(mst, from_str, to_str, direction)
        if cached_result is not None:
            return jsonify(cached_result)

        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    total_spend = 0.0
    total_tax = 0.0
    active_count = 0
    cancelled_count = 0

    vendor_stats = {}
    tax_breakdown = {"0%": 0.0, "5%": 0.0, "8%": 0.0, "10%": 0.0, "khac": 0.0}

    for inv in invoices:
        amount = inv.get("amount", 0.0)
        is_cancelled = inv.get("is_cancelled", False)

        if is_cancelled:
            cancelled_count += 1
        else:
            active_count += 1
            total_spend += amount

            vendor = inv.get("issuer", "Khong ro")
            if vendor not in vendor_stats:
                vendor_stats[vendor] = {"spend": 0.0, "count": 0}
            vendor_stats[vendor]["spend"] += amount
            vendor_stats[vendor]["count"] += 1

            line_items = inv.get("line_items", [])
            for item in line_items:
                rate = str(item.get("tax_rate", "10%")).strip()
                tax_amt = item.get("tax_amount", 0.0)
                total_tax += tax_amt

                if "10" in rate:
                    tax_breakdown["10%"] += tax_amt
                elif "8" in rate:
                    tax_breakdown["8%"] += tax_amt
                elif "5" in rate:
                    tax_breakdown["5%"] += tax_amt
                elif "0" in rate:
                    tax_breakdown["0%"] += tax_amt
                else:
                    tax_breakdown["khac"] += tax_amt

    top_vendors = []
    for vendor, data in vendor_stats.items():
        top_vendors.append({"name": vendor, "spend": data["spend"], "count": data["count"]})
    top_vendors.sort(key=lambda x: x["spend"], reverse=True)
    top_vendors = top_vendors[:5]

    response_payload = {
        "total_spend": total_spend,
        "total_tax": total_tax,
        "active_count": active_count,
        "cancelled_count": cancelled_count,
        "top_vendors": top_vendors,
        "tax_breakdown": tax_breakdown,
    }
    
    # Store calculated stats in hybrid cache
    set_cached_stats(mst, from_str, to_str, direction, response_payload)

    return jsonify(response_payload)

@invoices_blueprint.get("/api/export-excel")
def api_export_excel():
    """Export invoice search results to an Excel workbook download."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        cancelled_only = request.args.get("cancelled_only", "false").lower() == "true"
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, cancelled_only, direction))
        workbook_bytes = generate_excel_workbook(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    filename = f"invoices_{parsed_from.isoformat()}_{parsed_to.isoformat()}.xlsx"
    return send_file(
        BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@invoices_blueprint.get("/api/erp/export/misa")
def api_erp_export_misa():
    """Export selected or all invoices to a MISA-compatible Excel template."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    ids_str = request.args.get("ids", "")
    from invoices.models import Invoice
    if ids_str:
        invoices = Invoice.query.filter(Invoice.id.in_(ids_str.split(","))).all()
    else:
        invoices = Invoice.query.all()

    from invoices.erp_service import generate_misa_export
    try:
        excel_bytes = generate_misa_export(invoices)
        filename = "misa_export.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất MISA: {str(e)}"}), 500

@invoices_blueprint.get("/api/erp/export/odoo")
def api_erp_export_odoo():
    """Export selected or all invoices to an Odoo-compatible CSV template."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    ids_str = request.args.get("ids", "")
    from invoices.models import Invoice
    if ids_str:
        invoices = Invoice.query.filter(Invoice.id.in_(ids_str.split(","))).all()
    else:
        invoices = Invoice.query.all()

    from invoices.erp_service import generate_odoo_export
    try:
        csv_str = generate_odoo_export(invoices)
        filename = "odoo_export.csv"
        return send_file(
            BytesIO(csv_str.encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Odoo: {str(e)}"}), 500

@invoices_blueprint.get("/api/partners")
def api_partners():
    """Extract and return corporate business partners and their statistics, or return catalog if no date range is provided."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    if not date_from and not date_to:
        from invoices.models import Partner
        try:
            partners = Partner.query.order_by(Partner.mst).all()
            return jsonify([p.to_dict() for p in partners])
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    try:
        parsed_from, parsed_to = validate_date_range(
            date_from,
            date_to,
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        partners = extract_partners_from_invoices(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"partners": partners})

@invoices_blueprint.get("/api/partners/<mst>/status")
def api_partner_status(mst):
    """Force an on-demand MST tax status verification."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.mst_service import check_mst_status
    try:
        result = check_mst_status(mst, force_refresh=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/usage")
def api_reports_usage():
    """Aggregate and return BC26 Vietnamese tax invoice usage compliance tables."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "sold")  # Default to sold for business output tracking
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        report = generate_tax_usage_report(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    return jsonify({"report": report})

@invoices_blueprint.get("/api/invoices/<invoice_id>/pdf-view")
def api_invoice_pdf_view(invoice_id):
    """Render a beautiful, printable official-style HTML/CSS Vietnamese electronic invoice."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    # Locate invoice
    current_app.config["CURRENT_JWT"] = session.get("jwt")
    current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
    try:
        from datetime import date
        invoice = current_app.config["CURRENT_INVOICE_LOOKUP"].get(invoice_id)
        
        if not invoice:
            invoices_purchase = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "purchase"))
            invoice = build_invoice_lookup(invoices_purchase).get(invoice_id)
            
        if not invoice:
            invoices_sold = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "sold"))
            invoice = build_invoice_lookup(invoices_sold).get(invoice_id)

        if not invoice:
            from invoices.service import get_local_invoices
            local_db = get_local_invoices()
            for item in local_db:
                if item["id"] == invoice_id:
                    invoice = item
                    break
            
        if not invoice:
            return "Khong tim thay hoa don yeu cau.", 404

        line_items = fetch_invoice_line_items(invoice_id)
    except FileNotFoundError:
        return "Khong tim thay hoa don yeu cau.", 404
    except Exception as error:
        return f"Loi he thong: {str(error)}", 500
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    # Calculate sums
    sum_before_tax = sum(item.get("amount_before_tax", 0.0) for item in line_items)
    sum_tax = sum(item.get("tax_amount", 0.0) for item in line_items)
    total_payable = sum_before_tax + sum_tax

    # Auto buyer/seller properties
    user_company = {
        "name": "CONG TY CO PHAN CONG NGHE GDT INVOICE HUB",
        "mst": "0109999999",
        "address": "Toa nha Technopark, Gia Lam, TP. Ha Noi",
        "phone": "1900 8888",
    }
    
    partner_details = {
        "Cong ty A": {"mst": "0101234567", "address": "So 10 Pho Hue, Quan Hai Ba Trung, Ha Noi"},
        "Cong ty B": {"mst": "0209876543", "address": "250 Nguyen Thi Minh Khai, Quan 3, TP. Ho Chi Minh"},
        "Cong ty C": {"mst": "0301122334", "address": "15 Le Loi, Quan Hai Chau, Da Nang"},
    }
    
    if "seller_name" in invoice:
        seller = {
            "name": invoice.get("seller_name", ""),
            "mst": invoice.get("seller_mst", ""),
            "address": invoice.get("seller_address", ""),
            "phone": invoice.get("seller_phone", ""),
        }
        buyer = {
            "name": invoice.get("buyer_name", ""),
            "mst": invoice.get("buyer_mst", ""),
            "address": invoice.get("buyer_address", ""),
        }
    else:
        issuer = invoice.get("issuer", "Doi tac khac")
        partner = partner_details.get(
            issuer,
            {
                "mst": f"0{abs(hash(issuer)) % 1000000000:09d}",
                "address": f"Khu cong nghiep Binh Duong, Tinh Binh Duong",
            }
        )
        partner["name"] = issuer

        # If it is a purchase invoice, issuer is Seller, user_company is Buyer
        if invoice.get("direction", "purchase") == "purchase":
            seller = partner
            buyer = user_company
        else:
            seller = user_company
            buyer = partner

    return render_template(
        "invoice_pdf.html",
        invoice=invoice,
        line_items=line_items,
        seller=seller,
        buyer=buyer,
        sum_before_tax=sum_before_tax,
        sum_tax=sum_tax,
        total_payable=total_payable,
    )

@invoices_blueprint.post("/api/invoices/batch-download")
def api_batch_download_invoices():
    """Fetch and package all GDT invoices for a month in a ZIP archive (Asynchronous)."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    month = payload.get("month", "").strip()  # Format: YYYY-MM
    direction = payload.get("direction", "purchase").strip()
    duplicate_strategy = payload.get("duplicate_strategy", "overwrite").strip()

    if not month:
        return jsonify({"error": "Vui long chon thang can tai."}), 400

    task_id = str(uuid.uuid4())

    with DOWNLOAD_TASKS_LOCK:
        DOWNLOAD_TASKS[task_id] = {
            "task_id": task_id,
            "status": "pending",
            "progress": 0,
            "completed_count": 0,
            "total": 0,
            "imported_count": 0,
            "skipped_count": 0,
            "overwritten_count": 0,
            "failed_count": 0,
            "error": None,
            "zip_bytes": None,
            "created_at": datetime.now().isoformat()
        }

    # Fetch configuration for child thread execution
    jwt_token = session.get("jwt")
    username = session.get("username")
    encrypted_password = session.get("encrypted_password")
    invoice_lookup = session.get("invoice_lookup", {})
    gdt_use_mock = current_app.config.get("GDT_USE_MOCK", True)
    gdt_base_url = current_app.config.get("GDT_BASE_URL")
    gdt_timeout = current_app.config.get("GDT_TIMEOUT_SECONDS", 10)
    app_instance = current_app._get_current_object()

    def run_task():
        with app_instance.app_context():
            # Inject active context configs
            app_instance.config["CURRENT_JWT"] = jwt_token
            app_instance.config["CURRENT_USERNAME"] = username
            app_instance.config["CURRENT_ENCRYPTED_PASSWORD"] = encrypted_password
            app_instance.config["CURRENT_INVOICE_LOOKUP"] = invoice_lookup
            app_instance.config["GDT_USE_MOCK"] = gdt_use_mock
            app_instance.config["GDT_BASE_URL"] = gdt_base_url
            app_instance.config["GDT_TIMEOUT_SECONDS"] = gdt_timeout

            def on_progress(completed, total, status, error=None, zip_bytes=None, imported=0, skipped=0, overwritten=0, failed=0):
                with DOWNLOAD_TASKS_LOCK:
                    if task_id in DOWNLOAD_TASKS:
                        task = DOWNLOAD_TASKS[task_id]
                        task["completed_count"] = completed
                        task["total"] = total
                        task["status"] = status
                        task["error"] = error
                        task["imported_count"] = imported
                        task["skipped_count"] = skipped
                        task["overwritten_count"] = overwritten
                        task["failed_count"] = failed
                        if zip_bytes:
                            task["zip_bytes"] = zip_bytes
                        if total > 0:
                            task["progress"] = int((completed / total) * 100)

            try:
                from invoices.service import batch_download_invoices
                batch_download_invoices(month, direction, on_progress=on_progress, duplicate_strategy=duplicate_strategy)
            except Exception as e:
                on_progress(0, 0, "failed", error=str(e))
            finally:
                # Clean thread configurations
                app_instance.config["CURRENT_JWT"] = None
                app_instance.config["CURRENT_USERNAME"] = None
                app_instance.config["CURRENT_ENCRYPTED_PASSWORD"] = None
                app_instance.config["CURRENT_INVOICE_LOOKUP"] = {}


    thread = threading.Thread(target=run_task, name=f"BatchDownloadThread-{task_id}")
    thread.daemon = True
    thread.start()

    return jsonify({"task_id": task_id, "status": "pending"}), 202

@invoices_blueprint.get("/api/invoices/batch-download/status/<task_id>")
def api_batch_download_status(task_id):
    """Check progress of a batch download task."""

    with DOWNLOAD_TASKS_LOCK:
        task = DOWNLOAD_TASKS.get(task_id)
        if not task:
            return jsonify({"error": "Khong tim thay thong tin tien trinh tai."}), 404

        return jsonify({
            "task_id": task["task_id"],
            "status": task["status"],
            "progress": task["progress"],
            "completed_count": task["completed_count"],
            "total": task["total"],
            "imported_count": task.get("imported_count", 0),
            "skipped_count": task.get("skipped_count", 0),
            "overwritten_count": task.get("overwritten_count", 0),
            "failed_count": task.get("failed_count", 0),
            "error": task["error"]
        })

@invoices_blueprint.get("/api/invoices/batch-download/download/<task_id>")
def api_batch_download_retrieve(task_id):
    """Retrieve the generated ZIP file for a completed batch download task."""

    with DOWNLOAD_TASKS_LOCK:
        task = DOWNLOAD_TASKS.get(task_id)
        if not task:
            return jsonify({"error": "Khong tim thay file tai ve cho phien nay."}), 404

        if task["status"] != "completed" or not task.get("zip_bytes"):
            return jsonify({"error": f"Tien trinh chua hoan thanh. Trang thai: {task['status']}"}), 400

        zip_bytes = task["zip_bytes"]
        # Clear ZIP memory to prevent leaks
        del DOWNLOAD_TASKS[task_id]

    import io
    return send_file(
        io.BytesIO(zip_bytes),
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"GDT_Invoices_Batch.zip"
    )

@invoices_blueprint.post("/api/invoices/upload")
@roles_required("admin", "auditor")
def api_upload_invoices():
    """Import drag-and-drop XML/ZIP invoices and run smart MISA audits."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "files" not in request.files:
        return jsonify({"error": "Khong tim thay tep tin duoc tai len."}), 400

    files = request.files.getlist("files")
    if not files or files[0].filename == "":
        return jsonify({"error": "Khong co tep tin nao duoc chon."}), 400

    duplicate_strategy = request.form.get("duplicate_strategy", "overwrite").strip()

    from invoices.service import import_xml_invoice
    import zipfile

    imported_count = 0
    skipped_count = 0
    overwritten_count = 0
    errors = []

    for file in files:
        filename = file.filename
        try:
            file_bytes = file.read()
            if filename.lower().endswith(".xml"):
                res = import_xml_invoice(file_bytes, filename, duplicate_strategy=duplicate_strategy)
                status = res.get("import_status", "imported")
                if status == "skipped":
                    skipped_count += 1
                elif status == "overwritten":
                    overwritten_count += 1
                else:
                    imported_count += 1
            elif filename.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                    for zinfo in z.infolist():
                        if zinfo.filename.lower().endswith(".xml") and not zinfo.is_dir():
                            xml_content = z.read(zinfo.filename)
                            base_xml_name = os.path.basename(zinfo.filename)
                            res = import_xml_invoice(xml_content, base_xml_name, duplicate_strategy=duplicate_strategy)
                            status = res.get("import_status", "imported")
                            if status == "skipped":
                                skipped_count += 1
                            elif status == "overwritten":
                                overwritten_count += 1
                            else:
                                imported_count += 1
            else:
                errors.append(f"Tep {filename} khong dung dinh dang XML hoac ZIP.")
        except Exception as e:
            errors.append(f"Loi khi nhap tep {filename}: {str(e)}")

    return jsonify({
        "status": "success",
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "overwritten_count": overwritten_count,
        "errors": errors
    })

@invoices_blueprint.get("/api/invoices/local")
def api_get_local_invoices():
    """Retrieve all smart-audited locally stored invoices."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.service import get_local_invoices
    mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    if mst == "all" or not mst:
        mst = None
    return jsonify({"invoices": get_local_invoices(mst)})

@invoices_blueprint.get("/api/invoices/local/export-excel")
@roles_required("admin", "auditor")
def api_export_local_excel():
    """Export the local audited database to an Excel workbook download, filtered by active corporate taxpayer profile."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from invoices.service import get_local_invoices
        mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
        if mst == "all" or not mst:
            mst = None
        invoices = get_local_invoices(mst)
        workbook_bytes = generate_local_excel_workbook(invoices)
    except Exception as error:
        return jsonify({"error": str(error)}), 500

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"audited_invoices_{timestamp}.xlsx"
    return send_file(
        BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@invoices_blueprint.get("/api/invoices/local/items")
def api_search_local_items():
    """Global search across line items of locally imported invoices, filtered by active corporate taxpayer profile."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    q = request.args.get("q", "").strip()
    mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    if mst == "all" or not mst:
        mst = None
    from invoices.service import search_local_items
    return jsonify({"items": search_local_items(q, mst)})

@invoices_blueprint.delete("/api/invoices/local/clear")
@roles_required("admin", "auditor")
def api_clear_local_invoices():
    """Clear all records from local SQLite database and remove XML storage files."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice, LineItem
    from invoices.service import XML_DIR
    import shutil
    import os

    try:
        # Delete all records from Invoice and LineItem tables
        LineItem.query.delete()
        Invoice.query.delete()
        db.session.commit()

        from invoices.security_audit_service import log_security_event
        log_security_event("DELETE", "Cleared all records from local SQLite database and removed XML storage files.")

        if os.path.exists(XML_DIR):
            shutil.rmtree(XML_DIR)
            os.makedirs(XML_DIR, exist_ok=True)

        return jsonify({"status": "success", "message": "Da lam sach co so du lieu cuc bo."})

    except Exception as e:
        return jsonify({"error": f"Loi khi lam sach du lieu: {str(e)}"}), 500

@invoices_blueprint.delete("/api/invoices/local/<invoice_id>")
@roles_required("admin", "auditor")
def api_delete_local_invoice(invoice_id):
    """Delete a single local invoice and its XML file."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.service import delete_local_invoice
    success = delete_local_invoice(invoice_id)
    if not success:
        return jsonify({"error": "Không tìm thấy hóa đơn cần xóa."}), 404

    from invoices.security_audit_service import log_security_event
    log_security_event("DELETE", f"Deleted local invoice: {invoice_id}")

    return jsonify({"status": "success", "message": "Đã xóa hóa đơn thành công."})

@invoices_blueprint.patch("/api/invoices/local/<invoice_id>")
def api_adjust_local_invoice(invoice_id):
    """Adjust fields of a local invoice and update its smart auditing warnings."""

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}

    from invoices.service import adjust_local_invoice
    try:
        updated_invoice = adjust_local_invoice(invoice_id, payload)
        return jsonify({
            "status": "success",
            "message": "Đã điều chỉnh hóa đơn thành công.",
            "invoice": updated_invoice
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Lỗi không xác định: {str(e)}"}), 500

@invoices_blueprint.post("/api/invoices/realtime/trigger")
@roles_required("admin", "auditor")
def api_trigger_realtime_sync():
    """Manually trigger background real-time sync immediately."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.scheduler import load_scheduler_settings
    settings = load_scheduler_settings()

    # Trigger async thread so we return immediately and keep the UI responsive!
    from flask import current_app
    from invoices.scheduler import _scheduler_thread

    if _scheduler_thread and _scheduler_thread.is_alive():
        # Spin up a daemon thread to run the sync
        import threading
        def worker(app):
            with app.app_context():
                try:
                    _scheduler_thread.execute_realtime_sync(settings)
                except Exception as ex:
                    app.logger.error(f"Manual real-time sync failed: {ex}")

        t = threading.Thread(target=worker, args=(current_app._get_current_object(),), daemon=True)
        t.start()
        return jsonify({"status": "success", "message": "Đã kích hoạt đồng bộ hóa thời gian thực chạy ngầm."})
    else:
        return jsonify({"error": "Không thể kết nối với dịch vụ background scheduler."}), 500

@invoices_blueprint.get("/api/invoices/realtime/stream")
def api_realtime_stream():
    """SSE streaming endpoint to push downloaded invoice events in real-time."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import queue
    from invoices.scheduler import REALTIME_CLIENT_QUEUES, REALTIME_QUEUES_LOCK

    q = queue.Queue(maxsize=100)
    with REALTIME_QUEUES_LOCK:
        REALTIME_CLIENT_QUEUES.append(q)

    def event_generator():
        try:
            # Send initial keepalive
            yield f"data: {json.dumps({'event': 'connected'})}\n\n"
            while True:
                try:
                    # Wait for an event with a 20-second timeout for heartbeat/keepalive
                    event_data = q.get(timeout=20)
                    yield f"data: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                except queue.Empty:
                    # Send keepalive heartbeat to prevent connection timeout
                    yield "data: {\"event\": \"keepalive\"}\n\n"
        finally:
            with REALTIME_QUEUES_LOCK:
                if q in REALTIME_CLIENT_QUEUES:
                    REALTIME_CLIENT_QUEUES.remove(q)

    from flask import Response
    return Response(event_generator(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "Transfer-Encoding": "chunked",
        "Connection": "keep-alive"
    })

@invoices_blueprint.get("/api/invoices/<invoice_id>/pdf")
@roles_required("admin", "auditor")
def api_invoice_pdf_download(invoice_id):
    """Download printable official-style PDF electronic invoice using xhtml2pdf."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    # Locate invoice
    current_app.config["CURRENT_JWT"] = session.get("jwt")
    current_app.config["CURRENT_INVOICE_LOOKUP"] = session.get("invoice_lookup", {})
    try:
        from datetime import date
        invoice = current_app.config["CURRENT_INVOICE_LOOKUP"].get(invoice_id)
        
        if not invoice:
            invoices_purchase = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "purchase"))
            invoice = build_invoice_lookup(invoices_purchase).get(invoice_id)
            
        if not invoice:
            invoices_sold = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 20), False, "sold"))
            invoice = build_invoice_lookup(invoices_sold).get(invoice_id)

        if not invoice:
            from invoices.service import get_local_invoices
            local_db = get_local_invoices()
            for item in local_db:
                if item["id"] == invoice_id:
                    invoice = item
                    break
            
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn yêu cầu."}), 404

        line_items = fetch_invoice_line_items(invoice_id)
    except FileNotFoundError:
        return jsonify({"error": "Không tìm thấy hóa đơn yêu cầu."}), 404
    except Exception as error:
        return jsonify({"error": f"Lỗi hệ thống: {str(error)}"}), 500
    finally:
        current_app.config["CURRENT_JWT"] = None
        current_app.config["CURRENT_INVOICE_LOOKUP"] = {}

    # Calculate sums
    sum_before_tax = sum(item.get("amount_before_tax", 0.0) for item in line_items)
    sum_tax = sum(item.get("tax_amount", 0.0) for item in line_items)
    total_payable = sum_before_tax + sum_tax

    # Convert total payable to words
    from invoices.service import doc_so_tien_vietnam
    total_payable_words = doc_so_tien_vietnam(total_payable)

    # Auto buyer/seller properties
    user_company = {
        "name": "CONG TY CO PHAN CONG NGHE GDT INVOICE HUB",
        "mst": "0109999999",
        "address": "Toa nha Technopark, Gia Lam, TP. Ha Noi",
        "phone": "1900 8888",
    }
    
    partner_details = {
        "Cong ty A": {"mst": "0101234567", "address": "So 10 Pho Hue, Quan Hai Ba Trung, Ha Noi"},
        "Cong ty B": {"mst": "0209876543", "address": "250 Nguyen Thi Minh Khai, Quan 3, TP. Ho Chi Minh"},
        "Cong ty C": {"mst": "0301122334", "address": "15 Le Loi, Quan Hai Chau, Da Nang"},
    }
    
    if "seller_name" in invoice:
        seller = {
            "name": invoice.get("seller_name", ""),
            "mst": invoice.get("seller_mst", ""),
            "address": invoice.get("seller_address", ""),
            "phone": invoice.get("seller_phone", ""),
        }
        buyer = {
            "name": invoice.get("buyer_name", ""),
            "mst": invoice.get("buyer_mst", ""),
            "address": invoice.get("buyer_address", ""),
        }
    else:
        issuer = invoice.get("issuer", "Doi tac khac")
        partner = partner_details.get(
            issuer,
            {
                "mst": f"0{abs(hash(issuer)) % 1000000000:09d}",
                "address": f"Khu cong nghiep Binh Duong, Tinh Binh Duong",
            }
        )
        partner["name"] = issuer

        # If it is a purchase invoice, issuer is Seller, user_company is Buyer
        if invoice.get("direction", "purchase") == "purchase":
            seller = partner
            buyer = user_company
        else:
            seller = user_company
            buyer = partner

    html_content = render_template(
        "invoice_pdf_export.html",
        invoice=invoice,
        line_items=line_items,
        seller=seller,
        buyer=buyer,
        sum_before_tax=sum_before_tax,
        sum_tax=sum_tax,
        total_payable=total_payable,
        total_payable_words=total_payable_words
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"invoice_{invoice_id}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

@invoices_blueprint.get("/api/reports/partners/pdf")
@roles_required("admin", "auditor")
def api_reports_partners_pdf():
    """Export the Business Partner Directory matching the dashboard as a PDF report."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        partners = extract_partners_from_invoices(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    date_now = datetime.now().strftime("%d/%m/%Y %H:%M")
    html_content = render_template(
        "report_partners_pdf.html",
        partners=partners,
        from_date=parsed_from.strftime("%d/%m/%Y"),
        to_date=parsed_to.strftime("%d/%m/%Y"),
        direction=direction,
        date_now=date_now
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"partner_directory_{parsed_from.isoformat()}_{parsed_to.isoformat()}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

@invoices_blueprint.get("/api/reports/usage/pdf")
@roles_required("admin", "auditor")
def api_reports_usage_pdf():
    """Export the BC26 Tax Compliance and invoice usage report as a PDF report."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "sold")
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        report = generate_tax_usage_report(invoices)
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except GDTIntegrationNotReadyError as error:
        return jsonify({"error": str(error)}), 503
    finally:
        current_app.config["CURRENT_JWT"] = None

    date_now = datetime.now().strftime("%d/%m/%Y %H:%M")
    html_content = render_template(
        "report_usage_pdf.html",
        report=report,
        from_date=parsed_from.strftime("%d/%m/%Y"),
        to_date=parsed_to.strftime("%d/%m/%Y"),
        direction=direction,
        date_now=date_now
    )

    try:
        pdf_stream = render_html_to_pdf(html_content)
        filename = f"bc26_usage_report_{parsed_from.isoformat()}_{parsed_to.isoformat()}.pdf"
        return send_file(
            pdf_stream,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

@invoices_blueprint.get("/api/ai/chat/sessions")
def api_chat_sessions():
    """Retrieve all conversational sessions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession
    try:
        sessions = AIChatSession.query.order_by(AIChatSession.created_at.desc()).all()
        # Return list directly to satisfy legacy unit test (we will make main.js support both formats)
        return jsonify([s.to_dict() for s in sessions]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/chat/sessions")
def api_create_chat_session():
    """Create a new chat session."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession, AIChatMessage, Invoice
    import uuid
    from datetime import datetime
    try:
        data = request.get_json() or {}
        title = data.get("title", "Cuộc hội thoại mới")
        invoice_id = data.get("invoice_id")
        
        invoice = None
        if invoice_id:
            invoice = db.session.get(Invoice, invoice_id)
            if not invoice:
                return jsonify({"error": "Không tìm thấy hóa đơn liên kết."}), 404
            title = f"Tham vấn hóa đơn {invoice.number}"
            
        session_id = str(uuid.uuid4())
        new_session = AIChatSession(
            id=session_id,
            title=title,
            invoice_id=invoice_id,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(new_session)
        
        if invoice:
            import json
            welcome_content = (
                f"### [Hệ thống Kiểm soát Tuân thủ Thuế AI - GDT Hub]\n\n"
                f"Kính chào Quý khách, tôi là **Cố vấn Thuế cấp cao AI**. Tôi đã tiếp nhận yêu cầu tham vấn về hóa đơn điện tử sau:\n\n"
                f"- **Nhà cung cấp:** {invoice.seller_name} (MST: `{invoice.seller_mst}`)\n"
                f"- **Số hóa đơn:** `{invoice.number}` | **Ký hiệu:** `{invoice.symbol}` | **Ngày lập:** {invoice.date}\n"
                f"- **Tổng tiền thanh toán (sau thuế):** {invoice.total_amount:,.2f} {invoice.currency or 'VND'}\n"
                f"- **Chỉ số tuân thủ T-Score:** **{invoice.t_score}/100** ({invoice.t_rating})\n\n"
                f"**Đánh giá tuân thủ sơ bộ (Nghị định 123/2020/NĐ-CP & Thông tư 219/2013/TT-BTC):**\n"
            )
            
            warnings = []
            if invoice.warnings_json:
                try:
                    warnings = json.loads(invoice.warnings_json)
                except Exception:
                    pass
            
            cash_threshold = 20000000.0
            if invoice.total_amount >= cash_threshold and invoice.payment_method == "Tiền mặt":
                warnings.append("Hóa đơn thanh toán bằng Tiền mặt từ 20 triệu đồng trở lên có nguy cơ không được khấu trừ thuế GTGT đầu vào và không được tính vào chi phí được trừ khi quyết toán thuế TNDN (Điều 15 Thông tư 219/2013/TT-BTC).")
            
            if warnings:
                welcome_content += "⚠️ **Các điểm cần lưu ý/rủi ro:**\n"
                for w in warnings:
                    welcome_content += f"- {w}\n"
            else:
                welcome_content += "✅ Hóa đơn không phát hiện lỗi cấu trúc hay cảnh báo rủi ro rác/rủi ro hình thức nào trọng yếu.\n"
                
            welcome_content += (
                f"\nQuý khách có thể bắt đầu đặt câu hỏi cho tôi về việc **tính hợp lệ chi phí được trừ (Thuế TNDN)**, "
                f"**khấu trừ thuế GTGT đầu vào**, hoặc **cơ sở pháp lý** liên quan đến hóa đơn này."
            )
            
            welcome_msg = AIChatMessage(
                session_id=session_id,
                role="assistant",
                content=welcome_content,
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.session.add(welcome_msg)
            
        db.session.commit()
        # Dual compatibility: return details both directly and nested under 'session'
        res_dict = new_session.to_dict()
        res_dict["session"] = new_session.to_dict()
        return jsonify(res_dict), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/chat/sessions/<session_id>/message")
def api_send_chat_message(session_id):
    """Send a user message to the session and get the AI assistant response."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession, AIChatMessage
    from invoices.ai_service import AIChatAgent
    from datetime import datetime
    try:
        session = db.session.get(AIChatSession, session_id)
        if not session:
            return jsonify({"error": "Không tìm thấy phiên hội thoại."}), 404

        data = request.get_json() or {}
        content = data.get("message", "").strip()
        if not content:
            return jsonify({"error": "Nội dung tin nhắn trống."}), 400

        # Save user message
        user_msg = AIChatMessage(
            session_id=session_id,
            role="user",
            content=content,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(user_msg)
        db.session.commit()

        # Update session title if it was default
        if session.title == "Cuộc hội thoại mới" or session.title == "Cuộc trò chuyện mới":
            session.title = content[:30] + ("..." if len(content) > 30 else "")
            db.session.commit()

        # Call AI assistant agent
        agent = AIChatAgent()
        ai_response = agent.ask(session_id, content)

        # Save assistant response
        assistant_msg = AIChatMessage(
            session_id=session_id,
            role="assistant",
            content=ai_response,
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(assistant_msg)
        db.session.commit()

        # Dual compatibility: return both legacy fields and 'reply' field
        return jsonify({
            "user_message": user_msg.to_dict(),
            "assistant_message": assistant_msg.to_dict(),
            "session_title": session.title,
            "reply": ai_response
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.delete("/api/ai/chat/sessions/<session_id>")
def api_delete_chat_session(session_id):
    """Delete a chat session."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import AIChatSession
    try:
        session = db.session.get(AIChatSession, session_id)
        if not session:
            return jsonify({"error": "Không tìm thấy phiên hội thoại."}), 404
        db.session.delete(session)
        db.session.commit()
        return jsonify({"success": True, "message": "Đã xóa phiên hội thoại thành công."}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/classify-items")
@roles_required("admin", "auditor")
def api_classify_invoice_items():
    """Classify items inside an invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice, LineItem
    from invoices.ai_service import AIExpenseClassifier

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        if not invoice_id:
            return jsonify({"error": "Thiếu mã hóa đơn invoice_id."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        classifier = AIExpenseClassifier()
        items = LineItem.query.filter_by(invoice_id=invoice.id).all()
        if not items:
            return jsonify({"success": True, "classified_items": []})

        classifications = classifier.classify_line_items(items)

        # Save results to DB
        for item in items:
            if item.id in classifications:
                item.expense_category = classifications[item.id]
        db.session.commit()

        return jsonify({
            "success": True,
            "classified_items": [
                {
                    "item_id": item.id,
                    "item_name": item.item_name,
                    "category": item.expense_category
                } for item in items
            ]
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/update-item-category")
@roles_required("admin", "auditor")
def api_update_item_category():
    """Manually update the expense category of a specific line item."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import LineItem
    try:
        data = request.get_json() or {}
        item_id = data.get("item_id")
        new_category = data.get("category")

        if not item_id or not new_category:
            return jsonify({"error": "Thiếu thông tin item_id hoặc category."}), 400

        item = db.session.get(LineItem, item_id)
        if not item:
            return jsonify({"error": "Không tìm thấy mặt hàng."}), 404

        # Validate category matches standard list
        from invoices.ai_service import AIExpenseClassifier
        if new_category not in AIExpenseClassifier.CATEGORIES:
            return jsonify({"error": "Danh mục chi phí không hợp lệ."}), 400

        item.expense_category = new_category
        db.session.commit()

        return jsonify({"success": True, "item_id": item.id, "category": item.expense_category})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/repair-metadata")
@roles_required("admin", "auditor")
def api_repair_metadata():
    """Analyze and generate AI suggestions for repairing invoice metadata."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice
    from invoices.ai_service import AIDataRepairer

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        if not invoice_id:
            return jsonify({"error": "Thiếu mã hóa đơn invoice_id."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        repairer = AIDataRepairer()
        suggestions = repairer.repair_metadata(invoice)

        before = {
            "seller_name": invoice.seller_name or "",
            "buyer_name": invoice.buyer_name or "",
            "buyer_address": invoice.buyer_address or "",
            "amount_in_words": invoice.amount_in_words or ""
        }
        
        differences = []
        for key in ["seller_name", "buyer_name", "buyer_address", "amount_in_words"]:
            val_before = before[key].strip()
            val_after = suggestions.get(key, "").strip()
            if val_before != val_after and val_after:
                differences.append(key)

        return jsonify({
            "success": True,
            "invoice_id": invoice.id,
            "before": before,
            "after": suggestions,
            "differences": differences
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ai/apply-repair")
@roles_required("admin", "auditor")
def api_apply_repair():
    """Apply selected AI repair suggestions to persistent SQLite database."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice

    try:
        data = request.get_json() or {}
        invoice_id = data.get("invoice_id")
        fields_to_apply = data.get("fields", [])

        if not invoice_id or not fields_to_apply:
            return jsonify({"error": "Thiếu thông tin invoice_id hoặc fields để áp dụng."}), 400

        invoice = db.session.get(Invoice, invoice_id)
        if not invoice:
            return jsonify({"error": "Không tìm thấy hóa đơn."}), 404

        allowed_fields = ["seller_name", "buyer_name", "buyer_address", "amount_in_words"]
        applied = []
        for field in fields_to_apply:
            if field in allowed_fields:
                val = data.get(field)
                if val:
                    setattr(invoice, field, val)
                    applied.append(field)

        if applied:
            db.session.commit()
            from invoices.security_audit_service import log_security_event
            log_security_event("REPAIR", f"Applied AI repair to invoice {invoice_id} for fields: {', '.join(applied)}")

        return jsonify({
            "success": True,
            "invoice_id": invoice.id,
            "applied_fields": applied
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/invoices/summary-by-seller")
def api_summary_by_seller():
    """Aggregate input invoices by month or quarter, grouped by seller."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from extensions import db
    from invoices.models import Invoice
    from sqlalchemy import func

    try:
        period_type = request.args.get("period_type", "monthly")
        year_filter = request.args.get("year", "")

        # Base query for all invoices
        query = db.session.query(
            Invoice.seller_mst,
            Invoice.seller_name,
            Invoice.date,
            func.count(Invoice.id).label("invoice_count"),
            func.sum(Invoice.amount_before_tax).label("total_before_tax"),
            func.sum(Invoice.tax_amount).label("total_tax"),
            func.sum(Invoice.total_amount).label("total_amount")
        )

        # Apply year filter if provided
        if year_filter:
            query = query.filter(Invoice.date.like(f"{year_filter}-%"))
        
        # Pull raw grouped results and aggregate/group them cleanly in Python
        results = query.group_by(
            Invoice.seller_mst,
            Invoice.seller_name,
            func.substr(Invoice.date, 1, 7)
        ).all()

        period_map = {}

        for row in results:
            mst, name, date_val, count, before_tax, tax, total = row
            if not date_val or len(date_val) < 7:
                continue
            
            row_year = date_val[0:4]
            if year_filter and row_year != year_filter:
                continue

            row_month = date_val[5:7]

            if period_type == "quarterly":
                try:
                    m_int = int(row_month)
                except ValueError:
                    m_int = 1
                if m_int in [1, 2, 3]:
                    period = f"Quý 1 / {row_year}"
                elif m_int in [4, 5, 6]:
                    period = f"Quý 2 / {row_year}"
                elif m_int in [7, 8, 9]:
                    period = f"Quý 3 / {row_year}"
                else:
                    period = f"Quý 4 / {row_year}"
            else:
                period = f"Tháng {row_month} / {row_year}"

            if period not in period_map:
                period_map[period] = {}

            # Aggregate sellers within the same period
            seller_key = mst or "UNKNOWN"
            if seller_key not in period_map[period]:
                period_map[period][seller_key] = {
                    "seller_mst": mst or "Không rõ",
                    "seller_name": name or "Không rõ",
                    "invoice_count": 0,
                    "total_before_tax": 0.0,
                    "total_tax": 0.0,
                    "total_amount": 0.0
                }

            entry = period_map[period][seller_key]
            entry["invoice_count"] += count
            entry["total_before_tax"] += before_tax or 0.0
            entry["total_tax"] += tax or 0.0
            entry["total_amount"] += total or 0.0

        # Format and sort periods
        data = []
        for period, sellers_dict in period_map.items():
            sellers_list = list(sellers_dict.values())
            # Sort sellers by total amount descending
            sellers_list.sort(key=lambda x: x["total_amount"], reverse=True)
            data.append({
                "period": period,
                "sellers": sellers_list,
                "total_before_tax": sum(s["total_before_tax"] for s in sellers_list),
                "total_tax": sum(s["total_tax"] for s in sellers_list),
                "total_amount": sum(s["total_amount"] for s in sellers_list)
            })

        # Sort periods. Monthly: "Tháng 12 / 2026" -> "Tháng 01 / 2026" descending.
        # Format key is period title itself.
        data.sort(key=lambda x: x["period"], reverse=True)

        return jsonify({
            "success": True,
            "period_type": period_type,
            "year": year_filter or "Tất cả",
            "data": data
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/invoices/supplier-pivot")
def api_supplier_pivot():
    """Aggregate input invoices by supplier and month/year in a pivot structure."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        year_filter = request.args.get("year", "2026")
        value_type = request.args.get("value_type", "total_amount")
        mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
        if mst == "all":
            mst = None

        data = get_supplier_pivot_data(mst, year_filter, value_type)
        return jsonify({"success": True, **data})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@invoices_blueprint.get("/api/invoices/supplier-pivot/export")
def api_supplier_pivot_export():
    """Export the supplier pivot table to a beautifully formatted Excel sheet."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from export.formatter import auto_adjust_column_widths

    try:
        year_filter = request.args.get("year", "2026")
        value_type = request.args.get("value_type", "total_amount")
        mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
        if mst == "all":
            mst = None

        data = get_supplier_pivot_data(mst, year_filter, value_type)

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Pivot NCC"
        ws.views.sheetView[0].showGridLines = True

        # Titles
        title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        info_font = Font(name="Calibri", size=11, italic=True)
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        total_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        double_bottom_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='double', color='1F4E78')
        )

        ws["A1"] = f"BẢNG TỔNG HỢP HOÁ ĐƠN ĐẦU VÀO THEO NHÀ CUNG CẤP - NĂM {data['year']}"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 25

        value_type_titles = {
            "total_amount": "Tổng tiền thanh toán (đồng)",
            "amount_before_tax": "Doanh số trước thuế (đồng)",
            "tax_amount": "Tiền thuế GTGT (đồng)",
            "invoice_count": "Số lượng hóa đơn (tờ)"
        }
        value_title = value_type_titles.get(data["value_type"], "Tổng tiền thanh toán")

        mst_str = mst if mst else "Tất cả Doanh nghiệp"
        ws["A2"] = f"Mã số thuế Doanh nghiệp: {mst_str} | Chỉ số: {value_title}"
        ws["A2"].font = info_font
        ws.row_dimensions[2].height = 20

        # Headers on row 4
        headers = ["Mã số thuế", "Tên nhà cung cấp"]
        for m in data["months"]:
            if len(m) == 2:
                headers.append(f"Tháng {m}")
            else:
                headers.append(m)
        headers.append("Tổng cộng")

        ws.append([]) # row 3 blank
        ws.append(headers) # row 4
        ws.row_dimensions[4].height = 25

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Rows starting at row 5
        current_row = 5
        for row_idx, r in enumerate(data["rows"]):
            row_data = [r["seller_mst"], r["seller_name"]]
            for m in data["months"]:
                row_data.append(r["monthly_values"].get(m, 0.0))
            row_data.append(r["row_total"])

            ws.append(row_data)
            ws.row_dimensions[current_row].height = 20

            # Format cells
            is_even = row_idx % 2 == 1
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = regular_font
                elif col_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.font = regular_font
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if value_type == "invoice_count":
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0"
                    cell.font = regular_font

                if is_even:
                    cell.fill = zebra_fill

            current_row += 1

        # Totals Row at current_row
        total_row_data = ["TỔNG CỘNG", ""]
        for m in data["months"]:
            total_row_data.append(data["column_totals"].get(m, 0.0))
        total_row_data.append(data["grand_total"])

        ws.append(total_row_data)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.row_dimensions[current_row].height = 22

        for col_idx in range(1, len(total_row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
            cell.fill = total_fill
            cell.border = double_bottom_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if value_type == "invoice_count":
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto adjust column widths
        auto_adjust_column_widths(ws)

        # Set specific widths for MST and Name columns to look beautiful
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 38

        # Output
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_bytes = excel_file.getvalue()

        filename = f"Pivot_NCC_{value_type}_{data['year']}.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/vat-declaration")
def api_reports_vat_declaration():
    """Generate a draft of the Vietnamese VAT Return Mẫu 01/GTGT and list of disputed/high-risk input invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            # Default to current or last completed month/quarter
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        # Standardizing month string to "02" instead of "2" for monthly
        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice, LineItem

        query = Invoice.query.filter(Invoice.is_cancelled == False)

        # Apply date filters
        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        # Initialize output VAT rate aggregates (Sold)
        output_exempt_val = 0.0
        output_0_val = 0.0
        output_5_val = 0.0
        output_5_vat = 0.0
        output_10_val = 0.0
        output_10_vat = 0.0
        
        # Input Aggregates (Purchase)
        input_total_value = 0.0
        input_total_vat = 0.0
        input_deductible_vat = 0.0
        
        disputed_invoices = []

        for inv in invoices:
            if inv.invoice_type == "sold":
                has_items = len(inv.items) > 0
                if has_items:
                    for item in inv.items:
                        rate = (item.tax_rate or "").strip().lower()
                        val = item.amount_before_tax or 0.0
                        tax = item.tax_amount or 0.0
                        
                        if "không chịu" in rate or "khong chiu" in rate:
                            output_exempt_val += val
                        elif "0%" in rate or rate == "0":
                            output_0_val += val
                        elif "5%" in rate or rate == "5":
                            output_5_val += val
                            output_5_vat += tax
                        else:
                            # 8% and 10% grouped into standard rate
                            output_10_val += val
                            output_10_vat += tax
                else:
                    # Fallback to invoice totals if no line items exist
                    val = inv.amount_before_tax or 0.0
                    tax = inv.tax_amount or 0.0
                    output_10_val += val
                    output_10_vat += tax

            elif inv.invoice_type == "purchase":
                val = inv.amount_before_tax or 0.0
                tax = inv.tax_amount or 0.0
                
                input_total_value += val
                input_total_vat += tax
                
                # Combine traditional parsing warnings and Gemma-4 AI auditor warnings
                warnings = list(inv.warnings) if inv.warnings else []
                ai_warnings = [f"[AI: {w.warning_type}] {w.explanation}" for w in inv.ai_audit_results]
                warnings.extend(ai_warnings)
                
                is_disputed = len(warnings) > 0
                
                if is_disputed:
                    warning_msg = "; ".join(warnings)
                    disputed_invoices.append({
                        "id": inv.id,
                        "number": inv.number or "Không số",
                        "date": inv.date or "Không ngày",
                        "seller_name": inv.seller_name or "Không rõ",
                        "seller_mst": inv.seller_mst or "Không rõ",
                        "amount_before_tax": val,
                        "tax_amount": tax,
                        "total_amount": inv.total_amount or (val + tax),
                        "warning": warning_msg
                    })
                else:
                    input_deductible_vat += tax

        output_taxable_val = output_0_val + output_5_val + output_10_val
        output_total_value = output_exempt_val + output_taxable_val
        output_total_vat = output_5_vat + output_10_vat

        vat_payable = max(0.0, output_total_vat - input_deductible_vat)
        vat_carried_forward = max(0.0, input_deductible_vat - output_total_vat)

        return jsonify({
            "success": True,
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "outputs": {
                "exempt_val": output_exempt_val,
                "tax_0_val": output_0_val,
                "tax_5_val": output_5_val,
                "tax_5_vat": output_5_vat,
                "tax_10_val": output_10_val,
                "tax_10_vat": output_10_vat,
                "taxable_val": output_taxable_val,
                "total_val": output_total_value,
                "total_vat": output_total_vat
            },
            "inputs": {
                "total_value": input_total_value,
                "total_vat": input_total_vat,
                "deductible_vat": input_deductible_vat
            },
            "calculations": {
                "vat_payable": vat_payable,
                "vat_carried_forward": vat_carried_forward
            },
            "disputed_invoices": disputed_invoices
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/analytics/top-items")
def analytics_top_items():
    """Return top 20 most-purchased line item names for autocomplete."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from extensions import db
        from invoices.models import LineItem
        rows = (
            db.session.query(LineItem.item_name, db.func.count(LineItem.id).label("cnt"))
            .group_by(db.func.lower(LineItem.item_name))
            .order_by(db.desc("cnt"))
            .limit(20)
            .all()
        )
        items = [{"name": r[0], "count": r[1]} for r in rows]
        return jsonify({"success": True, "items": items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/analytics/supplier-price-trends")
def analytics_supplier_price_trends():
    """Return monthly unit price data for a given item, grouped by seller."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    item_name = request.args.get("item_name", "").strip()
    year_filter = request.args.get("year", "").strip()

    if not item_name:
        return jsonify({"error": "Vui lòng cung cấp tên mặt hàng (item_name)."}), 400

    try:
        from extensions import db
        from invoices.models import LineItem, Invoice as Inv

        query = (
            db.session.query(
                LineItem.item_name,
                LineItem.unit_price,
                LineItem.amount_before_tax,
                Inv.date,
                Inv.seller_name,
                Inv.seller_mst,
            )
            .join(Inv, LineItem.invoice_id == Inv.id)
            .filter(db.func.lower(LineItem.item_name).like(f"%{item_name.lower()}%"))
            .filter(Inv.is_cancelled == False)
        )

        if year_filter:
            query = query.filter(Inv.date.like(f"{year_filter}%"))

        rows = query.all()

        if not rows:
            return jsonify({"success": True, "item_name": item_name, "sellers": [], "months": [], "series": [], "anomalies": []})

        # Compute global average unit price for anomaly detection
        all_prices = [r[1] for r in rows if r[1] and r[1] > 0]
        avg_global = sum(all_prices) / len(all_prices) if all_prices else 0

        # Build month × seller matrix
        seller_map = {}  # seller_mst → {name, months: {YYYY-MM: [prices]}}
        months_set = set()

        for item_name_val, unit_price, amount, inv_date, seller_name, seller_mst in rows:
            if not inv_date:
                continue
            month_key = inv_date[:7]  # YYYY-MM
            months_set.add(month_key)
            mst = seller_mst or "unknown"
            if mst not in seller_map:
                seller_map[mst] = {"seller_mst": mst, "seller_name": seller_name or mst, "months": {}}
            seller_map[mst]["months"].setdefault(month_key, []).append(unit_price or 0)

        months_sorted = sorted(months_set)

        # Build series for chart
        series = []
        anomalies = []
        for mst, info in seller_map.items():
            prices_by_month = []
            for m in months_sorted:
                month_prices = info["months"].get(m, [])
                if month_prices:
                    avg_m = sum(month_prices) / len(month_prices)
                    prices_by_month.append(round(avg_m, 0))
                    if avg_global > 0 and avg_m > avg_global * 1.20:
                        anomalies.append({
                            "month": m,
                            "seller_name": info["seller_name"],
                            "seller_mst": mst,
                            "price": round(avg_m, 0),
                            "avg_global": round(avg_global, 0),
                            "pct_above": round((avg_m / avg_global - 1) * 100, 1),
                        })
                else:
                    prices_by_month.append(None)

            # Summary stats
            flat = [p for p in prices_by_month if p is not None]
            series.append({
                "seller_mst": mst,
                "seller_name": info["seller_name"],
                "prices": prices_by_month,
                "avg_price": round(sum(flat) / len(flat), 0) if flat else 0,
                "min_price": round(min(flat), 0) if flat else 0,
                "max_price": round(max(flat), 0) if flat else 0,
                "purchase_count": sum(len(info["months"].get(m, [])) for m in months_sorted),
            })

        # Sort series by avg_price ascending (cheapest first)
        series.sort(key=lambda s: s["avg_price"])

        return jsonify({
            "success": True,
            "item_name": item_name,
            "avg_global": round(avg_global, 0),
            "months": months_sorted,
            "sellers": [{"seller_mst": s["seller_mst"], "seller_name": s["seller_name"]} for s in series],
            "series": series,
            "anomalies": anomalies,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/analytics/vat-forecast")
def analytics_vat_forecast():
    """Return monthly actual VAT net (output-input) and 2-month linear forecast."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import datetime as _dt
    year_filter = request.args.get("year", str(_dt.now().year)).strip()

    try:
        from extensions import db
        from invoices.models import Invoice
        invoices_all = Invoice.query.filter(Invoice.is_cancelled == False).all()

        # Build monthly buckets for the selected year
        monthly = {}
        for inv in invoices_all:
            if not inv.date or not inv.date.startswith(year_filter):
                continue
            month_key = inv.date[:7]  # YYYY-MM
            if month_key not in monthly:
                monthly[month_key] = {"output_vat": 0.0, "input_vat": 0.0}
            if inv.invoice_type == "sold":
                monthly[month_key]["output_vat"] += inv.tax_amount or 0.0
            elif inv.invoice_type == "purchase":
                monthly[month_key]["input_vat"] += inv.tax_amount or 0.0

        # All 12 months of the selected year
        all_months = [f"{year_filter}-{m:02d}" for m in range(1, 13)]
        actual = []
        for m in all_months:
            b = monthly.get(m, {"output_vat": 0.0, "input_vat": 0.0})
            net = b["output_vat"] - b["input_vat"]
            actual.append({
                "month": m,
                "output_vat": round(b["output_vat"], 0),
                "input_vat": round(b["input_vat"], 0),
                "net_vat": round(net, 0),
                "has_data": m in monthly,
            })

        # Identify last N months with real data for trend computation
        real_months = [a for a in actual if a["has_data"]]
        forecast = []

        if len(real_months) >= 2:
            # Linear trend: avg delta over last min(3, n) real months
            window = real_months[-3:] if len(real_months) >= 3 else real_months
            deltas = [window[i]["net_vat"] - window[i - 1]["net_vat"] for i in range(1, len(window))]
            avg_delta = sum(deltas) / len(deltas) if deltas else 0
            last_net = real_months[-1]["net_vat"]
            last_month_str = real_months[-1]["month"]

            # Build next 2 months after the last real month
            last_dt = _dt.strptime(last_month_str, "%Y-%m")
            for i in range(1, 3):
                if last_dt.month + i <= 12:
                    fdt = last_dt.replace(month=last_dt.month + i)
                else:
                    fdt = last_dt.replace(year=last_dt.year + 1, month=(last_dt.month + i) - 12)
                projected_net = last_net + avg_delta * i
                prev_net = last_net + avg_delta * (i - 1)
                warning = prev_net > 0 and projected_net > prev_net * 1.30
                forecast.append({
                    "month": fdt.strftime("%Y-%m"),
                    "net_vat_forecast": round(projected_net, 0),
                    "warning": warning,
                })

        # Compute year summary
        total_output = sum(a["output_vat"] for a in actual if a["has_data"])
        total_input = sum(a["input_vat"] for a in actual if a["has_data"])
        total_net = total_output - total_input

        return jsonify({
            "success": True,
            "year": year_filter,
            "actual": actual,
            "forecast": forecast,
            "summary": {
                "total_output_vat": round(total_output, 0),
                "total_input_vat": round(total_input, 0),
                "total_net_vat": round(total_net, 0),
                "months_with_data": len(real_months),
            },
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/budget/config")
def budget_config_get():
    """Return saved budget configuration for a given month."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    month = request.args.get("month", "").strip()
    if not month:
        month = datetime.now().strftime("%Y-%m")

    try:
        from invoices.models import SystemConfig
        key = f"budget_config_{month}"
        cfg = db.session.get(SystemConfig, key)
        if cfg:
            import json as _json
            configs = _json.loads(cfg.value)
        else:
            configs = []
        return jsonify({"success": True, "month": month, "configs": configs})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/budget/config")
@roles_required("admin", "auditor")
def budget_config_save():
    """Save budget configuration for a given month."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        import json as _json
        body = request.get_json(force=True) or {}
        month = body.get("month", "").strip()
        configs = body.get("configs", [])
        if not month:
            month = datetime.now().strftime("%Y-%m")

        from extensions import db
        from invoices.models import SystemConfig
        key = f"budget_config_{month}"
        cfg = db.session.get(SystemConfig, key)
        if cfg:
            cfg.value = _json.dumps(configs, ensure_ascii=False)
        else:
            cfg = SystemConfig(key=key, value=_json.dumps(configs, ensure_ascii=False))
            db.session.add(cfg)
        db.session.commit()
        return jsonify({"success": True, "month": month, "saved": len(configs)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/budget/actuals")
def budget_actuals():
    """Return actual spending per expense_category for a given month with budget vs. actual comparison."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    month = request.args.get("month", "").strip()
    if not month:
        month = datetime.now().strftime("%Y-%m")

    try:
        import json as _json
        from extensions import db
        from invoices.models import LineItem, Invoice as _Inv, SystemConfig

        # Aggregate actual spending by expense_category for the month
        rows = (
            db.session.query(
                LineItem.expense_category,
                db.func.sum(LineItem.amount_before_tax).label("actual_vnd"),
            )
            .join(_Inv, LineItem.invoice_id == _Inv.id)
            .filter(
                _Inv.invoice_type == "purchase",
                _Inv.is_cancelled == False,
                _Inv.date.like(f"{month}%"),
            )
            .group_by(LineItem.expense_category)
            .all()
        )

        actuals_map = {}
        for category, actual_vnd in rows:
            cat = category or "Chưa phân loại"
            actuals_map[cat] = round(actual_vnd or 0, 0)

        # Load budget config
        key = f"budget_config_{month}"
        cfg_rec = db.session.get(SystemConfig, key)
        budget_configs = _json.loads(cfg_rec.value) if cfg_rec else []
        budget_map = {c["category"]: c["limit_vnd"] for c in budget_configs}

        # Build response combining actuals + budgets
        all_categories = set(actuals_map.keys()) | set(budget_map.keys())
        actuals = []
        for cat in sorted(all_categories):
            actual_vnd = actuals_map.get(cat, 0)
            limit_vnd = budget_map.get(cat)
            if limit_vnd and limit_vnd > 0:
                pct = round(actual_vnd / limit_vnd * 100, 1)
                if pct >= 100:
                    status = "over_budget"
                elif pct >= 70:
                    status = "warning"
                else:
                    status = "ok"
            else:
                pct = None
                status = "no_budget"
            actuals.append({
                "category": cat,
                "actual_vnd": actual_vnd,
                "limit_vnd": limit_vnd,
                "pct_used": pct,
                "status": status,
            })

        any_over = any(a["status"] == "over_budget" for a in actuals)
        any_warning = any(a["status"] == "warning" for a in actuals)
        alert_level = "over_budget" if any_over else ("warning" if any_warning else "ok")

        return jsonify({
            "success": True,
            "month": month,
            "actuals": actuals,
            "alert_level": alert_level,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/aging/summary")
def aging_summary():
    """
    Return outstanding sold (receivables) and bought (payables) invoices classified into aging buckets.

    Buckets: Current / 1-30 / 31-60 / 61-90 / >90 days overdue.
    - Excludes invoices with paid_date set (already paid).
    - Excludes cancelled invoices.
    - Falls back to invoice.date when due_date is absent.
    - Accepts optional ?as_of=YYYY-MM-DD (default: today).
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import date as _date
    as_of_str = request.args.get("as_of", "").strip()
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    
    try:
        as_of = _date.fromisoformat(as_of_str) if as_of_str else _date.today()
    except ValueError:
        return jsonify({"error": "Định dạng as_of không hợp lệ. Dùng YYYY-MM-DD."}), 400

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv

        # Fetch all outstanding sold and bought invoices for active taxpayer
        query = _Inv.query.filter(
            _Inv.is_cancelled == False,
            _Inv.paid_date == None,
        )
        if mst:
            query = query.filter(_Inv.taxpayer_mst == mst)
        invoices = query.all()

        # Build empty bucket structures for receivables and payables
        def empty_buckets():
            return [
                {"label": "Chưa quá hạn (Current)", "min_days": None, "max_days": 0, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "1–30 ngày", "min_days": 1, "max_days": 30, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "31–60 ngày", "min_days": 31, "max_days": 60, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": "61–90 ngày", "min_days": 61, "max_days": 90, "count": 0, "total_amount": 0.0, "invoices": []},
                {"label": ">90 ngày", "min_days": 91, "max_days": None, "count": 0, "total_amount": 0.0, "invoices": []}
            ]

        ar_buckets = empty_buckets()
        ap_buckets = empty_buckets()

        for inv in invoices:
            # Determine reference date for aging (due_date preferred, fall back to invoice date)
            ref_date_str = inv.due_date or inv.date
            if not ref_date_str:
                continue
            try:
                ref_date = _date.fromisoformat(ref_date_str)
            except ValueError:
                continue

            age_days = (as_of - ref_date).days

            # Phase 3: Autonomous Categorization rules
            cat = "OPEX"
            if inv.seller_name and any(kw in inv.seller_name.lower() for kw in ["điện", "nước", "utility"]):
                cat = "UTILITIES"
            elif inv.invoice_type == "sold":
                cat = "REVENUE"

            inv_data = {
                "id": inv.id,
                "date": inv.date,
                "due_date": inv.due_date,
                "invoice_type": inv.invoice_type,
                "seller_name": inv.seller_name or "",
                "buyer_name": inv.buyer_name or "",
                "buyer_mst": inv.buyer_mst or "",
                "amount_before_tax": inv.amount_before_tax or 0.0,
                "total_amount": inv.total_amount or 0.0,
                "age_days": age_days,
                "ai_category": cat
            }

            target_buckets = ar_buckets if inv.invoice_type == "sold" else ap_buckets

            # Assign to correct bucket based on age_days
            if age_days <= 0:
                target_buckets[0]["count"] += 1
                target_buckets[0]["total_amount"] += inv.total_amount or 0.0
                target_buckets[0]["invoices"].append(inv_data)
            else:
                for bucket in target_buckets[1:]:
                    mn, mx = bucket["min_days"], bucket["max_days"]
                    if age_days >= mn and (mx is None or age_days <= mx):
                        bucket["count"] += 1
                        bucket["total_amount"] += inv.total_amount or 0.0
                        bucket["invoices"].append(inv_data)
                        break

        # Calculate totals
        total_ar = sum(b["total_amount"] for b in ar_buckets)
        total_ap = sum(b["total_amount"] for b in ap_buckets)

        # For backwards compatibility with standard dashboard charts, return overdue-only AR buckets
        legacy_buckets = []
        for label, mn, mx in _AGING_BUCKETS:
            # Find the corresponding ar_bucket
            corr = next((b for b in ar_buckets if b["label"] == label), None)
            if corr:
                legacy_buckets.append(corr)
            else:
                legacy_buckets.append({
                    "label": label,
                    "min_days": mn,
                    "max_days": mx,
                    "count": 0,
                    "total_amount": 0.0,
                    "invoices": []
                })

        return jsonify({
            "success": True,
            "as_of": as_of.isoformat(),
            "total_outstanding": total_ar,  # legacy name for AR total
            "total_count": sum(b["count"] for b in ar_buckets[1:]),
            "buckets": legacy_buckets,      # legacy overdue sold buckets
            "receivables": {
                "total_amount": total_ar,
                "total_count": sum(b["count"] for b in ar_buckets),
                "buckets": ar_buckets
            },
            "payables": {
                "total_amount": total_ap,
                "total_count": sum(b["count"] for b in ap_buckets),
                "buckets": ap_buckets
            }
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/cashflow/projection")
def cashflow_projection_endpoint():
    """
    US-083: Predictive Cash Projection & What-If late payment simulation.
    Projects optimistic vs pessimistic cash balance daily over the next 90 days.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import date as _date, timedelta
    as_of_str = request.args.get("as_of", "").strip()
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    
    # Custom simulation parameters
    late_days = int(request.args.get("late_days", "0").strip() or "0")
    client_mst = request.args.get("client_mst", "").strip()
    
    # Base balance override
    base_balance_str = request.args.get("base_balance", "").strip()
    
    try:
        as_of = _date.fromisoformat(as_of_str) if as_of_str else _date.today()
    except ValueError:
        return jsonify({"error": "Định dạng as_of không hợp lệ. Dùng YYYY-MM-DD."}), 400

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv, BankTransaction as _Tx

        # 1. Base Balance Dynamic Calculation
        if base_balance_str:
            try:
                base_balance = float(base_balance_str)
            except ValueError:
                base_balance = 500000000.0
        else:
            # Calculate base balance from actual bank transaction history
            tx_query = _Tx.query
            if mst:
                tx_query = tx_query.filter(_Tx.taxpayer_mst == mst)
            txs = tx_query.all()
            if txs:
                base_balance = sum(tx.amount for tx in txs)
            else:
                base_balance = 500000000.0  # Default fallback if no bank tx

        # 2. Fetch all unpaid outstanding invoices
        query = _Inv.query.filter(
            _Inv.is_cancelled == False,
            _Inv.paid_date == None
        )
        if mst:
            query = query.filter(_Inv.taxpayer_mst == mst)
        invoices = query.all()

        # 3. Simulate day-by-day cashflows for the next 90 days
        projections = []
        current_opt = base_balance
        current_sim = base_balance

        # We pre-calculate expected flows per day
        opt_inflows = {}
        sim_inflows = {}
        outflows = {}

        for i in range(91):
            day = as_of + timedelta(days=i)
            day_str = day.isoformat()
            opt_inflows[day_str] = 0.0
            sim_inflows[day_str] = 0.0
            outflows[day_str] = 0.0

        for inv in invoices:
            ref_date_str = inv.due_date or inv.date
            if not ref_date_str:
                continue
            try:
                ref_date = _date.fromisoformat(ref_date_str)
            except ValueError:
                continue

            amount = inv.total_amount or 0.0

            if inv.invoice_type == "sold":
                # Sales -> Inflows
                # Optimistic Expected Date (always due_date/date)
                opt_day = ref_date
                if opt_day < as_of:
                    opt_day = as_of # Already overdue: assume collected today
                opt_str = opt_day.isoformat()
                if opt_str in opt_inflows:
                    opt_inflows[opt_str] += amount

                # Simulated/Pessimistic Expected Date
                sim_day = ref_date
                if late_days > 0:
                    if not client_mst or inv.buyer_mst == client_mst:
                        sim_day = sim_day + timedelta(days=late_days)
                
                if sim_day < as_of:
                    sim_day = as_of # Overdue shift
                
                sim_str = sim_day.isoformat()
                if sim_str in sim_inflows:
                    sim_inflows[sim_str] += amount
            else:
                # Purchases -> Outflows
                out_day = ref_date
                if out_day < as_of:
                    out_day = as_of # Overdue payables must be paid today
                out_str = out_day.isoformat()
                if out_str in outflows:
                    outflows[out_str] += amount

        for i in range(91):
            day = as_of + timedelta(days=i)
            day_str = day.isoformat()

            in_opt = opt_inflows.get(day_str, 0.0)
            in_sim = sim_inflows.get(day_str, 0.0)
            out = outflows.get(day_str, 0.0)

            current_opt += (in_opt - out)
            current_sim += (in_sim - out)

            projections.append({
                "date": day_str,
                "day_label": day.strftime("%d/%m"),
                "inflow_opt": in_opt,
                "inflow_sim": in_sim,
                "outflow": out,
                "balance_opt": current_opt,
                "balance_sim": current_sim
            })

        return jsonify({
            "success": True,
            "base_balance": base_balance,
            "projections": projections
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/analytics/cashflow_forecast")
def cashflow_forecast():
    """US-062: Cashflow Forecasting API
    Predicts inflow/outflow based on unpaid invoices over the next 30 days.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from datetime import datetime, timedelta
        mst = request.args.get("mst") or session.get("active_taxpayer_mst")
        
        # Query unpaid invoices
        query = Invoice.query.filter(
            Invoice.is_cancelled == False,
            Invoice.paid_date == None
        )
        if mst:
            query = query.filter(Invoice.taxpayer_mst == mst)
            
        invoices = query.all()
        
        # Bucket by day (0 to 30 days ahead)
        today = datetime.now().date()
        forecast = []
        
        base_liquidity = 100000000  # Default base cash reserve
        current_liquidity = base_liquidity
        
        for i in range(30):
            target_date = today + timedelta(days=i)
            target_date_str = target_date.isoformat()
            
            inflow = 0
            outflow = 0
            
            for inv in invoices:
                due = inv.due_date or inv.date
                if due == target_date_str:
                    if inv.invoice_type == "sold":
                        inflow += inv.total_amount
                    else:
                        outflow += inv.total_amount
                        
            current_liquidity += (inflow - outflow)
            
            forecast.append({
                "date": target_date_str,
                "day_label": target_date.strftime("%d/%m"),
                "inflow": inflow,
                "outflow": outflow,
                "net_liquidity": current_liquidity
            })
            
        return jsonify({
            "success": True,
            "base_liquidity": base_liquidity,
            "forecast": forecast
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.patch("/api/invoices/<string:invoice_id>/payment")
def update_invoice_payment(invoice_id: str):
    """
    Update due_date and/or paid_date on an invoice.

    Body: {due_date?: "YYYY-MM-DD", paid_date?: "YYYY-MM-DD"}
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from extensions import db
        from invoices.models import Invoice as _Inv

        inv = db.session.get(_Inv, invoice_id)
        if not inv:
            return jsonify({"error": f"Không tìm thấy hóa đơn: {invoice_id}"}), 404

        body = request.get_json(force=True) or {}
        updated = False

        if "due_date" in body:
            inv.due_date = body["due_date"] or None
            updated = True
        if "paid_date" in body:
            inv.paid_date = body["paid_date"] or None
            updated = True

        if updated:
            inv.updated_at = datetime.now().isoformat()
            db.session.commit()

        return jsonify({
            "success": True,
            "id": invoice_id,
            "due_date": inv.due_date,
            "paid_date": inv.paid_date,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/profiles")
@roles_required("admin", "auditor", "viewer")
def get_taxpayer_profiles():
    """Retrieve all stored taxpayer profiles."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    try:
        profiles = TaxpayerProfile.query.order_by(TaxpayerProfile.mst).all()
        return jsonify([p.to_dict() for p in profiles])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/profiles")
@roles_required("admin", "auditor")
def create_taxpayer_profile():
    """Create or update a taxpayer profile with encrypted credentials."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    from auth.crypto import encrypt_password
    try:
        body = request.get_json() or {}
        mst = body.get("mst")
        company_name = body.get("company_name")
        gdt_username = body.get("gdt_username")
        gdt_password = body.get("gdt_password")

        if not all([mst, company_name, gdt_username, gdt_password]):
            return jsonify({"error": "Thiếu các thông tin bắt buộc."}), 400

        mst = str(mst).strip()
        if len(mst) not in [10, 14]:
            return jsonify({"error": "Mã số thuế không đúng định dạng (phải có 10 hoặc 14 ký tự)."}), 400

        encrypted_password = encrypt_password(gdt_password)

        profile = db.session.get(TaxpayerProfile, mst)
        is_update = profile is not None
        if profile:
            profile.company_name = company_name
            profile.gdt_username = gdt_username
            profile.gdt_password_encrypted = encrypted_password
        else:
            profile = TaxpayerProfile(
                mst=mst,
                company_name=company_name,
                gdt_username=gdt_username,
                gdt_password_encrypted=encrypted_password,
                is_active=True,
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.session.add(profile)

        db.session.commit()

        from invoices.security_audit_service import log_security_event
        if is_update:
            log_security_event("PROFILE", f"Updated taxpayer profile for MST: {mst}")
        else:
            log_security_event("PROFILE", f"Created new taxpayer profile for MST: {mst}")

        return jsonify({"success": True, "profile": profile.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.delete("/api/profiles/<mst>")
@roles_required("admin", "auditor")
def delete_taxpayer_profile(mst):
    """Delete a taxpayer profile and cascade delete all its invoices."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import TaxpayerProfile
    try:
        profile = db.session.get(TaxpayerProfile, mst)
        if not profile:
            return jsonify({"error": f"Không tìm thấy hồ sơ mã số thuế {mst}."}), 404

        db.session.delete(profile)
        db.session.commit()

        from invoices.security_audit_service import log_security_event
        log_security_event("PROFILE", f"Deleted taxpayer profile for MST: {mst}")

        return jsonify({"success": True, "message": f"Đã xóa hồ sơ và tất cả hóa đơn của MST {mst}."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/profiles/switch")
@roles_required("admin", "auditor", "viewer")
def switch_taxpayer_profile():
    """Switch the current active taxpayer profile in session."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        mst = body.get("mst")

        from invoices.security_audit_service import log_security_event
        if mst == "all" or not mst:
            session["active_taxpayer_mst"] = None
            log_security_event("PROFILE", "Switched active taxpayer profile to all (no filter)")
            return jsonify({"success": True, "active_taxpayer_mst": None})

        from invoices.models import TaxpayerProfile
        profile = db.session.get(TaxpayerProfile, mst)
        if not profile:
            return jsonify({"error": f"Không tìm thấy hồ sơ mã số thuế {mst}."}), 404

        session["active_taxpayer_mst"] = mst
        log_security_event("PROFILE", f"Switched active taxpayer profile to MST: {mst}")
        return jsonify({"success": True, "active_taxpayer_mst": mst})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/invoices/<invoice_id>/post-erp")
@roles_required("admin", "auditor")
def api_post_invoice_to_erp(invoice_id):
    """Manually post an invoice to the configured ERP system."""
    err = _ensure_logged_in()
    if err:
        return err

    from invoices.models import Invoice
    from invoices.erp_service import post_invoice_to_erp

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return jsonify({"error": "Không tìm thấy hóa đơn cần đồng bộ."}), 404

    success = post_invoice_to_erp(invoice)
    if success:
        return jsonify({
            "success": True,
            "message": "Đồng bộ hóa đơn lên ERP thành công.",
            "erp_synced": invoice.erp_synced,
            "erp_sync_date": invoice.erp_sync_date
        })
    else:
        return jsonify({
            "success": False,
            "message": "Đồng bộ hóa đơn lên ERP thất bại.",
            "error": invoice.erp_sync_error
        }), 400

@invoices_blueprint.get("/api/reports/fct-declaration")
def api_reports_fct_declaration():
    """
    Generate a draft of the Vietnamese Foreign Contractor Tax (FCT) Return Mẫu 01/NTNN.
    Identifies foreign e-commerce/digital giant suppliers and calculates withholding VAT/CIT.
    """
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice

        query = Invoice.query.filter(Invoice.is_cancelled == False, Invoice.invoice_type == "purchase")

        # Apply date filters
        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        fct_invoices = []
        total_revenue = 0.0
        total_vat_withheld = 0.0
        total_cit_withheld = 0.0

        for inv in invoices:
            seller_mst = (inv.seller_mst or "").strip()
            seller_name = (inv.seller_name or "").strip()
            
            # Match 900xxxxxxx or digital giants names
            is_fct = (
                seller_mst.startswith("900") or
                any(k in seller_name.lower() for k in ["google", "facebook", "meta", "amazon", "aws", "netflix", "zoom", "slack", "microsoft", "github", "digitalocean"])
            )
            
            if not is_fct:
                continue

            amount = inv.amount_before_tax or 0.0
            category = "Thương mại điện tử & Dịch vụ số khác"
            vat_rate = 0.05
            cit_rate = 0.05
            
            if inv.items and len(inv.items) > 0:
                first_item = inv.items[0].item_name
                category, vat_rate, cit_rate = classify_fct_item(first_item, seller_name)
            else:
                category, vat_rate, cit_rate = classify_fct_item("", seller_name)

            vat_withheld = amount * vat_rate
            cit_withheld = amount * cit_rate
            fct_total = vat_withheld + cit_withheld

            fct_invoices.append({
                "id": inv.id,
                "number": inv.number or "Không số",
                "date": inv.date or "Không ngày",
                "seller_name": seller_name,
                "seller_mst": seller_mst,
                "category": category,
                "amount": amount,
                "vat_rate": vat_rate,
                "vat_withheld": vat_withheld,
                "cit_rate": cit_rate,
                "cit_withheld": cit_withheld,
                "fct_total": fct_total
            })

            total_revenue += amount
            total_vat_withheld += vat_withheld
            total_cit_withheld += cit_withheld

        return jsonify({
            "success": True,
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "total_revenue": total_revenue,
            "total_vat_withheld": total_vat_withheld,
            "total_cit_withheld": total_cit_withheld,
            "total_fct_payable": total_vat_withheld + total_cit_withheld,
            "fct_invoices": fct_invoices
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/fct-declaration/export-excel")
def api_reports_fct_declaration_export_excel():
    """Export the Mẫu 01/NTNN draft to a beautifully formatted Excel spreadsheet."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        period_type = request.args.get("period_type", "monthly")
        period_value = request.args.get("period_value", "")
        year = request.args.get("year", datetime.now().strftime("%Y"))

        if not period_value:
            now = datetime.now()
            if period_type == "quarterly":
                period_value = str((now.month - 1) // 3 + 1)
            else:
                period_value = f"{now.month:02d}"

        if period_type == "monthly" and len(period_value) == 1:
            period_value = f"0{period_value}"

        from invoices.models import Invoice
        query = Invoice.query.filter(Invoice.is_cancelled == False, Invoice.invoice_type == "purchase")

        if period_type == "quarterly":
            try:
                q = int(period_value)
            except ValueError:
                q = 1
            if q == 1:
                months = ["01", "02", "03"]
            elif q == 2:
                months = ["04", "05", "06"]
            elif q == 3:
                months = ["07", "08", "09"]
            else:
                months = ["10", "11", "12"]
            
            from sqlalchemy import or_
            filters = [Invoice.date.like(f"{year}-{m}-%") for m in months]
            query = query.filter(or_(*filters))
        else:
            query = query.filter(Invoice.date.like(f"{year}-{period_value}-%"))

        invoices = query.all()

        fct_invoices = []
        total_revenue = 0.0
        total_vat_withheld = 0.0
        total_cit_withheld = 0.0

        for inv in invoices:
            seller_mst = (inv.seller_mst or "").strip()
            seller_name = (inv.seller_name or "").strip()
            
            is_fct = (
                seller_mst.startswith("900") or
                any(k in seller_name.lower() for k in ["google", "facebook", "meta", "amazon", "aws", "netflix", "zoom", "slack", "microsoft", "github", "digitalocean"])
            )
            
            if not is_fct:
                continue

            amount = inv.amount_before_tax or 0.0
            category = "Thương mại điện tử & Dịch vụ số khác"
            vat_rate = 0.05
            cit_rate = 0.05
            
            if inv.items and len(inv.items) > 0:
                first_item = inv.items[0].item_name
                category, vat_rate, cit_rate = classify_fct_item(first_item, seller_name)
            else:
                category, vat_rate, cit_rate = classify_fct_item("", seller_name)

            vat_withheld = amount * vat_rate
            cit_withheld = amount * cit_rate
            fct_total = vat_withheld + cit_withheld

            fct_invoices.append({
                "seller_name": seller_name,
                "seller_mst": seller_mst,
                "category": category,
                "amount": amount,
                "vat_rate": vat_rate,
                "vat_withheld": vat_withheld,
                "cit_rate": cit_rate,
                "cit_withheld": cit_withheld,
                "fct_total": fct_total
            })

            total_revenue += amount
            total_vat_withheld += vat_withheld
            total_cit_withheld += cit_withheld

        fct_data = {
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "total_revenue": total_revenue,
            "total_vat_withheld": total_vat_withheld,
            "total_cit_withheld": total_cit_withheld,
            "total_fct_payable": total_vat_withheld + total_cit_withheld,
            "fct_invoices": fct_invoices
        }

        excel_bytes = generate_fct_excel(fct_data)
        
        filename = f"ToKhai_01NTNN_{year}_{period_value}.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/vat-refund-eligibility")
@roles_required("admin", "auditor")
def api_vat_refund_eligibility():
    """Calculates taxpayer eligibility for VAT refund and returns a structured breakdown."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    if not mst:
        # Fallback to the first taxpayer profile
        from invoices.models import TaxpayerProfile
        first_profile = TaxpayerProfile.query.first()
        if first_profile:
            mst = first_profile.mst
        else:
            return jsonify({"error": "Không có mã số thuế hoạt động hoặc hồ sơ doanh nghiệp nào được đăng ký."}), 400

    try:
        from invoices.refund_service import VATRefundEligibilityEngine
        engine = VATRefundEligibilityEngine()
        result = engine.get_eligibility(mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi tính toán hoàn thuế: {str(e)}"}), 500

@invoices_blueprint.post("/api/reports/vat-refund-eligibility/dossier")
@roles_required("admin", "auditor")
def api_vat_refund_dossier():
    """Generates Circular 80 Mẫu 01/HT refund dossier and AI justification letter."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    mst = payload.get("mst") or session.get("active_taxpayer_mst")
    
    if not mst:
        from invoices.models import TaxpayerProfile
        first_profile = TaxpayerProfile.query.first()
        if first_profile:
            mst = first_profile.mst
        else:
            return jsonify({"error": "Không có mã số thuế hoạt động hoặc hồ sơ doanh nghiệp nào được đăng ký."}), 400

    try:
        from invoices.refund_service import VATRefundEligibilityEngine
        engine = VATRefundEligibilityEngine()
        result = engine.generate_dossier(mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi soạn hồ sơ AI: {str(e)}"}), 500

@invoices_blueprint.post("/api/reports/vat-refund-eligibility/dossier/export")
@roles_required("admin", "auditor")
def api_export_vat_refund_dossier():
    """Exports the generated dossier or justification letter to Word (.doc) or PDF."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    content = payload.get("content", "").strip()
    export_format = payload.get("format", "doc").lower()  # 'doc' or 'pdf'
    document_type = payload.get("type", "dossier")  # 'dossier' or 'justification'

    if not content:
        return jsonify({"error": "Nội dung tài liệu trống."}), 400

    filename = "Mau_01_HT_De_Nghi_Hoan_Thue" if document_type == "dossier" else "Bao_Cao_Bien_Phap_Bao_Ve_Ho_So"

    if export_format == "pdf":
        html_content = f"""
        <html>
        <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: a4;
                margin: 2cm;
            }}
            body {{
                font-family: 'Arial', sans-serif;
                font-size: 11px;
                line-height: 1.5;
            }}
            .bold {{ font-weight: bold; }}
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
            .title {{ font-size: 13px; font-weight: bold; text-align: center; margin-top: 15px; margin-bottom: 15px; }}
            p {{ margin-bottom: 6px; text-align: justify; }}
            pre {{
                font-family: 'Arial', sans-serif;
                white-space: pre-wrap;
                word-wrap: break-word;
            }}
        </style>
        </head>
        <body>
        <pre>{content}</pre>
        </body>
        </html>
        """
        try:
            pdf_buf = render_html_to_pdf(html_content)
            return send_file(
                pdf_buf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"{filename}.pdf"
            )
        except Exception as e:
            return jsonify({"error": f"Lỗi xuất PDF: {str(e)}"}), 500

    else:
        # Default to Word compatible HTML (.doc)
        html_content = f"""
        <html xmlns:o="urn:schemas-microsoft-com:office:office"
              xmlns:w="urn:schemas-microsoft-com:office:word"
              xmlns="http://www.w3.org/TR/REC-html40">
        <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: 8.27in 11.69in; /* A4 */
                margin: 1.0in 0.79in 1.0in 1.18in;
            }}
            body {{
                font-family: 'Times New Roman', serif;
                font-size: 11pt;
                line-height: 1.5;
            }}
            pre {{
                font-family: 'Times New Roman', serif;
                white-space: pre-wrap;
                word-wrap: break-word;
            }}
        </style>
        </head>
        <body>
        <pre>{content}</pre>
        </body>
        </html>
        """
        buf = BytesIO(html_content.encode("utf-8"))
        return send_file(
            buf,
            mimetype="application/msword",
            as_attachment=True,
            download_name=f"{filename}.doc"
        )

@invoices_blueprint.post("/api/audit/vat-refund-eligibility")
def api_post_vat_refund_eligibility():
    """Receives MST, input_invoice_ids, and customs_declarations, returning eligibility evaluation."""
    body = request.get_json(silent=True) or {}
    mst = body.get("mst") or session.get("active_taxpayer_mst")
    if not mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400
        
    try:
        from invoices.refund_service import VATRefundEligibilityEngine
        engine = VATRefundEligibilityEngine()
        
        input_invoice_ids = body.get("input_invoice_ids")
        customs_declarations = body.get("customs_declarations")
        
        result = engine.get_eligibility(
            mst,
            input_invoice_ids=input_invoice_ids,
            customs_declarations=customs_declarations
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi tính toán hoàn thuế: {str(e)}"}), 500

@invoices_blueprint.post("/api/audit/export-refund-xml")
def api_post_export_refund_xml():
    """Returns the GDT-compliant XML stream representing Form 01/ĐNHT."""
    body = request.get_json(silent=True) or {}
    mst = body.get("mst") or session.get("active_taxpayer_mst")
    invoice_ids = body.get("eligible_invoice_ids", [])
    bank_account = body.get("bank_account", "")
    bank_name = body.get("bank_name", "")
    reason_type = body.get("reason_type", "")
    
    if not mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400
    if not invoice_ids:
        return jsonify({"error": "No eligible invoices provided"}), 400
        
    try:
        from invoices.refund_service import generate_form_01_dnht_xml
        xml_content = generate_form_01_dnht_xml(mst, invoice_ids, bank_account, bank_name, reason_type)
        return Response(xml_content, mimetype="application/xml", headers={
            "Content-Disposition": f"attachment; filename=Form_01_DNHT_{mst}.xml"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/v1/invoices")
@require_api_signature
def api_v1_invoices():
    """REST API to fetch invoices securely with HMAC signature verification."""
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    if not mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400
        
    from invoices.models import Invoice
    invoice_type = request.args.get("invoice_type")
    query = Invoice.query.filter_by(taxpayer_mst=mst)
    if invoice_type:
        query = query.filter_by(invoice_type=invoice_type)
        
    invoices = query.all()
    return jsonify([inv.to_dict() for inv in invoices])

@invoices_blueprint.get("/api/v1/compliance-scores")
@require_api_signature
def api_v1_compliance_scores():
    """REST API to fetch compliance scores securely with HMAC signature verification."""
    mst = request.args.get("mst") or session.get("active_taxpayer_mst")
    if not mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400
        
    from invoices.models import TaxpayerProfile, Invoice
    profile = TaxpayerProfile.query.get(mst)
    if not profile:
        return jsonify({"error": f"Taxpayer profile not found for MST: {mst}"}), 404
        
    invoices = Invoice.query.filter_by(taxpayer_mst=mst).all()
    avg_t_score = 100.0
    if invoices:
        avg_t_score = sum(inv.t_score for inv in invoices) / len(invoices)
        
    return jsonify({
        "mst": mst,
        "company_name": profile.company_name,
        "average_t_score": avg_t_score,
        "risk_level": "Safe" if avg_t_score >= 80 else "Caution" if avg_t_score >= 50 else "High-Risk"
    })

@invoices_blueprint.post("/api/v1/webhooks/register")
def api_v1_webhooks_register():
    """Registers a new webhook subscription for the taxpayer."""
    body = request.get_json(silent=True) or {}
    url = body.get("url")
    secret = body.get("secret")
    mst = body.get("mst") or session.get("active_taxpayer_mst")
    event_topics = body.get("event_topics", [])
    
    if not url or not secret:
        return jsonify({"error": "Missing url or secret"}), 400
    if not mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400
    if not event_topics:
        return jsonify({"error": "Missing event_topics to subscribe"}), 400
        
    from invoices.models import WebhookSubscription
    import uuid
    
    created_subscriptions = []
    for topic in event_topics:
        sub_id = f"sub_{topic}_{uuid.uuid4().hex[:8]}"
        now_str = datetime.now().isoformat()
        sub = WebhookSubscription(
            id=sub_id,
            taxpayer_mst=mst,
            url=url,
            secret=secret,
            is_active=True,
            created_at=now_str
        )
        db.session.add(sub)
        created_subscriptions.append({
            "id": sub_id,
            "event_topic": topic,
            "url": url,
            "is_active": True
        })
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to save subscriptions: {str(e)}"}), 500
        
    return jsonify({
        "status": "success",
        "subscriptions": created_subscriptions
    }), 201

@invoices_blueprint.post("/api/v1/webhooks/dispatch-test")
def api_v1_webhooks_dispatch_test():
    """Triggers an async webhook dispatch test."""
    body = request.get_json(silent=True) or {}
    sub_id = body.get("subscription_id")
    payload = body.get("payload") or {"test": "data", "message": "Test event dispatch"}
    
    if not sub_id:
        return jsonify({"error": "Missing subscription_id"}), 400
        
    from invoices.models import WebhookSubscription
    sub = WebhookSubscription.query.get(sub_id)
    if not sub:
        return jsonify({"error": f"Webhook subscription not found for id: {sub_id}"}), 404
        
    try:
        from invoices.webhook_hub import WebhookHub
        hub = WebhookHub(db_session=db.session)
        topic = sub_id.split("_")[1] if "_" in sub_id else "test.dispatch"
        hub.trigger(
            url=sub.url,
            secret=sub.secret,
            event_topic=topic,
            payload=payload,
            subscription_id=sub.id
        )
        return jsonify({
            "status": "success",
            "message": f"Async test webhook dispatch triggered for topic: {topic}"
        })
    except Exception as e:
        return jsonify({"error": f"Failed to dispatch test: {str(e)}"}), 500

@invoices_blueprint.post("/api/audit/tax-rag-query")
def api_tax_rag_query():
    """Performs semantic RAG search over indexed tax regulations using Ollama/Fallback."""
    body = request.get_json(silent=True) or {}
    question = body.get("question")
    model = body.get("model", "gemma:2b")
    deep_research = body.get("deep_research", False)
    
    if not question:
        return jsonify({"error": "Missing question parameter"}), 400
        
    try:
        from invoices.ai_tax_advisor import query_local_tax_rag
        result = query_local_tax_rag(question, model_name=model, deep_research=deep_research)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi truy vấn RAG: {str(e)}"}), 500

@invoices_blueprint.post("/api/audit/draft-defense-letter")
def api_draft_defense_letter():
    """Drafts a formal tax defense letter template based on invoice anomalies citing Decree 125."""
    body = request.get_json(silent=True) or {}
    invoice_id = body.get("invoice_id", "INV-MOCK-99")
    issue_type = body.get("issue_type", "Chữ ký số không hợp lệ")
    seller = body.get("seller", "Công ty Cổ phần Mẫu")
    amount = body.get("amount", 25000000.0)
    mst = body.get("taxpayer_mst") or session.get("active_taxpayer_mst") or "0109998887"
    
    from invoices.models import TaxpayerProfile
    profile = TaxpayerProfile.query.get(mst)
    company_name = profile.company_name if profile else "DOANH NGHIEP"
    
    now = datetime.now()
    date_str = f"ngày {now.day} tháng {now.month} năm {now.year}"
    
    letter_template = f"""CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM
Độc lập - Tự do - Hạnh phúc
----------------

V/v: Giải trình chênh lệch/sai sót hóa đơn theo Nghị định 125/2020/NĐ-CP

Hà Nội, {date_str}

Kính gửi: Chi cục Thuế / Cục Thuế quản lý trực tiếp

1. Tên người nộp thuế: {company_name.upper()}
2. Mã số thuế: {mst}
3. Người đại diện theo pháp luật: Ban Giám đốc doanh nghiệp

Doanh nghiệp chúng tôi nhận được thông báo của Quý cơ quan về việc rà soát, giải trình các hóa đơn mua vào có dấu hiệu rủi ro. Cụ thể đối với hóa đơn mã số {invoice_id} phát hành bởi nhà cung cấp {seller} với giá trị giao dịch là {amount:,.0f} VND. Nội dung cảnh báo: {issue_type}.

Doanh nghiệp xin được giải trình cụ thể như sau:

I. Tình hình thực tế của giao dịch:
- Giao dịch mua bán hàng hóa/dịch vụ giữa hai bên là có thật, đã được hoàn thành bàn giao và có đầy đủ biên bản giao nhận hàng hóa, phiếu nhập kho, hợp đồng kinh tế đi kèm.
- Doanh nghiệp đã thực hiện thanh toán đầy đủ cho nhà cung cấp theo phương thức thanh toán thỏa thuận trong hợp đồng.

II. Căn cứ pháp lý theo Nghị định số 125/2020/NĐ-CP:
1. Đối với hành vi không cố ý hoặc do lỗi kỹ thuật chữ ký số của nhà cung cấp: Căn cứ theo Điều 9 Nghị định 125/2020/NĐ-CP quy định về các trường hợp không xử phạt vi phạm hành chính về thuế, hóa đơn đối với các sự cố khách quan hoặc lỗi hệ thống công nghệ thông tin của bên thứ ba.
2. Đối với chênh lệch thuế GTGT: Doanh nghiệp đã chủ động loại trừ các hóa đơn có rủi ro cao ra khỏi hồ sơ hoàn thuế để tự điều chỉnh theo quy định, không làm phát sinh số thuế thiếu hoặc trốn thuế quy định tại Điều 16 Nghị định 125/2020/NĐ-CP.

Doanh nghiệp xin cam đoan các thông tin giải trình nêu trên là đúng sự thật và kính mong Quý cơ quan xem xét, tạo điều kiện thuận lợi cho doanh nghiệp trong quá trình chấp hành pháp luật thuế.

ĐẠI DIỆN HỢP PHÁP CỦA DOANH NGHIỆP
(Ký, ghi rõ họ tên và đóng dấu)
"""
    return jsonify({
        "status": "success",
        "invoice_id": invoice_id,
        "draft_letter": letter_template
    })

@invoices_blueprint.post("/api/bank/reconcile/upload")
def api_bank_reconcile_upload():
    """Ingests a Techcombank/Vietcombank Excel statement and stores transactions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = request.form.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    bank_name = request.form.get("bank_name", "Generic")
    account_number = request.form.get("account_number", "")

    if not taxpayer_mst:
        return jsonify({"error": "Mã số thuế hoạt động trống."}), 400

    if "file" not in request.files:
        return jsonify({"error": "Không có tệp tải lên."}), 400

    uploaded_file = request.files["file"]
    if not uploaded_file.filename:
        return jsonify({"error": "Tên tệp không hợp lệ."}), 400

    # Save to a local temporary location in our designated temp dir
    temp_dir = os.path.join(current_app.root_path, "data", "temp")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, f"statement_{int(datetime.now().timestamp())}.xlsx")
    uploaded_file.save(temp_path)

    try:
        from invoices.bank_reconcile_service import parse_bank_statement
        parsed_txs = parse_bank_statement(temp_path, bank_name)
        
        from invoices.models import BankTransaction
        imported_count = 0
        skipped_count = 0
        
        for p in parsed_txs:
            # Check for reference duplicate
            exists = BankTransaction.query.filter_by(id=p["id"]).first()
            if exists:
                skipped_count += 1
                continue
                
            tx = BankTransaction(
                id=p["id"],
                taxpayer_mst=taxpayer_mst,
                bank_name=p["bank_name"],
                account_number=account_number,
                transaction_date=p["transaction_date"],
                reference_number=p["reference_number"],
                description=p["description"],
                amount=p["amount"],
                status="unreconciled",
                imported_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.session.add(tx)
            imported_count += 1
            
        if imported_count > 0:
            db.session.commit()
            
        return jsonify({
            "status": "success",
            "imported_count": imported_count,
            "skipped_count": skipped_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi xử lý tệp sổ phụ: {str(e)}"}), 500
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

@invoices_blueprint.get("/api/bank/reconcile/transactions")
def api_bank_reconcile_transactions():
    """Retrieve bank transactions list for active MST."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    status_filter = request.args.get("status", "all")  # 'all', 'unreconciled', 'matched'

    if not taxpayer_mst:
        return jsonify([])

    from invoices.models import BankTransaction, Invoice
    query = BankTransaction.query.filter_by(taxpayer_mst=taxpayer_mst)
    if status_filter != "all":
        query = query.filter_by(status=status_filter)

    transactions = query.order_by(BankTransaction.transaction_date.desc()).all()
    
    result = []
    for tx in transactions:
        tx_dict = tx.to_dict()
        if tx.matched_invoice_id:
            inv = Invoice.query.get(tx.matched_invoice_id)
            if inv:
                tx_dict["invoice_number"] = inv.number
                tx_dict["partner_name"] = inv.buyer_name if tx.amount > 0 else inv.seller_name
        else:
            tx_dict["invoice_number"] = ""
            tx_dict["partner_name"] = ""
        result.append(tx_dict)
        
    return jsonify(result)

@invoices_blueprint.post("/api/bank/reconcile/auto")
def api_bank_reconcile_auto():
    """Triggers autonomous Soundex/Phonetic matching reconciliation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    taxpayer_mst = payload.get("taxpayer_mst") or session.get("active_taxpayer_mst")

    if not taxpayer_mst:
        return jsonify({"error": "Mã số thuế hoạt động trống."}), 400

    try:
        from invoices.bank_reconcile_service import execute_auto_reconciliation
        res = execute_auto_reconciliation(taxpayer_mst)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": f"Lỗi đối chiếu tự động: {str(e)}"}), 500

@invoices_blueprint.post("/api/bank/reconcile/manual")
def api_bank_reconcile_manual():
    """Manual reconciliation override by ledger accountant."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    transaction_id = payload.get("transaction_id")
    invoice_id = payload.get("invoice_id")

    if not transaction_id or not invoice_id:
        return jsonify({"error": "Thiếu mã giao dịch hoặc mã hóa đơn khớp."}), 400

    from invoices.models import BankTransaction, Invoice
    tx = BankTransaction.query.get(transaction_id)
    inv = Invoice.query.get(invoice_id)

    if not tx or not inv:
        return jsonify({"error": "Không tìm thấy giao dịch ngân hàng hoặc hóa đơn tương ứng."}), 404

    try:
        tx.matched_invoice_id = invoice_id
        tx.confidence_score = 1.0  # Manual matching has 100% confidence
        tx.status = "matched"
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Đối chiếu thủ công thành công.",
            "details": {
                "transaction_id": tx.id,
                "matched_invoice_id": invoice_id,
                "invoice_number": inv.number,
                "partner_name": inv.buyer_name if tx.amount > 0 else inv.seller_name
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi đối chiếu thủ công: {str(e)}"}), 500

@invoices_blueprint.get("/issue-invoice")
def issue_invoice_page():
    """Render the e-invoice draft builder page."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("issue_invoice.html")

@invoices_blueprint.post("/api/invoices/issue/draft")
def api_issue_draft():
    """Create a new e-invoice draft."""
    import os
    from datetime import datetime

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    buyer_mst = payload.get("buyer_mst", "").strip()
    buyer_name = payload.get("buyer_name", "").strip()
    buyer_address = payload.get("buyer_address", "").strip()
    items_data = payload.get("items") or []

    if not buyer_mst or not buyer_name:
        return jsonify({"error": "Thiếu mã số thuế hoặc tên đơn vị mua hàng."}), 400

    if not items_data:
        return jsonify({"error": "Danh sách hàng hóa dịch vụ không được trống."}), 400

    # Get active taxpayer
    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        return jsonify({"error": "Vui lòng chọn doanh nghiệp hoạt động trước."}), 400

    from invoices.models import TaxpayerProfile, Invoice, LineItem
    seller = TaxpayerProfile.query.filter_by(mst=active_mst).first()
    seller_name = seller.company_name if seller else "Công ty Phát hành Mẫu"
    seller_address = "123 Đường Phát Hành, Hà Nội"

    # Auto-increment number
    symbol = payload.get("symbol", "1C26TYY").strip()
    
    # Query count to auto-increment number
    count = Invoice.query.filter(Invoice.seller_mst == active_mst, Invoice.symbol == symbol).count()
    number = f"{count + 1:07d}"
    
    invoice_id = f"{active_mst}-{symbol}-{number}"

    # Calculate totals
    amount_before_tax = 0.0
    tax_amount = 0.0
    
    line_items = []
    
    for idx, item in enumerate(items_data):
        name = item.get("item_name", "").strip()
        unit = item.get("unit", "").strip()
        try:
            qty = float(item.get("quantity") or 0.0)
            price = float(item.get("unit_price") or 0.0)
        except ValueError:
            return jsonify({"error": f"Số lượng hoặc đơn giá của mục {idx+1} không hợp lệ."}), 400
            
        tax_rate_str = item.get("tax_rate", "10%")
        
        # Calculate item totals
        item_amt = qty * price
        
        # Calculate tax
        if "10" in tax_rate_str:
            item_tax = 0.10 * item_amt
        elif "8" in tax_rate_str:
            item_tax = 0.08 * item_amt
        elif "5" in tax_rate_str:
            item_tax = 0.05 * item_amt
        else:
            item_tax = 0.0
            
        amount_before_tax += item_amt
        tax_amount += item_tax
        
        line_items.append({
            "item_name": name,
            "unit": unit,
            "quantity": qty,
            "unit_price": price,
            "amount_before_tax": item_amt,
            "tax_rate": tax_rate_str,
            "tax_amount": item_tax
        })

    total_amount = amount_before_tax + tax_amount
    
    # Spell money in Vietnamese
    from invoices.ai_service import spell_money_vietnamese
    amount_in_words = spell_money_vietnamese(total_amount)

    try:
        inv = Invoice(
            id=invoice_id,
            filename=f"invoice_{invoice_id}.xml",
            invoice_type="sold",
            template_code="1",
            symbol=symbol,
            number=number,
            date=datetime.now().strftime("%Y-%m-%d"),
            currency="VND",
            seller_name=seller_name,
            seller_mst=active_mst,
            seller_address=seller_address,
            buyer_name=buyer_name,
            buyer_mst=buyer_mst,
            buyer_address=buyer_address,
            amount_before_tax=amount_before_tax,
            tax_amount=tax_amount,
            total_amount=total_amount,
            has_signature=False,
            amount_in_words=amount_in_words,
            imported_at=datetime.now().isoformat(),
            import_status="draft",
            invoice_status="draft",
            taxpayer_mst=active_mst
        )
        db.session.add(inv)
        
        for item_data in line_items:
            li = LineItem(
                invoice_id=invoice_id,
                item_name=item_data["item_name"],
                unit=item_data["unit"],
                quantity=item_data["quantity"],
                unit_price=item_data["unit_price"],
                amount_before_tax=item_data["amount_before_tax"],
                tax_rate=item_data["tax_rate"],
                tax_amount=item_data["tax_amount"],
                expense_category="REVENUE"
            )
            db.session.add(li)
            
        db.session.commit()
        return jsonify({
            "status": "success",
            "message": "Tạo hóa đơn nháp thành công.",
            "invoice": inv.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi lưu hóa đơn nháp: {str(e)}"}), 500

@invoices_blueprint.post("/api/invoices/issue/sign")
def api_issue_sign():
    """Digital sign draft e-invoice using mock USB Token."""
    import os
    import json
    from datetime import datetime

    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_id = payload.get("invoice_id")

    if not invoice_id:
        return jsonify({"error": "Thiếu mã hóa đơn cần ký."}), 400

    from invoices.models import Invoice, LineItem
    inv = Invoice.query.get(invoice_id)
    if not inv:
        return jsonify({"error": "Không tìm thấy hóa đơn tương ứng."}), 404

    if inv.invoice_status != "draft":
        return jsonify({"error": "Hóa đơn này đã được phát hành và ký số."}), 400

    try:
        # Build GDT Circular 78 Compliant XML DSHDon list
        items_xml = ""
        for idx, item in enumerate(inv.items):
            items_xml += f"""        <HHDVu>
          <TChat>1</TChat>
          <STT>{idx + 1}</STT>
          <Ten>{item.item_name}</Ten>
          <DVT>{item.unit or 'Lần'}</DVT>
          <SLuong>{item.quantity}</SLuong>
          <DGia>{item.unit_price}</DGia>
          <ThTien>{item.amount_before_tax}</ThTien>
          <TSuat>{item.tax_rate}</TSuat>
          <TThue>{item.tax_amount}</TThue>
        </HHDVu>"""

        # Compile Canonical DLHDon XML
        dlhdon_xml = f"""<DLHDon Id="HD_{inv.id}">
      <TTChung>
        <PBan>1.0.0</PBan>
        <THDon>Hóa đơn giá trị gia tăng</THDon>
        <KHHDon>{inv.symbol}</KHHDon>
        <SHDon>{inv.number}</SHDon>
        <NLap>{inv.date}</NLap>
        <DVTTe>{inv.currency or 'VND'}</DVTTe>
        <TGia>1.0</TGia>
      </TTChung>
      <NDHDon>
        <NBan>
          <Ten>{inv.seller_name}</Ten>
          <MST>{inv.seller_mst}</MST>
          <DChi>{inv.seller_address or ''}</DChi>
        </NBan>
        <NMua>
          <Ten>{inv.buyer_name}</Ten>
          <MST>{inv.buyer_mst}</MST>
          <DChi>{inv.buyer_address or ''}</DChi>
        </NMua>
        <DSHDon>
{items_xml}
        </DSHDon>
        <TToan>
          <TgTCThue>{inv.amount_before_tax}</TgTCThue>
          <TgTThue>{inv.tax_amount}</TgTThue>
          <TgTTTBSo>{inv.total_amount}</TgTTTBSo>
          <TgTTTBChu>{inv.amount_in_words}</TgTTTBChu>
        </TToan>
      </NDHDon>
    </DLHDon>"""

        # Perform SHA-256 + RSA-2048 mock USB token cryptographic signing
        import hashlib
        import base64
        
        # Calculate digest
        digest = hashlib.sha256(dlhdon_xml.encode("utf-8")).digest()
        digest_b64 = base64.b64encode(digest).decode("utf-8")
        
        # Simulate USB Token RSA signature
        sig_b64 = base64.b64encode(hashlib.sha256(digest).digest() * 2).decode("utf-8")[:172] + "=="
        
        # Mock certificate x509
        mock_cert = "MIIDuTCCAqGgAwIBAgIUdT6ySjZ+N...MOCK_GDT_CIRCULAR_78_CERTIFICATE..."
        
        # Complete Circular 78 XML Package
        full_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<HDon>
    {dlhdon_xml}
    <Signature xmlns="http://www.w3.org/2000/09/xmldsig#">
        <SignedInfo>
            <CanonicalizationMethod Algorithm="http://www.w3.org/TR/2001/REC-xml-c14n-20010315"/>
            <SignatureMethod Algorithm="http://www.w3.org/2000/09/xmldsig#rsa-sha256"/>
            <Reference URI="#HD_{inv.id}">
                <Transforms>
                    <Transform Algorithm="http://www.w3.org/2000/09/xmldsig#enveloped-signature"/>
                </Transforms>
                <DigestMethod Algorithm="http://www.w3.org/2001/04/xmlenc#sha256"/>
                <DigestValue>{digest_b64}</DigestValue>
            </Reference>
        </SignedInfo>
        <SignatureValue>{sig_b64}</SignatureValue>
        <KeyInfo>
            <X509Data>
                <X509Certificate>{mock_cert}</X509Certificate>
            </X509Data>
        </KeyInfo>
    </Signature>
</HDon>"""

        # Store signed XML file locally
        from invoices.service import XML_DIR
        safe_filename = f"invoice_{inv.id}.xml"
        xml_path = os.path.join(XML_DIR, safe_filename)
        with open(xml_path, "w", encoding="utf-8") as f:
            f.write(full_xml)

        # Update database model state
        inv.has_signature = True
        inv.signing_date = datetime.now().strftime("%Y-%m-%d")
        inv.import_status = "imported"
        inv.invoice_status = "Gốc"
        
        # Also generate mock signature JSON for frontend display
        inv.signature_details_json = json.dumps({
            "subject": f"C=VN, ST=Hanoi, O={inv.seller_name}, CN={inv.seller_name}",
            "issuer": "VNPT CA / GDT Root CA",
            "valid_from": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "valid_to": "2029-12-31 23:59:59",
            "serial": "18392098487293847293847",
            "algo": "sha256RSA"
        }, ensure_ascii=False)
        
        db.session.commit()
        
        # Trigger an SSE update stream event for newly issued invoice
        try:
            from invoices.sync_daemon import push_sync_event
            push_sync_event("invoice_downloaded", {
                "id": inv.id,
                "seller": inv.seller_name,
                "buyer": inv.buyer_name,
                "amount": inv.total_amount
            })
        except Exception:
            pass

        return jsonify({
            "status": "success",
            "message": "Ký số hóa đơn thành công thông qua USB Token.",
            "invoice_id": inv.id,
            "xml_preview": full_xml[:1000] + "..."
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Lỗi ký số hóa đơn: {str(e)}"}), 500

@invoices_blueprint.get("/api/audit/ledger")
def get_audit_ledger():
    """List audit ledger blocks with pagination (US-090)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.models import AuditBlock

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    per_page = min(per_page, 200)

    query = AuditBlock.query.order_by(AuditBlock.block_id.desc())
    total = query.count()
    blocks = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "blocks": [b.to_dict() for b in blocks],
    })

@invoices_blueprint.post("/api/audit/verify")
def verify_audit_ledger():
    """Run full-chain cryptographic integrity verification (US-091)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.audit_ledger_service import verify_ledger_integrity
    from invoices.models import AuditBlock

    is_valid, corrupted_id, error_msg = verify_ledger_integrity()
    total_blocks = AuditBlock.query.count()

    return jsonify({
        "is_valid": is_valid,
        "total_blocks": total_blocks,
        "corrupted_block_id": corrupted_id,
        "error_message": error_msg,
    })

@invoices_blueprint.get("/api/audit/stats")
def get_audit_stats():
    """Return summary statistics for the audit ledger dashboard (US-091)."""
    guard = _ensure_logged_in()
    if guard:
        return guard

    from invoices.models import AuditBlock
    from sqlalchemy import func

    total = AuditBlock.query.count()
    action_counts = (
        db.session.query(AuditBlock.action_type, func.count(AuditBlock.block_id))
        .group_by(AuditBlock.action_type)
        .all()
    )

    latest = AuditBlock.query.order_by(AuditBlock.block_id.desc()).first()

    return jsonify({
        "total_blocks": total,
        "action_breakdown": {action: count for action, count in action_counts},
        "latest_block": latest.to_dict() if latest else None,
    })

@invoices_blueprint.post("/api/analytics/forecast")
@roles_required("admin", "auditor", "viewer")
def api_forecast_tax():
    """Forecast future tax liability using moving averages (US-110, US-111)."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        historical_data = body.get("historical_data")
        projected_period = body.get("projected_period", datetime.now().strftime("%Y-%m"))
        alpha = body.get("alpha", 0.7)
        window_size = body.get("window_size", 3)
        budget_limit = body.get("budget_limit", 500000000.0)

        active_mst = session.get("active_taxpayer_mst")
        if not active_mst:
            from invoices.models import TaxpayerProfile
            prof = TaxpayerProfile.query.filter_by(is_active=True).first()
            if prof:
                active_mst = prof.mst

        # Query from DB if not provided in request body
        if historical_data is None:
            if not active_mst:
                return jsonify({"error": "Không tìm thấy mã số thuế hoạt động để truy vấn dữ liệu."}), 400

            from invoices.models import Invoice
            sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
            purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

            from collections import defaultdict
            period_map = defaultdict(lambda: {"output_vat": 0.0, "input_vat": 0.0})
            
            for s in sales:
                if not s.date or len(s.date) < 7:
                    continue
                period_map[s.date[:7]]["output_vat"] += s.tax_amount
            for p in purchases:
                if not p.date or len(p.date) < 7:
                    continue
                period_map[p.date[:7]]["input_vat"] += p.tax_amount

            historical_data = []
            for p in sorted(period_map.keys()):
                historical_data.append({
                    "period": p,
                    "output_vat": period_map[p]["output_vat"],
                    "input_vat": period_map[p]["input_vat"]
                })

        from invoices.tax_forecaster import forecast_next_period_tax, TaxAlertManager
        forecast = forecast_next_period_tax(
            historical_data,
            projected_period=projected_period,
            alpha=alpha,
            window_size=window_size
        )

        # Run alerts evaluation
        alert_manager = TaxAlertManager(budget_limit=budget_limit)
        forecast_evaluated = alert_manager.evaluate_forecast(forecast)

        return jsonify({
            "taxpayer_mst": active_mst,
            "forecast": forecast_evaluated.to_dict()
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/invoices/batch-parse")
@roles_required("admin", "auditor")
def api_batch_parse_invoices():
    """Concurrently parse XML invoices, decompress if zipped, and import to DB (US-112, US-113)."""
    err = _ensure_logged_in()
    if err:
        return err

    try:
        body = request.get_json() or {}
        invoice_items = body.get("invoices", [])
        duplicate_strategy = body.get("duplicate_strategy", "overwrite")
        active_mst = session.get("active_taxpayer_mst")

        # 1. Prepare batch (decompress if needed)
        parsed_batch_inputs = []
        for idx, item in enumerate(invoice_items):
            filename = item.get("filename", f"invoice_{idx}.xml")
            content = item.get("content", "")
            compressed = item.get("compressed", False)

            if not content:
                continue

            try:
                if compressed:
                    import base64
                    try:
                        byte_data = base64.b64decode(content)
                    except Exception:
                        byte_data = bytes.fromhex(content)
                    
                    from invoices.batch_parser import decompress_xml
                    xml_str = decompress_xml(byte_data)
                else:
                    xml_str = content

                parsed_batch_inputs.append((filename, xml_str))
            except Exception as e:
                pass

        # 2. Parallel Parse
        from invoices.batch_parser import parse_batch_xml
        parse_results = parse_batch_xml(parsed_batch_inputs)

        # 3. Serial Import to DB
        from invoices.service import import_xml_invoice
        db_results = []
        for res, (filename, xml_str) in zip(parse_results, parsed_batch_inputs):
            if not res.success:
                db_results.append({
                    "filename": filename,
                    "success": False,
                    "error_message": res.error_message
                })
                continue
            
            try:
                xml_bytes = xml_str.encode("utf-8")
                imported_dict = import_xml_invoice(
                    xml_bytes,
                    filename,
                    duplicate_strategy=duplicate_strategy,
                    taxpayer_mst=active_mst
                )
                db_results.append({
                    "filename": filename,
                    "success": True,
                    "invoice_id": imported_dict.get("id"),
                    "invoice_number": imported_dict.get("number"),
                    "total_amount": imported_dict.get("total_amount")
                })
            except Exception as e:
                db_results.append({
                    "filename": filename,
                    "success": False,
                    "error_message": str(e)
                })

        return jsonify({
            "total_processed": len(db_results),
            "results": db_results
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/analytics/kpis")
@roles_required("admin", "auditor", "viewer")
def api_get_financial_kpis():
    """Retrieve financial health metrics (Gross Margin, Tax Ratios, Clearance times) (US-114)."""
    err = _ensure_logged_in()
    if err:
        return err

    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        from invoices.models import TaxpayerProfile
        prof = TaxpayerProfile.query.filter_by(is_active=True).first()
        if prof:
            active_mst = prof.mst

    if not active_mst:
        return jsonify({"error": "Không tìm thấy mã số thuế hoạt động. Vui lòng chọn hồ sơ MST."}), 400

    from invoices.models import Invoice
    sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
    purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

    sales_dicts = [
        {"id": s.id, "amount_before_tax": s.amount_before_tax, "tax_amount": s.tax_amount, "date": s.date}
        for s in sales
    ]
    purchases_dicts = [
        {"id": p.id, "amount_before_tax": p.amount_before_tax, "tax_amount": p.tax_amount, "date": p.date}
        for p in purchases
    ]

    clearances = []
    for inv in sales + purchases:
        if inv.paid_date:
            clearances.append({
                "invoice_id": inv.id,
                "clearance_date": inv.paid_date
            })

    from invoices.financial_kpi import calculate_financial_kpis
    kpi = calculate_financial_kpis(sales_dicts, purchases_dicts, clearances)

    from collections import defaultdict
    sales_by_month = defaultdict(list)
    purchases_by_month = defaultdict(list)
    clearances_by_month = defaultdict(list)

    for s in sales_dicts:
        month = s["date"][:7] if s.get("date") else "unknown"
        sales_by_month[month].append(s)
    for p in purchases_dicts:
        month = p["date"][:7] if p.get("date") else "unknown"
        purchases_by_month[month].append(p)
    for c in clearances:
        inv_date = None
        for s in sales_dicts:
            if s["id"] == c["invoice_id"]:
                inv_date = s["date"]
                break
        if not inv_date:
            for p in purchases_dicts:
                if p["id"] == c["invoice_id"]:
                    inv_date = p["date"]
                    break
        month = inv_date[:7] if inv_date else "unknown"
        clearances_by_month[month].append(c)

    all_months = set(sales_by_month.keys()).union(purchases_by_month.keys())
    all_months.discard("unknown")
    
    monthly_trends = {}
    for month in sorted(all_months):
        m_kpi = calculate_financial_kpis(
            sales_by_month[month],
            purchases_by_month[month],
            clearances_by_month[month]
        )
        monthly_trends[month] = m_kpi.to_dict()

    return jsonify({
        "taxpayer_mst": active_mst,
        "overall": kpi.to_dict(),
        "monthly_trends": monthly_trends
    })

@invoices_blueprint.get("/api/analytics/kpis/export")
@roles_required("admin", "auditor", "viewer")
def api_export_financial_kpis():
    """Export monthly financial KPIs to a downloadable CSV file (US-115)."""
    err = _ensure_logged_in()
    if err:
        return err

    active_mst = session.get("active_taxpayer_mst")
    if not active_mst:
        from invoices.models import TaxpayerProfile
        prof = TaxpayerProfile.query.filter_by(is_active=True).first()
        if prof:
            active_mst = prof.mst

    if not active_mst:
        return jsonify({"error": "Không tìm thấy mã số thuế hoạt động. Vui lòng chọn hồ sơ MST."}), 400

    from invoices.models import Invoice
    sales = Invoice.query.filter_by(seller_mst=active_mst, is_cancelled=False).all()
    purchases = Invoice.query.filter_by(buyer_mst=active_mst, is_cancelled=False).all()

    sales_dicts = [
        {"id": s.id, "amount_before_tax": s.amount_before_tax, "tax_amount": s.tax_amount, "date": s.date}
        for s in sales
    ]
    purchases_dicts = [
        {"id": p.id, "amount_before_tax": p.amount_before_tax, "tax_amount": p.tax_amount, "date": p.date}
        for p in purchases
    ]

    clearances = []
    for inv in sales + purchases:
        if inv.paid_date:
            clearances.append({
                "invoice_id": inv.id,
                "clearance_date": inv.paid_date
            })

    from collections import defaultdict
    sales_by_month = defaultdict(list)
    purchases_by_month = defaultdict(list)
    clearances_by_month = defaultdict(list)

    for s in sales_dicts:
        month = s["date"][:7] if s.get("date") else "unknown"
        sales_by_month[month].append(s)
    for p in purchases_dicts:
        month = p["date"][:7] if p.get("date") else "unknown"
        purchases_by_month[month].append(p)
    for c in clearances:
        inv_date = None
        for s in sales_dicts:
            if s["id"] == c["invoice_id"]:
                inv_date = s["date"]
                break
        if not inv_date:
            for p in purchases_dicts:
                if p["id"] == c["invoice_id"]:
                    inv_date = p["date"]
                    break
        month = inv_date[:7] if inv_date else "unknown"
        clearances_by_month[month].append(c)

    all_months = set(sales_by_month.keys()).union(purchases_by_month.keys())
    all_months.discard("unknown")
    
    from invoices.financial_kpi import calculate_financial_kpis, export_kpi_to_csv
    
    period_metrics = {}
    for month in sorted(all_months):
        period_metrics[month] = calculate_financial_kpis(
            sales_by_month[month],
            purchases_by_month[month],
            clearances_by_month[month]
        )

    csv_content = export_kpi_to_csv(period_metrics)

    from flask import Response
    response = Response(csv_content, mimetype="text/csv")
    filename = f"kpi_report_{active_mst}_{datetime.now().strftime('%Y%m%d')}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@invoices_blueprint.get("/api/compliance/rulebook")
@roles_required("admin", "auditor", "viewer")
def api_get_compliance_rulebook():
    """Retrieve the currently active compliance rulebook (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import json
    from invoices.models import ComplianceRulebook
    active_mst = session.get("taxpayer_mst")
    
    rulebook = None
    if active_mst:
        rulebook = ComplianceRulebook.query.filter_by(taxpayer_mst=active_mst, is_active=True).first()
    
    if not rulebook:
        rulebook = ComplianceRulebook.query.filter_by(id="rulebook_default").first()

    if not rulebook:
        default_rulebook_json = {
            "name": "Default Compliance Rulebook",
            "rules": [
                {
                    "id": "rule_cash_limit",
                    "name": "Verify cash transactions over 20M limit",
                    "severity": "critical",
                    "channels": ["in_app"],
                    "expression": {
                        "and": [
                            {"field": "payment_method", "op": "contains", "value": "Tiền mặt"},
                            {"field": "total_amount", "op": ">=", "value": 20000000}
                        ]
                    }
                }
            ]
        }
        return jsonify({
            "status": "success",
            "rulebook": default_rulebook_json
        })

    try:
        data = json.loads(rulebook.rulebook_json)
    except Exception:
        data = {}

    return jsonify({
        "status": "success",
        "rulebook": data,
        "updated_at": rulebook.updated_at
    })

@invoices_blueprint.post("/api/compliance/rulebook")
@roles_required("admin", "auditor")
def api_update_compliance_rulebook():
    """Update or upload the active compliance rulebook DSL (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    rulebook_data = payload.get("rulebook")
    if not rulebook_data:
        return jsonify({"error": "Dữ liệu rulebook trống."}), 400

    import json
    from invoices.compliance_hub import validate_rulebook_dsl
    ok, err = validate_rulebook_dsl(rulebook_data)
    if not ok:
        return jsonify({"error": f"Lỗi cú pháp DSL Rulebook: {err}"}), 400

    from invoices.models import ComplianceRulebook
    active_mst = session.get("taxpayer_mst")
    rulebook_id = f"rulebook_{active_mst}" if active_mst else "rulebook_default"
    
    rulebook = db.session.get(ComplianceRulebook, rulebook_id)
    now = datetime.now().isoformat()
    
    if not rulebook:
        rulebook = ComplianceRulebook(
            id=rulebook_id,
            taxpayer_mst=active_mst,
            name=rulebook_data.get("name", "Custom Rulebook"),
            rulebook_json=json.dumps(rulebook_data, ensure_ascii=False),
            is_active=True,
            updated_at=now
        )
        db.session.add(rulebook)
    else:
        rulebook.name = rulebook_data.get("name", rulebook.name)
        rulebook.rulebook_json = json.dumps(rulebook_data, ensure_ascii=False)
        rulebook.updated_at = now

    db.session.commit()

    from invoices.security_audit_service import log_security_event
    log_security_event("UPDATE", f"Updated compliance rulebook DSL: {rulebook.name}")

    return jsonify({
        "status": "success",
        "message": "Cập nhật DSL Rulebook thành công.",
        "updated_at": now
    })

@invoices_blueprint.post("/api/compliance/evaluate")
@roles_required("admin", "auditor", "viewer")
def api_evaluate_compliance():
    """Evaluate compliance of specified invoices against the active rulebook (US-120)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_ids = payload.get("invoice_ids", [])
    
    if not invoice_ids:
        return jsonify({"error": "Danh sách invoice_ids trống."}), 400

    from invoices.models import Invoice, ComplianceRulebook
    from invoices.compliance_hub import ComplianceEngine
    import json

    active_mst = session.get("taxpayer_mst")
    rulebook = None
    if active_mst:
        rulebook = ComplianceRulebook.query.filter_by(taxpayer_mst=active_mst, is_active=True).first()
    if not rulebook:
        rulebook = ComplianceRulebook.query.filter_by(id="rulebook_default").first()

    engine = ComplianceEngine()
    if rulebook:
        try:
            rulebook_data = json.loads(rulebook.rulebook_json)
            engine.set_rulebook(rulebook_data)
        except Exception:
            pass

    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
    all_alerts = []
    
    for inv in invoices:
        # Convert invoice model to dictionary format suited for ComplianceEngine
        inv_dict = inv.to_dict()
        alerts = engine.evaluate_invoice(inv_dict)
        all_alerts.extend([a.to_dict() for a in alerts])

    return jsonify({
        "status": "success",
        "alerts": all_alerts
    })

@invoices_blueprint.post("/api/compliance/map-ifrs")
@roles_required("admin", "auditor", "viewer")
def api_map_ifrs_compliance():
    """Map invoices to standard IFRS & calculate FCT liabilities dynamically (US-121)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    invoice_ids = payload.get("invoice_ids", [])
    reporting_currency = payload.get("reporting_currency", "USD").strip().upper()
    fct_category = payload.get("fct_category", "services").strip().lower()

    if not invoice_ids:
        return jsonify({"error": "Danh sách invoice_ids trống."}), 400

    from invoices.models import Invoice
    from invoices.tax_mapping import TaxMappingEngine
    from dataclasses import asdict

    engine = TaxMappingEngine()
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
    
    mapped_results = []
    for inv in invoices:
        inv_dict = inv.to_dict()
        mapping = engine.map_to_ifrs(inv_dict, reporting_currency=reporting_currency, fct_category=fct_category)
        mapped_results.append(asdict(mapping))

    return jsonify({
        "status": "success",
        "reporting_currency": reporting_currency,
        "mapped_invoices": mapped_results
    })

@invoices_blueprint.post("/api/compliance/ias12-deferred-tax")
@roles_required("admin", "auditor", "viewer")
def api_ias12_deferred_tax():
    """Calculate IAS 12 deferred taxes for a given MST and year."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    active_mst = payload.get("mst") or session.get("active_taxpayer_mst") or session.get("taxpayer_mst")
    year = payload.get("year")
    
    if not active_mst:
        return jsonify({"error": "Mã số thuế không được để trống"}), 400
    if not year:
        year = datetime.now().year

    try:
        from invoices.ifrs_engine import IFRSTranslationService
        service = IFRSTranslationService()
        records = service.calculate_ias12_deferred_tax(active_mst, int(year))
        return jsonify({
            "status": "success",
            "mst": active_mst,
            "year": year,
            "records": records
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/ifrs16-lease-schedule")
@roles_required("admin", "auditor", "viewer")
def api_ifrs16_lease_schedule():
    """Generate IFRS 16 lease amortization schedule month-by-month."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    lease_id = payload.get("lease_id")
    monthly_payment = payload.get("monthly_payment")
    discount_rate = payload.get("discount_rate")
    lease_term_months = payload.get("lease_term_months")

    if not lease_id:
        return jsonify({"error": "lease_id không được để trống"}), 400
    if monthly_payment is None or discount_rate is None or lease_term_months is None:
        return jsonify({"error": "Thiếu các thông số tính toán amortization"}), 400

    try:
        from invoices.ifrs_engine import IFRSTranslationService
        service = IFRSTranslationService()
        schedule = service.calculate_ifrs16_amortization(
            lease_id, float(monthly_payment), float(discount_rate), int(lease_term_months)
        )
        return jsonify({
            "status": "success",
            "lease_id": lease_id,
            "schedule": schedule
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/pillar-two-estimate")
@roles_required("admin", "auditor", "viewer")
def api_pillar_two_estimate():
    """Estimate consolidated OECD Pillar Two GloBE top-up taxes."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    payload = request.get_json(silent=True) or {}
    parent_mst = payload.get("parent_mst") or session.get("active_taxpayer_mst") or session.get("taxpayer_mst")
    group_msts = payload.get("group_msts")
    year = payload.get("year")

    if not parent_mst:
        return jsonify({"error": "parent_mst không được để trống"}), 400
    if not group_msts or not isinstance(group_msts, list):
        return jsonify({"error": "group_msts phải là danh sách MST hợp lệ"}), 400
    if not year:
        year = datetime.now().year

    try:
        from invoices.ifrs_engine import IFRSTranslationService
        service = IFRSTranslationService()
        estimate = service.estimate_pillar_two_topup(parent_mst, group_msts, int(year))
        return jsonify({
            "status": "success",
            "estimate": estimate
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/tax-risk-scoreboard")
@roles_required("admin", "auditor", "viewer")
def api_tax_risk_scoreboard():
    """Retrieve tax compliance audit warning distribution and supplier risk scoreboard."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.supplier_risk_service import get_all_suppliers_risk_radar
    active_mst = session.get("taxpayer_mst")

    # Fetch dynamic supplier risk radar data
    radar_data = get_all_suppliers_risk_radar(active_mst)

    # Map fields for UI compatibility
    for s in radar_data["suppliers"]:
        s["average_t_score"] = s["risk_score"]
        s["warnings_count"] = len(s["flags"])
        s["is_blacklisted"] = "BLACKLISTED" in s["flags"] or s.get("gdt_status") == "BLACKLISTED"

    # Only include suppliers with warnings or blacklisted in high_risk list for the view
    high_risk_suppliers = [s for s in radar_data["suppliers"] if s["warnings_count"] > 0 or s["is_blacklisted"]]

    return jsonify({
        "status": "success",
        "summary": radar_data["summary"],
        "suppliers": high_risk_suppliers
    })

@invoices_blueprint.get("/api/reports/supplier-risk-radar")
@roles_required("admin", "auditor", "viewer")
def api_supplier_risk_radar():
    """Retrieve all suppliers and summarize the risk radar statistics."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    
    from invoices.supplier_risk_service import get_all_suppliers_risk_radar
    active_mst = session.get("taxpayer_mst")
    
    radar_data = get_all_suppliers_risk_radar(active_mst)
    return jsonify(radar_data)

@invoices_blueprint.post("/api/reports/supplier-risk-radar/blacklist")
@roles_required("admin", "auditor")
def api_add_supplier_blacklist():
    """Add a supplier MST to the blacklist."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    payload = request.json or {}
    mst = payload.get("mst", "").strip()
    reason = payload.get("reason", "").strip()
    if not mst:
        return jsonify({"error": "Mã số thuế không được để trống"}), 400
        
    from invoices.models import BlacklistedMST
    from extensions import db
    import datetime
    
    existing = db.session.get(BlacklistedMST, mst)
    if existing:
        existing.reason = reason
        existing.blacklisted_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    else:
        item = BlacklistedMST(
            mst=mst,
            reason=reason,
            blacklisted_at=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.session.add(item)
    db.session.commit()
    return jsonify({"status": "success", "message": "Đã thêm nhà cung cấp vào danh sách đen."})

@invoices_blueprint.delete("/api/reports/supplier-risk-radar/blacklist/<mst>")
@roles_required("admin", "auditor")
def api_delete_supplier_blacklist(mst):
    """Remove a supplier MST from the blacklist."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import BlacklistedMST
    from extensions import db
    
    item = db.session.get(BlacklistedMST, mst)
    if not item:
        return jsonify({"error": "Không tìm thấy nhà cung cấp trong danh sách đen."}), 404
        
    db.session.delete(item)
    db.session.commit()
    return jsonify({"status": "success", "message": "Đã xóa nhà cung cấp khỏi danh sách đen."})

@invoices_blueprint.get("/api/harness/summary")
def api_harness_summary():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        conn = get_harness_db()
        cur = conn.cursor()

        # Get status counts for stories
        cur.execute("SELECT status, COUNT(*) as cnt FROM story GROUP BY status")
        story_status_rows = cur.fetchall()
        story_status = {r["status"]: r["cnt"] for r in story_status_rows}

        # Get risk lane counts for stories
        cur.execute("SELECT risk_lane, COUNT(*) as cnt FROM story GROUP BY risk_lane")
        story_lane_rows = cur.fetchall()
        story_lane = {r["risk_lane"]: r["cnt"] for r in story_lane_rows}

        # Get decision status counts
        cur.execute("SELECT status, COUNT(*) as cnt FROM decision GROUP BY status")
        decision_status_rows = cur.fetchall()
        decision_status = {r["status"]: r["cnt"] for r in decision_status_rows}

        # Get backlog status counts
        cur.execute("SELECT status, COUNT(*) as cnt FROM backlog GROUP BY status")
        backlog_status_rows = cur.fetchall()
        backlog_status = {r["status"]: r["cnt"] for r in backlog_status_rows}

        # Get trace counts
        cur.execute("SELECT COUNT(*) as cnt FROM trace")
        trace_count = cur.fetchone()["cnt"]

        # Fetch all stories
        cur.execute("SELECT id, title, created_at, risk_lane, contract_doc, status, unit_proof, integration_proof, e2e_proof, platform_proof, evidence, notes FROM story ORDER BY id DESC")
        stories = [dict(r) for r in cur.fetchall()]

        # Fetch all decisions
        cur.execute("SELECT id, title, created_at, status, doc_path, verify_command, last_verified_at, last_verified_result, predicted_impact, actual_outcome, notes FROM decision ORDER BY id DESC")
        decisions = [dict(r) for r in cur.fetchall()]

        # Fetch recent traces (last 30)
        cur.execute("SELECT id, created_at, task_summary, intake_id, story_id, agent, actions_taken, files_read, files_changed, decisions_made, errors, outcome, duration_seconds, token_estimate, harness_friction, notes, git_hash FROM trace ORDER BY id DESC LIMIT 30")
        traces = [dict(r) for r in cur.fetchall()]

        # Fetch all backlog items
        cur.execute("SELECT id, created_at, title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, actual_outcome, implemented_at, notes FROM backlog ORDER BY id DESC")
        backlog = [dict(r) for r in cur.fetchall()]

        # Build stats struct
        stats = {
            "stories": {
                "total": sum(story_status.values()),
                "status": story_status,
                "lanes": story_lane
            },
            "decisions": {
                "total": sum(decision_status.values()),
                "status": decision_status
            },
            "backlog": {
                "total": sum(backlog_status.values()),
                "status": backlog_status
            },
            "traces": {
                "total": trace_count
            }
        }

        return jsonify({
            "stats": stats,
            "stories": stories,
            "decisions": decisions,
            "traces": traces,
            "backlog": backlog
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.post("/api/harness/story")
def api_harness_story_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        story_id = body.get("id", "").strip()
        title = body.get("title", "").strip()
        lane = body.get("lane", "normal").strip()
        contract = (body.get("contract") or body.get("contract_doc") or "").strip()
        status = body.get("status", "planned").strip()
        notes = body.get("notes", "").strip()

        if not story_id or not title:
            return jsonify({"error": "Mã (id) và Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO story (id, title, risk_lane, contract_doc, status, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (story_id, title, lane, contract, status, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Story {story_id} added successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.post("/api/harness/story/update")
def api_harness_story_update():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        story_id = body.get("id", "").strip()
        status = body.get("status", "planned").strip()
        evidence = body.get("evidence", "").strip()
        
        proofs = body.get("proofs") or {}
        unit = body.get("unit") if body.get("unit") is not None else proofs.get("unit")
        integration = body.get("integration") if body.get("integration") is not None else proofs.get("integration")
        e2e = body.get("e2e") if body.get("e2e") is not None else proofs.get("e2e")
        platform = body.get("platform") if body.get("platform") is not None else proofs.get("platform")

        if not story_id:
            return jsonify({"error": "Mã (id) là bắt buộc."}), 400

        # convert potential empty strings or convert type
        try:
            unit = int(unit) if unit is not None and str(unit).strip() != "" else None
        except Exception:
            unit = None
        try:
            integration = int(integration) if integration is not None and str(integration).strip() != "" else None
        except Exception:
            integration = None
        try:
            e2e = int(e2e) if e2e is not None and str(e2e).strip() != "" else None
        except Exception:
            e2e = None
        try:
            platform = int(platform) if platform is not None and str(platform).strip() != "" else None
        except Exception:
            platform = None

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE story
            SET status = ?, evidence = COALESCE(?, evidence),
                unit_proof = COALESCE(?, unit_proof),
                integration_proof = COALESCE(?, integration_proof),
                e2e_proof = COALESCE(?, e2e_proof),
                platform_proof = COALESCE(?, platform_proof)
            WHERE id = ?
            """,
            (status, evidence, unit, integration, e2e, platform, story_id)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Story {story_id} updated successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.post("/api/harness/decision")
def api_harness_decision_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        decision_id = body.get("id", "").strip()
        title = body.get("title", "").strip()
        status = body.get("status", "proposed").strip()
        doc = (body.get("doc") or body.get("doc_path") or "").strip()
        verify = (body.get("verify") or body.get("verify_command") or body.get("verify_cmd") or "").strip()
        predicted = (body.get("predicted") or body.get("predicted_impact") or "").strip()
        notes = body.get("notes", "").strip()

        if not decision_id or not title:
            return jsonify({"error": "Mã (id) và Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO decision (id, title, status, doc_path, verify_command, predicted_impact, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (decision_id, title, status, doc, verify, predicted, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": f"Decision {decision_id} recorded successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.post("/api/harness/backlog")
def api_harness_backlog_add():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    conn = None
    try:
        body = request.get_json(force=True) or {}
        title = body.get("title", "").strip()
        discovered_while = body.get("discovered_while", "").strip()
        current_pain = body.get("current_pain", "").strip()
        suggested_improvement = body.get("suggested_improvement", "").strip()
        risk = body.get("risk", "normal").strip()
        status = body.get("status", "open").strip()
        predicted_impact = body.get("predicted_impact", "").strip()
        notes = body.get("notes", "").strip()

        if not title:
            return jsonify({"error": "Tiêu đề (title) là bắt buộc."}), 400

        conn = get_harness_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO backlog (title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (title, discovered_while, current_pain, suggested_improvement, risk, status, predicted_impact, notes)
        )
        conn.commit()
        return jsonify({"success": True, "message": "Backlog item added successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.get("/api/harness/agent/stream")
def api_agent_stream():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    provider = request.args.get("provider", "gemini").strip()
    model = request.args.get("model", "gemini-2.5-flash").strip()
    goal = request.args.get("goal", "").strip()
    story_id = request.args.get("story_id", "").strip()

    if not goal:
        return jsonify({"error": "Goal is required"}), 400

    from flask import Response

    def generate():
        import subprocess
        import os
        import json

        env = os.environ.copy()
        env["AGENT_PROVIDER"] = provider
        env["AGENT_MODEL"] = model
        env["AGENT_GOAL"] = goal
        if story_id:
            env["AGENT_STORY_ID"] = story_id

        cmd = ["node", "scripts/agent-harness/run-agent.js"]

        proc = subprocess.Popen(
            cmd,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1
        )

        while True:
            line = proc.stdout.readline()
            if not line:
                break
            line_str = line.strip()
            if line_str:
                yield f"data: {line_str}\n\n"

        err = proc.stderr.read()
        if err:
            try:
                # Try parsing as JSON error from run-agent.js
                err_data = json.loads(err.strip())
                yield f"data: {json.dumps(err_data)}\n\n"
            except Exception:
                yield f"data: {json.dumps({'type': 'error', 'message': err.strip()})}\n\n"

        proc.wait()

        if proc.returncode != 0:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Process exited with code {proc.returncode}'})}\n\n"
        else:
            if story_id:
                try:
                    git_hash = "unknown"
                    try:
                        git_hash = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True).strip()
                        status_out = subprocess.check_output(["git", "status", "--porcelain"], text=True).strip()
                        if status_out:
                            git_hash += " (dirty)"
                    except Exception:
                        pass

                    conn = None
                    try:
                        conn = get_harness_db()
                        cur = conn.cursor()
                        cur.execute(
                            """
                            INSERT INTO trace (task_summary, story_id, agent, outcome, git_hash, created_at, actions_taken, notes)
                            VALUES (?, ?, ?, 'completed', ?, datetime('now'), ?, ?)
                            """,
                            (f"Autonomous Run: {goal[:50]}...", story_id, "SkawldAgent", git_hash, '["run-agent.js"]', f"Goal: {goal}")
                        )
                        conn.commit()
                    finally:
                        if conn:
                            conn.close()
                except Exception as db_err:
                    print(f"Error logging trace to DB: {db_err}")

    return Response(generate(), mimetype="text/event-stream")

@invoices_blueprint.post("/api/harness/risk/evaluate")
def api_harness_risk_evaluate():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    body = request.get_json(force=True) or {}
    text = body.get("text", "").strip()
    if not text:
        return jsonify({"error": "No spec text provided"}), 400

    text_lower = text.lower()
    
    checklist = {
        "auth": ["auth", "login", "logout", "session", "password", "token"],
        "authorization": ["role", "permission", "tenant", "access control"],
        "data_model": ["schema", "migration", "sqlite", "table", "column", "drop table"],
        "security": ["audit", "security", "privacy", "access log", "secret", "oauth"],
        "external": ["email", "payment", "sdk", "webhook", "queue", "api", "request", "http", "vietqr", "gdt"],
        "contract": ["api shape", "response envelope", "client-visible", "contract"],
        "cross_platform": ["desktop", "mobile", "browser", "native", "deep link"],
        "existing_behavior": ["refactor", "change", "fix", "patch"],
        "weak_proof": ["untested", "missing tests", "no test"],
        "multi_domain": ["multi-domain", "multiple domain"]
    }
    
    flags_found = []
    for flag, kw_list in checklist.items():
        if any(kw in text_lower for kw in kw_list):
            flags_found.append(flag)
            
    hard_gates = ["auth", "authorization", "data_model", "security", "external"]
    has_hard_gate = any(fg in hard_gates for fg in flags_found)
    
    num_flags = len(flags_found)
    if has_hard_gate or num_flags >= 4:
        lane = "high_risk"
    elif num_flags >= 2:
        lane = "normal"
    else:
        lane = "tiny"
        
    return jsonify({
        "suggested_lane": lane,
        "flags_found": flags_found,
        "has_hard_gate": has_hard_gate,
        "flag_count": num_flags
    })

@invoices_blueprint.get("/api/harness/db/stats")
def api_harness_db_stats():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    
    db_file = "harness.db"
    size_mb = 0.0
    last_modified = "unknown"
    if os.path.exists(db_file):
        size_bytes = os.path.getsize(db_file)
        size_mb = round(size_bytes / (1024 * 1024), 2)
        mtime = os.path.getmtime(db_file)
        last_modified = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")

    conn = None
    try:
        conn = get_harness_db()
        cur = conn.cursor()
        
        tables = {}
        for tbl in ["story", "decision", "backlog", "trace"]:
            cur.execute(f"SELECT COUNT(*) as count FROM {tbl}")
            tables[tbl] = cur.fetchone()["count"]
            
        return jsonify({
            "file_name": db_file,
            "size_mb": size_mb,
            "last_modified": last_modified,
            "table_counts": tables
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@invoices_blueprint.post("/api/harness/db/backup")
def api_harness_db_backup():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    
    import shutil
    db_file = "harness.db"
    if not os.path.exists(db_file):
        return jsonify({"error": "Database file not found"}), 404
        
    try:
        backup_dir = "data/backup"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(backup_dir, f"harness_backup_{timestamp}.db")
        shutil.copy2(db_file, backup_file)
        return jsonify({
            "success": True, 
            "message": f"Successfully backed up database to {backup_file}",
            "backup_file": backup_file
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/harness/db/download")
def api_harness_db_download():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    
    db_file = "harness.db"
    if not os.path.exists(db_file):
        return jsonify({"error": "Database file not found"}), 404
        
    return send_file(db_file, as_attachment=True, download_name="harness.db")

@invoices_blueprint.get("/api/harness/validate/stream")
def api_harness_validate_stream():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from flask import Response
    
    def generate_validation():
        import subprocess
        import os
        import json
        
        validate_script = os.path.join("scripts", "validate.bat")
        
        if not os.path.exists(validate_script):
            yield f"data: {json.dumps({'type': 'error', 'message': 'Validation script scripts/validate.bat not found.'})}\n\n"
            return
            
        proc = subprocess.Popen(
            [validate_script],
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )
        
        yield f"data: {json.dumps({'type': 'status', 'message': 'Running system validation checks...'})}\n\n"
        
        while True:
            line = proc.stdout.readline()
            if not line and proc.poll() is not None:
                break
            if line:
                yield f"data: {json.dumps({'type': 'output', 'text': line})}\n\n"
                
        proc.wait()
        
        if proc.returncode == 0:
            yield f"data: {json.dumps({'type': 'status', 'message': 'Validation PASSED. All tests and checks passed successfully.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'success': True})}\n\n"
        else:
            yield f"data: {json.dumps({'type': 'status', 'message': f'Validation FAILED with exit code {proc.returncode}.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'success': False})}\n\n"
            
    return Response(generate_validation(), mimetype="text/event-stream")

@invoices_blueprint.get("/api/harness/plugins")
def api_harness_plugins():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from pathlib import Path
    skills_dir = Path("C:/Users/THUAN/.gemini/antigravity/skills")
    plugins = []
    if skills_dir.exists() and skills_dir.is_dir():
        for item in skills_dir.iterdir():
            if item.is_dir():
                skill_md = item / "SKILL.md"
                if skill_md.exists():
                    try:
                        content = skill_md.read_text(encoding="utf-8", errors="ignore")
                        # ponytail: simplest possible parser using simple string split or regex
                        meta = {
                            "id": item.name,
                            "name": item.name,
                            "description": "No description provided.",
                            "version": "1.0.0",
                            "license": "N/A"
                        }
                        parts = content.split("---")
                        if len(parts) >= 3:
                            frontmatter = parts[1]
                            for line in frontmatter.splitlines():
                                if ":" in line:
                                    k, v = line.split(":", 1)
                                    k = k.strip().lower()
                                    v = v.strip()
                                    # Strip quotes if any
                                    if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
                                        v = v[1:-1]
                                    if k in meta:
                                        meta[k] = v
                        plugins.append(meta)
                    except Exception as e:
                        plugins.append({
                            "id": item.name,
                            "name": item.name,
                            "description": f"Failed to parse skill: {e}",
                            "version": "N/A",
                            "license": "N/A"
                        })
    return jsonify({"plugins": plugins})

@invoices_blueprint.get("/api/harness/plugins/install")
def api_harness_plugins_install():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    repo_url = request.args.get("repo_url", "").strip()
    if not repo_url:
        return jsonify({"error": "GitHub Repository URL is required"}), 400

    from flask import Response

    def generate():
        import subprocess
        import os
        import json

        # ponytail: Invoke scripts/agy.py directly using local venv python execution
        workspace_dir = "d:/LearnAnyThing/Webapp XML"
        python_exe = os.path.join(workspace_dir, "venv", "Scripts", "python.exe")
        agy_script = os.path.join(workspace_dir, "scripts", "agy.py")
        
        cmd = [python_exe, agy_script, "plugin", "install", repo_url]

        proc = subprocess.Popen(
            cmd,
            cwd=workspace_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )

        yield f"data: {json.dumps({'type': 'status', 'message': f'Starting plugin installation from {repo_url}...'})}\n\n"

        while True:
            line = proc.stdout.readline()
            if not line and proc.poll() is not None:
                break
            if line:
                yield f"data: {json.dumps({'type': 'output', 'text': line})}\n\n"

        proc.wait()

        if proc.returncode == 0:
            yield f"data: {json.dumps({'type': 'status', 'message': 'Plugin installation completed successfully.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'success': True})}\n\n"
        else:
            yield f"data: {json.dumps({'type': 'status', 'message': f'Installation failed with exit code {proc.returncode}.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'success': False})}\n\n"

    return Response(generate(), mimetype="text/event-stream")

@invoices_blueprint.get("/api/harness/plugins/ponytail/debt")
def api_harness_plugins_ponytail_debt():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    import re
    from pathlib import Path
    
    workspace_dir = Path("d:/LearnAnyThing/Webapp XML")
    search_dirs = ["invoices", "templates", "static", "tests", "scripts"]
    search_files = ["app.py", "config.py"]
    
    debt_items = []
    # ponytail: simple loop using Path.rglob and regex
    pattern = re.compile(r"ponytail:\s*(.*)", re.IGNORECASE)
    
    def scan_file(file_path: Path):
        try:
            content = file_path.read_text(encoding="utf-8", errors="ignore")
            for idx, line in enumerate(content.splitlines(), start=1):
                match = pattern.search(line)
                if match:
                    # Clean up the line to remove comments syntax like #, //, <!--, -->
                    desc = match.group(1).strip()
                    if desc.endswith("-->"):
                        desc = desc[:-3].strip()
                    debt_items.append({
                        "file": file_path.relative_to(workspace_dir).as_posix(),
                        "line": idx,
                        "description": desc
                    })
        except Exception:
            pass

    for d in search_dirs:
        dir_path = workspace_dir / d
        if dir_path.exists() and dir_path.is_dir():
            for ext in ["*.py", "*.html", "*.js", "*.css"]:
                for file_path in dir_path.rglob(ext):
                    scan_file(file_path)
                    
    for f in search_files:
        file_path = workspace_dir / f
        if file_path.exists() and file_path.is_file():
            scan_file(file_path)

    return jsonify({"debt": debt_items})

@invoices_blueprint.get("/api/harness/plugins/ponytail/audit")
def api_harness_plugins_ponytail_audit():
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from pathlib import Path
    workspace_dir = Path("d:/LearnAnyThing/Webapp XML")
    
    audit_results = {
        "score": 100,
        "findings": [],
        "total_files_scanned": 0,
        "total_lines_scanned": 0
    }
    
    search_dirs = ["invoices", "templates", "static", "tests", "scripts"]
    search_files = ["app.py", "config.py"]
    
    for d in search_dirs:
        dir_path = workspace_dir / d
        if dir_path.exists() and dir_path.is_dir():
            for ext in ["*.py", "*.html", "*.js", "*.css"]:
                for file_path in dir_path.rglob(ext):
                    if "venv" in file_path.parts or ".pytest_cache" in file_path.parts or "__pycache__" in file_path.parts:
                        continue
                    try:
                        audit_results["total_files_scanned"] += 1
                        content = file_path.read_text(encoding="utf-8", errors="ignore")
                        lines = content.splitlines()
                        num_lines = len(lines)
                        audit_results["total_lines_scanned"] += num_lines
                        
                        rel_path = file_path.relative_to(workspace_dir).as_posix()
                        
                        # Rule 1: File size audit (Bloat)
                        if num_lines > 500 and file_path.suffix == ".py":
                            penalty = min(15, (num_lines - 500) // 100 + 5)
                            audit_results["score"] -= penalty
                            audit_results["findings"].append({
                                "file": rel_path,
                                "type": "Bloat",
                                "severity": "Medium" if penalty < 10 else "High",
                                "message": f"File is too long ({num_lines} lines). Ponytail suggests dividing into focused utilities or reducing nested logic."
                            })
                        elif num_lines > 1000 and file_path.suffix == ".html":
                            penalty = min(10, (num_lines - 1000) // 200 + 3)
                            audit_results["score"] -= penalty
                            audit_results["findings"].append({
                                "file": rel_path,
                                "type": "Bloat",
                                "severity": "Medium",
                                "message": f"Template has too many lines ({num_lines}). Propose breaking up into smaller sub-templates using Flask include."
                            })
                            
                        # Rule 2: Deep nesting audit
                        deep_lines = []
                        for idx, line in enumerate(lines, start=1):
                            indent = len(line) - len(line.lstrip())
                            if (line.startswith(" ") and indent >= 16) or (line.startswith("\t") and indent >= 4):
                                if line.strip() and not line.strip().startswith("#") and not line.strip().startswith("//"):
                                    deep_lines.append(idx)
                                    
                        if deep_lines:
                            penalty = min(8, len(deep_lines) // 2 + 1)
                            audit_results["score"] -= penalty
                            audit_results["findings"].append({
                                "file": rel_path,
                                "type": "Nesting Complexity",
                                "severity": "Medium",
                                "message": f"Deep indentation found on line(s): {', '.join(map(str, deep_lines[:5]))}. Ponytail suggests extracting inner blocks to helper functions."
                            })
                            
                        # Rule 3: Complex external library usages (Stdlib candidates)
                        for idx, line in enumerate(lines, start=1):
                            if "import requests" in line:
                                audit_results["findings"].append({
                                    "file": rel_path,
                                    "type": "Dependency Bloat",
                                    "severity": "Low",
                                    "message": f"Line {idx}: Imports 'requests'. Ponytail hints: python standard library 'urllib.request' can sometimes do this in one line."
                                })
                            if "import os" in line and "import pathlib" in line:
                                audit_results["findings"].append({
                                    "file": rel_path,
                                    "type": "Redundant Libraries",
                                    "severity": "Low",
                                    "message": f"Line {idx}: Both 'os' and 'pathlib' are imported. Standardize on 'pathlib' for cleaner path handling."
                                })
                    except Exception:
                        pass
                        
    audit_results["score"] = max(10, audit_results["score"])
    return jsonify(audit_results)

@invoices_blueprint.post("/api/bctc/compile")
@roles_required("admin", "auditor")
def api_bctc_compile():
    """Compile BCTC B01-DN, B02-DN, B03-DN from Trial Balance ledger data (US-200)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    balances = {}
    metadata = request.json or {}
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
                # Populate metadata from request form fields if present
                metadata = {
                    "mst": request.form.get("mst", "0109998887"),
                    "company_name": request.form.get("company_name", "CONG TY TNHH MOCK"),
                    "year": int(request.form.get("year", datetime.now().year)),
                    "reporting_period_type": request.form.get("reporting_period_type", "N"),
                    "dividends_paid": float(request.form.get("dividends_paid", 0.0))
                }
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        balances = metadata.get("balances", {})
        
    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400
        
    try:
        from invoices.bctc_service import compile_bctc
        xml_str, warnings = compile_bctc(balances, metadata)
        return jsonify({
            "status": "success" if not warnings else "warning",
            "xml": xml_str,
            "warnings": warnings
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/bctc/audit-ledger")
@roles_required("admin", "auditor")
def api_bctc_audit_ledger():
    """Cross-reference General Ledger entries with e-invoices for compliance (US-201)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    balances = {}
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
                taxpayer_mst = request.form.get("taxpayer_mst") or taxpayer_mst
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        balances = payload.get("balances", {})
        taxpayer_mst = payload.get("taxpayer_mst") or taxpayer_mst
        
    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400
        
    try:
        from invoices.bctc_service import audit_ledger_against_invoices
        report = audit_ledger_against_invoices(balances, taxpayer_mst)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/payments/tax-slip")
@roles_required("admin", "auditor")
def api_payments_tax_slip():
    """Generate GDT Form 711/MB Tax Payment Slip XML and VietQR code (US-202)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    payload = request.json or {}
    mst = payload.get("mst") or session.get("taxpayer_mst") or "0109998887"
    company_name = payload.get("company_name", "CONG TY TNHH MOCK")
    tax_type = payload.get("tax_type")
    amount = payload.get("amount")
    
    if not tax_type or not amount:
        return jsonify({"error": "Thieu thong tin loai thue hoac so tien."}), 400
        
    try:
        amount_val = float(amount)
    except ValueError:
        return jsonify({"error": "So tien khong hop le."}), 400
        
    chapter_type = payload.get("chapter_type", "domestic_private")
    treasury_name = payload.get("treasury_name", "Kho bac Nha nuoc Quan Cau Giay")
    treasury_account = payload.get("treasury_account", "111222333444")
    bank_bin = payload.get("bank_bin", "970415")
    
    try:
        from invoices.tax_payment_service import generate_tax_payment_slip
        slip = generate_tax_payment_slip(
            mst=mst,
            company_name=company_name,
            tax_type=tax_type,
            amount=amount_val,
            chapter_type=chapter_type,
            treasury_name=treasury_name,
            treasury_account=treasury_account,
            bank_bin=bank_bin
        )
        return jsonify(slip)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/payments/bank-recon")
@roles_required("admin", "auditor")
def api_payments_bank_recon():
    """Standard bank statement parsing, fuzzy matching, and cash payment compliance auditing (US-203)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    # Process uploaded bank statement if present
    results = {}
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                content = file.read().decode("utf-8")
                from invoices.reconciliation_service import ReconciliationEngine
                engine = ReconciliationEngine()
                engine.process_csv(content)
                results = engine.run_matching()
            except Exception as e:
                return jsonify({"error": f"Loi xu ly file sao ke: {str(e)}"}), 400
                
    # Run cash compliance checks for invoices >= 20M VND
    try:
        from invoices.bank_reconcile_service import check_cash_payment_compliance
        compliance_flags = check_cash_payment_compliance(taxpayer_mst)
        return jsonify({
            "status": "success",
            "reconciliation_summary": results,
            "compliance_warnings": compliance_flags
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ecommerce/sync")
@roles_required("admin", "auditor")
def api_ecommerce_sync():
    """Parse platform reports and record daily consolidated revenue & fees in the database (US-204)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    platform = request.args.get("platform", "shopee").strip()
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    orders = []
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.ecommerce_service import parse_ecommerce_sheet
                orders = parse_ecommerce_sheet(file.read(), platform)
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        orders = payload.get("orders", [])
        taxpayer_mst = payload.get("taxpayer_mst") or taxpayer_mst
        platform = payload.get("platform") or platform
        
    if not orders:
        return jsonify({"error": "Thieu du lieu don hang e-commerce."}), 400
        
    try:
        from invoices.ecommerce_service import sync_ecommerce_orders
        res = sync_ecommerce_orders(orders, taxpayer_mst, platform)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/ecommerce/reconcile")
@roles_required("admin", "auditor", "viewer")
def api_ecommerce_reconcile():
    """Reconcile Shopee/TikTok Shop order logs with output invoices (US-205)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    taxpayer_mst = request.args.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109998887"
    
    import json
    orders_json = request.args.get("orders")
    platform_orders = []
    if orders_json:
        try:
            platform_orders = json.loads(orders_json)
        except Exception:
            pass
            
    if not platform_orders:
        platform_orders = session.get("normalized_orders", [])
        
    if not platform_orders:
        platform_orders = [
            {"order_id": "ORD-SHOPEE-1001", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 500000.0, "commission_fee": 15000.0, "service_fee": 5000.0},
            {"order_id": "ORD-SHOPEE-1002", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 1200000.0, "commission_fee": 36000.0, "service_fee": 12000.0},
            {"order_id": "ORD-SHOPEE-1003", "date": datetime.now().strftime("%Y-%m-%d"), "gross_revenue": 850000.0, "commission_fee": 25500.0, "service_fee": 8500.0}
        ]
        
    try:
        from invoices.ecommerce_service import reconcile_ecommerce_tax
        report = reconcile_ecommerce_tax(taxpayer_mst, platform_orders)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/audit-trail")
@roles_required("admin", "auditor")
def audit_trail_page():
    """Render the Audit Trail Viewer UI."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return redirect(url_for("index"))
    return render_template("audit_trail.html",
                           logged_in=session.get("logged_in"),
                           session_username=session.get("display_name") or session.get("username"))

@invoices_blueprint.get("/advanced-audit")
@roles_required("admin", "auditor")
def advanced_audit_page():
    """Render the Advanced Audit & Fraud Detection Page."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return redirect(url_for("index"))
    return render_template("advanced_audit.html",
                           logged_in=session.get("logged_in"),
                           session_username=session.get("display_name") or session.get("username"))

@invoices_blueprint.get("/api/audit-logs")
@roles_required("admin", "auditor")
def api_get_audit_logs():
    """Retrieve security audit logs with optional filtering."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog

    try:
        query = SecurityAuditLog.query

        # Filters
        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)

        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))

        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))

        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)

        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")

        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        # Pagination
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)
        per_page = min(per_page, 200)

        total = query.count()
        logs = query.order_by(SecurityAuditLog.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            "total": total,
            "page": page,
            "per_page": per_page,
            "logs": [
                {
                    "id": log.id,
                    "timestamp": log.timestamp,
                    "username": log.username,
                    "tax_code": log.tax_code,
                    "event_category": log.event_category,
                    "ip_address": log.ip_address,
                    "event_details": log.event_details,
                }
                for log in logs
            ],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/audit-logs/export/csv")
@roles_required("admin", "auditor")
def api_export_audit_logs_csv():
    """Export filtered audit logs as CSV file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog
    import csv
    import io

    try:
        query = SecurityAuditLog.query

        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)
        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))
        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))
        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)
        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")
        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        logs = query.order_by(SecurityAuditLog.id.desc()).limit(10000).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Timestamp", "Username", "Tax Code", "Category", "IP Address", "Details"])
        for log in logs:
            writer.writerow([
                log.id, log.timestamp, log.username, log.tax_code or "",
                log.event_category, log.ip_address or "", log.event_details or "",
            ])

        from datetime import datetime
        filename = f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/audit-logs/export/pdf")
@roles_required("admin", "auditor")
def api_export_audit_logs_pdf():
    """Export filtered audit logs as PDF file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import SecurityAuditLog
    from datetime import datetime

    try:
        query = SecurityAuditLog.query

        category = request.args.get("category")
        if category:
            query = query.filter(SecurityAuditLog.event_category == category)
        username = request.args.get("username")
        if username:
            query = query.filter(SecurityAuditLog.username.ilike(f"%{username}%"))
        tax_code = request.args.get("tax_code")
        if tax_code:
            query = query.filter(SecurityAuditLog.tax_code.ilike(f"%{tax_code}%"))
        date_from = request.args.get("date_from")
        if date_from:
            query = query.filter(SecurityAuditLog.timestamp >= date_from)
        date_to = request.args.get("date_to")
        if date_to:
            query = query.filter(SecurityAuditLog.timestamp <= date_to + "T23:59:59Z")
        keyword = request.args.get("keyword")
        if keyword:
            query = query.filter(SecurityAuditLog.event_details.ilike(f"%{keyword}%"))

        logs = query.order_by(SecurityAuditLog.id.desc()).limit(5000).all()

        # Generate HTML-based PDF using render_template
        html = render_template("audit_trail_pdf.html",
                               logs=logs,
                               generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                               total_records=len(logs),
                               filters={
                                   "category": category,
                                   "username": username,
                                   "tax_code": tax_code,
                                   "date_from": date_from,
                                   "date_to": date_to,
                                   "keyword": keyword,
                               })

        filename = f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        return Response(
            html,
            mimetype="text/html",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/reports/signed-compliance")
@roles_required("admin", "auditor")
def api_export_signed_compliance():
    """Export audited compliance report with embedded cryptographic signature."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        parsed_from, parsed_to = validate_date_range(
            request.args.get("from", ""),
            request.args.get("to", ""),
        )
        direction = request.args.get("direction", "purchase")
        
        # Get invoices from service
        current_app.config["CURRENT_JWT"] = session.get("jwt")
        invoices = fetch_invoices(InvoiceQuery(parsed_from, parsed_to, False, direction))
        
        # Get system secret key for hashing
        secret_key = current_app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
        
        from invoices.compliance_report_service import generate_signed_excel_report
        excel_bytes = generate_signed_excel_report(invoices, secret_key)
        
        # Log this administrative export event in the security audit ledger
        from invoices.security_audit_service import log_security_event
        log_security_event(
            username=session.get("username", "admin"),
            event_category="EXPORT",
            tax_code=session.get("tax_code", ""),
            ip_address=request.remote_addr,
            event_details=f"Exported cryptographically signed compliance report for period {parsed_from} to {parsed_to} ({len(invoices)} invoices)."
        )

        filename = f"signed_compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except DateValidationError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        current_app.config["CURRENT_JWT"] = None

@invoices_blueprint.post("/api/reports/verify-signed")
@roles_required("admin", "auditor")
def api_verify_signed_report():
    """Upload and verify a signed compliance report file."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy tệp tin báo cáo được tải lên."}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Tên tệp tin không hợp lệ."}), 400

    try:
        file_bytes = file.read()
        secret_key = current_app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
        
        from invoices.compliance_report_service import verify_excel_report
        result = verify_excel_report(file_bytes, secret_key)
        
        # Log security audit verification event
        status_str = "SUCCESS" if result.get("verified") else "FAILED"
        from invoices.security_audit_service import log_security_event
        log_security_event(
            username=session.get("username", "admin"),
            event_category="VERIFY",
            tax_code=session.get("tax_code", ""),
            ip_address=request.remote_addr,
            event_details=f"Performed cryptographic verification of compliance report file '{file.filename}'. Result: {status_str} ({result.get('invoices_count', 0)} invoices parsed)."
        )

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"Lỗi xử lý xác minh báo cáo: {str(e)}"}), 500

@invoices_blueprint.get("/api/sync/health")
@roles_required("admin", "auditor")
def api_sync_health():
    """Retrieve CAPTCHA solver statistics and overall crawler status."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        from auth.captcha_solver import captcha_analytics

        # CAPTCHA metrics
        stats = captcha_analytics.get_stats()

        # Crawler status
        crawler_status = "idle"
        queue_instance = current_app.extensions.get("resilient_sync_queue")
        if queue_instance:
            with queue_instance._lock:
                if any(job.status == "running" for job in queue_instance.jobs.values()):
                    crawler_status = "running"

        return jsonify({
            "status": "healthy",
            "crawler_status": crawler_status,
            "solver": stats,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/consolidated-dashboard")
@roles_required("admin", "auditor")
def consolidated_dashboard_page():
    """Render the corporate multi-entity consolidated dashboard."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("consolidated.html")

@invoices_blueprint.route("/api/tenant/groups", methods=["GET", "POST"])
@roles_required("admin", "auditor")
def api_tenant_groups():
    """GET/POST API to fetch or create corporate tenant groups."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import TenantGroup
    import json

    username = session.get("username", "admin")

    if request.method == "POST":
        try:
            data = request.get_json() or {}
            group_name = data.get("group_name")
            taxpayer_msts = data.get("taxpayer_msts", [])

            if not group_name:
                return jsonify({"error": "Tên tập đoàn không được để trống."}), 400
            if not isinstance(taxpayer_msts, list):
                return jsonify({"error": "Danh sách MST phải là một mảng."}), 400

            # Validate MSTs
            taxpayer_msts = [str(mst).strip() for mst in taxpayer_msts if mst]

            # Upsert group
            group = TenantGroup.query.filter_by(group_name=group_name).first()
            if group:
                group.taxpayer_msts = json.dumps(taxpayer_msts)
                group.admin_username = username
            else:
                group = TenantGroup(
                    group_name=group_name,
                    admin_username=username,
                    taxpayer_msts=json.dumps(taxpayer_msts)
                )
                db.session.add(group)

            db.session.commit()
            return jsonify({"status": "success", "group": group.to_dict()})
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    # GET
    try:
        groups = TenantGroup.query.filter_by(admin_username=username).all()
        # Fallback to all groups if admin or no group found
        if not groups and username == "admin":
            groups = TenantGroup.query.all()
        return jsonify([g.to_dict() for g in groups])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/tenant/consolidated")
@roles_required("admin", "auditor")
def api_tenant_consolidated():
    """Retrieve consolidated financial metrics and risk scores across a group's MSTs."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import TenantGroup
    from invoices.multitenant_service import get_tenant_consolidated_stats
    import json

    username = session.get("username", "admin")
    group_id = request.args.get("group_id")

    try:
        # 1. Resolve Group
        group = None
        if group_id:
            group = TenantGroup.query.get(group_id)
        else:
            group = TenantGroup.query.filter_by(admin_username=username).first()
            if not group and username == "admin":
                group = TenantGroup.query.first()

        if not group:
            return jsonify({
                "group_id": None,
                "group_name": "Không có nhóm",
                "summary": {
                    "total_invoices": 0,
                    "total_revenue": 0.0,
                    "vat_output": 0.0,
                    "vat_input": 0.0,
                    "average_t_score": 100.0
                },
                "entities": []
            })

        # 2. Query each member MST
        mst_list = group.get_mst_list()
        entities = []
        for mst in mst_list:
            stats = get_tenant_consolidated_stats(mst)
            entities.append(stats)

        # 3. Aggregate totals
        total_invoices = sum(e["total_invoices"] for e in entities)
        total_revenue = sum(e["total_revenue"] for e in entities)
        vat_output = sum(e["vat_output"] for e in entities)
        vat_input = sum(e["vat_input"] for e in entities)
        
        # Weighted average for T-Score
        t_score_sum = 0.0
        t_score_count = 0
        for e in entities:
            if e["total_invoices"] > 0:
                t_score_sum += e["average_t_score"] * e["total_invoices"]
                t_score_count += e["total_invoices"]
            else:
                t_score_sum += e["average_t_score"]
                t_score_count += 1
                
        average_t_score = round(t_score_sum / t_score_count, 1) if t_score_count > 0 else 100.0

        return jsonify({
            "group_id": group.id,
            "group_name": group.group_name,
            "summary": {
                "total_invoices": total_invoices,
                "total_revenue": total_revenue,
                "vat_output": vat_output,
                "vat_input": vat_input,
                "average_t_score": average_t_score
            },
            "entities": entities
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/finalize")
@roles_required("admin", "auditor")
def api_cit_finalize():
    """US-180: Compile CIT Finalization and generate Form 03/TNDN XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from datetime import datetime
    balances = {}
    metadata = {
        "mst": request.form.get("mst") or session.get("taxpayer_mst") or "0109998887",
        "company_name": request.form.get("company_name", "CONG TY TNHH MOCK"),
        "year": int(request.form.get("year", datetime.now().year)),
        "non_deductible_manual": float(request.form.get("non_deductible_manual", 0.0)),
        "loss_carry_forward": float(request.form.get("loss_carry_forward", 0.0)),
        "rd_allowance": float(request.form.get("rd_allowance", 0.0))
    }

    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            try:
                from invoices.bctc_service import parse_ledger_file
                balances = parse_ledger_file(file.read(), file.filename)
            except Exception as e:
                return jsonify({"error": f"Loi doc file: {str(e)}"}), 400
    else:
        payload = request.json or {}
        balances = payload.get("balances", {})
        metadata.update(payload.get("metadata", {}))

    if not balances:
        return jsonify({"error": "Thieu du lieu bang can doi phat sinh / so cai."}), 400

    try:
        from invoices.cit_service import finalize_cit
        result = finalize_cit(balances, metadata)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/simulate-scenario")
@roles_required("admin", "auditor")
def api_cit_simulate_scenario():
    """US-181: Simulate what-if tax scenarios based on slider adjustments."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        payload = request.get_json() or {}
        base_data = payload.get("base_data", {})
        adjustments = payload.get("adjustments", {})
        
        from invoices.cit_service import simulate_cit_scenario
        result = simulate_cit_scenario(base_data, adjustments)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.route("/api/finance/cashflow")
def api_finance_cashflow():
    """US-150: Return rolling 30/60/90-day cash-flow projections."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("mst")
        from invoices.cashflow_service import calculate_cashflow_projection
        result = calculate_cashflow_projection(taxpayer_mst=taxpayer_mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.route("/api/finance/simulate", methods=["POST"])
def api_finance_simulate():
    """US-151: Stateless scenario simulation with adjustable parameters."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    try:
        payload = request.get_json() or {}
        delay_days = int(payload.get("delay_days", 0))
        rejection_rate = float(payload.get("rejection_rate", 0.0))
        vat_adjustment = float(payload.get("vat_adjustment", 0.0))
        taxpayer_mst = session.get("active_taxpayer_mst") or payload.get("mst")

        from invoices.cashflow_service import simulate_scenario
        result = simulate_scenario(
            taxpayer_mst=taxpayer_mst,
            delay_days=delay_days,
            rejection_rate=rejection_rate,
            vat_adjustment=vat_adjustment,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/partners/<mst>/decree-132")
@roles_required("admin", "auditor")
def update_partner_decree_132(mst):
    """Update Decree 132 relationship code for a specific partner."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    from invoices.models import Partner
    from extensions import db
    try:
        partner = db.session.get(Partner, mst)
        if not partner:
            return jsonify({"error": f"Không tìm thấy đối tác với MST {mst}"}), 404
        
        body = request.get_json(silent=True) or {}
        relationship = body.get("decree_132_relationship")
        
        if relationship is not None:
            relationship = str(relationship).strip()
            if relationship == "":
                relationship = None
            else:
                valid_codes = {chr(c) for c in range(ord('A'), ord('L') + 1)}
                if relationship.upper() not in valid_codes:
                    return jsonify({"error": "Mã liên kết không hợp lệ. Phải thuộc từ A đến L."}), 400
                relationship = relationship.upper()
                
        partner.decree_132_relationship = relationship
        db.session.commit()
        
        return jsonify({
            "success": True,
            "partner": partner.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/send")
def api_agents_send():
    """US-320: Post a message from one AI agent to another."""
    import json
    from datetime import datetime, timezone
    from invoices.models import AgentMessage
    from extensions import db

    body = request.get_json(silent=True) or {}
    sender = body.get("sender_agent")
    receiver = body.get("receiver_agent")
    subject = body.get("subject")
    payload = body.get("payload", {})

    if not sender or not receiver or not subject:
        return jsonify({"error": "sender_agent, receiver_agent, and subject are required."}), 400

    try:
        if isinstance(payload, (dict, list)):
            payload_str = json.dumps(payload)
        else:
            payload_str = str(payload)

        msg = AgentMessage(
            sender_agent=str(sender).strip(),
            receiver_agent=str(receiver).strip(),
            subject=str(subject).strip(),
            payload=payload_str,
            status="pending",
            timestamp=datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        )
        db.session.add(msg)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Message sent successfully.",
            "data": msg.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/agents/inbox/<agent_name>")
def api_agents_inbox(agent_name):
    """US-320: Get pending and processed messages for a specific agent."""
    from invoices.models import AgentMessage
    try:
        status_filter = request.args.get("status", "pending")
        query = AgentMessage.query.filter_by(receiver_agent=agent_name)
        if status_filter:
            query = query.filter_by(status=status_filter)
        messages = query.order_by(AgentMessage.id.desc()).all()
        return jsonify([msg.to_dict() for msg in messages])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/update-status/<int:message_id>")
def api_agents_update_status(message_id):
    """US-320: Update the processing status of an agent message."""
    from invoices.models import AgentMessage
    from extensions import db
    try:
        msg = db.session.get(AgentMessage, message_id)
        if not msg:
            return jsonify({"error": f"Message with ID {message_id} not found."}), 404

        body = request.get_json(silent=True) or {}
        new_status = body.get("status")
        if new_status not in ["pending", "processed", "failed"]:
            return jsonify({"error": "Invalid status. Must be pending, processed, or failed."}), 400

        msg.status = new_status
        db.session.commit()
        return jsonify({
            "success": True,
            "message": "Status updated successfully.",
            "data": msg.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/audit-coordinator")
def api_agents_audit_coordinator():
    """US-321: Run the multi-agent joint audit coordinator swarm."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    user_prompt = body.get("user_prompt")

    if not taxpayer_mst or not user_prompt:
        return jsonify({"error": "taxpayer_mst and user_prompt are required."}), 400

    from invoices.agent_swarm import JointAuditCoordinator
    try:
        coordinator = JointAuditCoordinator()
        result = coordinator.execute_swarm(taxpayer_mst=taxpayer_mst, user_prompt=user_prompt)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/bank/ingest")
def api_bank_ingest():
    """US-322: Ingest bank statement feed files."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    file_content = body.get("file_content")
    bank_name = body.get("bank_name", "Vietcombank")
    file_type = body.get("file_type", "csv")

    if not taxpayer_mst or not file_content:
        return jsonify({"error": "taxpayer_mst and file_content are required."}), 400

    from invoices.bank_stream_service import BankStreamService
    try:
        service = BankStreamService()
        count = service.ingest_bank_statement(file_content, taxpayer_mst, bank_name, file_type)
        return jsonify({"success": True, "inserted_count": count}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/bank/match")
def api_bank_match():
    """US-323: Execute automated matching of transactions with invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")

    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.bank_stream_service import BankStreamService
    try:
        service = BankStreamService()
        result = service.execute_transaction_matching(taxpayer_mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/bank/transactions")
def api_bank_transactions():
    """List bank transactions."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    match_status = request.args.get("match_status")
    from invoices.models import BankTransaction
    query = BankTransaction.query.filter_by(taxpayer_mst=taxpayer_mst)
    if match_status:
        query = query.filter_by(match_status=match_status)

    transactions = query.all()
    return jsonify([tx.to_dict() for tx in transactions])

@invoices_blueprint.get("/api/fraud/network")
def api_fraud_network():
    """US-330: Fetch directed supplier-buyer transaction network graph."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.graph_service import TaxpayerNetworkGraphGenerator
    try:
        graph = TaxpayerNetworkGraphGenerator.build_network_graph(taxpayer_mst)
        formatted_nodes = [node for node in graph["nodes"].values()]
        formatted_edges = [edge for edge in graph["edges"].values()]
        return jsonify({
            "status": "success",
            "nodes": formatted_nodes,
            "edges": formatted_edges
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/fraud/alerts")
def api_fraud_alerts():
    """US-331: Get VAT circular invoicing loop alerts and authority score outliers."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.graph_service import TaxpayerNetworkGraphGenerator, VATFraudRingNetworkDetector
    try:
        graph = TaxpayerNetworkGraphGenerator.build_network_graph(taxpayer_mst)
        detector = VATFraudRingNetworkDetector(graph)
        alerts = detector.detect_fraud_networks()
        return jsonify({
            "status": "success",
            "alerts": alerts
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ledger/verify")
def api_ledger_verify():
    """US-332: Verify the cryptographic Merkle Ledger integrity for a taxpayer."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.merkle_service import verify_ledger_integrity, rebuild_and_write_merkle_roots
    try:
        rebuild_and_write_merkle_roots(taxpayer_mst)
        is_valid, tampered_ids = verify_ledger_integrity(taxpayer_mst)
        return jsonify({
            "status": "success",
            "is_valid": is_valid,
            "tampered_invoice_ids": tampered_ids,
            "message": "Không phát hiện hành vi can thiệp dữ liệu." if is_valid else f"Phát hiện dữ liệu bị sửa đổi ở các hóa đơn: {', '.join(tampered_ids)}"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ledger/zkp-prove")
def api_ledger_zkp_prove():
    """US-333: Generate ZKP proof of compliance for a given invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    invoice_id = body.get("invoice_id")
    if not invoice_id:
        return jsonify({"error": "invoice_id is required."}), 400

    from invoices.models import Invoice
    invoice = Invoice.query.filter_by(id=invoice_id).first()
    if not invoice:
        return jsonify({"error": "Invoice not found."}), 404

    rate_percent = 10
    if invoice.amount_before_tax > 0:
        calculated_rate = (invoice.tax_amount / invoice.amount_before_tax) * 100
        rate_percent = int(round(calculated_rate))

    from invoices.zkp_service import generate_vat_compliance_proof
    try:
        proof = generate_vat_compliance_proof(
            amount_before_tax=invoice.amount_before_tax,
            tax_amount=invoice.tax_amount,
            rate_percent=rate_percent
        )
        return jsonify({
            "status": "success",
            "invoice_id": invoice_id,
            "proof": proof
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ledger/zkp-verify")
def api_ledger_zkp_verify():
    """US-333: Verify a ZKP proof of compliance."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    proof_data = body.get("proof_data")
    if not proof_data:
        return jsonify({"error": "proof_data is required."}), 400

    from invoices.zkp_service import verify_vat_compliance_proof
    try:
        is_valid = verify_vat_compliance_proof(proof_data)
        return jsonify({
            "status": "success",
            "is_valid": is_valid,
            "message": "Chứng minh tuân thủ thuế GTGT hợp lệ (ZKP Verified)." if is_valid else "Chứng minh tuân thủ không hợp lệ."
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/customs/upload")
@roles_required("admin", "auditor")
def api_customs_upload():
    """US-334: Import VNACCS/VCIS Customs XML import declarations."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file or not file.filename.endswith(".xml"):
        return jsonify({"error": "Only XML files are supported"}), 400

    try:
        xml_bytes = file.read()
        from invoices.customs_service import CustomsReconciliationEngine
        decl = CustomsReconciliationEngine.ingest_declaration(xml_bytes)
        return jsonify({
            "status": "success",
            "message": "Customs declaration imported successfully.",
            "declaration": decl.to_dict()
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/customs/reconcile")
@roles_required("admin", "auditor")
def api_customs_reconcile():
    """US-335: Compare customs declarations with domestic/import VAT input invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.customs_service import CustomsReconciliationEngine
    try:
        results = CustomsReconciliationEngine.run_reconciliation(taxpayer_mst)
        return jsonify({
            "status": "success",
            "results": results
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/customs/declarations")
def api_customs_declarations():
    """US-334: List imported customs declarations."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    taxpayer_mst = session.get("active_taxpayer_mst") or request.args.get("taxpayer_mst")
    if not taxpayer_mst:
        return jsonify({"error": "taxpayer_mst is required."}), 400

    from invoices.models import CustomsDeclaration
    try:
        decls = CustomsDeclaration.query.filter_by(taxpayer_mst=taxpayer_mst).all()
        return jsonify({
            "status": "success",
            "declarations": [d.to_dict() for d in decls]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/predictive/tax-forecast")
def api_tax_forecast():
    """US-324: Retrieve predictive tax liability reports using ML trend + seasonality forecasting."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    months_ahead = int(body.get("months_ahead", 12))

    # Allow client to supply historical data, or aggregate from DB
    historical = body.get("historical_data")
    if historical is None:
        if not taxpayer_mst:
            return jsonify({"error": "taxpayer_mst is required to retrieve database history."}), 400

        from invoices.models import Invoice
        try:
            invoices = Invoice.query.filter_by(taxpayer_mst=taxpayer_mst, is_cancelled=False).all()
            from collections import defaultdict
            monthly_data = defaultdict(lambda: {
                "output_vat": 0.0,
                "input_vat": 0.0,
                "revenue": 0.0,
                "expenses": 0.0,
            })

            for inv in invoices:
                if not inv.date or len(inv.date) < 7:
                    continue
                period = inv.date[:7]  # YYYY-MM
                if "-" not in period:
                    continue

                if inv.seller_mst == taxpayer_mst:
                    monthly_data[period]["revenue"] += inv.amount_before_tax
                    monthly_data[period]["output_vat"] += inv.tax_amount
                elif inv.buyer_mst == taxpayer_mst:
                    monthly_data[period]["expenses"] += inv.amount_before_tax
                    monthly_data[period]["input_vat"] += inv.tax_amount

            historical = []
            for period, vals in monthly_data.items():
                vat_pay = max(0.0, vals["output_vat"] - vals["input_vat"])
                pretax = vals["revenue"] - vals["expenses"]
                cit_pay = max(0.0, pretax * 0.20)
                fct_pay = max(0.0, vals["expenses"] * 0.10 * 0.05)
                
                historical.append({
                    "period": period,
                    "vat_payable": vat_pay,
                    "cit_payable": cit_pay,
                    "fct_payable": fct_pay,
                })
        except Exception as e:
            return jsonify({"error": f"Failed to retrieve history: {str(e)}"}), 500

    from invoices.tax_forecaster import ml_forecast_tax_liabilities
    try:
        forecast = ml_forecast_tax_liabilities(historical, months_ahead=months_ahead)
        return jsonify({
            "status": "success",
            "forecast": forecast
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/predictive/simulate-scenario")
def api_simulate_scenario():
    """US-325: Execute stateless comparative tax scenario calculations."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = session.get("active_taxpayer_mst") or body.get("taxpayer_mst")
    
    adjustments = body.get("adjustments", {})
    base_data = body.get("base_data")

    # If base data is not supplied, build it from DB aggregates or default mocks
    if base_data is None:
        if not taxpayer_mst:
            return jsonify({"error": "taxpayer_mst is required to aggregate baseline data."}), 400

        from invoices.models import Invoice
        try:
            invoices = Invoice.query.filter_by(taxpayer_mst=taxpayer_mst, is_cancelled=False).all()
            output_vat_base = 0.0
            input_vat_base = 0.0
            revenue_base = 0.0
            expenses_base = 0.0

            for inv in invoices:
                if inv.seller_mst == taxpayer_mst:
                    revenue_base += inv.amount_before_tax
                    output_vat_base += inv.tax_amount
                elif inv.buyer_mst == taxpayer_mst:
                    expenses_base += inv.amount_before_tax
                    input_vat_base += inv.tax_amount

            base_data = {
                "output_vat_base": output_vat_base,
                "input_vat_base": input_vat_base,
                "revenue_base": revenue_base,
                "expenses_base": expenses_base,
                "fct_base_amount": expenses_base * 0.10,
                "related_party_interest_base": expenses_base * 0.05,
                "depreciation_base": expenses_base * 0.08,
            }
        except Exception as e:
            return jsonify({"error": f"Failed to build baseline data: {str(e)}"}), 500

    from invoices.tax_forecaster import simulate_tax_scenario
    try:
        result = simulate_tax_scenario(base_data, adjustments)
        return jsonify({
            "status": "success",
            "result": result
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/audit/calculate-penalties")
def api_calculate_penalties():
    """US-340: Calculate GDT tax penalties and daily late payment interest."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
    
    body = request.get_json(silent=True) or {}
    underpaid_tax = float(body.get("underpaid_tax", 0.0))
    due_date = body.get("due_date")
    payment_date = body.get("payment_date")
    evasion_multiplier = float(body.get("evasion_multiplier", 0.0))
    has_mitigating_factors = bool(body.get("has_mitigating_factors", False))
    
    if not due_date or not payment_date:
        return jsonify({"error": "Thieu thong tin ngay den han hoac ngay nop tien thuc te."}), 400
        
    try:
        from invoices.tax_audit_service import calculate_audit_penalties
        result = calculate_audit_penalties(
            underpaid_tax=underpaid_tax,
            due_date=due_date,
            payment_date=payment_date,
            evasion_multiplier=evasion_multiplier,
            has_mitigating_factors=has_mitigating_factors
        )
        return jsonify({
            "status": "success",
            "calculation": result
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/audit/generate-explanation")
def api_generate_explanation():
    """US-341: Generate statutory Vietnamese letters citing compliance laws."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    body = request.get_json(silent=True) or {}
    risk_type = body.get("risk_type", "")
    taxpayer_name = body.get("taxpayer_name", "CONG TY TNHH MOCK")
    taxpayer_mst = body.get("taxpayer_mst") or session.get("active_taxpayer_mst") or "0109998887"
    details = body.get("details", {})
    
    if not risk_type:
        return jsonify({"error": "Thieu thong tin loai rui ro (risk_type)."}), 400
        
    try:
        from invoices.tax_audit_service import generate_audit_defense_letter
        letter = generate_audit_defense_letter(
            risk_type=risk_type,
            taxpayer_name=taxpayer_name,
            taxpayer_mst=taxpayer_mst,
            details=details
        )
        return jsonify({
            "status": "success",
            "letter": letter
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/ecommerce/normalize-orders")
def api_ecommerce_normalize_orders():
    """US-342: Map raw platform order fields from Shopee, Lazada, and TikTok Shop into standardized internal model."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    body = request.get_json(silent=True) or {}
    raw_orders = body.get("orders") or body.get("raw_logs") or []
    platform = body.get("platform", "shopee")
    
    if not raw_orders:
        return jsonify({"error": "Thieu danh sach don hang."}), 400
        
    try:
        from invoices.ecommerce_service import normalize_ecommerce_orders
        normalized = normalize_ecommerce_orders(raw_orders, platform)
        session["normalized_orders"] = normalized
        return jsonify({
            "status": "success",
            "count": len(normalized),
            "orders": normalized
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/payroll/audit-summary")
def api_payroll_audit_summary():
    """US-344: Verify PIT progressive tax tables (5%-35%) and social insurance rates."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    body = request.get_json(silent=True) or {}
    employees = body.get("employees") or []
    
    if not employees:
        employees = [
            {"id": "EMP001", "name": "Nguyễn Văn A", "mst": "8012345678", "gross_salary": 45000000.0, "dependents": 2, "withheld_pit": 2445000.0, "withheld_insurance": 4725000.0},
            {"id": "EMP002", "name": "Trần Thị B", "mst": "8012345679", "gross_salary": 12000000.0, "dependents": 0, "withheld_pit": 50000.0, "withheld_insurance": 1260000.0},
            {"id": "EMP003", "name": "Lê Văn C", "mst": "8012345680", "gross_salary": 85000000.0, "dependents": 1, "withheld_pit": 12500000.0, "withheld_insurance": 4914000.0},
            {"id": "EMP004", "name": "Phạm Thị D", "mst": "8012345681", "gross_salary": 25000000.0, "dependents": 3, "withheld_pit": 0.0, "withheld_insurance": 2625000.0},
            {"id": "EMP005", "name": "Hoàng Văn E", "mst": "8012345682", "gross_salary": 60000000.0, "dependents": 1, "withheld_pit": 7000000.0, "withheld_insurance": 4914000.0}
        ]
        
    try:
        from invoices.payroll_pit_service import audit_payroll_register
        report = audit_payroll_register(employees)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/payroll/export-pit-xml")
def api_payroll_export_pit_xml():
    """US-345: Scaffold GDT-compliant year-end PIT finalization Form 05/QTT-TNCN XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized
        
    body = request.get_json(silent=True) or {}
    metadata = body.get("metadata") or {}
    employees = body.get("employees") or []
    
    if not metadata:
        metadata = {
            "mst": body.get("taxpayer_mst") or "0109998887",
            "company_name": "Công ty TNHH GDT Invoice Hub",
            "year": body.get("tax_year") or datetime.now().year
        }
        
    if not employees:
        employees = [
            {"id": "EMP001", "name": "Nguyễn Văn A", "mst": "8012345678", "gross_salary": 45000000.0, "dependents": 2, "withheld_pit": 2445000.0, "withheld_insurance": 4725000.0},
            {"id": "EMP002", "name": "Trần Thị B", "mst": "8012345679", "gross_salary": 12000000.0, "dependents": 0, "withheld_pit": 50000.0, "withheld_insurance": 1260000.0},
            {"id": "EMP003", "name": "Lê Văn C", "mst": "8012345680", "gross_salary": 85000000.0, "dependents": 1, "withheld_pit": 12500000.0, "withheld_insurance": 4914000.0},
            {"id": "EMP004", "name": "Phạm Thị D", "mst": "8012345681", "gross_salary": 25000000.0, "dependents": 3, "withheld_pit": 0.0, "withheld_insurance": 2625000.0},
            {"id": "EMP005", "name": "Hoàng Văn E", "mst": "8012345682", "gross_salary": 60000000.0, "dependents": 1, "withheld_pit": 7000000.0, "withheld_insurance": 4914000.0}
        ]
        
    try:
        from invoices.payroll_pit_service import generate_form_05_qtt_tncn_xml
        xml_str = generate_form_05_qtt_tncn_xml(metadata, employees)
        return jsonify({
            "status": "success",
            "xml": xml_str
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/invoices/scaffold-xml")
def api_invoices_scaffold_xml():
    """US-361: Generate GDT-compliant e-invoice XML draft from OCR fields."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    ocr_data = body.get("ocr_data") or body
    
    if not ocr_data:
        return jsonify({"error": "Thieu thong tin OCR de tao XML."}), 400
        
    try:
        from invoices.v24_compliance_service import scaffold_xml_from_ocr_data
        xml_bytes = scaffold_xml_from_ocr_data(ocr_data)
        return jsonify({
            "status": "success",
            "xml": xml_bytes.decode("utf-8")
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/invoices/sign-hsm")
def api_invoices_sign_hsm():
    """US-362: Cryptographically sign invoice XML using simulated HSM certificate."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    xml_str = body.get("xml")
    ocr_data = body.get("ocr_data")
    
    try:
        from invoices.v24_compliance_service import scaffold_xml_from_ocr_data, generate_hsm_mock_certificate, sign_xml_invoice
        
        if not xml_str and ocr_data:
            xml_bytes = scaffold_xml_from_ocr_data(ocr_data)
            company_name = ocr_data.get("seller_name") or "Cong Ty Mau"
            mst = ocr_data.get("seller_mst") or "0100112233"
        elif xml_str:
            xml_bytes = xml_str.encode("utf-8")
            # Parse XML to extract company name and mst for certificate
            import lxml.etree
            root = lxml.etree.fromstring(xml_bytes)
            seller_mst_nodes = root.xpath("//*[local-name()='NBan']/*[local-name()='MST']")
            seller_name_nodes = root.xpath("//*[local-name()='NBan']/*[local-name()='Ten']")
            company_name = seller_name_nodes[0].text if seller_name_nodes else "Cong Ty Mau"
            mst = seller_mst_nodes[0].text if seller_mst_nodes else "0100112233"
        else:
            return jsonify({"error": "Yeu cau thieu xml hoac ocr_data."}), 400

        cert_der, priv_key = generate_hsm_mock_certificate(company_name, mst)
        signed_bytes = sign_xml_invoice(xml_bytes, cert_der, priv_key)
        
        return jsonify({
            "status": "success",
            "signed_xml": signed_bytes.decode("utf-8"),
            "certificate_issuer": "MISA-CA Root Authority"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/gdt-sandbox/transmit")
def api_gdt_sandbox_transmit():
    """US-363: Transmit signed XML to GDT Sandbox Gateway and verify compliance."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    xml_str = body.get("signed_xml") or body.get("xml")
    
    if not xml_str:
        return jsonify({"error": "Yeu cau thieu xml da ky."}), 400
        
    try:
        from invoices.v24_compliance_service import transmit_to_gdt_sandbox
        result = transmit_to_gdt_sandbox(xml_str.encode("utf-8"))
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.route("/api/compliance/decree132-checklist", methods=["GET", "POST"])
def api_compliance_decree132_checklist():
    """US-364: Check related party thresholds under Decree 132/2020/NĐ-CP."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        taxpayer_mst = body.get("taxpayer_mst")
        start_date = body.get("start_date")
        end_date = body.get("end_date")
    else:
        taxpayer_mst = request.args.get("taxpayer_mst")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

    if not taxpayer_mst or not start_date or not end_date:
        return jsonify({"error": "Yeu cau thieu taxpayer_mst, start_date hoac end_date."}), 400

    try:
        from invoices.v24_compliance_service import calculate_related_party_disclosure
        checklist = calculate_related_party_disclosure(taxpayer_mst, start_date, end_date)
        return jsonify(checklist)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/transfer-pricing-risk")
def api_compliance_transfer_pricing_risk():
    """US-365: Compare operating margins against statistical sector benchmarks."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    transactions = body.get("transactions") or []
    sector = body.get("sector") or "Manufacturing"
    
    if not transactions:
        # Provide sample template transactions if empty
        transactions = [
            {"id": "TX001", "partner_name": "Cong ty Lien ket A", "revenue": 50000000000.0, "cogs": 48500000000.0},
            {"id": "TX002", "partner_name": "Cong ty Lien ket B", "revenue": 12000000000.0, "cogs": 11000000000.0}
        ]
        
    try:
        from invoices.v24_compliance_service import analyze_transfer_pricing_risk
        analysis = analyze_transfer_pricing_risk(transactions, sector)
        return jsonify(analysis)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/compliance/gdt-status")
def api_compliance_gdt_status():
    """US-371: Fetch invoice GDT verification status, search & filter."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    status_filter = request.args.get("status")
    search_query = request.args.get("q")
    
    from invoices.models import Invoice
    query = Invoice.query
    
    if status_filter:
        query = query.filter(Invoice.invoice_status == status_filter)
    if search_query:
        query = query.filter(
            Invoice.id.contains(search_query) | 
            Invoice.number.contains(search_query) | 
            Invoice.seller_mst.contains(search_query) | 
            Invoice.buyer_mst.contains(search_query)
        )
        
    invoices = query.order_by(Invoice.updated_at.desc()).all()
    
    return jsonify({
        "status": "success",
        "invoices": [{
            "id": inv.id,
            "number": inv.number,
            "symbol": inv.symbol,
            "template_code": inv.template_code,
            "date": inv.date,
            "seller_mst": inv.seller_mst,
            "seller_name": inv.seller_name,
            "buyer_mst": inv.buyer_mst,
            "buyer_name": inv.buyer_name,
            "total_amount": inv.total_amount,
            "payment_method": inv.payment_method,
            "has_signature": inv.has_signature,
            "invoice_status": inv.invoice_status or "pending",
            "notes": inv.notes,
            "updated_at": inv.updated_at
        } for inv in invoices]
    })

@invoices_blueprint.post("/api/compliance/gdt-sync")
def api_compliance_gdt_sync():
    """US-370 / US-371: Trigger GDT Sync Agent or sync specific invoice IDs."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    invoice_ids = body.get("invoice_ids")
    
    from invoices.v25_compliance_service import run_portal_sync_agent, sync_gdt_verification_status
    try:
        if invoice_ids:
            result = sync_gdt_verification_status(invoice_ids)
        else:
            result = run_portal_sync_agent()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/generate-correction-xml")
def api_compliance_generate_correction_xml():
    """US-372: Generate Decree 123 conforming XML for corrected/replaced invoice."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    original_invoice_id = body.get("original_invoice_id")
    type_change = body.get("type_change") # "correction" or "replacement"
    new_data = body.get("new_data") or {}

    if not original_invoice_id or not type_change:
        return jsonify({"error": "Thiếu original_invoice_id hoặc type_change."}), 400

    from invoices.models import Invoice
    orig_inv = Invoice.query.get(original_invoice_id)
    if not orig_inv:
        return jsonify({"error": f"Không tìm thấy hóa đơn gốc {original_invoice_id}."}), 404

    from invoices.v25_compliance_service import generate_correction_or_replacement_xml
    try:
        xml_bytes = generate_correction_or_replacement_xml(orig_inv, new_data, type_change)
        return jsonify({
            "status": "success",
            "xml": xml_bytes.decode("utf-8"),
            "filename": f"{type_change}_invoice_{orig_inv.number}.xml"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/transmit-form-04ss")
def api_compliance_transmit_form_04ss():
    """US-373: Scaffold, sign with HSM, and transmit Form 04/SS-HĐĐT."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst")
    company_name = body.get("company_name")
    bad_invoices = body.get("bad_invoices") or []

    if not taxpayer_mst or not company_name or not bad_invoices:
        return jsonify({"error": "Thiếu taxpayer_mst, company_name hoặc danh sách bad_invoices."}), 400

    from invoices.v25_compliance_service import generate_form_04_ss_xml
    from invoices.v24_compliance_service import generate_hsm_mock_certificate, sign_xml_invoice, transmit_to_gdt_sandbox
    try:
        # 1. Scaffold Form 04/SS XML
        xml_bytes = generate_form_04_ss_xml(taxpayer_mst, company_name, bad_invoices)
        
        # 2. Sign XML using mock HSM certificate
        cert_der, priv_key = generate_hsm_mock_certificate(company_name, taxpayer_mst)
        signed_xml_bytes = sign_xml_invoice(xml_bytes, cert_der, priv_key)
        
        # 3. Transmit signed XML to GDT sandbox
        transmission_result = transmit_to_gdt_sandbox(signed_xml_bytes)
        
        return jsonify({
            "status": "success",
            "raw_xml": xml_bytes.decode("utf-8"),
            "signed_xml": signed_xml_bytes.decode("utf-8"),
            "transmission_result": transmission_result
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/tax-optimization")
def api_compliance_tax_optimization():
    """US-374: Run corporate tax optimization and scenario simulations."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    scenarios_config = body.get("scenarios")

    if not scenarios_config:
        # Fallback to standard scenario checklist if empty
        scenarios_config = [
            {
                "name": "Kịch bản tối ưu 1: Thuế suất ưu đãi 10% & Miễn thuế 2 năm",
                "preferential_rate": 0.10,
                "holiday_exempt_years": 2,
                "holiday_reduce_years": 4,
                "reduce_loan_interest": True,
                "enforce_bank_transfer": True
            },
            {
                "name": "Kịch bản tối ưu 2: Thuế suất ưu đãi 15% & Giảm thuế 50% trong 2 năm",
                "preferential_rate": 0.15,
                "holiday_exempt_years": 0,
                "holiday_reduce_years": 2,
                "reduce_loan_interest": False,
                "enforce_bank_transfer": False
            }
        ]

    from invoices.v25_compliance_service import calculate_corporate_tax_optimization
    try:
        report = calculate_corporate_tax_optimization(taxpayer_mst, scenarios_config)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/insurance-audit")
def api_compliance_insurance_audit():
    """US-380: Audit payroll trích đóng BHXH/BHYT/BHTN against statutory rates."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    basic_salary = request.args.get("basic_salary", type=float) or 2340000.0

    # Mock employee payroll records for auditing
    mock_payroll = [
        {"id": "EMP-001", "name": "Nguyễn Văn A", "gross_salary": 15000000.0, "withheld_insurance": 1575000.0},
        {"id": "EMP-002", "name": "Trần Thị B", "gross_salary": 25000000.0, "withheld_insurance": 2300000.0},  # Mismatch (Statutory is 2,625,000)
        {"id": "EMP-003", "name": "Lê Văn C", "gross_salary": 55000000.0, "withheld_insurance": 4914000.0}   # Capped at 46,800,000 (Statutory 4,914,000)
    ]

    from invoices.v26_service import audit_social_insurance
    try:
        result = audit_social_insurance(mock_payroll, basic_salary)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/compliance/insurance-export-csv")
def api_compliance_insurance_export_csv():
    """US-381: Export social insurance audit discrepancies as a CSV report."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))

    basic_salary = request.args.get("basic_salary", type=float) or 2340000.0

    mock_payroll = [
        {"id": "EMP-001", "name": "Nguyễn Văn A", "gross_salary": 15000000.0, "withheld_insurance": 1575000.0},
        {"id": "EMP-002", "name": "Trần Thị B", "gross_salary": 25000000.0, "withheld_insurance": 2300000.0},
        {"id": "EMP-003", "name": "Lê Văn C", "gross_salary": 55000000.0, "withheld_insurance": 4914000.0}
    ]

    from invoices.v26_service import audit_social_insurance, export_si_reconciliation_csv
    try:
        audit_result = audit_social_insurance(mock_payroll, basic_salary)
        csv_data = export_si_reconciliation_csv(audit_result)
        
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=si_reconciliation_report.csv"}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/tax-ledger-reconcile")
def api_compliance_tax_ledger_reconcile():
    """US-382: Sync taxpayer e-Tax ledger and reconcile against local journals."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    local_payments = body.get("local_payments") or []
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"

    from invoices.v26_service import reconcile_tax_ledger
    try:
        recompiled = reconcile_tax_ledger(taxpayer_mst, local_payments)
        return jsonify({
            "status": "success",
            "recompiled": recompiled
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/vietqr-generate")
def api_compliance_vietqr_generate():
    """US-383: Generate Napas-compliant dynamic VietQR tax payment code."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    tax_type = body.get("tax_type", "VAT")
    amount = body.get("amount", 1000.0)
    taxpayer_mst = session.get("taxpayer_mst") or "0109999999"

    from invoices.v26_service import generate_napas_vietqr_payload
    try:
        payload = generate_napas_vietqr_payload(tax_type, amount, taxpayer_mst)
        return jsonify(payload)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/vietqr-confirm")
def api_compliance_vietqr_confirm():
    """US-383: Confirm dynamic tax payment transaction (status change simulation)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    tx_id = body.get("transaction_id")
    if not tx_id:
        return jsonify({"error": "Thiếu transaction_id"}), 400

    return jsonify({
        "transaction_id": tx_id,
        "status": "paid",
        "message": f"Giao dịch nộp thuế {tx_id} đã được xác nhận khớp lệnh với Kho bạc Nhà nước.",
        "completed_at": datetime.now().isoformat()
    })

@invoices_blueprint.get("/api/compliance/kg-query")
def api_compliance_kg_query():
    """US-384: Query Vietnamese Tax Law Knowledge Graph."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    query = request.args.get("query", "")
    from invoices.v26_service import TaxLawKnowledgeGraph
    try:
        kg = TaxLawKnowledgeGraph()
        results = kg.keyword_search(query)
        return jsonify({"results": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/defense-compose")
def api_compliance_defense_compose():
    """US-385: Dynamic AI Audit Defense Document Composer."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    warning_type = body.get("warning_type")
    context = body.get("context") or {}

    profile = {
        "mst": session.get("taxpayer_mst") or "0109999999",
        "company_name": session.get("company_name") or "Công ty TNHH Giải pháp Phần mềm Ánh Sáng",
        "district": "Cục Thuế Thành phố Hà Nội",
        "representative": "Giám Đốc"
    }

    from invoices.v26_service import compose_audit_defense_letter
    try:
        letter_html = compose_audit_defense_letter(profile, warning_type, context)
        return jsonify({
            "status": "success",
            "letter_html": letter_html
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/pxk-parse")
def api_compliance_pxk_parse():
    """US-390: Parse and validate official electronic delivery note XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    xml_content = body.get("xml_content", "")
    if not xml_content:
        return jsonify({"error": "Thiếu dữ liệu xml_content"}), 400

    from invoices.v27_service import parse_delivery_note_xml
    try:
        parsed = parse_delivery_note_xml(xml_content)
        return jsonify(parsed)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/pxk-reconcile")
def api_compliance_pxk_reconcile():
    """US-391: Reconcile delivery note items against commercial invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    delivery_notes = body.get("delivery_notes") or []
    invoices = body.get("invoices") or []

    from invoices.v27_service import reconcile_delivery_to_invoice
    try:
        report = reconcile_delivery_to_invoice(delivery_notes, invoices)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/pxk-export-csv")
def api_compliance_pxk_export_csv():
    """US-391: Export reconciliation differences as a CSV report."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    delivery_notes = body.get("delivery_notes") or []
    invoices = body.get("invoices") or []

    from invoices.v27_service import reconcile_delivery_to_invoice, export_delivery_reconciliation_csv
    try:
        report = reconcile_delivery_to_invoice(delivery_notes, invoices)
        csv_data = export_delivery_reconciliation_csv(report)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=pxk_reconciliation_report.csv"}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/pre-audit-risk")
def api_compliance_pre_audit_risk():
    """US-392: Calculate the pre-audit corporate tax risk scorecard."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    invoices = body.get("invoices") or []
    related_party_context = body.get("related_party_context") or {}
    profile = {
        "mst": session.get("taxpayer_mst") or "0109999999",
        "company_name": session.get("company_name") or "Công ty TNHH Ánh Sáng"
    }

    from invoices.v27_service import calculate_pre_audit_risk
    try:
        report = calculate_pre_audit_risk(profile, invoices, related_party_context)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/risk-radar-svg")
def api_compliance_risk_radar_svg():
    """US-393: Generate dynamic SVG risk radar chart markup."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    scores = body.get("scores") or {}

    from invoices.v27_service import generate_svg_radar_chart
    try:
        svg_markup = generate_svg_radar_chart(scores)
        return jsonify({"svg_markup": svg_markup})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/econtract-parse")
def api_compliance_econtract_parse():
    """US-394: Parse electronic contract structured metadata."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    json_content = body.get("json_content", "")
    if not json_content:
        return jsonify({"error": "Thiếu dữ liệu json_content"}), 400

    from invoices.v27_service import parse_econtract_metadata
    try:
        parsed = parse_econtract_metadata(json_content)
        return jsonify(parsed)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/econtract-reconcile")
def api_compliance_econtract_reconcile():
    """US-394: Reconcile contract milestones with invoices and payments."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    contract = body.get("contract") or {}
    invoices = body.get("invoices") or []
    payments = body.get("payments") or []

    from invoices.v27_service import reconcile_contract_milestones
    try:
        report = reconcile_contract_milestones(contract, invoices, payments)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/treasury-forecast")
def api_compliance_treasury_forecast():
    """US-395: Smart treasury forecast scenario simulation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    milestones = body.get("milestones") or []
    invoices = body.get("invoices") or []
    starting_cash = body.get("starting_cash", 1000000000.0)
    delay_days = body.get("delay_days", 0)
    cit_discount = body.get("cit_discount", 0.0)

    from invoices.v27_service import simulate_treasury_forecast
    try:
        forecast = simulate_treasury_forecast(milestones, invoices, starting_cash, delay_days, cit_discount)
        return jsonify(forecast)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/xml-audit")
def api_compliance_xml_audit():
    """US-397: Audit invoice XML compliance and structure."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    xml_content = body.get("xml_content", "")
    if not xml_content:
        return jsonify({"error": "Thiếu dữ liệu xml_content"}), 400

    from invoices.v28_service import audit_xml_compliance
    try:
        result = audit_xml_compliance(xml_content)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/xml-auto-repair")
def api_compliance_xml_auto_repair():
    """US-397: Auto-repair schema errors, tags order, and generate sign-off XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    xml_content = body.get("xml_content", "")
    if not xml_content:
        return jsonify({"error": "Thiếu dữ liệu xml_content"}), 400

    from invoices.v28_service import repair_xml_invoice
    try:
        result = repair_xml_invoice(xml_content)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-chat")
def api_agents_swarm_chat():
    """US-396: Run interactive collaborative agent swarm simulation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    query = body.get("query", "")
    if not query:
        return jsonify({"error": "Thiếu câu hỏi rà soát"}), 400

    taxpayer_mst = session.get("taxpayer_mst") or "0109998887"
    
    from invoices.v28_service import simulate_swarm_step_by_step, JointAuditCoordinator
    try:
        # Simulate swarm communication logs
        chat_steps = simulate_swarm_step_by_step(taxpayer_mst, query)
        
        # Also invoke the actual JointAuditCoordinator to get the final generated report markdown
        coordinator = JointAuditCoordinator(taxpayer_mst=taxpayer_mst)
        report_markdown = coordinator.execute_swarm(query)
        
        return jsonify({
            "status": "success",
            "chat_steps": chat_steps,
            "report_markdown": report_markdown
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/ghost-check")
def api_compliance_ghost_check():
    """US-400: Ghost Company Blacklist Scraper & Probability Index Engine."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    seller_mst = body.get("seller_mst", "")
    seller_name = body.get("seller_name", "")
    invoice_value = float(body.get("invoice_value", 0))

    if not seller_mst:
        return jsonify({"error": "Thiếu dữ liệu seller_mst"}), 400

    from invoices.v29_service import check_ghost_company
    try:
        result = check_ghost_company(seller_mst, seller_name, invoice_value)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/defense-letter")
def api_compliance_defense_letter():
    """US-401: Generate Tax Audit Defense Letter & Rectification Plan."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    seller_mst = body.get("seller_mst", "")
    seller_name = body.get("seller_name", "")
    invoice_value = float(body.get("invoice_value", 0))
    payment_method = body.get("payment_method", "Chuyển khoản qua Ngân hàng thương mại")

    if not seller_mst or not seller_name:
        return jsonify({"error": "Thiếu dữ liệu nhà cung cấp"}), 400

    from invoices.v29_service import generate_audit_mitigation_letter
    try:
        letter_text = generate_audit_mitigation_letter(seller_mst, seller_name, invoice_value, payment_method)
        return jsonify({
            "success": True,
            "letter": letter_text
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/compliance/tax-knowledge-graph")
def api_compliance_tax_knowledge_graph():
    """US-402: Return Vietnamese Tax Regulations Knowledge Graph."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.v29_service import get_tax_knowledge_graph
    try:
        graph_data = get_tax_knowledge_graph()
        return jsonify(graph_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v29-chat")
def api_agents_swarm_v29_chat():
    """US-401 Swarm: Run simulated multi-agent swarm discussion for invoice compliance defense."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    seller_mst = body.get("seller_mst", "")
    seller_name = body.get("seller_name", "")
    invoice_value = float(body.get("invoice_value", 0))
    taxpayer_mst = body.get("mst", "0109998887")

    if not seller_mst:
        return jsonify({"error": "Thiếu dữ liệu seller_mst"}), 400

    from invoices.v29_service import SwarmV29Advisor, generate_audit_mitigation_letter
    try:
        advisor = SwarmV29Advisor(taxpayer_mst=taxpayer_mst)
        chat_steps = advisor.simulate_defense_chat(seller_mst, seller_name, invoice_value)
        report_markdown = generate_audit_mitigation_letter(seller_mst, seller_name, invoice_value, "Chuyển khoản qua ngân hàng (CK)")
        return jsonify({
            "status": "success",
            "chat_steps": chat_steps,
            "report_markdown": report_markdown
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/transfer-pricing-check")
def api_compliance_transfer_pricing_check():
    """US-410: Related-Party Transaction Markup & Interquartile (IQR) Margin Risk Analyzer."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    sector = body.get("sector", "manufacturing")
    markup_pct = float(body.get("markup_pct", 0.0))
    cost_of_goods = float(body.get("cost_of_goods", 0.0))

    from invoices.v30_service import calculate_transfer_pricing_risk
    try:
        result = calculate_transfer_pricing_risk(markup_pct, cost_of_goods, sector)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v30-chat")
def api_agents_swarm_v30_chat():
    """US-412 Swarm: Run simulated multi-agent swarm discussion for related-party transfer pricing audit preparation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst", "0109998887")
    taxpayer_name = body.get("taxpayer_name", "Doanh nghiệp mẫu")
    sector = body.get("sector", "manufacturing")
    markup_pct = float(body.get("markup_pct", 0.0))
    cost_of_goods = float(body.get("cost_of_goods", 0.0))

    from invoices.v30_service import SwarmV30Advisor, calculate_transfer_pricing_risk, generate_tp_audit_dossier
    try:
        advisor = SwarmV30Advisor(taxpayer_mst=taxpayer_mst)
        chat_steps = advisor.simulate_tp_defense_chat(sector, markup_pct, cost_of_goods)
        
        risk_details = calculate_transfer_pricing_risk(markup_pct, cost_of_goods, sector)
        dossier = generate_tp_audit_dossier(taxpayer_name, taxpayer_mst, sector, markup_pct, cost_of_goods, risk_details)
        
        return jsonify({
            "status": "success",
            "chat_steps": chat_steps,
            "dossier": dossier
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/vat-reconciliation")
def api_compliance_vat_reconciliation():
    """US-420: Multi-Period VAT Reconciliation Engine with Input/Output VAT Balancing."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    periods = body.get("periods")

    from invoices.v31_service import vat_reconciliation_multi_period
    try:
        result = vat_reconciliation_multi_period(taxpayer_mst, periods)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/form01-gtgt-xml")
def api_compliance_form01_gtgt_xml():
    """US-421: Automated Form 01/GTGT VAT Declaration XML Builder."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    taxpayer_name = body.get("taxpayer_name") or "Công ty TNHH Giải pháp Phần mềm"
    period = body.get("period") or ""
    output_vat = float(body.get("output_vat", 0))
    input_vat = float(body.get("input_vat", 0))
    carry_forward_prev = float(body.get("carry_forward_prev", 0))

    if not period:
        return jsonify({"error": "Thiếu kỳ kê khai (period)"}), 400

    from invoices.v31_service import build_form01_gtgt_xml
    try:
        result = build_form01_gtgt_xml(
            taxpayer_mst, taxpayer_name, period,
            output_vat, input_vat, carry_forward_prev,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v31-chat")
def api_agents_swarm_v31_chat():
    """US-422: AI VAT Anomaly Detection Swarm and Cross-Period Audit Advisory."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    taxpayer_name = body.get("taxpayer_name") or "Doanh nghiệp phân tích"

    from invoices.v31_service import run_vat_anomaly_swarm
    try:
        result = run_vat_anomaly_swarm(taxpayer_mst, taxpayer_name)
        return jsonify({
            "status": "success",
            "chat_steps": result["chat_steps"],
            "report_markdown": result["report_markdown"],
            "reconciliation": result["reconciliation"],
            "risk_level": result["risk_level"],
            "total_anomalies": result["total_anomalies"],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v32-chat")
def api_agents_swarm_v32_chat():
    """US-432: AI Swarm VAT Refund Justification Compiler & Multi-Agent Debate."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("active_taxpayer_mst")
    taxpayer_name = body.get("taxpayer_name") or "Doanh nghiệp hoàn thuế"
    eligible_invoice_ids = body.get("eligible_invoice_ids")
    customs_declarations = body.get("customs_declarations")

    if not taxpayer_mst:
        return jsonify({"error": "Missing taxpayer MST"}), 400

    from invoices.v32_service import run_refund_audit_swarm
    try:
        result = run_refund_audit_swarm(
            taxpayer_mst=taxpayer_mst,
            taxpayer_name=taxpayer_name,
            eligible_invoice_ids=eligible_invoice_ids,
            customs_declarations=customs_declarations
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/cit-quarterly")
def api_compliance_cit_quarterly():
    """US-450: CIT Quarterly Provisional Tax Calculation Engine."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    quarter = int(body.get("quarter", 1))
    year = int(body.get("year", 2026))
    revenue = float(body.get("revenue", 0))
    cogs = float(body.get("cogs", 0))
    operating_expenses = float(body.get("operating_expenses", 0))
    other_income = float(body.get("other_income", 0))
    other_expenses = float(body.get("other_expenses", 0))
    preferential_rate = body.get("preferential_rate")
    if preferential_rate is not None:
        preferential_rate = float(preferential_rate)
    carry_forward_loss = float(body.get("carry_forward_loss", 0))

    from invoices.v33_service import calculate_cit_quarterly
    try:
        result = calculate_cit_quarterly(
            taxpayer_mst, quarter, year, revenue, cogs,
            operating_expenses, other_income, other_expenses,
            preferential_rate, carry_forward_loss,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/form01a-tndn-xml")
def api_compliance_form01a_tndn_xml():
    """US-450: Generate Form 01A/TNDN HTKK-compatible XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    taxpayer_name = body.get("taxpayer_name") or "Công ty TNHH Giải pháp Phần mềm"
    quarter = int(body.get("quarter", 1))
    year = int(body.get("year", 2026))
    revenue = float(body.get("revenue", 0))
    cogs = float(body.get("cogs", 0))
    operating_expenses = float(body.get("operating_expenses", 0))
    other_income = float(body.get("other_income", 0))
    other_expenses = float(body.get("other_expenses", 0))
    preferential_rate = body.get("preferential_rate")
    if preferential_rate is not None:
        preferential_rate = float(preferential_rate)
    carry_forward_loss = float(body.get("carry_forward_loss", 0))

    from invoices.v33_service import build_form01a_tndn_xml
    try:
        result = build_form01a_tndn_xml(
            taxpayer_mst, taxpayer_name, quarter, year, revenue, cogs,
            operating_expenses, other_income, other_expenses,
            preferential_rate, carry_forward_loss,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/compliance/tax-calendar")
def api_compliance_tax_calendar():
    """US-451: Return Vietnamese tax compliance calendar for a given year."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    year = request.args.get("year", type=int) or 2026

    from invoices.v33_service import get_tax_compliance_calendar
    try:
        cal = get_tax_compliance_calendar(year)
        return jsonify(cal)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v33-chat")
def api_agents_swarm_v33_chat():
    """US-451: CIT Optimization Swarm Advisory for quarterly declaration."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    taxpayer_name = body.get("taxpayer_name") or "Doanh nghiệp phân tích"
    quarter = int(body.get("quarter", 1))
    year = int(body.get("year", 2026))
    revenue = float(body.get("revenue", 0))
    cogs = float(body.get("cogs", 0))
    operating_expenses = float(body.get("operating_expenses", 0))

    from invoices.v33_service import run_cit_optimization_swarm
    try:
        result = run_cit_optimization_swarm(
            taxpayer_mst, taxpayer_name, quarter, year,
            revenue, cogs, operating_expenses,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/invoice-aging")
def api_compliance_invoice_aging():
    """US-460: Invoice Aging Analysis for AR and AP."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    as_of_date = body.get("as_of_date")

    from invoices.v34_service import analyze_invoice_aging
    try:
        result = analyze_invoice_aging(taxpayer_mst, as_of_date)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/aging-heatmap")
def api_compliance_aging_heatmap():
    """US-461: Generate aging heatmap grid data."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    as_of_date = body.get("as_of_date")

    from invoices.v34_service import analyze_invoice_aging, generate_aging_heatmap_data
    try:
        aging = analyze_invoice_aging(taxpayer_mst, as_of_date)
        heatmap = generate_aging_heatmap_data(aging)
        return jsonify(heatmap)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v34-chat")
def api_agents_swarm_v34_chat():
    """US-461: AR/AP Debt Collection Swarm Advisory."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    taxpayer_name = body.get("taxpayer_name") or "Doanh nghiệp phân tích"

    from invoices.v34_service import run_aging_advisory_swarm
    try:
        result = run_aging_advisory_swarm(taxpayer_mst, taxpayer_name)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/stress-test")
def api_compliance_stress_test():
    """US-471: Run tax audit risk stress simulation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    scan_rate = float(body.get("scan_rate", 0.5))
    strictness = body.get("strictness", "medium")

    from invoices.v35_service import run_tax_stress_simulation
    try:
        result = run_tax_stress_simulation(taxpayer_mst, scan_rate, strictness)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/compliance/defense-package")
def api_compliance_defense_package():
    """US-472: Generate and download defense briefcase ZIP."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"
    invoice_ids = body.get("invoice_ids", [])

    from invoices.v35_service import build_defense_briefcase
    import os
    try:
        zip_path = build_defense_briefcase(taxpayer_mst, invoice_ids)
        from flask import send_file
        return send_file(
            zip_path,
            mimetype="application/zip",
            as_attachment=True,
            download_name=os.path.basename(zip_path)
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/agents/swarm-v35-chat")
def api_agents_swarm_v35_chat():
    """US-474: AI Swarm Defense Chat mock debate."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    taxpayer_mst = body.get("taxpayer_mst") or session.get("taxpayer_mst") or "0109999999"

    from invoices.v35_service import run_v35_swarm
    try:
        result = run_v35_swarm(taxpayer_mst)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/calculate")
def api_cit_calculate():
    """US-480: Calculate CIT liability."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    revenue = float(body.get("revenue", 0.0))
    cogs = float(body.get("cogs", 0.0))
    selling_expenses = float(body.get("selling_expenses", 0.0))
    admin_expenses = float(body.get("admin_expenses", 0.0))
    non_deductible_adjustments = float(body.get("non_deductible_adjustments", 0.0))
    loss_offset = float(body.get("loss_offset", 0.0))
    cit_rate = float(body.get("cit_rate", 0.20))
    holiday_discount = float(body.get("holiday_discount", 0.0))

    from invoices.v36_service import CITFinalizationService
    try:
        result = CITFinalizationService.calculate_cit(
            revenue, cogs, selling_expenses, admin_expenses, 
            non_deductible_adjustments, loss_offset, cit_rate, holiday_discount
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/optimize-losses")
def api_cit_optimize_losses():
    """US-481: Compute optimal carry-forward matrix."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    # Parse historical losses: keys should be ints
    hist_losses_raw = body.get("historical_losses", {})
    hist_losses = {int(k): float(v) for k, v in hist_losses_raw.items()}
    
    # Parse projected profits: keys should be ints
    proj_profits_raw = body.get("projected_profits", {})
    proj_profits = {int(k): float(v) for k, v in proj_profits_raw.items()}
    
    # Parse holidays: keys should be ints
    holidays_raw = body.get("tax_holidays", {})
    holidays = {}
    for k, v in holidays_raw.items():
        holidays[int(k)] = {
            "tax_free": bool(v.get("tax_free", False)),
            "reduction": float(v.get("reduction", 0.0))
        }
        
    cit_rate = float(body.get("cit_rate", 0.20))

    from invoices.v36_service import CITFinalizationService
    try:
        result = CITFinalizationService.optimize_loss_carry_forward(
            hist_losses, proj_profits, holidays, cit_rate
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/export-xml")
def api_cit_export_xml():
    """US-482: Generate GDT Form 03/TNDN XML."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    mst = body.get("mst") or session.get("taxpayer_mst") or "0102030405"
    taxpayer_name = body.get("taxpayer_name") or "CÔNG TY CỔ PHẦN CÔNG NGHỆ ANTIGRAVITY"
    year = int(body.get("year", 2026))
    
    cit_data = body.get("cit_data", {})
    loss_data = body.get("loss_data", {})

    from invoices.v36_service import CITFinalizationService
    try:
        xml_content = CITFinalizationService.generate_cit_xml(
            mst, taxpayer_name, year, cit_data, loss_data
        )
        return Response(
            xml_content,
            mimetype="application/xml",
            headers={"Content-Disposition": f'attachment; filename="Form_03_TNDN_{year}.xml"'}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/cit/swarm-chat")
def api_cit_swarm_chat():
    """US-484: AI Swarm Consensus debate simulation."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    cit_data = body.get("cit_data", {})
    loss_data = body.get("loss_data", {})

    from invoices.v36_service import CITFinalizationService
    try:
        debate, memo = CITFinalizationService.simulate_cit_swarm_debate(cit_data, loss_data)
        return jsonify({
            "debate": debate,
            "memo": memo
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/ceo-dashboard")
def api_ceo_dashboard():
    """US-490: Get financial indicators, health score and commentary."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("taxpayer_mst") or "0102030405"
    from invoices.v37_service import CEOIntelligenceService
    try:
        health = CEOIntelligenceService.calculate_financial_health_score(mst)
        commentary = CEOIntelligenceService.generate_management_commentary(mst)
        return jsonify({
            "status": "success",
            "health_score": health["overall_score"],
            "sub_scores": health["sub_scores"],
            "commentary": commentary,
            "mst": mst
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/ceo-dashboard/sankey")
def api_ceo_dashboard_sankey():
    """US-490: Generate Sankey diagram node/link structures."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("taxpayer_mst") or "0102030405"
    from invoices.v37_service import CEOIntelligenceService
    try:
        sankey_data = CEOIntelligenceService.generate_sankey_data(mst)
        return jsonify(sankey_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/tax-planning/projection")
def api_tax_planning_projection():
    """US-491: Linear regression tax projection and NPV optimization analysis."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("taxpayer_mst") or "0102030405"
    years = int(request.args.get("years", 3))
    rev_growth = float(request.args.get("rev_growth", 0.10))
    cost_inflation = float(request.args.get("cost_inflation", 0.05))
    discount_rate = float(request.args.get("discount_rate", 0.08))

    from invoices.v37_service import MultiYearTaxPlanningService
    try:
        proj = MultiYearTaxPlanningService.generate_tax_projection(mst, years, rev_growth, cost_inflation)
        npv_opt = MultiYearTaxPlanningService.optimize_tax_npv(proj, discount_rate)
        return jsonify({
            "status": "success",
            "projection": proj["projection"],
            "npv_optimization": npv_opt
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/tax-planning/calendar")
def api_tax_planning_calendar():
    """US-492: Get or populate compliance deadline records."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    year = int(request.args.get("year", datetime.now().year))
    from invoices.v37_service import TaxFilingCalendarService
    from invoices.models import TaxFilingRecord
    try:
        # Populate calendar for current year if none exist
        count = TaxFilingRecord.query.filter(TaxFilingRecord.period.like(f"{year}%")).count()
        if count == 0:
            TaxFilingCalendarService.populate_calendar_db(year)

        records = TaxFilingRecord.query.all()
        compliance_score = TaxFilingCalendarService.calculate_compliance_score()
        return jsonify({
            "status": "success",
            "calendar": [r.to_dict() for r in records],
            "compliance_score": compliance_score
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/tax-planning/calendar/mark-filed")
def api_tax_planning_calendar_mark_filed():
    """US-492: Mark a filing task as filed."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    record_id = body.get("record_id")
    filed_date = body.get("filed_date") or date.today().isoformat()
    xml_path = body.get("xml_file_path")

    if not record_id:
        return jsonify({"error": "Missing record_id"}), 400

    from invoices.v37_service import TaxFilingCalendarService
    try:
        success = TaxFilingCalendarService.mark_filed(record_id, filed_date, xml_path)
        if success:
            return jsonify({"status": "success", "message": "Marked tax record as filed successfully."})
        return jsonify({"error": "Filing record not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/assets")
def api_assets_list():
    """US-493: List all registered fixed assets."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import FixedAsset
    try:
        assets = FixedAsset.query.all()
        return jsonify({
            "status": "success",
            "assets": [a.to_dict() for a in assets]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/assets")
def api_assets_create():
    """US-493: Register a new fixed asset."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    code = body.get("asset_code")
    name = body.get("name")
    category = body.get("category")
    acq_date = body.get("acquisition_date") or date.today().isoformat()
    cost = float(body.get("original_cost", 0.0))
    residual = float(body.get("residual_value", 0.0))
    life = int(body.get("useful_life_months", 36))
    method = body.get("depreciation_method", "straight_line")
    linked_inv = body.get("linked_invoice_id")

    if not code or not name or not category:
        return jsonify({"error": "Missing required fields: code, name, or category"}), 400

    from invoices.models import FixedAsset
    try:
        asset = FixedAsset(
            asset_code=code,
            name=name,
            category=category,
            acquisition_date=acq_date,
            original_cost=cost,
            residual_value=residual,
            useful_life_months=life,
            depreciation_method=method,
            linked_invoice_id=linked_inv,
            status="active"
        )
        db.session.add(asset)
        db.session.commit()

        # Seed depreciation schedule immediately in DB
        from invoices.v37_service import FixedAssetDepreciationEngine
        schedule = FixedAssetDepreciationEngine.generate_depreciation_schedule(asset.id)
        from invoices.models import DepreciationEntry
        for entry in schedule:
            db_entry = DepreciationEntry(
                asset_id=asset.id,
                period=entry["period"],
                depreciation_amount=entry["depreciation_amount"],
                accumulated_depreciation=entry["accumulated_depreciation"],
                net_book_value=entry["net_book_value"]
            )
            db.session.add(db_entry)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Fixed asset registered successfully.",
            "asset": asset.to_dict()
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/assets/schedule")
def api_assets_schedule():
    """US-493: Get full depreciation schedule for an asset."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    asset_id = request.args.get("asset_id")
    if not asset_id:
        return jsonify({"error": "Missing asset_id"}), 400

    from invoices.models import DepreciationEntry
    try:
        entries = DepreciationEntry.query.filter_by(asset_id=int(asset_id)).order_by(DepreciationEntry.period.asc()).all()
        return jsonify({
            "status": "success",
            "schedule": [e.to_dict() for e in entries]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/assets/dispose")
def api_assets_dispose():
    """US-493: Dispose of an asset and record salvage gain/loss."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    body = request.get_json(silent=True) or {}
    asset_id = body.get("asset_id")
    disposed_date = body.get("disposed_date") or date.today().isoformat()
    proceeds = float(body.get("disposal_proceeds", 0.0))

    if not asset_id:
        return jsonify({"error": "Missing asset_id"}), 400

    from invoices.v37_service import FixedAssetDepreciationEngine
    try:
        res = FixedAssetDepreciationEngine.dispose_asset(int(asset_id), disposed_date, proceeds)
        if "error" in res:
            return jsonify(res), 404
        return jsonify({
            "status": "success",
            "message": "Asset disposed successfully.",
            "result": res
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/assets/auto-detect")
def api_assets_auto_detect():
    """US-494: Auto-detect fixed asset candidates from purchase invoices."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    mst = request.args.get("mst") or session.get("taxpayer_mst") or "0102030405"
    from invoices.v37_service import AIInvoiceAssetLinker
    try:
        candidates = AIInvoiceAssetLinker.auto_detect_fixed_assets(mst)
        return jsonify({
            "status": "success",
            "candidates": candidates
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/assets/validate")
def api_assets_validate():
    """US-494: Check asset depreciation compliance against TT45 limits."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    asset_id = request.args.get("asset_id")
    if not asset_id:
        return jsonify({"error": "Missing asset_id"}), 400

    from invoices.models import FixedAsset
    from invoices.v37_service import AIInvoiceAssetLinker
    try:
        asset = db.session.get(FixedAsset, int(asset_id))
        if not asset:
            return jsonify({"error": "Asset not found"}), 404

        val_res = AIInvoiceAssetLinker.validate_depreciation_compliance(asset)
        return jsonify({
            "status": "success",
            "validation": val_res
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/group-fund")
def api_create_group_fund():
    """Create a new group fund (PRD-FUND-E1-S1)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import GroupFund, TenantGroup
    import json

    data = request.get_json() or {}
    group_id = data.get("group_id")
    name = data.get("name")
    currency = data.get("currency", "VND")

    if not name or not name.strip():
        return jsonify({"error": "Tên quỹ không được để trống."}), 400

    username = session.get("username", "admin")
    if not group_id:
        # Resolve group
        group = TenantGroup.query.filter_by(admin_username=username).first()
        if not group and username == "admin":
            group = TenantGroup.query.first()
        if not group:
            return jsonify({"error": "Không tìm thấy nhóm tương ứng để tạo quỹ."}), 400
        group_id = group.id

    # Check if a fund already exists for this group
    existing_fund = GroupFund.query.filter_by(group_id=group_id).first()
    if existing_fund:
        return jsonify({
            "status": "success",
            "message": "Nhóm đã có quỹ.",
            "fund": existing_fund.to_dict()
        })

    try:
        new_fund = GroupFund(
            group_id=group_id,
            name=name.strip(),
            currency=currency,
            created_at=datetime.utcnow().isoformat() + "Z"
        )
        db.session.add(new_fund)
        db.session.commit()
        return jsonify({
            "status": "success",
            "message": "Tạo quỹ nhóm thành công.",
            "fund": new_fund.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/group-fund")
def api_get_group_fund():
    """Fetch details and current balance of a group fund (PRD-FUND-E3-S1)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import GroupFund, FundTransaction, TenantGroup
    from sqlalchemy import func

    group_id = request.args.get("group_id")
    username = session.get("username", "admin")

    if not group_id:
        # Resolve default group
        group = TenantGroup.query.filter_by(admin_username=username).first()
        if not group and username == "admin":
            group = TenantGroup.query.first()
        if not group:
            return jsonify({"error": "Không tìm thấy nhóm tương ứng."}), 404
        group_id = group.id

    fund = GroupFund.query.filter_by(group_id=group_id).first()
    if not fund:
        return jsonify({
            "fund_exists": False,
            "message": "Nhóm chưa có quỹ."
        }), 200

    # Calculate balance
    # Total deposits
    deposit_sum = db.session.query(func.sum(FundTransaction.amount)).filter(
        FundTransaction.fund_id == fund.id,
        FundTransaction.transaction_type == "deposit"
    ).scalar() or 0.0

    # Total expenses
    expense_sum = db.session.query(func.sum(FundTransaction.amount)).filter(
        FundTransaction.fund_id == fund.id,
        FundTransaction.transaction_type == "expense"
    ).scalar() or 0.0

    balance = deposit_sum - expense_sum

    fund_dict = fund.to_dict()
    fund_dict["balance"] = balance
    fund_dict["total_deposits"] = deposit_sum
    fund_dict["total_expenses"] = expense_sum
    fund_dict["fund_exists"] = True

    return jsonify(fund_dict), 200

@invoices_blueprint.post("/api/group-fund/deposit")
def api_log_deposit():
    """Record a deposit transaction to a fund (PRD-FUND-E2-S1)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import GroupFund, FundTransaction

    data = request.get_json() or {}
    fund_id = data.get("fund_id")
    payer = data.get("payer")
    amount_raw = data.get("amount")
    date = data.get("date")

    if not fund_id:
        return jsonify({"error": "Thiếu fund_id."}), 400
    if not payer or not payer.strip():
        return jsonify({"error": "Tên người nộp không được để trống."}), 400
    if not date or not date.strip():
        return jsonify({"error": "Ngày nộp không được để trống."}), 400

    try:
        amount = float(amount_raw)
    except (ValueError, TypeError):
        return jsonify({"error": "Số tiền nộp không hợp lệ hoặc để trống."}), 400

    if amount <= 0:
        return jsonify({"error": "Số tiền nộp phải lớn hơn 0."}), 400

    fund = GroupFund.query.get(fund_id)
    if not fund:
        return jsonify({"error": "Quỹ không tồn tại."}), 404

    try:
        tx = FundTransaction(
            fund_id=fund.id,
            transaction_type="deposit",
            payer=payer.strip(),
            amount=amount,
            date=date.strip(),
            created_at=datetime.utcnow().isoformat() + "Z"
        )
        db.session.add(tx)
        db.session.commit()
        return jsonify({
            "status": "success",
            "message": "Ghi nhận khoản nộp thành công.",
            "transaction": tx.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.post("/api/group-fund/expense")
def api_log_expense():
    """Record an expense transaction from a fund (PRD-FUND-E2-S2)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import GroupFund, FundTransaction

    data = request.get_json() or {}
    fund_id = data.get("fund_id")
    description = data.get("description")
    amount_raw = data.get("amount")
    date = data.get("date")

    if not fund_id:
        return jsonify({"error": "Thiếu fund_id."}), 400
    if not description or not description.strip():
        return jsonify({"error": "Mô tả khoản chi không được để trống."}), 400
    if not date or not date.strip():
        return jsonify({"error": "Ngày chi không được để trống."}), 400

    try:
        amount = float(amount_raw)
    except (ValueError, TypeError):
        return jsonify({"error": "Số tiền chi không hợp lệ hoặc để trống."}), 400

    if amount <= 0:
        return jsonify({"error": "Số tiền chi phải lớn hơn 0."}), 400

    fund = GroupFund.query.get(fund_id)
    if not fund:
        return jsonify({"error": "Quỹ không tồn tại."}), 404

    try:
        tx = FundTransaction(
            fund_id=fund.id,
            transaction_type="expense",
            description=description.strip(),
            amount=amount,
            date=date.strip(),
            created_at=datetime.utcnow().isoformat() + "Z"
        )
        db.session.add(tx)
        db.session.commit()
        return jsonify({
            "status": "success",
            "message": "Ghi nhận khoản chi thành công.",
            "transaction": tx.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/api/group-fund/transactions")
def api_get_transactions():
    """Retrieve transaction history of a group fund (PRD-FUND-E3-S2)."""
    unauthorized = _ensure_logged_in()
    if unauthorized:
        return unauthorized

    from invoices.models import GroupFund, FundTransaction

    fund_id = request.args.get("fund_id")
    if not fund_id:
        return jsonify({"error": "Thiếu fund_id."}), 400

    fund = GroupFund.query.get(fund_id)
    if not fund:
        return jsonify({"error": "Quỹ không tồn tại."}), 404

    try:
        # Sort by date descending, then by id descending
        txs = FundTransaction.query.filter_by(fund_id=fund.id).order_by(
            FundTransaction.date.desc(), FundTransaction.id.desc()
        ).all()
        return jsonify([t.to_dict() for t in txs]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@invoices_blueprint.get("/group-fund")
def group_fund_page():
    """Render the Group Fund (Sổ Quỹ) UI page."""
    if not session.get("logged_in"):
        return redirect(url_for("auth.login_page"))
    return render_template("fund.html")
