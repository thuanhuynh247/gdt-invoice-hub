from __future__ import annotations
from io import BytesIO
from flask import Blueprint, Response, current_app, jsonify, redirect, render_template, request, send_file, session, url_for
from export.excel import generate_excel_workbook, generate_local_excel_workbook
from invoices.parser import DateValidationError, validate_date_range
from invoices.service import (
    GDTIntegrationNotReadyError,
    InvoiceQuery,
    build_invoice_lookup,
    download_invoice_xml,
    fetch_invoices,
    resolve_live_download_name,
    fetch_invoice_line_items,
    extract_partners_from_invoices,
    generate_tax_usage_report,
)
from extensions import db
from auth.decorators import roles_required
import os
import uuid
import threading
from datetime import datetime
from flask import send_file
import io

"""Invoice-facing routes for search, download and export."""

def _ensure_logged_in():
    """Return a 401 JSON response when the session is missing."""

    if not session.get("logged_in"):
        return jsonify({"error": "Phien dang nhap da het han. Vui long dang nhap lai."}), 401
    return None

def render_html_to_pdf(html_content: str) -> io.BytesIO:
    """Helper to compile HTML content to PDF using xhtml2pdf with pre-registered Vietnamese fonts."""
    # Find a suitable Vietnamese TrueType Font
    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    font_path = None
    for path in candidates:
        if os.path.exists(path):
            font_path = path
            break
            
    if font_path:
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            pdfmetrics.registerFont(TTFont('Arial', font_path))
            
            # Also attempt to register Bold font
            bold_path = font_path.replace("arial.ttf", "arialbd.ttf").replace("tahoma.ttf", "tahomabd.ttf").replace("Arial.ttf", "Arial Bold.ttf")
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont('Arial-Bold', bold_path))
            else:
                pdfmetrics.registerFont(TTFont('Arial-Bold', font_path))
        except Exception:
            pass

    pdf_buffer = io.BytesIO()
    from xhtml2pdf import pisa
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
    if pisa_status.err:
        raise RuntimeError("xhtml2pdf rendering failed.")
    pdf_buffer.seek(0)
    return pdf_buffer

def get_supplier_pivot_data(mst, year_filter, value_type):
    from extensions import db
    from invoices.models import Invoice
    from sqlalchemy import func

    # Query input invoices (invoice_type = 'purchase')
    query = db.session.query(
        Invoice.seller_mst,
        Invoice.seller_name,
        func.substr(Invoice.date, 1, 7).label("month_str"),
        func.count(Invoice.id).label("count"),
        func.sum(Invoice.amount_before_tax).label("amount_before_tax"),
        func.sum(Invoice.tax_amount).label("tax_amount"),
        func.sum(Invoice.total_amount).label("total_amount")
    ).filter(
        Invoice.invoice_type == 'purchase'
    )

    if mst and mst != "all":
        query = query.filter(Invoice.taxpayer_mst == mst)

    if year_filter:
        query = query.filter(Invoice.date.like(f"{year_filter}-%"))

    results = query.group_by(
        Invoice.seller_mst,
        Invoice.seller_name,
        func.substr(Invoice.date, 1, 7)
    ).all()

    # Build pivot structure
    if year_filter:
        months_list = [f"{i:02d}" for i in range(1, 13)]
    else:
        months_set = set()
        for r in results:
            if r.month_str and len(r.month_str) == 7:
                months_set.add(r.month_str)
        months_list = sorted(list(months_set))

    sellers_map = {}
    for r in results:
        seller_mst = r.seller_mst or "UNKNOWN"
        seller_name = r.seller_name or "Không rõ"
        month_key = r.month_str
        if year_filter and month_key:
            month_key = month_key.split("-")[1]

        val = 0.0
        if value_type == "amount_before_tax":
            val = r.amount_before_tax or 0.0
        elif value_type == "tax_amount":
            val = r.tax_amount or 0.0
        elif value_type == "invoice_count":
            val = r.count or 0
        else:
            val = r.total_amount or 0.0

        if seller_mst not in sellers_map:
            sellers_map[seller_mst] = {
                "seller_mst": seller_mst,
                "seller_name": seller_name,
                "monthly_values": {m: 0.0 for m in months_list},
                "row_total": 0.0
            }
        
        if month_key in sellers_map[seller_mst]["monthly_values"]:
            sellers_map[seller_mst]["monthly_values"][month_key] = val
            sellers_map[seller_mst]["row_total"] += val

    rows = list(sellers_map.values())
    rows.sort(key=lambda x: x["row_total"], reverse=True)

    column_totals = {m: 0.0 for m in months_list}
    grand_total = 0.0

    for r in rows:
        for m in months_list:
            val = r["monthly_values"].get(m, 0.0)
            column_totals[m] += val
            grand_total += val

    return {
        "year": year_filter or "Tất cả",
        "value_type": value_type,
        "months": months_list,
        "rows": rows,
        "column_totals": column_totals,
        "grand_total": grand_total
    }

_AGING_BUCKETS = [
    ("1–30 ngày",   1,  30),
    ("31–60 ngày",  31, 60),
    ("61–90 ngày",  61, 90),
    (">90 ngày",    91, None),
]

def classify_fct_item(item_name: str, seller_name: str) -> tuple[str, float, float]:
    """
    Classify transaction based on Circular 103/2014/TT-BTC for e-commerce and digital services.
    Returns: (Category, VAT_rate, CIT_rate)
    """
    name = (item_name or "").lower() + " " + (seller_name or "").lower()
    
    # 1. Digital Advertising (e.g. Google Ads, Meta Ads)
    if any(k in name for k in ["ads", "quảng cáo", "quang cao", "marketing", "facebook ads", "google ads", "adwords"]):
        return "Dịch vụ Quảng cáo trực tuyến (Online Advertising)", 0.05, 0.05
        
    # 2. Cloud & Hosting (e.g. AWS, Azure)
    if any(k in name for k in ["cloud", "hosting", "aws", "amazon web services", "azure", "server", "vps", "lưu trữ", "digitalocean"]):
        return "Dịch vụ Điện toán đám mây & Lưu trữ (Cloud & Hosting)", 0.05, 0.05
        
    # 3. Software License / SaaS (VAT Exempt under VN Tax Law, CIT 5%)
    if any(k in name for k in ["phần mềm", "phan mem", "software", "license", "bản quyền", "ban quyen", "zoom", "slack", "subscription"]):
        return "Bản quyền & Phần mềm SaaS (Software & License)", 0.00, 0.05
        
    # Default fallback (General Digital Services)
    return "Thương mại điện tử & Dịch vụ số khác", 0.05, 0.05

def generate_fct_excel(fct_data: dict) -> bytes:
    """Generate a highly polished Excel sheet conforming to Mẫu số 01/NTNN layout."""
    import openpyxl
    from openpyxl import Workbook
    from export.formatter import auto_adjust_column_widths, HEADER_FILL, MED_RISK_FILL, OK_FILL

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Tờ khai 01-NTNN"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.append(["TỜ KHAI THUẾ NHÀ THẦU NƯỚC NGOÀI (Mẫu số 01/NTNN)"])
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.font = openpyxl.styles.Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Period Info
    period_str = f"Tháng {fct_data['period_value']}" if fct_data['period_type'] == "monthly" else f"Quý {fct_data['period_value']}"
    ws.append([f"Kỳ kê khai: {period_str} năm {fct_data['year']}", "", "", "", "", "", "", "", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    ws.merge_cells("A2:H2")
    ws["A2"].font = openpyxl.styles.Font(italic=True, color="555555")
    ws["I2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws.append([])  # Blank row

    # Headers
    headers = [
        "STT",
        "Tên Nhà thầu nước ngoài",
        "Mã số thuế",
        "Nội dung dịch vụ",
        "Doanh thu tính thuế (₫)",
        "Tỷ lệ GTGT (%)",
        "Thuế GTGT phải nộp (₫)",
        "Tỷ lệ TNDN (%)",
        "Thuế TNDN phải nộp (₫)",
    ]
    ws.append(headers)
    
    # Format Headers
    for col_idx in range(1, 10):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Data Rows
    invoices = fct_data.get("fct_invoices", [])
    for idx, inv in enumerate(invoices, 1):
        row = [
            idx,
            inv["seller_name"],
            inv["seller_mst"],
            inv["category"],
            inv["amount"],
            f"{inv['vat_rate'] * 100}%",
            inv["vat_withheld"],
            f"{inv['cit_rate'] * 100}%",
            inv["cit_withheld"]
        ]
        ws.append(row)
        
        # Zebra styling
        row_cells = ws[ws.max_row]
        for cell in row_cells:
            cell.fill = OK_FILL if idx % 2 == 0 else openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.font = openpyxl.styles.Font(size=10)

    # Total Row
    ws.append([
        "TỔNG CỘNG", "", "", "",
        fct_data["total_revenue"],
        "",
        fct_data["total_vat_withheld"],
        "",
        fct_data["total_cit_withheld"]
    ])
    ws.merge_cells(f"A{ws.max_row}:D{ws.max_row}")
    
    for col_idx in range(1, 10):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True, size=11, color="1F4E78")
        cell.fill = MED_RISK_FILL
    ws.row_dimensions[ws.max_row].height = 24

    # Alignments & Formats
    for row in range(5, ws.max_row + 1):
        ws[f"A{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"C{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"F{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"H{row}"].alignment = openpyxl.styles.Alignment(horizontal="center")
        
        for col_idx in [5, 7, 9]:
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '#,##0 "VND"'
            cell.alignment = openpyxl.styles.Alignment(horizontal="right")

    auto_adjust_column_widths(ws)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def require_api_signature(f):
    from functools import wraps
    import time
    @wraps(f)
    def decorated(*args, **kwargs):
        signature = request.headers.get("X-GDT-Signature")
        timestamp_str = request.headers.get("X-GDT-Timestamp")
        
        if not signature or not timestamp_str:
            return jsonify({"error": "Missing signature or timestamp headers"}), 401
            
        try:
            timestamp = int(timestamp_str)
            if abs(time.time() - timestamp) > 300:
                return jsonify({"error": "Signature timestamp expired or invalid"}), 401
        except Exception:
            return jsonify({"error": "Invalid timestamp format"}), 401
            
        if request.method == "GET":
            payload_str = request.query_string.decode("utf-8")
        else:
            payload_str = request.get_data(as_text=True)
            
        secret = current_app.config.get("SECRET_KEY", "super-secret-key")
        
        import hmac
        import hashlib
        message = f"{timestamp_str}.{payload_str}".encode("utf-8")
        computed = hmac.new(secret.encode("utf-8"), message, hashlib.sha256).hexdigest()
        expected = f"sha256={computed}"
        
        if not hmac.compare_digest(signature, expected):
            return jsonify({"error": "Invalid signature verification failed"}), 401
            
        return f(*args, **kwargs)
    return decorated

def get_harness_db():
    import sqlite3
    import os
    
    conn = sqlite3.connect("harness.db", timeout=10.0)
    conn.text_factory = lambda x: str(x, encoding="utf-8", errors="replace")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA journal_mode = WAL;")
    
    # Auto-initialize tables if story table doesn't exist
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='story';")
    if not cur.fetchone():
        conn.execute("""
        CREATE TABLE IF NOT EXISTS story (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            risk_lane TEXT NOT NULL,
            contract_doc TEXT,
            status TEXT NOT NULL,
            notes TEXT,
            unit_proof INTEGER DEFAULT 0,
            integration_proof INTEGER DEFAULT 0,
            e2e_proof INTEGER DEFAULT 0,
            platform_proof INTEGER DEFAULT 0,
            evidence TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS decision (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            status TEXT NOT NULL,
            doc_path TEXT,
            verify_command TEXT,
            last_verified_at TIMESTAMP,
            last_verified_result TEXT,
            predicted_impact TEXT,
            actual_outcome TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS backlog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            discovered_while TEXT,
            current_pain TEXT,
            suggested_improvement TEXT,
            risk TEXT,
            status TEXT NOT NULL,
            predicted_impact TEXT,
            actual_outcome TEXT,
            implemented_at TIMESTAMP,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.execute("""
        CREATE TABLE IF NOT EXISTS trace (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_summary TEXT NOT NULL,
            intake_id TEXT,
            story_id TEXT,
            agent TEXT,
            actions_taken TEXT,
            files_read TEXT,
            files_changed TEXT,
            decisions_made TEXT,
            errors TEXT,
            outcome TEXT,
            duration_seconds INTEGER,
            token_estimate INTEGER,
            harness_friction TEXT,
            notes TEXT,
            git_hash TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
    return conn
