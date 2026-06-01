"""Service for generating and verifying cryptographically signed compliance reports (US-145)."""

from __future__ import annotations

import hashlib
import json
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Curated, premium soft colors for Excel styling
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FOOTER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
DATA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def calculate_report_hash(invoices: list[dict], secret_key: str) -> str:
    """
    Calculate a SHA-256 digest of the exported invoices list.
    Concatenates sorted records (invoice ID + total_amount + date) and hashes with a secret key.
    """
    # Sort by invoice ID to guarantee deterministic ordering
    sorted_invoices = sorted(invoices, key=lambda x: str(x.get("id", "")))
    
    parts = []
    for inv in sorted_invoices:
        inv_id = str(inv.get("id", ""))
        # Format total_amount to 2 decimal places to prevent float formatting differences
        try:
            total_amount = float(inv.get("total_amount", 0.0))
        except (TypeError, ValueError):
            total_amount = 0.0
            
        inv_date = str(inv.get("date", ""))
        parts.append(f"{inv_id}:{total_amount:.2f}:{inv_date}")
        
    raw_string = "|".join(parts)
    payload = f"{raw_string}::{secret_key}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()

def generate_signed_excel_report(invoices: list[dict], secret_key: str) -> bytes:
    """
    Generate an Excel workbook representing the audited compliance ledger.
    Appends a SHA-256 integrity signature block at the footer.
    """
    # Calculate the SHA-256 signature first
    signature = calculate_report_hash(invoices, secret_key)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Compliance Ledger"
    worksheet.views.sheetView[0].showGridLines = True
    
    # Title Block
    worksheet.append(["BÁO CÁO TUÂN THỦ THUẾ & NHẬT KÝ KIỂM TOÁN TÍCH HỢP CHỮ KÝ SỐ"])
    worksheet.merge_cells("A1:G1")
    title_cell = worksheet["A1"]
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 35
    
    # Metadata Row
    worksheet.append([
        "Ngày xuất bản:", 
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"), 
        "", 
        "Hệ thống:", 
        "GDT Invoice Hub", 
        "", 
        "Trạng thái: ĐÃ XÁC MINH"
    ])
    worksheet.merge_cells("B2:C2")
    worksheet.merge_cells("E2:F2")
    worksheet["A2"].font = Font(italic=True, color="555555")
    worksheet["G2"].font = Font(bold=True, color="008000")
    worksheet.append([])  # Spacer row
    
    # Table Headers
    headers = [
        "Mã Hóa Đơn (ID)", 
        "Ký Hiệu", 
        "Số Hóa Đơn", 
        "Ngày Lập", 
        "Tên Đối Tác", 
        "Tổng Cộng (VND)", 
        "Đánh Giá"
    ]
    worksheet.append(headers)
    
    # Style Header Row
    header_row_idx = 4
    for col_idx in range(1, 8):
        cell = worksheet.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[header_row_idx].height = 25
    
    # Write Invoice Data
    for inv in invoices:
        try:
            total_amount = float(inv.get("total_amount", 0.0))
        except (TypeError, ValueError):
            total_amount = 0.0
            
        row = [
            inv.get("id", ""),
            inv.get("symbol", ""),
            inv.get("number", ""),
            inv.get("date", ""),
            inv.get("seller_name", "") if inv.get("direction") == "purchase" else inv.get("buyer_name", ""),
            total_amount,
            "Đạt yêu cầu" if inv.get("is_valid", True) else "Cần rà soát"
        ]
        worksheet.append(row)
        
        # Style amount column format
        worksheet.cell(row=worksheet.max_row, column=6).number_format = '#,##0 "VND"'
        
    # Spacer row before signature
    worksheet.append([])
    
    # Append the Cryptographic Integrity Signature Block
    sig_row_idx = worksheet.max_row + 1
    worksheet.append(["MÃ HÓA TOÀN VẸN DỮ LIỆU (SHA-256 INTEGRITY BLOCK)"])
    worksheet.merge_cells(f"A{sig_row_idx}:G{sig_row_idx}")
    label_cell = worksheet[f"A{sig_row_idx}"]
    label_cell.font = Font(bold=True, size=10, color="1F4E78")
    label_cell.alignment = Alignment(horizontal="center")
    
    hash_row_idx = sig_row_idx + 1
    worksheet.append([signature])
    worksheet.merge_cells(f"A{hash_row_idx}:G{hash_row_idx}")
    hash_cell = worksheet[f"A{hash_row_idx}"]
    hash_cell.font = Font(name="Courier New", size=11, bold=True, color="FF0000")
    hash_cell.fill = FOOTER_FILL
    hash_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[hash_row_idx].height = 30
    
    from openpyxl.utils import get_column_letter
    
    # Auto-adjust columns widths
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title and signature rows for auto-width estimation
            if cell.row in (1, sig_row_idx, hash_row_idx):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def verify_excel_report(file_bytes: bytes, secret_key: str) -> dict:
    """
    Read an uploaded signed Excel report, extract the invoices, 
    re-calculate the SHA-256 digest, and compare with the footer signature.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        
        # Read invoices and signature block
        invoices = []
        signature_found = None
        
        # Find header row index
        header_row_idx = 4 # Default standard
        
        # Scan worksheet row by row to locate data and signature
        max_row = worksheet.max_row
        data_rows = []
        
        # Scan from row 5 (after header) to max_row
        for r_idx in range(5, max_row + 1):
            cell_a = worksheet.cell(row=r_idx, column=1).value
            
            # Check if this row is the start of the signature block
            if cell_a == "MÃ HÓA TOÀN VẸN DỮ LIỆU (SHA-256 INTEGRITY BLOCK)":
                # The signature string is located in the next row, column 1
                signature_found = worksheet.cell(row=r_idx + 1, column=1).value
                break
                
            # If it's a spacer row or empty, skip
            if not cell_a:
                continue
                
            # Parse invoice row data
            inv_id = str(cell_a).strip()
            symbol = str(worksheet.cell(row=r_idx, column=2).value or "").strip()
            number = str(worksheet.cell(row=r_idx, column=3).value or "").strip()
            date = str(worksheet.cell(row=r_idx, column=4).value or "").strip()
            partner_name = str(worksheet.cell(row=r_idx, column=5).value or "").strip()
            
            total_val = worksheet.cell(row=r_idx, column=6).value
            try:
                total_amount = float(total_val)
            except (TypeError, ValueError):
                total_amount = 0.0
                
            is_valid_str = str(worksheet.cell(row=r_idx, column=7).value or "").strip()
            is_valid = is_valid_str == "Đạt yêu cầu"
            
            invoices.append({
                "id": inv_id,
                "symbol": symbol,
                "number": number,
                "date": date,
                "seller_name": partner_name,
                "total_amount": total_amount,
                "is_valid": is_valid
            })
            
        if not signature_found:
            return {
                "verified": False,
                "error": "Không tìm thấy mã chữ ký số bảo vệ ở chân trang báo cáo.",
                "invoices_count": 0
            }
            
        # Clean signature if it contains extra characters
        signature_found = str(signature_found).strip()
        
        # Calculate expected signature
        expected_sig = calculate_report_hash(invoices, secret_key)
        
        # Compare
        is_verified = (signature_found == expected_sig)
        
        return {
            "verified": is_verified,
            "signature_found": signature_found,
            "signature_expected": expected_sig,
            "invoices_count": len(invoices),
            "invoices": invoices
        }
        
    except Exception as e:
        return {
            "verified": False,
            "error": f"Lỗi đọc định dạng tệp báo cáo Excel: {str(e)}",
            "invoices_count": 0
        }
