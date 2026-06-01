"""Integration tests for US-145: Signed Compliance Report Exporter and Verification."""

from __future__ import annotations

import io
import hashlib
from datetime import date
from io import BytesIO
import openpyxl
import pytest

from extensions import db
from invoices.models import Invoice, TaxpayerProfile, SecurityAuditLog
from invoices.compliance_report_service import calculate_report_hash, generate_signed_excel_report, verify_excel_report


def _seed_test_invoices(app):
    """Seed test invoices inside the database context."""
    with app.app_context():
        # Clear tables
        Invoice.query.delete()
        TaxpayerProfile.query.delete()
        SecurityAuditLog.query.delete()
        db.session.commit()

        # Seed profile
        profile = TaxpayerProfile(
            mst="0101234567",
            company_name="Valid Enterprise",
            gdt_username="valid_user",
            gdt_password_encrypted="pass_hash",
            is_active=True,
            created_at="2026-05-30T00:00:00Z"
        )
        db.session.add(profile)
        db.session.commit()

        # Seed invoices
        inv1 = Invoice(
            id="0101234567-1C26TBA-0000001",
            filename="inv1.xml",
            invoice_type="Hóa đơn GTGT",
            symbol="1C26TBA",
            number="0000001",
            date="2026-05-20",
            currency="VND",
            seller_name="Partner Corp A",
            seller_mst="0101234567",
            buyer_name="My Corp",
            buyer_mst="0202345678",
            amount_before_tax=10000000.0,
            tax_amount=1000000.0,
            total_amount=11000000.0,
            imported_at="2026-05-30T01:00:00Z",
            import_status="imported",
            taxpayer_mst="0101234567"
        )
        inv2 = Invoice(
            id="0101234567-1C26TBA-0000002",
            filename="inv2.xml",
            invoice_type="Hóa đơn GTGT",
            symbol="1C26TBA",
            number="0000002",
            date="2026-05-21",
            currency="VND",
            seller_name="Partner Corp B",
            seller_mst="0101234567",
            buyer_name="My Corp",
            buyer_mst="0202345678",
            amount_before_tax=20000000.0,
            tax_amount=2000000.0,
            total_amount=22000000.0,
            imported_at="2026-05-30T01:00:00Z",
            import_status="imported",
            taxpayer_mst="0101234567"
        )
        db.session.add(inv1)
        db.session.add(inv2)
        db.session.commit()
        db.session.remove()


class TestSignedComplianceReport:
    """Test Suite covering Signed Compliance Report Excel generation and verification (US-145)."""

    def test_calculate_report_hash_deterministic(self):
        """Verify that report hash calculation is deterministic regardless of list order."""
        secret_key = "test-secret"
        invoices = [
            {"id": "inv-A", "total_amount": 10000.0, "date": "2026-05-01"},
            {"id": "inv-B", "total_amount": 20000.0, "date": "2026-05-02"},
        ]
        
        # Order 1
        hash_1 = calculate_report_hash(invoices, secret_key)
        # Order 2 (reversed)
        hash_2 = calculate_report_hash(list(reversed(invoices)), secret_key)
        
        assert hash_1 == hash_2
        assert len(hash_1) == 64  # SHA-256 hex digest length

    def test_signed_excel_generation(self):
        """Verify that signed Excel contains expected layout and footer signature."""
        secret_key = "test-secret"
        invoices = [
            {"id": "inv-1", "symbol": "AB/26T", "number": "1", "date": "2026-05-01", "total_amount": 500000.0, "seller_name": "Vendor A", "direction": "purchase", "is_valid": True},
            {"id": "inv-2", "symbol": "AB/26T", "number": "2", "date": "2026-05-02", "total_amount": 900000.0, "seller_name": "Vendor B", "direction": "purchase", "is_valid": False},
        ]
        
        excel_bytes = generate_signed_excel_report(invoices, secret_key)
        assert len(excel_bytes) > 0
        
        # Read back using openpyxl
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
        ws = wb.active
        assert ws.title == "Compliance Ledger"
        
        # Verify title
        assert "BÁO CÁO TUÂN THỦ THUẾ" in ws["A1"].value
        
        # Scan for signature block
        found_signature = None
        for row in range(5, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val == "MÃ HÓA TOÀN VẸN DỮ LIỆU (SHA-256 INTEGRITY BLOCK)":
                found_signature = ws.cell(row=row + 1, column=1).value
                break
                
        assert found_signature is not None
        assert found_signature == calculate_report_hash(invoices, secret_key)

    def test_report_verification_valid(self):
        """Verify that a freshly generated report passes verification successfully."""
        secret_key = "test-secret"
        invoices = [
            {"id": "inv-1", "symbol": "A1", "number": "100", "date": "2026-05-01", "total_amount": 100.0, "seller_name": "Vendor A", "direction": "purchase", "is_valid": True},
        ]
        excel_bytes = generate_signed_excel_report(invoices, secret_key)
        
        result = verify_excel_report(excel_bytes, secret_key)
        assert result["verified"] is True
        assert result["invoices_count"] == 1
        assert result["invoices"][0]["id"] == "inv-1"
        assert result["invoices"][0]["total_amount"] == 100.0

    def test_report_verification_tampered(self):
        """Verify that any modification/tampering in the Excel file causes verification to fail."""
        secret_key = "test-secret"
        invoices = [
            {"id": "inv-1", "symbol": "A1", "number": "100", "date": "2026-05-01", "total_amount": 100.0, "seller_name": "Vendor A", "direction": "purchase", "is_valid": True},
        ]
        excel_bytes = generate_signed_excel_report(invoices, secret_key)
        
        # Load and tamper the total_amount cell
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # Row 5 column 6 is the total_amount for "inv-1"
        assert ws.cell(row=5, column=6).value == 100.0
        ws.cell(row=5, column=6, value=999999.0)  # Modifying total amount!
        
        tampered_output = BytesIO()
        wb.save(tampered_output)
        tampered_bytes = tampered_output.getvalue()
        
        # Verify tampered file
        result = verify_excel_report(tampered_bytes, secret_key)
        assert result["verified"] is False
        assert "signature_expected" in result
        assert result["signature_found"] != result["signature_expected"]


class TestSignedComplianceReportRoutes:
    """Test Suite covering the API endpoints for exporting and verifying signed reports."""

    def test_export_signed_compliance_route_auth(self, client):
        """Verify that export API requires authentication."""
        r = client.get("/api/reports/signed-compliance?from=2026-05-01&to=2026-05-30")
        assert r.status_code == 401

    def test_export_signed_compliance_route_admin(self, logged_in_client, app):
        """Verify that authenticated admin can successfully download the Excel report."""
        _seed_test_invoices(app)
        
        r = logged_in_client.get("/api/reports/signed-compliance?from=2026-05-01&to=2026-05-30&direction=purchase")
        assert r.status_code == 200
        assert r.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "Content-Disposition" in r.headers
        assert "signed_compliance_report_" in r.headers["Content-Disposition"]
        assert len(r.data) > 0

        # Check that SecurityAuditLog event is automatically recorded
        with app.app_context():
            logs = SecurityAuditLog.query.filter_by(event_category="EXPORT").all()
            assert len(logs) == 1
            assert "Exported cryptographically signed compliance report" in logs[0].event_details

    def test_verify_signed_report_route_success(self, logged_in_client, app):
        """Verify uploading and verifying a valid signed report via API."""
        _seed_test_invoices(app)
        
        # Generate the report bytes first
        with app.app_context():
            from invoices.service import fetch_invoices, InvoiceQuery
            invoices = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 30), False, "purchase"))
            secret_key = app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
            excel_bytes = generate_signed_excel_report(invoices, secret_key)
            
        # Call verification endpoint via client upload
        data = {
            "file": (BytesIO(excel_bytes), "signed_report.xlsx")
        }
        r = logged_in_client.post(
            "/api/reports/verify-signed",
            data=data,
            content_type="multipart/form-data"
        )
        assert r.status_code == 200
        result = r.get_json()
        assert result["verified"] is True
        assert result["invoices_count"] == len(invoices)

        # Check verification audit log recorded
        with app.app_context():
            logs = SecurityAuditLog.query.filter_by(event_category="VERIFY").all()
            assert len(logs) == 1
            assert "Result: SUCCESS" in logs[0].event_details

    def test_verify_signed_report_route_tampered(self, logged_in_client, app):
        """Verify uploading a tampered report via API returns verified=False."""
        _seed_test_invoices(app)
        
        # Generate and tamper the report bytes
        with app.app_context():
            from invoices.service import fetch_invoices, InvoiceQuery
            invoices = fetch_invoices(InvoiceQuery(date(2026, 5, 1), date(2026, 5, 30), False, "purchase"))
            secret_key = app.config.get("SECRET_KEY", "compliance-system-secret-key-12345")
            excel_bytes = generate_signed_excel_report(invoices, secret_key)
            
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # Tamper row 5 number (total payable)
        ws.cell(row=5, column=6, value=999.0)
        
        tampered_output = BytesIO()
        wb.save(tampered_output)
        tampered_bytes = tampered_output.getvalue()

        data = {
            "file": (BytesIO(tampered_bytes), "signed_report_tampered.xlsx")
        }
        r = logged_in_client.post(
            "/api/reports/verify-signed",
            data=data,
            content_type="multipart/form-data"
        )
        assert r.status_code == 200
        result = r.get_json()
        assert result["verified"] is False
        assert result["invoices_count"] == len(invoices)

        # Check verification failure log recorded
        with app.app_context():
            logs = SecurityAuditLog.query.filter_by(event_category="VERIFY").all()
            assert len(logs) == 1
            assert "Result: FAILED" in logs[0].event_details
