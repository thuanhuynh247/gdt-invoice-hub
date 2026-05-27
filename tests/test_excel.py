"""Excel export tests."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from export.excel import generate_excel_workbook


def test_generate_excel_workbook_has_headers():
    """Workbook should always include the expected header row."""

    workbook_bytes = generate_excel_workbook([])
    workbook = load_workbook(BytesIO(workbook_bytes))
    headers = [cell.value for cell in workbook.active[1]]
    assert headers[:5] == ["ID", "Date", "Amount", "Status", "Issuer"]


def test_generate_excel_workbook_contains_rows():
    """Workbook should include invoice data rows."""

    workbook_bytes = generate_excel_workbook(
        [
            {
                "id": "INV-1",
                "date": "2026-05-01",
                "amount": 100000,
                "status": "valid",
                "issuer": "Cong ty Test",
                "description": "Demo",
                "is_cancelled": False,
                "cancellation_date": None,
                "cancellation_reason": None,
            }
        ]
    )
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.active["A2"].value == "INV-1"
    assert workbook.active["B2"].value == "01/05/2026"
